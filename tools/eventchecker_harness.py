#!/usr/bin/env python3
"""Lightweight harness for catching repeat regressions in EventInspector.

Run this before pushing changes:

    python3 tools/eventchecker_harness.py

The harness intentionally focuses on the brittle parts of the app:
- update manifest shape
- exact-match parsing contracts
- installation-id state transitions
- package-code mapping
- release payload/source sync for every build target

It does not start the UI or connect to devices.
"""

from __future__ import annotations

import argparse
import copy
import functools
import http.server
import importlib.util
import json
import os
import re
import hashlib
import subprocess
import sys
import tempfile
import threading
import types
from dataclasses import dataclass
from pathlib import Path, PosixPath
from typing import Callable, Iterable, List
from urllib.parse import urlparse

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
CURRENT_RELEASE_VERSION = "2026-09-04-1-2.5.0-56"
CURRENT_RELEASE_BUILD = 56
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import Log_checker as lc  # noqa: E402
# desktop_app only needs webview after main() starts; keep the harness usable
# in a plain Python environment that has not installed the UI dependency.
sys.modules.setdefault("webview", types.ModuleType("webview"))
import desktop_app as desktop  # noqa: E402


@dataclass
class HarnessResult:
    name: str
    passed: bool
    message: str = ""


def _reset_runtime_state() -> None:
    lc.active_platform = "android"
    lc.connected_devices_info = []
    lc.installation_id_state.clear()
    lc.sdk_check_runtime_state.clear()
    lc.sdk_check_expected_map.clear()
    lc.sdk_check_expected_order.clear()
    lc.sdk_check_search_list.clear()
    lc.sdk_check_current_network.clear()
    lc.is_paused = False


def _assert(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def _assert_equal(actual, expected, message: str) -> None:
    if actual != expected:
        raise AssertionError(f"{message}: expected={expected!r} actual={actual!r}")


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _payload_path_candidates(manifest_path: Path, item: dict) -> List[Path]:
    rel_path = str(item.get("source_path") or item.get("path", "")).strip()
    if not rel_path:
        return []
    payload_dir = manifest_path.parent
    return [ROOT / rel_path, payload_dir / rel_path]


def _valid_payload_urls(manifest_path: Path, rel_path: str, payload_path: Path) -> set[str]:
    repo_rel = payload_path.relative_to(ROOT).as_posix()
    branch = "main"
    return {
        f"https://github.com/trucbm/Eventchecker/raw/{branch}/{repo_rel}",
        f"https://raw.githubusercontent.com/trucbm/Eventchecker/{branch}/{repo_rel}",
        f"https://cdn.jsdelivr.net/gh/trucbm/Eventchecker@{branch}/{repo_rel}",
    }


def _manifest_paths() -> List[Path]:
    candidates = [ROOT / "Updates_2_5" / "remote_manifest.json"]
    return [path for path in candidates if path.exists()]


def test_manifest_contract() -> None:
    manifests = _manifest_paths()
    _assert(manifests, "No release manifest files found")
    for manifest_path in manifests:
        data = json.loads(manifest_path.read_text(encoding="utf-8"))
        _assert("version" in data and str(data["version"]).strip(), f"{manifest_path.name} missing version")
        files = data.get("files") or []
        _assert(files, f"{manifest_path.name} missing files list")
        log_checker_files = [item for item in files if str(item.get("path", "")).strip() == "Log_checker.py"]
        _assert(log_checker_files, f"{manifest_path.name} missing Log_checker.py payload")
        for item in files:
            _assert(str(item.get("path", "")).strip(), f"{manifest_path.name} contains file entry without path")
            _assert(str(item.get("url", "")).strip(), f"{manifest_path.name} contains file entry without url")
            _assert(str(item.get("sha256", "")).strip(), f"{manifest_path.name} contains file entry without sha256")


def test_manifest_payload_integrity() -> None:
    manifests = _manifest_paths()
    _assert(manifests, "No release manifest files found")
    for manifest_path in manifests:
        data = json.loads(manifest_path.read_text(encoding="utf-8"))
        for item in data.get("files") or []:
            rel_path = str(item.get("path", "")).strip()
            if not rel_path:
                continue
            candidates = _payload_path_candidates(manifest_path, item)
            payload_path = next((candidate for candidate in candidates if candidate.exists()), None)
            _assert(payload_path is not None, f"{manifest_path.name} payload missing for {rel_path}")

            expected_sha = str(item.get("sha256", "")).strip().lower()
            actual_sha = _sha256_file(payload_path).lower()
            _assert_equal(actual_sha, expected_sha, f"{manifest_path.name} sha mismatch for {rel_path}")

            valid_urls = _valid_payload_urls(manifest_path, rel_path, payload_path)
            url = str(item.get("url", "")).strip()
            _assert(
                url in valid_urls,
                f"{manifest_path.name} url points outside {manifest_path.parent.name} for {rel_path}: {url}",
            )
            for extra_url in item.get("urls") or []:
                extra_url = str(extra_url).strip()
                _assert(
                    extra_url in valid_urls,
                    f"{manifest_path.name} urls entry points outside {manifest_path.parent.name} for {rel_path}: {extra_url}",
                )


def test_canonical_update_channel_contract() -> None:
    """Keep every current release check on the single main/v250 channel."""
    import remote_update as updater

    _assert_equal(updater.CHANNEL_ID, "v250", "current updater channel changed")
    _assert_equal(tuple(updater.KNOWN_CHANNELS), ("v250",), "current updater must have one channel")
    _assert_equal(updater.CONFIG_FILENAME, "remote_update_config_v250.json", "current updater config changed")
    _assert_equal(updater.STATE_FILENAME, "update_state_v250.json", "current updater state changed")
    _assert_equal(updater.UPDATES_DIRNAME, "updates_v250", "current updater payload directory changed")

    config_path = ROOT / "remote_update_config_v250.json"
    config = json.loads(config_path.read_text(encoding="utf-8"))
    _assert(
        config.get("manifest_url") in updater.DEFAULT_MANIFEST_URLS,
        "current manifest URL must use the canonical main/v250 URL set",
    )
    _assert_equal(
        set(config.get("manifest_urls") or []),
        set(updater.DEFAULT_MANIFEST_URLS),
        "current manifest URL fallbacks changed",
    )
    _assert_equal(config.get("min_interval_sec"), 0, "current updater must check without a stale throttle")

    manifest_path = ROOT / "Updates_2_5" / "remote_manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    for item in manifest.get("files") or []:
        urls = [item.get("url"), *(item.get("urls") or [])]
        _assert(
            all("/main/" in str(url) or "@main/" in str(url) for url in urls if url),
            f"current release payload must use main: {item.get('path')}",
        )


def test_no_pre_v25_release_artifacts() -> None:
    """The checked-out release surface must contain only the v2.5 channel."""
    forbidden_tokens = (
        "2" + ".3.0",
        "2" + ".4.0",
        "v" + "230",
        "v" + "240",
        "Updates_2_" + "3",
        "Updates_2_" + "4",
        "updates_v" + "230",
        "updates_v" + "240",
        "remote_update_config_v" + "230",
        "remote_update_config_v" + "240",
        "branch-2." + "3.0",
        "branch-2." + "4.0",
        "checkpoint-v2." + "3.0",
        "checkpoint-v2." + "4.0",
    )
    forbidden_paths = (
        ROOT / "Updates",
        ROOT / ("Updates_2_" + "1"),
        ROOT / ("Updates_2_" + "2"),
        ROOT / ("Updates_2_" + "3"),
        ROOT / ("Updates_2_" + "4"),
        ROOT / "remote_update_config.json",
        ROOT / ("remote_update_config_v" + "210.json"),
        ROOT / ("remote_update_config_v" + "220.json"),
        ROOT / ("remote_update_config_v" + "230.json"),
    )
    _assert(not any(path.exists() for path in forbidden_paths), "pre-v2.5 release files/directories are still present")

    ignored_parts = {".git", ".venv", "dist", "__pycache__"}
    for path in ROOT.rglob("*"):
        if not path.is_file() or path.suffix == ".pyc" or any(part in ignored_parts for part in path.parts):
            continue
        try:
            source = path.read_text(encoding="utf-8", errors="ignore")
        except OSError:
            continue
        for token in forbidden_tokens:
            _assert(token not in source, f"pre-v2.5 reference {token!r} remains in {path.relative_to(ROOT)}")


def test_package_code_mapping() -> None:
    expected = {
        "com.indiez.nonogram": "NG",
        "com.indiez.train.miner": "TM",
        "com.indiez.idletycoon.horse.racing": "HR",
        "com.indiez.solitaire.word.card.puzzle": "SW",
        "com.nostel.dot.line.puzzle": "KN",
        "com.nostel.parking.car": "CP",
        "com.afk.idle.cat.food.restaurent": "CR",
        "tap.monster.block.away": "TP",
    }
    for package_id, game_code in expected.items():
        _assert_equal(lc._game_code_for_package(package_id), game_code, f"wrong game code for {package_id}")
    _assert_equal(lc._game_code_for_package("com.example.unknown"), "Unknown", "unknown package should stay Unknown")


def test_installation_id_state_machine() -> None:
    _reset_runtime_state()
    device_id = "device-1"

    lc._update_installation_id_runtime(device_id, package_id="com.indiez.nonogram", installation_id="id-old", force_emit=True)
    _assert_equal(lc.installation_id_state[device_id]["installation_id"], "id-old", "initial installation id not stored")

    lc._update_installation_id_runtime(device_id, package_id="com.indiez.nonogram", installation_id="id-old-2")
    _assert_equal(lc.installation_id_state[device_id]["installation_id"], "id-old-2", "same package should accept the newest installation id")

    lc._update_installation_id_runtime(device_id, package_id="com.nostel.parking.car", installation_id="", force_emit=True)
    _assert_equal(lc.installation_id_state[device_id]["installation_id"], "", "package switch should clear stale installation id")
    _assert_equal(lc.installation_id_state[device_id]["game_code"], "CP", "package switch should refresh game code")

    lc._update_installation_id_runtime(device_id, package_id="com.nostel.parking.car", installation_id="id-new", force_emit=True)
    _assert_equal(lc.installation_id_state[device_id]["installation_id"], "id-new", "new package should accept a fresh installation id")


def test_installation_id_log_parsing() -> None:
    _reset_runtime_state()
    device_id = "device-2"

    original_get_package = lc._get_android_foreground_package
    try:
        lc._get_android_foreground_package = lambda _device_id: "com.indiez.nonogram"
        lc.process_installation_id_log(
            '16:31:37.056 SS S21 Unity 07-02 16:31:37.056  3425  3954 I Unity   : [Firebase] FirebaseInstallationIdPostInitHandler->_DebugInstallationId: {"idTask":{"idTask.IsFaulted":false,"idTask.Result":"d_vP_hhkSc2Z2c1b215kBR"}}',
            device_id,
        )
        _assert_equal(lc.installation_id_state[device_id]["installation_id"], "d_vP_hhkSc2Z2c1b215kBR", "json installation id log not parsed")

        lc._get_android_foreground_package = lambda _device_id: "com.nostel.parking.car"
        lc.process_installation_id_log("18:11:37.095\tUnity\tInstallations id fg88SZWfT6e_L4CxDX04Gp", device_id)
        _assert_equal(lc.installation_id_state[device_id]["installation_id"], "fg88SZWfT6e_L4CxDX04Gp", "plain installation id log not parsed")
        _assert_equal(lc.installation_id_state[device_id]["game_code"], "CP", "installation id log should follow active package")
    finally:
        lc._get_android_foreground_package = original_get_package


def test_sdk_exact_contracts() -> None:
    _reset_runtime_state()
    android_line = "Initializing Firebase Crashlytics 20.0.5"
    ios_line = "[Firebase/Crashlytics] Version 12.12.0"
    _assert("Initializing Firebase Crashlytics 20.0.5" in android_line, "android crashlytics contract changed")
    _assert("[Firebase/Crashlytics] Version 12.12.0" in ios_line, "ios crashlytics contract changed")
    _assert_equal(lc._extract_sdk_comparable_version(android_line), "20.0.5", "android crashlytics version should stay exact")
    _assert_equal(lc._extract_sdk_comparable_version(ios_line), "12.12.0", "ios crashlytics version should stay exact")


def test_cloudx_sdk_adapter_metadata() -> None:
    original_platform = lc.active_platform
    original_devices = lc.connected_devices_info
    original_sdk_active = lc.sdk_check_active
    original_search_list = list(lc.sdk_check_search_list)
    try:
        lc.active_platform = "android"
        lc.connected_devices_info = []
        lc.sdk_check_active = True
        lc.sdk_check_expected_map = {}
        lc.sdk_check_expected_order = []
        lc.sdk_check_runtime_state = {}
        lc.sdk_check_current_network = {}

        labels = [
            ("Digital Turbine (fyber) - Cloudx", "Fyber", "1.0.0"),
            ("InMobi - Cloudx", "InMobi", "2.0.0"),
            ("Liftoff Monetization (vungle) - Cloudx", "Vungle", "3.0.0"),
            ("Meta Audience Network - Cloudx", "Meta", "4.0.0"),
            ("Mintegral - Cloudx", "Mintegral", "5.0.0"),
            ("Mobilefuse - Cloudx", "Mobilefuse", "6.0.0"),
            ("Moloco - Cloudx", "Moloco", "7.0.0"),
            ("Pangle - Cloudx", "Pangle", "8.0.0"),
            ("UnityAds - Cloudx", "Unity", "9.0.0"),
            ("Verve / Pubnative - Cloudx", "Pubnative", "10.0.0"),
            ("TaurusX - Cloudx", "TaurusX", "11.0.0"),
        ]

        cloudx_parsed = lc._parse_sdk_expected_line("CloudX\t4.5.0\t4.5.0")
        _assert(cloudx_parsed is not None, "CloudX network label not accepted")
        _assert_equal(cloudx_parsed.get("source"), "", "CloudX must not be treated as a - Cloudx row")
        lc._register_sdk_expected("CloudX", adapter="4.5.0", sdk="4.5.0")
        cloudx_http_log = (
            '08-14 17:18:30.818 D CloudX [CloudXHttpClient] '
            '[{"sdk":{"sdkVersion":"4.5.0","pluginVersion":"unity-4.5.0"}}]'
        )
        _assert(
            lc._process_sdk_cloudx_http_metadata_line(cloudx_http_log, "cloudx-device"),
            "CloudX telemetry log was not parsed",
        )
        cloudx_state = lc.sdk_check_runtime_state["cloudx-device"]["cloudx"]
        _assert_equal(cloudx_state.get("adapter_version"), "4.5.0", "CloudX plugin version was not normalized")
        _assert_equal(cloudx_state.get("sdk_version"), "4.5.0", "CloudX SDK version was not parsed")
        _assert(
            not lc._process_sdk_cloudx_http_metadata_line(
                '[CloudXHttpClientExtra] [{"sdk":{"sdkVersion":"9.9.9","pluginVersion":"unity-9.9.9"}}]',
                "cloudx-device",
            ),
            "CloudX parser must require the exact CloudXHttpClient marker",
        )
        suffix_state = lc.sdk_check_runtime_state["cloudx-device"].get("mintegralcloudx")
        _assert(not suffix_state, "CloudX telemetry must not update a - Cloudx network")

        for display_name, _, expected_version in labels:
            parsed = lc._parse_sdk_expected_line(display_name)
            _assert(parsed is not None, f"Cloudx network label not accepted: {display_name}")
            _assert_equal(parsed.get("source"), "cloudx", f"Cloudx source missing for {display_name}")
            lc._register_sdk_expected(
                display_name,
                adapter=expected_version,
                source=parsed.get("source", ""),
                match_network=parsed.get("match_network", ""),
            )

        for _, actual_name, actual_version in labels:
            line = (
                "*[AdapterMetadataResolver] Discovered adapter metadata: "
                f"network={actual_name}, adapterVersion={actual_version},*"
            )
            lc._process_sdk_cloudx_adapter_metadata_line(line, "cloudx-device")

        _assert_equal(len(lc.sdk_check_expected_map), len(labels) + 1, "Cloudx network count mismatch")
        for display_name, _, expected_version in labels:
            expected_key = lc._normalize_sdk_network_name(display_name)
            actual = lc.sdk_check_runtime_state["cloudx-device"][expected_key]["adapter_version"]
            _assert_equal(actual, expected_version, f"Cloudx adapter version mismatch for {display_name}")

        cloudx_native_versions = {
            "digitalturbinefybercloudx": "8.4.7",
            "inmobicloudx": "11.4.0",
            "liftoffmonetizationvunglecloudx": "7.7.7",
            "metaaudiencenetworkcloudx": "6.22.0",
            "mintegralcloudx": "17.1.71",
            "mobilefusecloudx": "1.11.0",
            "molococloudx": "4.11.0",
            "panglecloudx": "8.2.0.4",
            "unityadscloudx": "4.19.0",
            "vervepubnativecloudx": "3.9.0",
            "taurusxcloudx": "1.18.3",
        }
        cloudx_actual_names = {
            "digitalturbinefybercloudx": "Fyber",
            "inmobicloudx": "InMobi",
            "liftoffmonetizationvunglecloudx": "Vungle",
            "metaaudiencenetworkcloudx": "Meta",
            "mintegralcloudx": "Mintegral",
            "mobilefusecloudx": "Mobilefuse",
            "molococloudx": "Moloco",
            "panglecloudx": "Pangle",
            "unityadscloudx": "Unity",
            "vervepubnativecloudx": "Pubnative",
            "taurusxcloudx": "TaurusX",
        }
        for expected_key, native_version in cloudx_native_versions.items():
            lc.sdk_check_expected_map[expected_key]["sdk"] = native_version

        mintegral_key = lc._normalize_sdk_network_name("Mintegral - Cloudx")
        lc._process_sdk_cloudx_adapter_metadata_line(
            "08-14 15:26:50.991 com.indiez.nonogram 20641 20791 D CloudX "
            "[AdapterMetadataResolver] Discovered adapter metadata: "
            "network=Mintegral, adapterVersion=17.1.71.0, "
            "networkSdkVersion=17.1.71, minimumSdkVersion=4.2.0, "
            "minimumSdkVersionCode=4002000, adapterApiVersion=1, extras=Bundle[]",
            "cloudx-device",
        )
        mintegral_cloudx_state = lc.sdk_check_runtime_state["cloudx-device"][mintegral_key]
        _assert_equal(mintegral_cloudx_state.get("adapter_version"), "17.1.71.0", "Mintegral Cloudx adapter version mismatch")
        _assert_equal(mintegral_cloudx_state.get("native_version"), "17.1.71", "Mintegral Cloudx native version was not parsed")

        for expected_key, native_version in cloudx_native_versions.items():
            display_name = lc.sdk_check_expected_map[expected_key]["display_name"]
            lc._process_sdk_cloudx_adapter_metadata_line(
                f"[AdapterMetadataResolver] Discovered adapter metadata: network={cloudx_actual_names[expected_key]}, "
                f"adapterVersion=1.0.0, networkSdkVersion={native_version}, minimumSdkVersion=1.0.0",
                "cloudx-native-device",
            )
            _assert_equal(
                lc.sdk_check_runtime_state["cloudx-native-device"][expected_key].get("native_version"),
                native_version,
                f"Cloudx native version mismatch for {display_name}",
            )

        emitted = []
        original_emit = lc.socketio.emit
        try:
            lc.connected_devices_info = [{"id": "cloudx-native-device", "name": "CloudX device"}]
            lc.socketio.emit = lambda event, payload: emitted.append((event, payload))
            lc._emit_sdk_check_results()
        finally:
            lc.socketio.emit = original_emit
        emitted_rows = next(payload for event, payload in emitted if event == "update_sdk_check_table")
        native_rows = [row for row in emitted_rows if row.get("display_text", "").startswith("Native Version")]
        _assert_equal(len(native_rows), len(cloudx_native_versions), "Cloudx native result row count changed")
        _assert(all(row.get("status") == "PASSED" for row in native_rows), "Cloudx native result must pass matching versions")

        lc._process_sdk_cloudx_adapter_metadata_line(
            "[AdapterMetadataResolver] Discovered adapter metadata: "
            "network=Mintegral, adapterVersion=17.1.71.0, networkSdkVersion=17.1.70",
            "cloudx-native-device",
        )
        emitted = []
        original_emit = lc.socketio.emit
        try:
            lc.socketio.emit = lambda event, payload: emitted.append((event, payload))
            lc._emit_sdk_check_results()
        finally:
            lc.socketio.emit = original_emit
        emitted_rows = next(payload for event, payload in emitted if event == "update_sdk_check_table")
        mintegral_native_rows = [
            row for row in emitted_rows
            if row.get("display_text", "").startswith("Native Version")
            and "17.1.70" in row.get("display_text", "")
        ]
        _assert_equal(len(mintegral_native_rows), 1, "Mintegral Cloudx native mismatch row missing")
        _assert_equal(mintegral_native_rows[0].get("status"), "FAILED", "wrong Cloudx native version must fail")

        lc._register_sdk_expected("Mintegral", adapter="5.0.0", sdk="17.1.71")
        lc._process_sdk_check_line("IntegrationHelper ----- Mintegral -----", "cloudx-device")
        lc._process_sdk_check_line("IntegrationHelper Adapter Version - 5.0.0", "cloudx-device")
        mintegral_standard_key = lc._normalize_sdk_network_name("Mintegral")
        _assert_equal(
            lc.sdk_check_runtime_state["cloudx-device"][mintegral_standard_key]["adapter_version"],
            "5.0.0",
            "IntegrationHelper Mintegral state was not kept separate",
        )
        _assert_equal(
            lc.sdk_check_runtime_state["cloudx-device"][mintegral_key].get("native_version"),
            "17.1.71",
            "IntegrationHelper Mintegral changed Cloudx native state",
        )

        lc._register_sdk_expected("Meta Audience Network", adapter="1.0.0")
        lc._process_sdk_check_line("IntegrationHelper ----- Meta Audience Network -----", "cloudx-device")
        lc._process_sdk_check_line("IntegrationHelper Adapter Version - 1.0.0", "cloudx-device")
        _assert_equal(
            lc.sdk_check_runtime_state["cloudx-device"]["metaaudiencenetwork"]["adapter_version"],
            "1.0.0",
            "IntegrationHelper state was not kept separate",
        )
        _assert_equal(
            lc.sdk_check_runtime_state["cloudx-device"]["metaaudiencenetworkcloudx"]["adapter_version"],
            "4.0.0",
            "Cloudx state was overwritten by IntegrationHelper",
        )

        _assert_equal(
            lc._match_sdk_expected_key("Liftoff Monetization"),
            "",
            "IntegrationHelper matcher must not fall through to Cloudx-only entries",
        )
        lc._register_sdk_expected("Liftoff Monetization (vungle)", adapter="12.0.0")
        liftoff_key = lc._normalize_sdk_network_name("Liftoff Monetization (vungle)")
        _assert_equal(
            lc._match_sdk_expected_key("Liftoff Monetization"),
            liftoff_key,
            "Liftoff Monetization alias did not match the list label",
        )
        _assert_equal(
            lc._match_sdk_expected_key("Liftoff Monetize"),
            liftoff_key,
            "Liftoff Monetize alias did not match the list label",
        )
        lc._process_sdk_check_line("IntegrationHelper ----- Liftoff Monetize -----", "cloudx-device")
        lc._process_sdk_check_line("IntegrationHelper Adapter Version - 12.0.0", "cloudx-device")
        _assert_equal(
            lc.sdk_check_runtime_state["cloudx-device"][liftoff_key]["adapter_version"],
            "12.0.0",
            "IntegrationHelper Liftoff alias did not update the standard entry",
        )
        _assert_equal(
            lc.sdk_check_runtime_state["cloudx-device"]["liftoffmonetizationvunglecloudx"].get("adapter_version"),
            "3.0.0",
            "IntegrationHelper Liftoff alias overwrote Cloudx state",
        )
    finally:
        lc.active_platform = original_platform
        lc.connected_devices_info = original_devices
        lc.sdk_check_active = original_sdk_active
        lc.sdk_check_search_list[:] = original_search_list
        lc.sdk_check_expected_map = {}
        lc.sdk_check_expected_order = []
        lc.sdk_check_runtime_state = {}
        lc.sdk_check_current_network = {}


def test_max_sdk_logs() -> None:
    original_platform = lc.active_platform
    original_devices = lc.connected_devices_info
    original_sdk_active = lc.sdk_check_active
    original_search_list = list(lc.sdk_check_search_list)
    original_expected_map = lc.sdk_check_expected_map
    original_expected_order = lc.sdk_check_expected_order
    original_runtime_state = lc.sdk_check_runtime_state
    original_current_network = lc.sdk_check_current_network
    original_emit = lc.socketio.emit
    try:
        lc.active_platform = "android"
        lc.connected_devices_info = []
        lc.sdk_check_active = True
        lc.sdk_check_search_list = []
        lc.sdk_check_expected_map = {}
        lc.sdk_check_expected_order = []
        lc.sdk_check_runtime_state = {}
        lc.sdk_check_current_network = {}

        c191_lines = lc._load_sdk_check_presets()["C-191-Android"]["lines"]
        max_expected = {}
        for line in c191_lines:
            if " - MAX" not in line:
                continue
            parsed = lc._parse_sdk_expected_line(line)
            _assert(parsed is not None, f"MAX entry cannot be parsed: {line}")
            _assert_equal(parsed.get("source"), "max", f"MAX entry source missing: {line}")
            max_expected[parsed["network"]] = parsed

        original_emit = lc.socketio.emit
        lc.socketio.emit = lambda *_args, **_kwargs: None
        try:
            lc.sdk_check({"text": "\n".join(c191_lines)})
        finally:
            lc.socketio.emit = original_emit

        core_key = lc._normalize_sdk_network_name("MAX / AppLovin - MAX")
        failed_core_line = (
            "09-03 14:38:39.685 com.indiez.nonogram 7428 8001 D AppLovinSdk "
            "[TaskInitializeSdk] AppLovin SDK 99.9.9 initialization failed"
        )
        _assert(not lc._process_sdk_max_line(failed_core_line, "max-device"), "failed MAX core init must be ignored")
        _assert(core_key not in lc.sdk_check_runtime_state.get("max-device", {}), "failed MAX core init created state")

        original_emit = lc.socketio.emit
        lc.socketio.emit = lambda *_args, **_kwargs: None
        try:
            core_line = (
                "09-03 14:38:39.685\tcom.indiez.nonogram\t7428\t8001\tD\tAppLovinSdk\t"
                "[TaskInitializeSdk] AppLovin SDK 13.6.4 initialization succeeded in 94ms"
            )
            lc._process_sdk_check_line(core_line, "max-device")
        finally:
            lc.socketio.emit = original_emit

        core_state = lc.sdk_check_runtime_state["max-device"][core_key]
        _assert_equal(core_state.get("sdk_version"), "13.6.4", "MAX core SDK version was not parsed")
        _assert_equal(core_state.get("adapter_version"), "", "MAX core line must update only one SDK version")

        started_core_line = core_line.replace("initialization succeeded", "initialization started").replace("13.6.4", "99.9.9")
        _assert(not lc._process_sdk_max_line(started_core_line, "max-device"), "MAX core parser accepted a non-succeeded line")
        _assert_equal(core_state.get("sdk_version"), "13.6.4", "non-succeeded MAX core line overwrote the version")

        facebook_line = (
            "09-03 13:58:14.758\tcom.indiez.nonogram\t24597\t25054\tD\tAppLovinSdk\t"
            "[HealthEventsReporter] Reporting signal_collection_success with extra parameters "
            "collection [{duration_ms=0, ad_format=REWARDED, ad_unit_id=08072f2e4324b79d, "
            "adapter_version=6.22.0.0, network_name=FACEBOOK_NETWORK, "
            "adapter_class=com.applovin.mediation.adapters.FacebookMediationAdapter}]"
        )
        original_emit = lc.socketio.emit
        lc.socketio.emit = lambda *_args, **_kwargs: None
        try:
            lc._process_sdk_check_line(facebook_line, "max-device")
        finally:
            lc.socketio.emit = original_emit

        facebook_key = lc._normalize_sdk_network_name("Meta Audience Network - MAX")
        facebook_state = lc.sdk_check_runtime_state["max-device"][facebook_key]
        _assert_equal(facebook_state.get("adapter_version"), "6.22.0.0", "FACEBOOK_NETWORK adapter version was not parsed")
        _assert_equal(facebook_state.get("observed_network_name"), "FACEBOOK_NETWORK", "MAX network name was not retained")
        _assert_equal(
            lc._sdk_result_status(facebook_state.get("adapter_version"), max_expected["Meta Audience Network - MAX"]["adapter"]),
            "PASSED",
            "FACEBOOK_NETWORK adapter version did not compare against the preset",
        )

        alias_cases = [
            ("IRONSOURCE_NETWORK", "LevelPlay / ironSource - MAX"),
            ("BIDMACHINE_NETWORK", "BidMachine - MAX"),
            ("BIGO_NETWORK", "Bigo Ads - MAX"),
            ("CHARTBOOST_NETWORK", "Chartboost - MAX"),
            ("DT_EXCHANGE", "Digital Turbine (fyber) - MAX"),
            ("GOOGLE_AD_MANAGER_NETWORK", "Google (AdMob and Ad Manager) - MAX"),
            ("INMOBI_NETWORK", "InMobi - MAX"),
            ("LINE_NETWORK", "LINE Ads - MAX"),
            ("LIFTOFF_NETWORK", "Liftoff Monetization (vungle) - MAX"),
            ("MINTEGRAL_NETWORK", "Mintegral - MAX"),
            ("MOBILEFUSE_NETWORK", "Mobilefuse - MAX"),
            ("MOLOCO_NETWORK", "Moloco - MAX"),
            ("OGURY_NETWORK", "Ogury - MAX"),
            ("PANGLE_NETWORK", "Pangle (Tiktok) - MAX"),
            ("PUBMATIC_NETWORK", "PubMatic (OpenWrap) - MAX"),
            ("UNITY_ADS_NETWORK", "UnityAds - MAX"),
            ("VERVE_NETWORK", "Verve / Pubnative - MAX"),
            ("YANDEX_NETWORK", "Yandex - MAX"),
            ("YSO_NETWORK", "YSO - MAX"),
            ("ASCENDX_NETWORK", "Ascendx - MAX"),
            ("MATICOO_NETWORK", "Yeahmobi/ Maticoo - MAX"),
            ("TAURUSX_NETWORK", "TaurusX - MAX"),
            ("PRADO_NETWORK", "Prado - MAX"),
            ("BIGO_BIDDING", "Bigo Ads - MAX"),
            ("BIGO_NATIVE_BIDDING", "Bigo Ads - MAX"),
            ("CHARTBOOST_BIDDING", "Chartboost - MAX"),
            ("BIDMACHINE_BIDDING", "BidMachine - MAX"),
            ("FYBER_BIDDING", "Digital Turbine (fyber) - MAX"),
            ("INMOBI_NATIVE_BIDDING", "InMobi - MAX"),
            ("MINTEGRAL_NATIVE_BIDDING", "Mintegral - MAX"),
            ("TIKTOK_NATIVE_BIDDING", "Pangle (Tiktok) - MAX"),
            ("OGURY_PRESAGE_BIDDING", "Ogury - MAX"),
            ("VUNGLE_NATIVE_BIDDING", "Liftoff Monetization (vungle) - MAX"),
            ("MOLOCO_NATIVE_BIDDING", "Moloco - MAX"),
        ]
        for observed_name, display_name in alias_cases:
            expected = max_expected[display_name]
            version = expected["adapter"]
            _assert(version, f"MAX adapter preset is empty: {display_name}")
            metadata_line = (
                "[HealthEventsReporter] Reporting signal_collection_success with extra parameters "
                f"collection [{{adapter_version={version}, network_name={observed_name}, "
                "adapter_class=com.applovin.mediation.adapters.TestMediationAdapter}}]"
            )
            _assert(lc._process_sdk_max_line(metadata_line, "alias-device"), f"MAX metadata did not update: {display_name}")
            expected_key = lc._normalize_sdk_network_name(display_name)
            actual = lc.sdk_check_runtime_state["alias-device"][expected_key].get("adapter_version")
            _assert_equal(actual, version, f"MAX alias mapping failed for {observed_name}")
            _assert_equal(lc._sdk_result_status(actual, version), "PASSED", f"MAX version compare failed for {display_name}")
    finally:
        lc.active_platform = original_platform
        lc.connected_devices_info = original_devices
        lc.sdk_check_active = original_sdk_active
        lc.sdk_check_search_list[:] = original_search_list
        lc.sdk_check_expected_map = original_expected_map
        lc.sdk_check_expected_order = original_expected_order
        lc.sdk_check_runtime_state = original_runtime_state
        lc.sdk_check_current_network = original_current_network
        lc.socketio.emit = original_emit


def test_sdk_check_preset_contract() -> None:
    presets = lc._load_sdk_check_presets()
    _assert("C-191-Android" in presets, "C-191 Android SDK preset is missing")
    _assert("C-190-iOS" in presets, "C-190 iOS SDK preset is missing")
    _assert("C-180-Android" in presets, "C-180 Android SDK preset is missing")
    _assert("C-180-iOS" in presets, "C-180 iOS SDK preset is missing")
    c191_lines = presets["C-191-Android"].get("lines") or []
    expected_c191_lines = [
        "Ads Network\tAdapter\tNative",
        "IronSource\t9.6.0\t9.6.0",
        "AppLovin\t5.9.0\t13.6.4",
        "BidMachine\t5.8.0\t3.8.0",
        "Bigo Ads\t5.11.0\t6.0.0",
        "Chartboost\t5.8.0\t9.13.0",
        "Digital Turbine (fyber)\t5.10.0\t8.4.7",
        "Google (AdMob and Ad Manager)\t5.9.0\t25.4.0",
        "HyprMX\t5.3.0\t6.4.6",
        "InMobi\t5.9.0\t11.4.1",
        "LINE Ads\t5.4.0\t3.1.1",
        "Liftoff Monetization (vungle)\t5.14.0\t7.7.8",
        "Meta Audience Network\t5.4.0\t6.22.0",
        "Mintegral\t5.19.0\t17.1.81",
        "Mobilefuse\t5.4.0\t1.12.0",
        "Moloco\t5.16.0\t4.11.1",
        "myTarget (VK Ads)\t5.6.0\t5.51.2",
        "Ogury\t5.5.0\t6.3.1",
        "Pangle\t5.22.0\t8.2.0.4",
        "PubMatic (OpenWrap)\t5.9.0\t5.3.0",
        "SuperAwesome\t5.5.0\t10.3.1",
        "UnityAds\t5.12.0\t4.20.0",
        "Verve / Pubnative\t5.7.0\t3.9.1",
        "Voodoo\t5.7.0\t4.29.2",
        "Yandex\t5.13.0\t8.3.0",
        "YSO\t5.6.0\t1.3.8",
        "LevelPlay / ironSource - MAX\t9.6.0.0.0",
        "MAX / AppLovin - MAX\t\t13.6.4",
        "BidMachine - MAX\t3.8.0.0",
        "Bigo Ads - MAX\t6.0.0.0",
        "Chartboost - MAX\t9.13.0.0",
        "Digital Turbine (fyber) - MAX\t8.4.7.0",
        "Google (AdMob and Ad Manager) - MAX\t25.4.0.0",
        "InMobi - MAX\t11.4.1.2",
        "LINE Ads - MAX\t3000.1.1.0",
        "Liftoff Monetization (vungle) - MAX\t7.7.8.0",
        "Meta Audience Network - MAX\t6.22.0.0",
        "Mintegral - MAX\t17.1.81.0",
        "Mobilefuse - MAX\t1.12.0.0",
        "Moloco - MAX\t4.11.1.0",
        "Ogury - MAX\t6.3.1.0",
        "Pangle (Tiktok) - MAX\t8.2.0.4.0",
        "PubMatic (OpenWrap) - MAX\t5.3.0.0",
        "UnityAds - MAX\t4.20.0.0",
        "Verve / Pubnative - MAX\t3.9.1.0",
        "Yandex - MAX\t8.3.0.0",
        "YSO - MAX\t1.3.8.0",
        "Ascendx - MAX\t1.11.1",
        "Yeahmobi/ Maticoo - MAX\t2.0.7.0",
        "TaurusX - MAX\t1.16.3.1",
        "Prado - MAX\t2.0.2",
        "Adverty\t5.2.9",
        "Gadsme\t1.12.6",
        "AudioMob\t10.2.3",
        "Adjust\t\t5.8.0",
        "Firebase Crashlytics\t\t20.1.0",
        "Facebook SDK\t\t18.3.0",
        "AppMetrica SDK\t\t8.4.1",
    ]
    _assert_equal(c191_lines, expected_c191_lines, "C-191 Android SDK preset does not match the requested list")
    _assert_equal(presets["C-191-Android"].get("platform"), "android", "C-191 Android preset platform changed")
    for line in c191_lines[1:]:
        parsed = lc._parse_sdk_expected_line(line)
        _assert(parsed is not None, f"C-191 Android entry cannot be parsed: {line}")
    ios_preset = presets["C-190-iOS"]
    _assert_equal(ios_preset.get("platform"), "ios", "C-190 iOS preset platform changed")
    _assert_equal(ios_preset.get("lines"), [], "C-190 iOS preset must remain empty")
    c180_ios_preset = presets["C-180-iOS"]
    _assert_equal(c180_ios_preset.get("platform"), "ios", "C-180 iOS preset platform changed")
    c180_ios_lines = c180_ios_preset.get("lines") or []
    _assert_equal(len(c180_ios_lines), 31, "C-180 iOS preset line count changed")
    _assert_equal(c180_ios_lines[0], "Ads Network\tAdapter\tSDK", "C-180 iOS preset header changed")
    _assert("ironSource\t9.4.2.1\t9.4.2" in c180_ios_lines, "C-180 iOS ironSource entry changed")
    _assert("Appsflyer\tremoved\tremoved" in c180_ios_lines, "C-180 iOS Appsflyer entry changed")
    _assert("Firebase Crashlytics\t\t12.15.0" in c180_ios_lines, "C-180 iOS Firebase Crashlytics entry changed")
    _assert(all("http://" not in line and "https://" not in line for line in c180_ios_lines), "C-180 iOS preset must not contain links")
    c180_preset = presets["C-180-Android"]
    _assert_equal(c180_preset.get("platform"), "android", "C-180 preset platform changed")
    c180_lines = c180_preset.get("lines") or []
    _assert_equal(len(c180_lines), 33, "C-180 preset line count changed")
    _assert_equal(c180_lines[0], "Ads Network\tAdapter\tNative", "C-180 preset header changed")
    _assert("Appsflyer\t\tRemoved" in c180_lines, "C-180 Appsflyer entry changed")
    _assert(not any(line.startswith("AdQuality\t") for line in c180_lines), "C-180 AdQuality entry must remain absent")
    _assert(all("http://" not in line and "https://" not in line for line in c180_lines), "C-180 preset must not contain links")
    for line in c180_lines[1:]:
        parsed = lc._parse_sdk_expected_line(line)
        _assert(parsed is not None, f"C-180 Android entry cannot be parsed: {line}")
    for line in c180_ios_lines[1:]:
        parsed = lc._parse_sdk_expected_line(line)
        _assert(parsed is not None, f"C-180 iOS entry cannot be parsed: {line}")

    source_text = (ROOT / "Log_checker.py").read_text(encoding="utf-8", errors="ignore")
    _assert("sdkCheckPresetPanel" in source_text, "SDK preset panel is missing")
    _assert("applySdkCheckPreset" in source_text, "SDK preset action is missing")
    _assert("/api/sdk-check-presets" in source_text, "remote SDK preset endpoint is missing")
    _assert("loadSdkCheckPresetsFromGit" in source_text, "SDK preset remote load action is missing")
    _assert("sdkCheckInput" in source_text, "manual SDK input fallback is missing")
    _assert(source_text.count('id="reloadSdkCheckPresetsBtn"') == 1, "SDK preset reload button must exist exactly once")
    _assert("clean_lines" in source_text and '"lines": clean_lines' in source_text, "empty SDK presets must remain valid")
    _assert('"Accept-Encoding": "identity"' in source_text, "SDK preset fetch must bypass stale compressed cache")
    _assert("def _fetch_sdk_check_presets(force_remote=False):" in source_text, "SDK preset force refresh support is missing")
    _assert("refresh_requested = request.args.get(\"refresh\"" in source_text, "SDK preset refresh query is missing")
    _assert("force_remote=refresh_requested" in source_text, "SDK preset endpoint must honor forced refresh")
    _assert("main/sdk_check_presets.json" in source_text, "SDK preset main fallback URL is missing")


def test_rendered_sdk_preset_javascript_contract() -> None:
    response = lc.app.test_client().get("/")
    _assert_equal(response.status_code, 200, "local UI route must render")
    html = response.get_data(as_text=True)
    expected_join = r"input.value = lines.join('\n');"
    broken_join = "input.value = lines.join('" + "\n" + "');"
    _assert(expected_join in html, "rendered SDK preset JavaScript must keep the newline escape")
    _assert(broken_join not in html, "rendered SDK preset JavaScript contains a literal newline inside a string")
    _assert(
        html.index('id="sdkCheckPresetPanel"') < html.index('id="startSdkCheckBtn"'),
        "SDK preset selector must remain above Start Checking",
    )
    _assert('id="reloadSdkCheckPresetsBtn"' in html, "rendered SDK preset reload button is missing")
    _assert("loadSdkCheckPresetsFromGit();" in html, "SDK presets must reload when the app UI starts")
    _assert("loadSdkCheckPresetsFromGit(true);" in html, "preset Reload button must force a fresh GitHub request")
    _assert("refreshQuery = force ? '&refresh=1'" in html, "SDK preset Reload must request a remote refresh")
    _assert("reloadSdkCheckPresetsBtn" in html and "loadSdkCheckPresetsFromGit" in html, "reload button handler is missing")


def test_default_ad_event_contract() -> None:
    normal_events = {
        "interstitial": [
            "ad_load",
            "ad_load_failed",
            "ad_load_succeeded",
            "ad_ready",
            "ad_not_ready",
            "ad_impression",
            "ad_show_succeeded",
            "ad_show_failed",
            "ad_clicked",
            "ad_closed",
        ],
        "rewarded": [
            "ad_load",
            "ad_load_failed",
            "ad_load_succeeded",
            "ad_ready",
            "ad_not_ready",
            "ad_impression",
            "ad_show_succeeded",
            "ad_show_failed",
            "ad_clicked",
            "ad_closed",
            "ad_rewarded",
        ],
        "banner": [
            "ad_load",
            "ad_load_failed",
            "ad_load_succeeded",
            "ad_impression",
            "ad_show_succeeded",
            "ad_show_failed",
            "ad_clicked",
        ],
        "mrec": [
            "ad_load",
            "ad_load_failed",
            "ad_load_succeeded",
            "ad_impression",
            "ad_show_succeeded",
            "ad_show_failed",
            "ad_clicked",
        ],
    }
    simple_events = {
        "audioads": {
            "odeeo": [
                "ad_load_succeeded",
                "ad_ready",
                "ad_not_ready",
                "ad_impression",
                "ad_show_succeeded",
                "ad_show_failed",
                "ad_clicked",
                "ad_closed",
                "ad_rewarded",
            ],
            "audiomob": [
                "ad_load",
                "ad_load_failed",
                "ad_impression",
                "ad_show_succeeded",
                "ad_show_failed",
                "ad_clicked",
            ],
        },
        "inplayads": {
            "adverty": ["ad_impression", "ad_show_succeeded", "ad_clicked"],
            "gadsme": ["ad_load_failed", "ad_load_succeeded", "ad_impression", "ad_show_succeeded", "ad_clicked"],
        },
    }

    workbook = Workbook()
    workbook.active.title = "events"

    def add_normal_sheet(provider: str, platform_value=None) -> None:
        worksheet = workbook.create_sheet(provider)
        platform_value = platform_value or provider
        for column, ad_format in zip((2, 6, 10, 14), lc.DEFAULT_AD_EVENT_FORMATS):
            worksheet.merge_cells(
                start_row=2,
                start_column=column,
                end_row=2,
                end_column=column + 2,
            )
            worksheet.cell(2, column).value = lc.DEFAULT_AD_EVENT_FORMAT_LABELS[ad_format]
            row = 3
            for event_name in normal_events[ad_format]:
                worksheet.cell(row, column).value = event_name
                worksheet.cell(row + 1, column + 1).value = "ad_platform"
                worksheet.cell(row + 1, column + 2).value = platform_value
                worksheet.cell(row + 2, column + 1).value = "ad_format"
                worksheet.cell(row + 2, column + 2).value = ad_format
                row += 3

    add_normal_sheet("Ironsource", "ironsource")
    add_normal_sheet("MAX", "max")
    add_normal_sheet("Cloudx")
    add_normal_sheet("Ascendx")

    def add_simple_sheet(category: str) -> None:
        worksheet = workbook.create_sheet(lc.DEFAULT_AD_EVENT_SIMPLE_CATEGORY_LABELS[category])
        for column, provider in zip((2, 6), lc.DEFAULT_AD_EVENT_SIMPLE_PROVIDERS[category]):
            worksheet.merge_cells(
                start_row=2,
                start_column=column,
                end_row=2,
                end_column=column + 2,
            )
            worksheet.cell(2, column).value = provider
            row = 3
            for event_name in simple_events[category][provider]:
                worksheet.cell(row, column).value = event_name
                worksheet.cell(row + 1, column + 1).value = "ad_platform"
                worksheet.cell(row + 1, column + 2).value = provider
                worksheet.cell(row + 2, column + 1).value = "ad_format"
                worksheet.cell(row + 2, column + 2).value = "audio_rewarded" if category == "audioads" else "in_play"
                row += 3

    add_simple_sheet("audioads")
    add_simple_sheet("inplayads")

    parsed_normal = lc._read_default_ad_event_config(workbook)
    expected_normal = {
        ad_format: {
            provider: list(normal_events[ad_format])
            for provider in lc.DEFAULT_AD_EVENT_PROVIDERS
        }
        for ad_format in lc.DEFAULT_AD_EVENT_FORMATS
    }
    _assert_equal(parsed_normal, expected_normal, "normal Default Ad Event sheets parsed incorrectly")
    parsed_simple = lc._read_default_ad_event_simple_config(workbook)
    _assert_equal(parsed_simple, simple_events, "AudioAds/InplayAds sheets parsed incorrectly")

    empty_workbook = Workbook()
    empty_workbook.active.title = "events"
    empty_normal = lc._read_default_ad_event_config(empty_workbook)
    empty_simple = lc._read_default_ad_event_simple_config(empty_workbook)
    _assert(
        not any(event_names for providers in empty_normal.values() for event_names in providers.values()),
        "missing provider sheets must not create normal ad-event rows",
    )
    _assert(
        not any(event_names for providers in empty_simple.values() for event_names in providers.values()),
        "missing AudioAds/InplayAds sheets must not create simple ad-event rows",
    )

    original_config = lc.default_ad_event_config
    original_simple_config = lc.default_ad_event_simple_config
    original_hits = set(lc.default_ad_event_hits)
    original_simple_hits = set(lc.default_ad_event_simple_hits)
    original_clients = set(lc.default_ad_event_clients)
    original_paused = lc.is_paused
    original_emit = lc.socketio.emit
    emitted = []
    try:
        lc.default_ad_event_config = parsed_normal
        lc.default_ad_event_simple_config = parsed_simple
        lc.default_ad_event_hits.clear()
        lc.default_ad_event_simple_hits.clear()
        lc.default_ad_event_clients.clear()
        lc.default_ad_event_clients.add("harness-client")
        lc.is_paused = False
        lc.socketio.emit = lambda event, payload: emitted.append((event, payload))

        line = (
            'Unity : TrackingService->Track: '
            '{"eventName":"ad_impression","e":{"ad_platform":"max","ad_format":"Rewarded"}}'
        )
        event_name, actual_params, _json_string = lc.find_and_parse_event(line)
        _assert_equal(event_name, "ad_impression", "generic ad event was not parsed")
        _assert_equal(
            actual_params,
            {"ad_platform": "max", "ad_format": "Rewarded"},
            "generic ad event parameters were not parsed",
        )
        lc._record_default_ad_event_hit(event_name, actual_params, "device-normal")
        lc._record_default_ad_event_hit(
            "ad_clicked",
            {"ad_platform": "Audiomob", "ad_format": "audio_skippable"},
            "device-simple",
        )
        lc._record_default_ad_event_hit(
            "ad_rewarded",
            {"ad_platform": "Cloudx", "ad_format": "rewarded"},
            "device-hidden-provider",
        )

        payload = lc._default_ad_event_payload()
        _assert_equal(
            payload["hits"],
            [{
                "device_id": "device-normal",
                "provider": "max",
                "ad_format": "rewarded",
                "event_name": "ad_impression",
            }],
            "MAX ad-event match did not use max ad_platform and ad_format",
        )
        _assert(
            not any(hit.get("provider") in {"cloudx", "ascendx"} for hit in payload["hits"]),
            "temporarily hidden providers must not create visible coverage hits",
        )
        _assert_equal(
            payload["simple_hits"],
            [{
                "device_id": "device-simple",
                "category": "audioads",
                "provider": "audiomob",
                "event_name": "ad_clicked",
            }],
            "simple ad-event match did not use ad_platform-only matching",
        )
        _assert_equal(len(emitted), 2, "only matching ad events should emit coverage updates")

        catalog = lc._default_ad_event_catalog_payload()
        _assert_equal(
            [provider["key"] for provider in catalog["providers"]],
            ["ironsource", "max"],
            "Default Ad Events must expose Ironsource and MAX only",
        )
        _assert_equal(
            [provider["label"] for provider in catalog["providers"]],
            ["Ironsource", "MAX"],
            "Default Ad Events provider labels changed",
        )

        rendered = lc.app.test_client().get("/")
        _assert_equal(rendered.status_code, 200, "Default Ad Events UI route must render")
        html = rendered.get_data(as_text=True)
        for marker in (
            ">Default Ad Events</button>",
            "defaultAdEventMatrix",
            "defaultAdEventHitKey",
            "defaultAdEventSimpleHitKey",
            "Odeeo",
            "Audiomob",
            "Adverty",
            "Gadsme",
            "MAX",
            "Simple providers are intentionally rendered as separate blocks",
        ):
            _assert(marker in html, f"Default Ad Events UI contract is missing: {marker}")
    finally:
        lc.default_ad_event_config = original_config
        lc.default_ad_event_simple_config = original_simple_config
        lc.default_ad_event_hits.clear()
        lc.default_ad_event_hits.update(original_hits)
        lc.default_ad_event_simple_hits.clear()
        lc.default_ad_event_simple_hits.update(original_simple_hits)
        lc.default_ad_event_clients.clear()
        lc.default_ad_event_clients.update(original_clients)
        lc.is_paused = original_paused
        lc.socketio.emit = original_emit


def test_installation_id_copy_contract() -> None:
    for source_path in (ROOT / "Log_checker.py",):
        source_text = source_path.read_text(encoding="utf-8", errors="ignore")
        _assert("async function copyTextToClipboard(text)" in source_text, f"copy helper missing in {source_path}")
        _assert("await navigator.clipboard.writeText(value);" in source_text, f"Clipboard API path missing in {source_path}")
        _assert("document.execCommand('copy')" in source_text, f"Windows clipboard fallback missing in {source_path}")
        _assert(
            "if (!await copyTextToClipboard(installationId))" in source_text,
            f"Installation ID button is not using the cross-platform copy helper in {source_path}",
        )
        _assert(
            "await navigator.clipboard.writeText(installationId);" not in source_text,
            f"Installation ID button still bypasses the cross-platform copy helper in {source_path}",
        )
        _assert(
            'id="installationIdValue" type="text" readonly' in source_text,
            f"Installation ID must be exposed as a selectable readonly text field in {source_path}",
        )
        _assert(
            "installationIdValueEl.value = installationId;" in source_text,
            f"Installation ID text field is not updated through its value in {source_path}",
        )
        _assert(
            "installationIdValueEl.select();" in source_text,
            f"Installation ID text field does not select its value on focus in {source_path}",
        )
        _assert(
            ".installation-id-value" in source_text and "user-select: text !important" in source_text,
            f"Installation ID text field is not explicitly selectable in {source_path}",
        )

    rendered = lc.app.test_client().get("/").get_data(as_text=True)
    _assert("async function copyTextToClipboard(text)" in rendered, "rendered copy helper is missing")
    _assert("document.execCommand('copy')" in rendered, "rendered Windows clipboard fallback is missing")
    _assert('id="installationIdValue" type="text" readonly' in rendered, "rendered Installation ID field is not selectable")
    _assert("id=\"clearUpdateCacheBtn\"" not in rendered, "Clear Cache button must stay removed from the release UI")


def test_sdk_failed_groups_sort_first() -> None:
    groups = [
        [{"status": "PASSED"}, {"status": "PASSED"}],
        [{"status": "FAILED"}],
        [{"status": "FOUND"}],
        [{"status": "PASSED"}, {"status": "FAILED"}],
    ]
    ordered = sorted(groups, key=lc._sdk_result_group_sort_key)
    _assert(ordered[0][0]["status"] == "FAILED", "failed SDK group must be first")
    _assert(ordered[1][1]["status"] == "FAILED", "mixed failed SDK group must remain before passed groups")
    _assert(ordered[-1][0]["status"] == "FOUND", "found-only SDK group must remain after passed groups")
    _assert_equal(lc._sdk_result_status("", "1.18.3.0"), "FAILED", "missing actual with expected version must fail")
    _assert_equal(lc._sdk_result_status("", ""), "NOT_FOUND", "empty expected version must be not found")
    _assert_equal(lc._sdk_result_status("NOT FOUND", "Removed"), "PASSED", "removed SDK with missing actual must pass")
    _assert_equal(lc._sdk_result_status("MISSING", "Removed"), "PASSED", "removed SDK with missing marker must pass")


def test_release_build_marker() -> None:
    text = (ROOT / "Log_checker.py").read_text(encoding="utf-8", errors="ignore")
    _assert(
        f"v2.5.0({CURRENT_RELEASE_BUILD})" in text,
        f"Log_checker.py must be prepared for v2.5.0({CURRENT_RELEASE_BUILD})",
    )


def test_rewarded_bidding_filter_contract() -> None:
    source_text = (ROOT / "Log_checker.py").read_text(encoding="utf-8", errors="ignore")
    needle = 'data-message-needle="[Ad,RewardedBidding"'
    _assert(source_text.count('value="rewarded_bidding"') == 1, "RewardedBidding filter must exist exactly once")
    _assert(source_text.count('value="rewardedcap"') == 1, "RewardedCap filter must exist exactly once")
    _assert(source_text.count(needle) == 1, "RewardedBidding must use the exact message needle")
    _assert('data-message-needle="[Ad,RewardedBidding,"' not in source_text, "RewardedBidding filter must not require a trailing comma")
    _assert('data-android-only="true"' in source_text, "RewardedBidding must be Android-only")
    _assert("exactMessage.includes(state.quickMessage)" in source_text, "message filter must remain case-sensitive")
    _assert("flex items-start justify-start gap-32" in source_text, "RewardedBidding layout must stay beside the first filter column")


def test_price_rotation_exact_parser() -> None:
    source_text = (ROOT / "Log_checker.py").read_text(encoding="utf-8", errors="ignore")
    _assert('onclick="switchTab(\'PriceRotation\')">Bidding</button>' in source_text, "Price Rotation tab label must be Bidding")
    _assert(source_text.count('id="priceRotationTypeMax"') == 1, "MAX filter must exist exactly once")
    _assert('value="max"' in source_text, "MAX filter value is missing")
    _assert('id="priceRotationTypeAscendx"' not in source_text, "Ascendx Bidding filter must be hidden")
    _assert('id="priceRotationTypeCloudx"' not in source_text, "Cloudx Bidding filter must be hidden")
    _assert("rotationType === 'ascendx' || rotationType === 'cloudx'" in source_text, "hidden Bidding types must stay filtered in the UI")
    _assert(source_text.count('id="priceRotationTypeWaterfall"') == 1, "Waterfall filter must exist exactly once")
    _assert(source_text.count('id="priceRotationTypeInterstitialCap"') == 1, "InterstitialCap filter must exist exactly once")
    _assert(source_text.count('id="priceRotationAdTypeAll"') == 1, "Ad type All filter must exist exactly once")
    _assert(source_text.count('id="priceRotationAdTypeRewarded"') == 1, "Rewarded ad type filter must exist exactly once")
    _assert(source_text.count('id="priceRotationAdTypeInterstitial"') == 1, "Interstitial ad type filter must exist exactly once")
    _assert('value="rewarded"' in source_text and 'value="interstitial"' in source_text, "Rewarded/Interstitial filters are missing")
    _assert('value="waterfall"' in source_text, "Waterfall filter value is missing")
    valid = 'Unity : [Ad,RewardedBidding, CloudX] Raise: {"act":"demandFailedObserved","error":"WinnerBid not found"}'
    parsed = lc._parse_price_rotation_log(valid, "device-price")
    _assert(parsed is not None, "Price Rotation parser must accept the exact marker")
    _assert_equal(parsed["type"], "CloudX", "Price Rotation type should come from the marker")
    _assert('"act": "demandFailedObserved"' in parsed["details"], "Price Rotation details should format JSON")
    waterfall = lc._parse_price_rotation_log('Unity : [Ad,RewardedBidding, Waterfall] Raise: {"act":"bid"}', "device-price")
    _assert(waterfall is not None, "Price Rotation parser must accept the Waterfall marker")
    _assert_equal(waterfall["type"], "Waterfall", "Waterfall type should come from the marker")
    cap = lc._parse_price_rotation_log('Unity : [Ad,RewardedBidding] RewardedCapService: {"act":"cap"}', "device-price")
    _assert(cap is not None, "Price Rotation parser must accept the RewardedCapService marker")
    _assert_equal(cap["type"], "RewardedCap", "RewardedCapService must use the RewardedCap type")
    plain = lc._parse_price_rotation_log('Unity : [Ad,RewardedBidding] Raise: {"act":"bid"}', "device-price")
    _assert(plain is not None, "Price Rotation parser must accept a plain RewardedBidding marker")
    _assert_equal(plain["type"], "RewardedBidding", "A plain marker must use the RewardedBidding type")
    interstitial_raw = 'Unity : [Ad,InterstitialBidding, LevelPlay] Raise: {"act":"bid"}'
    interstitial = lc._parse_price_rotation_log(interstitial_raw, "device-price")
    _assert(interstitial is not None, "Price Rotation parser must accept the InterstitialBidding marker")
    _assert_equal(interstitial["type"], "LevelPlay", "Interstitial type should come from the marker")
    interstitial_cap = lc._parse_price_rotation_log('Unity : [Ad,InterstitialBidding] InterstitialCapService: {"act":"cap"}', "device-price")
    _assert(interstitial_cap is not None, "Price Rotation parser must accept the InterstitialCapService marker")
    _assert_equal(interstitial_cap["type"], "InterstitialCap", "InterstitialCapService must use the InterstitialCap type")
    interstitial_plain = lc._parse_price_rotation_log('Unity : [Ad,InterstitialBidding] Raise: {"act":"bid"}', "device-price")
    _assert(interstitial_plain is not None, "Price Rotation parser must accept a plain InterstitialBidding marker")
    _assert_equal(interstitial_plain["type"], "InterstitialBidding", "A plain marker must use the InterstitialBidding type")
    _assert(lc._parse_price_rotation_log(valid.replace("[Ad,RewardedBidding,", "[Ad,RewardedBiddingX,"), "device-price") is None, "Price Rotation marker must be exact")
    _assert(lc._parse_price_rotation_log(interstitial_raw.replace("[Ad,InterstitialBidding,", "[Ad,InterstitialBiddingX,"), "device-price") is None, "Interstitial marker must be exact")

    original_paused = lc.is_paused
    original_emit = lc.socketio.emit
    original_rows = list(lc.price_rotation_logs)
    try:
        lc.is_paused = False
        lc.price_rotation_logs.clear()
        lc.socketio.emit = lambda *_args, **_kwargs: None
        max_rewarded = lc._parse_price_rotation_log(
            'Unity : [Ad,RewardedBidding, MAX] Raise: {"act":"bid"}',
            "device-price",
        )
        max_rewarded_applovin = lc._parse_price_rotation_log(
            'Unity : [Ad,RewardedBidding, applovin] Raise: {"act":"bid"}',
            "device-price",
        )
        max_interstitial = lc._parse_price_rotation_log(
            'Unity : [Ad,InterstitialBidding, MAX] Raise: {"act":"bid"}',
            "device-price",
        )
        max_interstitial_applovin = lc._parse_price_rotation_log(
            'Unity : [Ad,InterstitialBidding, applovin] Raise: {"act":"bid"}',
            "device-price",
        )
        _assert(
            all(parsed and parsed["type"] in {"MAX", "applovin"} for parsed in (
                max_rewarded,
                max_rewarded_applovin,
                max_interstitial,
                max_interstitial_applovin,
            )),
            "MAX markers must parse for Rewarded and Interstitial bidding",
        )
        for raw_log in (
            'Unity : [Ad,RewardedBidding, MAX] Raise: {"act":"bid"}',
            'Unity : [Ad,RewardedBidding, applovin] Raise: {"act":"bid"}',
            'Unity : [Ad,InterstitialBidding, MAX] Raise: {"act":"bid"}',
            'Unity : [Ad,InterstitialBidding, applovin] Raise: {"act":"bid"}',
        ):
            lc.process_price_rotation_log(raw_log, "device-price")
        _assert_equal(len(lc.price_rotation_logs), 4, "MAX Bidding variants must remain visible")

        lc.process_price_rotation_log(valid, "device-price")
        _assert_equal(len(lc.price_rotation_logs), 4, "hidden Cloudx Bidding logs must not be stored")
        lc.process_price_rotation_log(
            'Unity : [Ad,InterstitialBidding, Ascendx] Raise: {"act":"bid"}',
            "device-price",
        )
        _assert_equal(len(lc.price_rotation_logs), 4, "hidden Ascendx Bidding logs must not be stored")
    finally:
        lc.is_paused = original_paused
        lc.socketio.emit = original_emit
        lc.price_rotation_logs.clear()
        lc.price_rotation_logs.extend(original_rows)


def test_load_ads_provider_contract() -> None:
    original_platform = lc.active_platform
    original_recording_state = dict(lc.recording_states["LoadAdsExt"])
    original_emit = lc.socketio.emit
    try:
        lc.active_platform = "ios"
        lc.recording_states["LoadAdsExt"].update({"is_recording": True, "current_sheet": None})
        lc.load_ads_ext_events.clear()
        lc.unique_load_ads_ext.clear()
        emitted = []
        lc.socketio.emit = lambda event, payload: emitted.append((event, payload))

        ios_line = (
            'Appmetrica TrackAdRevenueEvent: '
            '{"AdRevenueValue":"0.01","Currency":"USD","AdNetwork":"unityads",'
            '"AdType":2,"Payload":{"ad_platform":"cloudx",'
            '"ad_network":"unityads","ad_format":"interstitial"},"Precision":null}'
        )
        lc.process_load_ads_ext_log(ios_line, "ios-load-ads")
        _assert_equal(len(lc.load_ads_ext_events), 1, "iOS AppMetrica Load Ads row was not recorded")
        _assert_equal(lc.load_ads_ext_events[0].get("provider"), "cloudx", "iOS provider was not extracted")

        # Provider is not unique: the same provider may appear for another network.
        second_ios_line = ios_line.replace('"AdNetwork":"unityads"', '"AdNetwork":"applovin"')
        second_ios_line = second_ios_line.replace('"ad_network":"unityads"', '"ad_network":"applovin"')
        lc.process_load_ads_ext_log(second_ios_line, "ios-load-ads")
        _assert_equal(len(lc.load_ads_ext_events), 2, "duplicate Provider must remain visible for another network")
        _assert(all(row.get("provider") == "cloudx" for row in lc.load_ads_ext_events), "provider values changed unexpectedly")

        # Same network/format from another provider is a distinct record.
        third_ios_line = ios_line.replace('"ad_platform":"cloudx"', '"ad_platform":"ascendx"')
        lc.process_load_ads_ext_log(third_ios_line, "ios-load-ads")
        _assert_equal(len(lc.load_ads_ext_events), 3, "different providers must not collide in dedup")
        _assert_equal(lc.load_ads_ext_events[-1].get("provider"), "ascendx", "second provider was not extracted")

        # Android's AdRevenue{...} parser must use the nested ad_platform too.
        lc.active_platform = "android"
        android_line = (
            'AdRevenue Received: AdRevenue{adType=rewarded,adNetwork=ironsource,'
            'payload={"ad_platform":"ironsource","ad_network":"ironsource",'
            '"ad_format":"rewarded"}}'
        )
        lc.process_load_ads_ext_log(android_line, "android-load-ads")
        android_rows = [row for row in lc.load_ads_ext_events if row.get("device_id") == "android-load-ads"]
        _assert_equal(len(android_rows), 1, "Android AppMetrica Load Ads row was not recorded")
        _assert_equal(android_rows[0].get("provider"), "ironsource", "Android provider was not extracted")

        rendered = lc.app.test_client().get("/").get_data(as_text=True)
        _assert('>Load Ads</button>' in rendered, "Load Ads tab label was not updated")
        _assert("Record Load Ads:" in rendered, "Load Ads recording label was not updated")
        _assert('>Provider</th>' in rendered, "Load Ads Provider column is missing")
        _assert('"provider": _normalize_load_ads_provider(provider)' in (ROOT / "Log_checker.py").read_text(encoding="utf-8"), "Sheet provider payload is missing")
    finally:
        lc.active_platform = original_platform
        lc.recording_states["LoadAdsExt"].clear()
        lc.recording_states["LoadAdsExt"].update(original_recording_state)
        lc.load_ads_ext_events.clear()
        lc.unique_load_ads_ext.clear()
        lc.socketio.emit = original_emit


def test_load_ads_max_contract() -> None:
    """MAX Load Ads rows must come from the exact MAX revenue callback fields."""
    original_platform = lc.active_platform
    original_recording_state = dict(lc.recording_states["LoadAdsExt"])
    original_emit = lc.socketio.emit
    original_send_to_sheet = lc.send_to_sheet
    original_rows = list(lc.load_ads_ext_events)
    original_unique = set(lc.unique_load_ads_ext)
    emitted = []
    sheet_rows = []
    try:
        lc.active_platform = "android"
        lc.recording_states["LoadAdsExt"].update({"is_recording": True, "current_sheet": "NG381"})
        lc.load_ads_ext_events.clear()
        lc.unique_load_ads_ext.clear()
        lc.socketio.emit = lambda event, payload: emitted.append((event, payload))
        lc.send_to_sheet = lambda *args: sheet_rows.append(args)

        samples = [
            "09-03 11:39:34.462\tcom.indiez.nonogram\t13912\t-\tD\tAppLovinSdk\t [MaxInterstitialAd] MaxAdRevenueListener.onAdRevenuePaid(ad=MediatedAd{thirdPartyAdPlacementId=945322059, adUnitId=test_mode_interstitial, format=INTER, networkName='Pangle'}), listener=null",
            "09-03 11:40:05.062\tcom.indiez.nonogram\t13912\t-\tD\tAppLovinSdk\t [MaxRewardedAd] MaxAdRevenueListener.onAdRevenuePaid(ad=MediatedAd{thirdPartyAdPlacementId=945322060, adUnitId=test_mode_rewarded_TIKTOK_BIDDING, format=REWARDED, networkName='Pangle'}), listener=null",
            "09-03 11:41:13.494\tcom.indiez.nonogram\t13912\t-\tD\tAppLovinSdk\t [MaxAdView] MaxAdRevenueListener.onAdRevenuePaid(ad=MediatedAd{thirdPartyAdPlacementId=980040498, adUnitId=test_mode_banner, format=BANNER, networkName='Pangle'}), listener=null",
            "09-03 11:41:23.051\tcom.indiez.nonogram\t13912\t-\tD\tAppLovinSdk\t [MaxAdView] MaxAdRevenueListener.onAdRevenuePaid(ad=MediatedAd{thirdPartyAdPlacementId=980040498, adUnitId=test_mode_banner, format=BANNER, networkName='Pangle'}), listener=null",
            "09-03 11:41:35.454\tcom.indiez.nonogram\t13912\t-\tD\tAppLovinSdk\t [MaxAdView] MaxAdRevenueListener.onAdRevenuePaid(ad=MediatedAd{thirdPartyAdPlacementId=982160535, adUnitId=test_mode_mrec, format=MREC, networkName='Pangle'}), listener=null",
        ]
        for line in samples:
            lc.process_load_ads_max_log(line, "device-max")

        rows = list(lc.load_ads_ext_events)
        _assert_equal(len(rows), 4, "MAX Load Ads must keep one row per network/format")
        _assert_equal({row.get("ad_format") for row in rows}, {"INTER", "REWARDED", "BANNER", "MREC"}, "MAX formats were not parsed from MediatedAd")
        _assert(all(row.get("provider") == "MAX" for row in rows), "MAX Load Ads provider was not recorded")
        _assert(all(row.get("ad_network") == "Pangle" for row in rows), "MAX networkName was not parsed from the raw log")
        _assert(all(lc.MAX_LOAD_ADS_REVENUE_KEYWORD in row.get("raw_log", "") for row in rows), "MAX raw callback log was not preserved")
        _assert_equal(len(emitted), 4, "MAX Load Ads rows were not emitted to the UI")
        _assert(all(event_name == "update_load_ads_ext" for event_name, _ in emitted), "MAX rows were emitted to the hidden Load Ads table")
        _assert_equal(len(sheet_rows), 4, "MAX Load Ads rows were not sent to the sheet")
        _assert(all(args[4:] == ("LoadAdsExt", "MAX") for args in sheet_rows), "MAX sheet payload used the wrong provider or tab")

        # Non-MAX diagnostics and disabled recording must never create rows.
        lc.process_load_ads_max_log("AppLovinSdk unrelated diagnostic format=BANNER networkName='Pangle'", "device-max")
        _assert_equal(len(lc.load_ads_ext_events), 4, "unrelated AppLovin diagnostics must be ignored")
        lc.recording_states["LoadAdsExt"]["is_recording"] = False
        lc.process_load_ads_max_log(samples[0], "device-max-2")
        _assert_equal(len(lc.load_ads_ext_events), 4, "MAX rows must respect Load Ads recording state")

        rendered = lc.app.test_client().get("/").get_data(as_text=True)
        _assert(rendered.count(">Provider</th>") >= 2, "Provider column must be present in both Load Ads tables")
    finally:
        lc.active_platform = original_platform
        lc.recording_states["LoadAdsExt"].clear()
        lc.recording_states["LoadAdsExt"].update(original_recording_state)
        lc.load_ads_ext_events.clear()
        lc.load_ads_ext_events.extend(original_rows)
        lc.unique_load_ads_ext.clear()
        lc.unique_load_ads_ext.update(original_unique)
        lc.socketio.emit = original_emit
        lc.send_to_sheet = original_send_to_sheet


def test_levelplay_impression_data_callback_contract() -> None:
    source_text = (ROOT / "Log_checker.py").read_text(encoding="utf-8", errors="ignore")
    for marker in (
        "LevelPlayAdImpressionEventMapper->Map",
        "LevelPlayImpressionData - {ad_unit_name}",
        'return "Banner"',
        'return "Rewarded"',
        'return "Interstitial"',
    ):
        _assert(marker in source_text, f"LevelPlay callback contract is missing: {marker}")

    original_paused = lc.is_paused
    original_emit = lc.socketio.emit
    original_rows = list(lc.callback_ad_logs)
    emitted = []
    try:
        lc.is_paused = False
        lc.callback_ad_logs.clear()
        lc.socketio.emit = lambda event, payload: emitted.append((event, payload))

        ad_units = {
            "Banner": {"mediationAdUnitName": "Banner", "adUnit": "banner", "adFormat": "banner"},
            "Rewarded": {"mediationAdUnitName": "Rewarded", "adUnit": "rewarded_video", "adFormat": "rewarded_video"},
            "Interstitial": {"mediationAdUnitName": "Interstitial", "adUnit": "interstitial", "adFormat": "interstitial"},
        }
        for expected_label, impression_data in ad_units.items():
            nested_json = json.dumps(impression_data, separators=(",", ":"))
            outer_json = json.dumps({"LevelPlayImpressionData": nested_json}, separators=(",", ":"))
            line = f"Unity LevelPlayAdImpressionEventMapper->Map: {outer_json}"
            lc.process_callback_and_ad_event_log(line, "device-levelplay")

        rows = list(lc.callback_ad_logs)
        _assert_equal(len(rows), 3, "LevelPlay impression mapper should add one callback row per ad unit")
        _assert_equal(
            [row["event_name"] for row in rows],
            ["LevelPlayImpressionData - Banner", "LevelPlayImpressionData - Rewarded", "LevelPlayImpressionData - Interstitial"],
            "LevelPlay impression callback names do not include the ad unit",
        )
        _assert(all(row["type"] == "Callback Levelplay" for row in rows), "LevelPlay impression mapper rows must use the LevelPlay callback filter")
        _assert(all("LevelPlayAdImpressionEventMapper->Map" in row["raw_log"] for row in rows), "LevelPlay impression raw logs were not preserved")
        _assert_equal(len(emitted), 3, "LevelPlay impression callbacks were not emitted to the Callback tab")

        # Existing LevelPlay listener callbacks must keep their old display name.
        lc.process_callback_and_ad_event_log("Unity LevelPlayInterstitialAdListener", "device-levelplay")
        _assert_equal(lc.callback_ad_logs[-1]["event_name"], "Interstitial", "existing LevelPlay listener name changed")
    finally:
        lc.is_paused = original_paused
        lc.callback_ad_logs.clear()
        lc.callback_ad_logs.extend(original_rows)
        lc.socketio.emit = original_emit


def test_ascendx_cloudx_callback_contract() -> None:
    source_text = (ROOT / "Log_checker.py").read_text(encoding="utf-8", errors="ignore")
    required_markers = (
        "_OnAdLoadedEvent",
        "_OnAdImpressionEvent",
        "_OnAdDisplayedEvent",
        "_OnAdClickedEvent",
        "_OnAdClosedEvent",
        "_OnAdReceivedRewardEvent",
        "OnAdLoadFailed",
        "OnAdLoadSuccess",
        "OnAdRevenuePaid",
        "OnAdClicked",
        "OnAdShowSuccess",
        "OnAdRewarded",
        "OnAdHidden",
        'value="ascendx_callback"',
        'value="cloudx_callback"',
        "Callback Ascendx",
        "Callback Cloudx",
        "Callback Levelplay",
        'id="callbackTypeAll" name="callbackType" type="checkbox"',
        'id="callbackTypeCallback" name="callbackType" type="checkbox"',
        'id="callbackTypeAscendx" name="callbackType" type="checkbox"',
        'id="callbackTypeCloudx" name="callbackType" type="checkbox"',
        'id="callbackTypeGadsme" name="callbackType" type="checkbox"',
        'id="callbackTypeAdverty5" name="callbackType" type="checkbox"',
        "selectedTypeFilters",
        "bindCallbackCheckboxGroup",
        'details-cell text-sm align-top"><div class="max-h-64 overflow-auto">${res.details}',
        'log-cell text-xs font-normal text-gray-600 align-top"><div class="max-h-64 overflow-auto whitespace-pre-wrap break-all">',
    )
    for marker in required_markers:
        _assert(marker in source_text, f"AscendX/CloudX callback contract is missing: {marker}")

    original_paused = lc.is_paused
    original_emit = lc.socketio.emit
    original_rows = list(lc.callback_ad_logs)
    emitted = []
    try:
        lc.is_paused = False
        lc.callback_ad_logs.clear()
        lc.socketio.emit = lambda event, payload: emitted.append((event, payload))

        for callback_event in lc.ASCENDX_CALLBACK_EVENTS:
            ad_unit = "Reward" if callback_event == "_OnAdReceivedRewardEvent" else "Interstitial"
            line = f"Unity [Ad,AscendX,{ad_unit}] {callback_event}"
            lc.process_callback_and_ad_event_log(line, "device-callback")

        for callback_event in lc.CLOUDX_CALLBACK_EVENTS:
            ad_unit = "Reward" if callback_event == "OnAdRewarded" else "Interstitial"
            line = f"Unity [Ad,CloudX,{ad_unit}] {callback_event}"
            lc.process_callback_and_ad_event_log(line, "device-callback")

        rows = list(lc.callback_ad_logs)
        _assert_equal(len(rows), len(lc.ASCENDX_CALLBACK_EVENTS) + len(lc.CLOUDX_CALLBACK_EVENTS), "provider callback row count changed")
        _assert_equal(
            [row["type"] for row in rows[:len(lc.ASCENDX_CALLBACK_EVENTS)]],
            ["Callback Ascendx"] * len(lc.ASCENDX_CALLBACK_EVENTS),
            "AscendX callback rows use the wrong filter type",
        )
        _assert_equal(
            [row["type"] for row in rows[len(lc.ASCENDX_CALLBACK_EVENTS):]],
            ["Callback Cloudx"] * len(lc.CLOUDX_CALLBACK_EVENTS),
            "CloudX callback rows use the wrong filter type",
        )
        _assert_equal(rows[0]["event_name"], "_OnAdLoadedEvent - Interstitial", "AscendX ad unit name was not added")
        _assert_equal(rows[5]["event_name"], "_OnAdReceivedRewardEvent - Reward", "AscendX reward name was not added")
        _assert_equal(rows[6]["event_name"], "OnAdLoadFailed - Interstitial", "CloudX ad unit name was not added")
        _assert_equal(rows[11]["event_name"], "OnAdRewarded - Reward", "CloudX reward name was not added")
        _assert(all(row["raw_log"].startswith("Unity [Ad,") for row in rows), "provider callback raw logs were not preserved")
        _assert_equal(len(emitted), len(rows), "provider callbacks were not emitted to the Callback tab")

        rendered = lc.app.test_client().get("/").get_data(as_text=True)
        for marker in (
            'id="callbackTypeAscendx"',
            'id="callbackTypeCloudx"',
            'value="ascendx_callback"',
            'value="cloudx_callback"',
        ):
            _assert(marker in rendered, f"rendered callback filter is missing: {marker}")
        ordered_filters = [
            rendered.index('id="callbackTypeCallback"'),
            rendered.index('id="callbackTypeAscendx"'),
            rendered.index('id="callbackTypeCloudx"'),
            rendered.index('id="callbackTypeGadsme"'),
            rendered.index('id="callbackTypeAdverty5"'),
        ]
        _assert_equal(ordered_filters, sorted(ordered_filters), "callback filters are not grouped in the requested order")
    finally:
        lc.is_paused = original_paused
        lc.callback_ad_logs.clear()
        lc.callback_ad_logs.extend(original_rows)
        lc.socketio.emit = original_emit


def test_release_payload_sync() -> None:
    source_text = (ROOT / "Log_checker.py").read_text(encoding="utf-8", errors="ignore")
    manifest_path = ROOT / "Updates_2_5" / "remote_manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    log_item = next(item for item in manifest["files"] if item.get("path") == "Log_checker.py")
    _assert_equal(_sha256_file(ROOT / "Log_checker.py"), log_item["sha256"], "source/release Log_checker.py drift detected")
    _assert("compat_sha256" not in log_item, "canonical v2.5 payload must not select the stale compatibility source")
    remote_item = next(item for item in manifest["files"] if item.get("path") == "remote_update.py")
    _assert("compat_sha256" not in remote_item, "canonical updater payload must not select the stale compatibility source")
    _assert_equal(manifest["version"], CURRENT_RELEASE_VERSION, "v2.5 release manifest version changed")

    markers = {
        "release_badge": r"v2\.5\.0\((\d+)\)",
        "html_title": rf"<title>Event Inspector v2\.5\.0\({CURRENT_RELEASE_BUILD}\)</title>",
        "socket_fallback": r"typeof window\.io === 'function'",
        "brightsdk_tab": r"switchTab\('BrightSDK'\)",
        "tm_ios_package": r'data-ios-value="([^"]+)"\s+data-ios-label="TM - ([^"]+)"',
        "check_update_call": r"result = remote_update\.check_for_updates\(\)",
    }

    for label, pattern in markers.items():
        source_match = re.search(pattern, source_text)
        _assert(source_match is not None, f"source missing {label}")

    for remote_contract in (
        "def _fetch_sdk_check_presets(force_remote=False):",
        "refresh_requested = request.args.get(\"refresh\"",
        "force_remote=refresh_requested",
        "const refreshQuery = force ? '&refresh=1'",
    ):
        _assert(remote_contract in source_text, f"canonical SDK preset refresh is missing: {remote_contract}")


def test_update_candidate_does_not_downgrade() -> None:
    candidates = [
        {"update_dir": "/tmp/build52", "build": 52, "source": "older"},
        {"update_dir": "/tmp/build53", "build": 53, "source": "current"},
        {"update_dir": "/tmp/build56", "build": 56, "source": "release"},
        {"update_dir": "/tmp/build54", "build": 54, "source": "older-release"},
    ]
    selected = desktop._select_prepared_update_candidate(candidates, bundled_build=53)
    _assert_equal(selected["build"], 56, "bundled build 53 must select the newest prepared update")
    _assert_equal(
        desktop._select_prepared_update_candidate(candidates[:2], bundled_build=53),
        None,
        "a prepared payload at or below the bundled build must be ignored",
    )
    _assert_equal(
        desktop._select_prepared_update_candidate(candidates[:1], bundled_build=None)["build"],
        52,
        "a client without a detected bundled build must keep update compatibility",
    )
    equal_build_candidate = [{"update_dir": "/tmp/build53-cache", "build": 53, "source": "same_build_cache"}]
    _assert_equal(
        desktop._select_prepared_update_candidate(equal_build_candidate, bundled_build=53),
        None,
        "a same-build prepared payload must not override the bundled source",
    )
    compatibility_candidate = [{"update_dir": "/tmp/build56", "build": 56, "source": "channel_state"}]
    _assert_equal(
        desktop._select_prepared_update_candidate(compatibility_candidate, bundled_build=53)["build"],
        CURRENT_RELEASE_BUILD,
        "a newer prepared payload must be accepted over the bundled build",
    )


def test_services_checker_gradle_mapping_contract() -> None:
    service_source = (ROOT / "services_checker" / "app.py").read_text(encoding="utf-8")
    gradle_mapping = json.loads((ROOT / "services_checker" / "gradle_lib_mapping.json").read_text(encoding="utf-8"))
    gradle_presets = json.loads((ROOT / "services_checker" / "gradle_check_presets.json").read_text(encoding="utf-8"))
    podfile_presets = json.loads((ROOT / "services_checker" / "podfile_check_presets.json").read_text(encoding="utf-8"))
    c191_gradle_lines = gradle_presets.get("C-191-Android", {}).get("lines") or []
    _assert("Voodoo (ADN) Adapter\t5.7.0" in c191_gradle_lines, "C-191 Voodoo adapter version changed")
    _assert("Voodoo (ADN) SDK\t4.29.2" in c191_gradle_lines, "C-191 Voodoo SDK version changed")
    _assert("LINE Ads adapter\t5.4.0" not in c191_gradle_lines, "C-191 Gradle preset still contains LINE Ads adapter")
    _assert("LINE Ads native\t3.1.1" not in c191_gradle_lines, "C-191 Gradle preset still contains LINE Ads native")
    expected_c191_max_lines = {
        "Mobilefuse - MAX\t1.12.0.0",
        "Ascendx - MAX\t1.11.1",
        "Yeahmobi/ Maticoo - MAX\t2.0.7.0",
        "TaurusX - MAX\t1.16.3.1",
    }
    _assert(
        expected_c191_max_lines.issubset(set(c191_gradle_lines)),
        "C-191 MAX Gradle preset entries are incomplete",
    )
    expected_c191_max_mapping = {
        "Mobilefuse - MAX": "com.applovin.mediation:mobilefuse-adapter",
        "Ascendx - MAX": "com.knorex:ascendx-mobile-sdk-max-custom-adapter",
        "Yeahmobi/ Maticoo - MAX": "io.github.maticooads:maticoo-adapter-max",
        "TaurusX - MAX": "com.applovin.mediation:taurusXAdapters",
        "Prado - MAX": "co.prado.sdk:prado-android-applovin-adapter",
    }
    for library_name, artifact_id in expected_c191_max_mapping.items():
        _assert_equal(
            gradle_mapping.get(library_name),
            artifact_id,
            f"C-191 MAX Gradle mapping changed for {library_name}",
        )
    _assert_equal(
        podfile_presets.get("C-180-iOS", {}).get("lines") or [],
        [
            "Kidoz Adapter\t2.2.0",
            "Kidoz SDK\t10.1.5",
            "Yeahmobi/ Maticoo Adapter\t2.2.0",
            "Yeahmobi/ Maticoo SDK\t2.2.0",
            "TaurusX SDK\t1.18.1",
            "TaurusX Adapter\t1.18.1",
            "Odeeo SDK\t3.10.0",
            "AppMetrica Analytics\t6.4.0",
            "Google UMP SDK\t3.1.0",
            "Ascendx adapter\t0.9.0",
            "Voodoo Adapter\t5.2.0.0",
            "Adjust/AdjustGoogleOd\t5.6.2",
        ],
        "C-180 iOS Podfile preset changed",
    )
    _assert("gradle_lib_mapping.json" in service_source, "Services Checker Gradle mapping file is missing")
    _assert("def _load_gradle_lib_mapping(" in service_source, "Services Checker Gradle mapping loader is missing")
    _assert("GRADLE_LIB_MAPPING =" not in service_source, "Gradle mapping must not remain hardcoded in app.py")
    _assert_equal(
        service_source.count("def scan_gradle_for_versions("),
        1,
        "Services Checker must have one authoritative Gradle scanner",
    )
    _assert("import json" in service_source.split("# --- Flask Routes ---", 1)[0], "Services Checker preset JSON import is missing")
    _assert_equal(
        service_source.count("def _load_build_check_presets("),
        1,
        "Services Checker preset loader must be defined in the active source block",
    )
    for loader_name in ("_load_gradle_check_presets", "_load_podfile_check_presets"):
        _assert_equal(service_source.count(f"def {loader_name}("), 1, f"Services Checker loader missing: {loader_name}")
    for preset_file in (
        "apk_check_presets.json",
        "gradle_check_presets.json",
        "podfile_check_presets.json",
        "manifest_check_presets.json",
    ):
        _assert(preset_file in service_source, f"Services Checker preset file is missing: {preset_file}")
    for remote_contract in (
        "import requests",
        "SERVICES_CHECKER_PRESET_BRANCHES",
        "SERVICES_CHECKER_GITHUB_API_URL",
        "def _fetch_services_checker_revision(",
        "SERVICES_CHECKER_COMMIT_SHA_PATTERN",
        "SERVICES_CHECKER_PRESET_FILENAMES",
        "SERVICES_CHECKER_REMOTE_DATA_FILENAMES",
        "def _remote_preset_urls(",
        "requests.get(",
        "response.raise_for_status()",
        "eventinspector_refresh={cache_bust}",
        '"Accept-Encoding": "identity"',
        "def _refresh_remote_preset_files(",
        "_remote_preset_refresh_lock = threading.Lock()",
        "with _remote_preset_refresh_lock:",
        "explicit_refresh = request.args.get(\"refresh\"",
        "refresh_requested = True",
        "forceRemote = false",
        "refreshQuery = forceRemote ? '&refresh=1'",
        "await loadBuildCheckPresets(true)",
        "function presetMatchesControl(control, preset)",
        "function syncBuildCheckPreset(control, select, clearManual = false)",
        "function restoreSelectedBuildCheckPreset(selectId)",
        "area.replaceChildren(messageDiv)",
        "let buildCheckPresetRequestSerial = 0;",
        "const requestSerial = ++buildCheckPresetRequestSerial;",
        "if (requestSerial !== buildCheckPresetRequestSerial) return;",
        "syncBuildCheckPreset(control, select)",
        "select.onchange = () => syncBuildCheckPreset(control, select, true)",
        "if (!response.ok) throw new Error('preset_request_failed:' + response.status)",
    ):
        _assert(remote_contract in service_source, f"Live Services Checker preset refresh is missing: {remote_contract}")
    _assert("'gradle_presets': gradle_presets" in service_source, "Gradle preset API group is missing")
    _assert("'podfile_presets': podfile_presets" in service_source, "Podfile preset API group is missing")
    required_mappings = {
        "Adjust Meta Referrer": "com.adjust.sdk:adjust-android-meta-referrer",
        "Adjust Samsung Referrer": "com.adjust.sdk:adjust-android-samsung-referrer",
        "Adjust Vivo Referrer": "com.adjust.sdk:adjust-android-vivo-referrer",
        "Adjust Xiaomi Referrer": "com.adjust.sdk:adjust-android-xiaomi-referrer",
        "Xiaomi Install Referrer": "com.miui.referrer:homereferrer",
        "Samsung Install Referrer": "store.galaxy.samsung.installreferrer:samsung_galaxystore_install_referrer",
    }
    for mapping_name, mapping_artifact in required_mappings.items():
        _assert_equal(
            gradle_mapping.get(mapping_name),
            mapping_artifact,
            f"Services Checker mapping changed or missing: {mapping_name}",
        )
    _assert("LINE Ads adapter" not in gradle_mapping, "LINE Ads mapping must remain removed")
    for control_id in (
        'id="apk-build-check-preset"',
        'id="gradle-build-check-preset"',
        'id="podfile-build-check-preset"',
        'id="manifest-check-preset"',
        'id="reload-apk-build-check-presets"',
        'id="reload-gradle-build-check-presets"',
        'id="reload-podfile-build-check-presets"',
        'id="reload-manifest-check-presets"',
    ):
        _assert(service_source.count(control_id) == 1, f"Services Checker preset control must exist once: {control_id}")
    _assert('platform: \'android\'' in service_source, "Android build preset filtering is missing")
    _assert('platform: \'ios\'' in service_source, "iOS build preset filtering is missing")
    _assert('presetMatchesControl(control, preset)' in service_source, "build preset platform filtering is missing")
    _assert('restoreSelectedBuildCheckPreset(\'apk-build-check-preset\')' in service_source, "APK preset must survive tab reset")
    _assert('restoreSelectedBuildCheckPreset(\'gradle-build-check-preset\')' in service_source, "Gradle preset must survive tab reset")
    _assert('restoreSelectedBuildCheckPreset(\'podfile-build-check-preset\')' in service_source, "Podfile preset must survive tab reset")
    _assert('SERVICES_CHECKER_PRESET_BRANCH = "main"' in service_source, "main must be the canonical preset branch")
    _assert(
        'SERVICES_CHECKER_PRESET_BRANCHES = (SERVICES_CHECKER_PRESET_BRANCH,)' in service_source,
        "preset refresh must not fall back to a stale release branch",
    )
    _assert("'refreshed_sources': refreshed_sources" in service_source, "preset reload must report the GitHub branch used")
    _assert('id="build-check-preset"' not in service_source, "obsolete global build preset selector must be removed")


def test_services_checker_live_preset_refresh_after_restart() -> None:
    """Catch the stale-cache regression with two fresh service imports."""
    import requests as service_requests

    service_path = ROOT / "services_checker" / "app.py"
    original_get = service_requests.get
    original_cache_dir = os.environ.get("EVENTINSPECTOR_PRESET_CACHE_DIR")
    revision = {"value": "old"}
    calls = []

    class FakeResponse:
        def __init__(self, payload):
            self.content = json.dumps(payload, separators=(",", ":")).encode("utf-8")

        def raise_for_status(self):
            return None

        def json(self):
            return json.loads(self.content.decode("utf-8"))

    def fake_get(url, **_kwargs):
        calls.append(url)
        if "api.github.com/repos/" in url:
            return FakeResponse({"sha": "a" * 40})
        _assert(
            "/services_checker/" in url
            and (
                "/main/services_checker/" in url
                or "/" + ("a" * 40) + "/services_checker/" in url
            ),
            f"preset refresh used a non-main source: {url}",
        )
        filename = url.split("/services_checker/", 1)[1].split("?", 1)[0]
        if filename == "gradle_lib_mapping.json":
            return FakeResponse({"Harness Library": f"com.example:harness-{revision['value']}"})
        payload = {
            "C-191-Android": {
                "platform": "ios" if filename == "podfile_check_presets.json" else "android",
                "lines": [f"{filename}:{revision['value']}"],
            }
        }
        return FakeResponse(payload)

    module_names = []
    try:
        with tempfile.TemporaryDirectory(prefix="eventinspector_services_presets_") as cache_dir:
            os.environ["EVENTINSPECTOR_PRESET_CACHE_DIR"] = cache_dir
            service_requests.get = fake_get

            def load_service_module(name):
                spec = importlib.util.spec_from_file_location(name, service_path)
                _assert(spec is not None and spec.loader is not None, "Services Checker module could not be loaded")
                module = importlib.util.module_from_spec(spec)
                sys.modules[name] = module
                module_names.append(name)
                spec.loader.exec_module(module)
                return module

            first = load_service_module("eventinspector_services_checker_refresh_one")
            first_response = first.app.test_client().get("/api/build-check-presets?ts=1")
            _assert_equal(first_response.status_code, 200, "first Services Checker preset request failed")
            first_data = first_response.get_json()
            _assert_equal(
                first_data["manifest_presets"]["C-191-Android"]["lines"],
                ["manifest_check_presets.json:old"],
                "first Service Checker start did not load the remote preset",
            )
            _assert_equal(
                first._load_gradle_lib_mapping().get("Harness Library"),
                "com.example:harness-old",
                "first Services Checker start did not load the remote Gradle mapping",
            )

            revision["value"] = "new"
            second = load_service_module("eventinspector_services_checker_refresh_two")
            second_response = second.app.test_client().get("/api/build-check-presets?ts=2")
            _assert_equal(second_response.status_code, 200, "restart Services Checker preset request failed")
            second_data = second_response.get_json()
            _assert_equal(
                second_data["manifest_presets"]["C-191-Android"]["lines"],
                ["manifest_check_presets.json:new"],
                "Services Checker restart kept the stale per-user preset cache",
            )
            _assert_equal(
                second._load_gradle_lib_mapping().get("Harness Library"),
                "com.example:harness-new",
                "Services Checker restart kept the stale Gradle mapping cache",
            )
            _assert_equal(
                set(second_data["refreshed_files"]),
                {
                    "apk_check_presets.json",
                    "gradle_check_presets.json",
                    "gradle_lib_mapping.json",
                    "podfile_check_presets.json",
                    "manifest_check_presets.json",
                },
                "restart did not refresh all Services Checker preset files",
            )
            preset_calls = [url for url in calls if "/services_checker/" in url]
            _assert(
                preset_calls
                and all("/" + ("a" * 40) + "/services_checker/" in url for url in preset_calls),
                "preset source was not immutable main commit",
            )
    finally:
        service_requests.get = original_get
        for name in module_names:
            sys.modules.pop(name, None)
        if original_cache_dir is None:
            os.environ.pop("EVENTINSPECTOR_PRESET_CACHE_DIR", None)
        else:
            os.environ["EVENTINSPECTOR_PRESET_CACHE_DIR"] = original_cache_dir


def test_services_checker_git_value_reload_from_real_commit() -> None:
    """Require a real Git commit change to reach the same client process.

    A mocked requests.get can prove only that code paths exist. This test
    creates a temporary Git repository, serves its checked-in files over HTTP,
    changes a committed preset value, and verifies that the already imported
    Services Checker returns the new value after Reload.
    """
    service_path = ROOT / "services_checker" / "app.py"
    preset_filenames = (
        "apk_check_presets.json",
        "gradle_check_presets.json",
        "gradle_lib_mapping.json",
        "podfile_check_presets.json",
        "manifest_check_presets.json",
    )
    original_cache_dir = os.environ.get("EVENTINSPECTOR_PRESET_CACHE_DIR")
    module_name = "eventinspector_services_checker_real_git_reload"

    def run_git(repo, *args):
        return subprocess.run(
            ["git", *args],
            cwd=repo,
            check=True,
            capture_output=True,
            text=True,
        ).stdout.strip()

    class QuietHandler(http.server.SimpleHTTPRequestHandler):
        def log_message(self, _format, *_args):
            return

    server = None
    try:
        with tempfile.TemporaryDirectory(prefix="eventinspector_git_presets_") as temp_dir:
            repo = Path(temp_dir) / "remote"
            cache_dir = Path(temp_dir) / "client-cache"
            (repo / "services_checker").mkdir(parents=True)
            run_git(repo, "init", "-q", "-b", "main")
            run_git(repo, "config", "user.email", "harness@example.invalid")
            run_git(repo, "config", "user.name", "EventInspector Harness")

            def write_presets(revision):
                for filename in preset_filenames:
                    if filename == "gradle_lib_mapping.json":
                        payload = {"Harness Library": f"com.example:harness-{revision}"}
                    else:
                        payload = {
                            "C-191-Android": {
                                "platform": "ios" if filename == "podfile_check_presets.json" else "android",
                                "lines": [f"{filename}:{revision}"],
                                "harness_marker": revision,
                            }
                        }
                    (repo / "services_checker" / filename).write_text(
                        json.dumps(payload, sort_keys=True),
                        encoding="utf-8",
                    )

            write_presets("commit-one")
            run_git(repo, "add", "services_checker")
            run_git(repo, "commit", "-qm", "preset commit one")
            first_commit = run_git(repo, "rev-parse", "HEAD")

            server = http.server.ThreadingHTTPServer(
                ("127.0.0.1", 0),
                functools.partial(QuietHandler, directory=str(repo)),
            )
            threading.Thread(target=server.serve_forever, daemon=True).start()
            base_url = f"http://127.0.0.1:{server.server_port}/services_checker"

            os.environ["EVENTINSPECTOR_PRESET_CACHE_DIR"] = str(cache_dir)
            spec = importlib.util.spec_from_file_location(module_name, service_path)
            _assert(spec is not None and spec.loader is not None, "Services Checker module could not be loaded")
            module = importlib.util.module_from_spec(spec)
            sys.modules[module_name] = module
            spec.loader.exec_module(module)
            module._fetch_services_checker_revision = lambda: None
            module._remote_preset_urls = lambda filename, _revision=None: (
                ("git-fixture", f"{base_url}/{filename}")
                for _ in (0,)
            )

            client = module.app.test_client()
            first_response = client.get("/api/build-check-presets?refresh=1&ts=commit-one")
            _assert_equal(first_response.status_code, 200, "real Git preset request failed")
            first_data = first_response.get_json()
            _assert_equal(
                first_data["manifest_presets"]["C-191-Android"]["harness_marker"],
                "commit-one",
                "client did not receive the first committed Git value",
            )
            _assert_equal(
                module._load_gradle_lib_mapping().get("Harness Library"),
                "com.example:harness-commit-one",
                "client did not receive the first committed Gradle mapping",
            )
            first_digest = first_data["loaded_preset_files"]["manifest_check_presets.json"]["sha256"]
            _assert_equal(
                first_data["refreshed_sources"]["manifest_check_presets.json"],
                "git-fixture",
                "wrong Git source",
            )

            write_presets("commit-two")
            run_git(repo, "add", "services_checker")
            run_git(repo, "commit", "-qm", "preset commit two")
            second_commit = run_git(repo, "rev-parse", "HEAD")
            _assert(first_commit != second_commit, "harness Git fixture did not create a new commit")

            second_response = client.get("/api/build-check-presets?refresh=1&ts=commit-two")
            _assert_equal(second_response.status_code, 200, "real Git reload request failed")
            second_data = second_response.get_json()
            _assert_equal(
                second_data["manifest_presets"]["C-191-Android"]["harness_marker"],
                "commit-two",
                "client kept the old preset value after a real Git commit changed",
            )
            _assert_equal(
                module._load_gradle_lib_mapping().get("Harness Library"),
                "com.example:harness-commit-two",
                "client kept the old Gradle mapping after a real Git commit changed",
            )
            second_digest = second_data["loaded_preset_files"]["manifest_check_presets.json"]["sha256"]
            _assert(first_digest != second_digest, "client preset hash did not change after Git commit")
            _assert_equal(
                set(second_data["refreshed_files"]),
                set(preset_filenames),
                "Reload did not download every Git-backed Service Checker preset",
            )
            _assert_equal(
                second_data["preset_revision"] != first_data["preset_revision"],
                True,
                "client revision marker did not change after Git commit",
            )
    finally:
        if server is not None:
            server.shutdown()
            server.server_close()
        sys.modules.pop(module_name, None)
        if original_cache_dir is None:
            os.environ.pop("EVENTINSPECTOR_PRESET_CACHE_DIR", None)
        else:
            os.environ["EVENTINSPECTOR_PRESET_CACHE_DIR"] = original_cache_dir


def test_sdk_preset_git_value_reload_from_real_commit() -> None:
    """Require a real Git commit change to reach the SDK preset client.

    This intentionally uses the real SDK endpoint and HTTP transport. The
    temporary repository gives the test two actual committed revisions without
    changing the production repository.
    """
    preset_filename = "sdk_check_presets.json"
    original_urls = list(lc.SDK_CHECK_PRESETS_REMOTE_URLS)

    def run_git(repo, *args):
        return subprocess.run(
            ["git", *args],
            cwd=repo,
            check=True,
            capture_output=True,
            text=True,
        ).stdout.strip()

    class QuietHandler(http.server.SimpleHTTPRequestHandler):
        def log_message(self, _format, *_args):
            return

    server = None
    try:
        with tempfile.TemporaryDirectory(prefix="eventinspector_sdk_git_presets_") as temp_dir:
            repo = Path(temp_dir) / "remote"
            repo.mkdir()
            run_git(repo, "init", "-q", "-b", "main")
            run_git(repo, "config", "user.email", "harness@example.invalid")
            run_git(repo, "config", "user.name", "EventInspector Harness")

            def write_preset(revision):
                payload = {
                    "C-191-Android": {
                        "platform": "android",
                        "lines": [f"Ads Network\tHarness-{revision}"],
                        "harness_marker": revision,
                    }
                }
                (repo / preset_filename).write_text(
                    json.dumps(payload, sort_keys=True),
                    encoding="utf-8",
                )

            write_preset("commit-one")
            run_git(repo, "add", preset_filename)
            run_git(repo, "commit", "-qm", "SDK preset commit one")
            first_commit = run_git(repo, "rev-parse", "HEAD")

            server = http.server.ThreadingHTTPServer(
                ("127.0.0.1", 0),
                functools.partial(QuietHandler, directory=str(repo)),
            )
            threading.Thread(target=server.serve_forever, daemon=True).start()
            lc.SDK_CHECK_PRESETS_REMOTE_URLS[:] = [
                f"http://127.0.0.1:{server.server_port}/{preset_filename}"
            ]

            client = lc.app.test_client()
            first_response = client.get("/api/sdk-check-presets?refresh=1&ts=commit-one")
            _assert_equal(first_response.status_code, 200, "real SDK Git preset request failed")
            first_data = first_response.get_json()
            _assert_equal(first_data.get("source"), "github", "SDK preset did not come from Git HTTP")
            _assert_equal(
                first_data["presets"]["C-191-Android"]["lines"],
                ["Ads Network\tHarness-commit-one"],
                "client did not receive the first committed SDK preset value",
            )

            write_preset("commit-two")
            run_git(repo, "add", preset_filename)
            run_git(repo, "commit", "-qm", "SDK preset commit two")
            second_commit = run_git(repo, "rev-parse", "HEAD")
            _assert(first_commit != second_commit, "harness SDK Git fixture did not create a new commit")

            second_response = client.get("/api/sdk-check-presets?refresh=1&ts=commit-two")
            _assert_equal(second_response.status_code, 200, "real SDK Git reload request failed")
            second_data = second_response.get_json()
            _assert_equal(
                second_data["presets"]["C-191-Android"]["lines"],
                ["Ads Network\tHarness-commit-two"],
                "client kept the old SDK preset value after a real Git commit changed",
            )
    finally:
        if server is not None:
            server.shutdown()
            server.server_close()
        lc.SDK_CHECK_PRESETS_REMOTE_URLS[:] = original_urls


def test_canonical_manifest_build_contract() -> None:
    manifest_path = ROOT / "Updates_2_5" / "remote_manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    _assert_equal(manifest.get("version"), CURRENT_RELEASE_VERSION, "canonical manifest must target the current build")
    for item in manifest.get("files") or []:
        _assert("compat_sha256" not in item, "canonical manifest must not carry a compatibility payload hash")
        for url in [item.get("url"), *(item.get("urls") or [])]:
            _assert(
                "/main/" in str(url) or "@main/" in str(url),
                f"canonical manifest URL must use the main payload: {url}",
            )


def test_update_flow_canonical_v25() -> None:
    import remote_update as updater

    manifest_path = ROOT / "Updates_2_5" / "remote_manifest.json"
    manifest_bytes = manifest_path.read_bytes()
    manifest = json.loads(manifest_bytes)
    payloads = {
        str(item["path"]): (ROOT / str(item["path"])).read_bytes()
        for item in manifest.get("files") or []
        if item.get("path")
    }
    original_home = os.environ.get("HOME")
    original_bundle_build = os.environ.get("EVENTINSPECTOR_BUNDLED_BUILD")
    original_bundle_source = os.environ.get("EVENTINSPECTOR_BUNDLED_BUILD_SOURCE")
    original_download_first = updater._download_first
    original_download_verified = updater._download_verified
    try:
        with tempfile.TemporaryDirectory(prefix="eventinspector_harness_") as temp_home:
            os.environ["HOME"] = temp_home
            os.environ["EVENTINSPECTOR_BUNDLED_BUILD"] = "1"
            os.environ["EVENTINSPECTOR_BUNDLED_BUILD_SOURCE"] = "detected"

            def fake_download_first(_urls, _timeout):
                return manifest_bytes, "harness://remote_manifest.json"

            def fake_download_verified(urls, _timeout, _expected_sha256=""):
                rel_path = next((path for path in payloads if any(url.endswith("/" + path) for url in urls)), None)
                _assert(rel_path is not None, f"unexpected update payload URL list: {urls}")
                return payloads[rel_path], f"harness://{rel_path}"

            updater._download_first = fake_download_first
            updater._download_verified = fake_download_verified

            first = updater.check_for_updates(force_refresh=True)
            _assert_equal(first.get("status"), "updated", "v2.5 client must prepare the release payload")
            _assert_equal(first.get("version"), CURRENT_RELEASE_VERSION, "prepared v2.5 payload version mismatch")
            prepared = updater.get_prepared_update_info()
            _assert_equal(prepared.get("build"), CURRENT_RELEASE_BUILD, "prepared v2.5 payload build mismatch")
            _assert(os.path.exists(os.path.join(prepared["update_dir"], "Log_checker.py")), "prepared Log_checker.py missing")
            _assert(os.path.exists(os.path.join(prepared["update_dir"], "sdk_check_presets.json")), "prepared SDK preset file missing")
            _assert(
                os.path.exists(os.path.join(prepared["update_dir"], "services_checker", "bundletool-all-1.18.1.jar")),
                "prepared bundletool asset missing from the remote update payload",
            )
            _assert(
                os.path.exists(os.path.join(prepared["update_dir"], "services_checker", "manifest_check_presets.json")),
                "prepared manifest preset payload missing",
            )
            _assert(
                os.path.exists(os.path.join(prepared["update_dir"], "services_checker", "gradle_lib_mapping.json")),
                "prepared Gradle mapping payload missing",
            )
            _assert(
                os.path.exists(os.path.join(prepared["update_dir"], "services_checker", "axml_fallback.py")),
                "prepared dependency-free AXML fallback payload missing",
            )

            selected = desktop._select_prepared_update_candidate(
                [{
                    "update_dir": prepared["update_dir"],
                    "build": prepared["build"],
                    "source": "harness-prepared-release",
                }],
                bundled_build=50,
            )
            _assert(selected is not None, f"bundle build 50 must accept the prepared build {CURRENT_RELEASE_BUILD} payload")
            _assert_equal(selected.get("build"), CURRENT_RELEASE_BUILD, f"prepared build {CURRENT_RELEASE_BUILD} was not selected")
            updated_source = Path(prepared["update_dir"], "Log_checker.py").read_text(
                encoding="utf-8", errors="ignore"
            )
            _assert("Default Ad Events" in updated_source, "prepared payload does not contain Default Ad Events")
            _assert(
                f"v2.5.0({CURRENT_RELEASE_BUILD})" in updated_source,
                "prepared payload carries the wrong UI build marker",
            )

            second = updater.check_for_updates()
            _assert_equal(second.get("status"), "up_to_date", "same v2.5 payload must not download repeatedly")

    finally:
        updater._download_first = original_download_first
        updater._download_verified = original_download_verified
        if original_home is None:
            os.environ.pop("HOME", None)
        else:
            os.environ["HOME"] = original_home
        if original_bundle_build is None:
            os.environ.pop("EVENTINSPECTOR_BUNDLED_BUILD", None)
        else:
            os.environ["EVENTINSPECTOR_BUNDLED_BUILD"] = original_bundle_build
        if original_bundle_source is None:
            os.environ.pop("EVENTINSPECTOR_BUNDLED_BUILD_SOURCE", None)
        else:
            os.environ["EVENTINSPECTOR_BUNDLED_BUILD_SOURCE"] = original_bundle_source


def test_windows_portable_update_staging() -> None:
    """Windows must stage a new payload beside, not over, the active payload."""
    import remote_update as updater

    manifest_path = ROOT / "Updates_2_5" / "remote_manifest.json"
    manifest_bytes = manifest_path.read_bytes()
    manifest = json.loads(manifest_bytes)
    payloads = {
        str(item["path"]): (ROOT / str(item["path"])).read_bytes()
        for item in manifest.get("files") or []
        if item.get("path")
    }
    original_os_name = updater.os.name
    original_user_data_dir = updater._user_data_dir
    original_download_first = updater._download_first
    original_download_verified = updater._download_verified
    try:
        with tempfile.TemporaryDirectory(prefix="eventinspector_windows_harness_") as temp_support:
            updater.os.name = "nt"
            updater._user_data_dir = lambda: temp_support

            active_dir = os.path.join(temp_support, "updates_v250")
            os.makedirs(active_dir, exist_ok=True)
            with open(os.path.join(active_dir, "active_payload.locked"), "w", encoding="utf-8") as handle:
                handle.write("keep")
            updater._save_state({
                "last_check": 0,
                "version": "2026-08-27-1-2.5.0-52",
                "update_dir": str(active_dir),
                "files": ["active_payload.locked"],
            })

            def fake_download_first(_urls, _timeout):
                return manifest_bytes, "harness://remote_manifest.json"

            def fake_download_verified(urls, _timeout, _expected_sha256=""):
                rel_path = next((path for path in payloads if any(url.endswith("/" + path) for url in urls)), None)
                _assert(rel_path is not None, f"unexpected update payload URL list: {urls}")
                return payloads[rel_path], f"harness://{rel_path}"

            updater._download_first = fake_download_first
            updater._download_verified = fake_download_verified

            result = updater.check_for_updates()
            _assert_equal(result.get("status"), "updated", "Windows portable update must complete without replacing the active directory")
            prepared = updater.get_prepared_update_info()
            prepared_dir = prepared["update_dir"]
            _assert(prepared_dir != active_dir, "Windows update must use a new payload directory")
            _assert(os.path.basename(prepared_dir).startswith("updates_v250_"), "Windows staged payload name must identify the v2.5 channel")
            _assert(os.path.exists(os.path.join(active_dir, "active_payload.locked")), "active Windows payload was modified during update")
            _assert(os.path.exists(os.path.join(prepared_dir, "Log_checker.py")), "Windows staged Log_checker.py is missing")
            _assert_equal(prepared.get("build"), CURRENT_RELEASE_BUILD, "Windows staged payload build mismatch")
            selected = desktop._select_prepared_update_candidate(
                [{
                    "update_dir": prepared_dir,
                    "build": prepared.get("build"),
                    "source": "windows-channel-state",
                }],
                bundled_build=53,
            )
            _assert(selected is not None, "Windows restart must see the staged build as a usable update")
            _assert_equal(selected.get("update_dir"), prepared_dir, "Windows restart selected the wrong payload directory")

            second = updater.check_for_updates()
            _assert_equal(second.get("status"), "up_to_date", "Windows staged payload must not download repeatedly")
    finally:
        updater.os.name = original_os_name
        updater._user_data_dir = original_user_data_dir
        updater._download_first = original_download_first
        updater._download_verified = original_download_verified


def test_build_scripts_clean_outputs() -> None:
    mac_script = (ROOT / "build" / "macos" / "build_macos.sh").read_text(encoding="utf-8", errors="ignore")
    win_portable_script = (ROOT / "build" / "windows" / "build_portable.bat").read_text(encoding="utf-8", errors="ignore")
    win_installer_script = (ROOT / "build" / "windows" / "build_windows.bat").read_text(encoding="utf-8", errors="ignore")

    mac_expected = [
        'rm -rf "dist/EventInspector.app"',
        'rm -f "dist/EventInspector.dmg"',
        'rm -rf "build/EventInspector"',
    ]
    for needle in mac_expected:
        _assert(needle in mac_script, f"build_macos.sh must clean stale artifact: {needle}")

    _assert('if [ -z "${MACOS_TARGET_ARCH:-}" ]; then' in mac_script, "macOS build must resolve a target architecture")
    _assert('arm64) MACOS_TARGET_ARCH="arm64"' in mac_script, "macOS build must support Apple Silicon")
    _assert('x86_64) MACOS_TARGET_ARCH="x86_64"' in mac_script, "macOS build must support Intel")
    _assert('--target-arch "$MACOS_TARGET_ARCH"' in mac_script, "macOS build must pass the target architecture")
    _assert('--exclude-module "markupsafe._speedups"' in mac_script, "macOS universal build must avoid the arm64-only MarkupSafe speedup")
    _assert('--add-data "Log_checker.py:."' in mac_script, "macOS build must package the bundled release marker")
    _assert('--add-data "sdk_check_presets.json:."' in mac_script, "macOS build must package SDK presets")
    _assert('--collect-submodules "androguard.core"' in mac_script, "macOS build must package manifest parser submodules")
    _assert('--collect-data "androguard"' in mac_script, "macOS build must package manifest parser data")
    _assert('--hidden-import "androguard.core.axml"' in mac_script, "macOS build must include AXMLPrinter explicitly")
    for service_asset in (
        "services_checker/app.py",
        "services_checker/axml_fallback.py",
        "services_checker/bundletool-all-1.18.1.jar",
        "services_checker/apk_check_presets.json",
        "services_checker/gradle_check_presets.json",
        "services_checker/gradle_lib_mapping.json",
        "services_checker/podfile_check_presets.json",
        "services_checker/manifest_check_presets.json",
        "services_checker/my-key.keystore",
    ):
        _assert(service_asset in mac_script, f"macOS build must package {service_asset}")
    _assert('--add-data "sdk_check_presets.json;."' in win_portable_script, "Windows portable build must package SDK presets")
    _assert('--add-data "sdk_check_presets.json;."' in win_installer_script, "Windows installer build must package SDK presets")
    _assert('--add-data "Log_checker.py;."' in win_portable_script, "Windows portable build must package the release source")
    _assert('--add-data "Log_checker.py;."' in win_installer_script, "Windows installer build must package the release source")
    _assert('verify_release_source.ps1' in win_portable_script, "Windows portable build must validate the release source")
    _assert('verify_release_source.ps1' in win_installer_script, "Windows installer build must validate the release source")
    _assert('verify_bundle.ps1' in win_portable_script, "Windows portable build must validate the built bundle")
    _assert('verify_bundle.ps1' in win_installer_script, "Windows installer build must validate the built bundle")
    _assert('--collect-submodules "androguard.core"' in win_portable_script, "Windows portable build must package manifest parser submodules")
    _assert('--collect-submodules "androguard.core"' in win_installer_script, "Windows installer build must package manifest parser submodules")
    _assert('--collect-data "androguard"' in win_portable_script, "Windows portable build must package manifest parser data")
    _assert('--collect-data "androguard"' in win_installer_script, "Windows installer build must package manifest parser data")
    _assert('--hidden-import "androguard.core.axml"' in win_portable_script, "Windows portable build must include AXMLPrinter explicitly")
    _assert('--hidden-import "androguard.core.axml"' in win_installer_script, "Windows installer build must include AXMLPrinter explicitly")
    for service_asset in (
        "services_checker\\app.py",
        "services_checker\\axml_fallback.py",
        "services_checker\\bundletool-all-1.18.1.jar",
        "services_checker\\apk_check_presets.json",
        "services_checker\\gradle_check_presets.json",
        "services_checker\\gradle_lib_mapping.json",
        "services_checker\\podfile_check_presets.json",
        "services_checker\\manifest_check_presets.json",
        "services_checker\\my-key.keystore",
    ):
        _assert(service_asset in win_portable_script, f"Windows portable build must package {service_asset}")
        _assert(service_asset in win_installer_script, f"Windows installer build must package {service_asset}")
    spec_text = (ROOT / "EventInspector.spec").read_text(encoding="utf-8", errors="ignore")
    _assert("('sdk_check_presets.json', '.')" in spec_text, "PyInstaller spec must package SDK presets")
    _assert("collect_data_files('androguard')" in spec_text, "PyInstaller spec must collect manifest parser data")
    _assert("collect_submodules('androguard.core')" in spec_text, "PyInstaller spec must include manifest parser submodules")
    _assert("'androguard.core.axml'" in spec_text, "PyInstaller spec must include AXMLPrinter explicitly")
    _assert("('services_checker', 'services_checker')" not in spec_text, "PyInstaller spec must not package runtime upload files")
    for service_asset in (
        "services_checker/app.py",
        "services_checker/axml_fallback.py",
        "services_checker/bundletool-all-1.18.1.jar",
        "services_checker/apk_check_presets.json",
        "services_checker/gradle_check_presets.json",
        "services_checker/gradle_lib_mapping.json",
        "services_checker/podfile_check_presets.json",
        "services_checker/manifest_check_presets.json",
        "services_checker/my-key.keystore",
    ):
        _assert(service_asset in spec_text, f"PyInstaller spec must package {service_asset}")

    win_expected = [
        'if exist "dist\\EventInspector" rmdir /s /q "dist\\EventInspector"',
        'if exist "build\\EventInspector" rmdir /s /q "build\\EventInspector"',
    ]
    for needle in win_expected:
        _assert(needle in win_portable_script, f"build_portable.bat must clean stale artifact: {needle}")
        _assert(needle in win_installer_script, f"build_windows.bat must clean stale artifact: {needle}")
    _assert(
        'if exist "build\\windows\\Output" rmdir /s /q "build\\windows\\Output"' in win_installer_script,
        "build_windows.bat must clean stale installer output",
    )


def test_windows_release_build_version_contract() -> None:
    workflow = (ROOT / ".github" / "workflows" / "windows-build.yml").read_text(encoding="utf-8", errors="ignore")
    source_guard = (ROOT / "build" / "windows" / "verify_release_source.ps1").read_text(encoding="utf-8", errors="ignore")
    bundle_guard = (ROOT / "build" / "windows" / "verify_bundle.ps1").read_text(encoding="utf-8", errors="ignore")
    installer_script = (ROOT / "build" / "windows" / "build_windows.bat").read_text(encoding="utf-8", errors="ignore")
    iss_script = (ROOT / "build" / "windows" / "EventChecker.iss").read_text(encoding="utf-8", errors="ignore")
    smoke_source = (ROOT / "tools" / "windows_update_smoke.py").read_text(encoding="utf-8", errors="ignore")

    _assert('default: "main"' in workflow, "Windows workflow must default to the main source ref")
    _assert("ref: ${{ inputs.source_ref || 'main' }}" in workflow, "Windows workflow must checkout the requested main source ref")
    _assert("Verify Portable ZIP version" in workflow, "Windows workflow must verify the portable ZIP")
    _assert("Verify Windows portable updater" in workflow, "Windows workflow must run the portable updater smoke test")
    _assert("python tools\\windows_update_smoke.py" in workflow, "Windows workflow must execute the Windows updater smoke test")
    _assert("sys.path.insert(0, str(ROOT))" in smoke_source, "Windows updater smoke must import the root updater from CI")
    _assert("EVENTINSPECTOR_WINDOWS_SMOKE_SIMULATE" in smoke_source, "Windows updater smoke must support local code-path testing")
    _assert("EventInspector-Windows-v${{ steps.release.outputs.release_version }}" in workflow, "Windows artifacts must be versioned")
    _assert('[switch]$PrintVersion' in source_guard, "Windows source guard must expose the resolved version")
    _assert('ExpectedSeries = "2.5.0"' in source_guard, "Windows source guard must enforce the v2.5 series")
    _assert("$match.Groups['build'].Value" in source_guard, "Windows source guard must derive the build number from source")
    _assert("bundleMarker -ne $sourceMarker" in bundle_guard, "Windows bundle guard must reject stale bundled source")
    _assert("EventInspector.exe" in bundle_guard, "Windows bundle guard must require the executable")
    _assert("-PrintVersion" in installer_script, "Windows installer must derive its version from the source")
    _assert("/DMyAppVersion=%EVENTINSPECTOR_RELEASE_VERSION%" in installer_script, "Inno Setup must receive the source version")
    _assert(
        f'#define MyAppVersion "2.5.0.{CURRENT_RELEASE_BUILD}"' in iss_script,
        "Inno Setup fallback must match the current v2.5 release",
    )


def test_windows_update_http_smoke() -> None:
    """Run the same real-HTTP updater smoke used by the Windows release workflow."""
    smoke_path = ROOT / "tools" / "windows_update_smoke.py"
    environment = os.environ.copy()
    if os.name != "nt":
        environment["EVENTINSPECTOR_WINDOWS_SMOKE_SIMULATE"] = "1"
    completed = subprocess.run(
        [sys.executable, str(smoke_path)],
        cwd=str(ROOT),
        env=environment,
        capture_output=True,
        text=True,
        timeout=120,
    )
    output = (completed.stdout or "") + (completed.stderr or "")
    _assert_equal(completed.returncode, 0, f"Windows updater HTTP smoke failed:\n{output}")
    _assert("windows_update_smoke: PASS" in output, "Windows updater HTTP smoke did not report PASS")


def test_windows_update_recovery_script() -> None:
    script_path = ROOT / "tools" / "reset_update_state_windows.bat"
    text = script_path.read_text(encoding="utf-8", errors="ignore")
    _assert(
        f"TARGET_VERSION={CURRENT_RELEASE_VERSION}" in text,
        "windows recovery script must target the current release",
    )
    _assert('updates_%%C' in text and 'v250' in text, "windows recovery script must clear every update channel")
    _assert('Updates_2_5/remote_manifest.json' in text, "windows recovery script must target the v2.5 manifest")
    _assert('services_checker/bundletool-all-1.18.1.jar' in text, "windows recovery script must preserve the Services Checker payload")
    stale_scripts = sorted((ROOT / "tools").glob("bootstrap_windows_to_v*.bat"))
    _assert(not stale_scripts, f"remove stale Windows bootstrap scripts: {[p.name for p in stale_scripts]}")


def test_services_checker_bridge_contract() -> None:
    source = (ROOT / "Log_checker.py").read_text(encoding="utf-8")
    for payload_source in (source,):
        _assert("EVENTINSPECTOR_ALLOW_EXTERNAL_SERVICES_CHECKER" in payload_source, "bundled Services Checker source switch missing")
        _assert("_services_checker_external_sources_enabled" in payload_source, "bundled Services Checker source policy missing")
        _assert("EVENTINSPECTOR_SERVICES_COMMAND" in payload_source, "Services Checker command override missing")
        _assert("AndroidTool.command" in payload_source, "Drive launcher fallback missing")
        _assert("Androidchecker.cmd" in payload_source, "Windows Drive launcher fallback missing")
        _assert("SERVICES_CHECKER_APP_RELATIVE_PATH" in payload_source, "shared Drive app path resolver missing")
        _assert("_services_checker_uses_shared_keystore" in payload_source, "Drive keystore source check missing")
        _assert("force_drive_import" in payload_source, "stale Services Checker source must be bypassed")
        _assert("subprocess.Popen" in payload_source, "Services Checker launcher missing")
        _assert("_services_checker_ready" in payload_source, "Services Checker readiness check missing")
        _assert("_prepare_services_checker_runtime" in payload_source, "Services Checker bundle import preparation missing")
        _assert("_services_checker_resource_roots" in payload_source, "Services Checker resource layout resolver missing")
        _assert("_services_checker_bundled_resource_root" in payload_source, "Services Checker complete resource contract missing")
        _assert("EVENTINSPECTOR_BUNDLETOOL_PATH" in payload_source, "Services Checker bundletool path pin missing")
        _assert("SERVICES_CHECKER_REQUIRED_RESOURCES" in payload_source, "Services Checker required resource list missing")
        _assert('importlib.import_module("androguard.core.axml")' in payload_source, "Services Checker androguard preload missing")
        _assert("_services_checker_saved_host_path" in payload_source, "saved host override missing")
        _assert("/api/services-checker/host" in payload_source, "host replacement API missing")
        _assert("/api/services-checker/reload" in payload_source, "Services Checker reload API missing")
        _assert("servicesCheckerReloadBtn" not in payload_source, "obsolete Services Checker reload button must stay removed")
        _assert('id="servicesCheckerStatus"' in payload_source, "Services Checker status UI missing")
        _assert('id="servicesCheckerError"' in payload_source, "Services Checker error UI missing")
        _assert("Source: bundled with Event Inspector" not in payload_source, "obsolete Services Checker source label must stay removed")
        _assert("servicesCheckerReplaceHostBtn" not in payload_source, "host replacement UI must not be exposed in bundled mode")
        _assert('"services_checker", "app.py"' in payload_source, "bundled Services Checker fallback missing")

        file_candidates_start = payload_source.index("def _services_checker_file_candidates")
        command_candidates_start = payload_source.index("def _services_checker_command_candidates")
        file_candidates_source = payload_source[file_candidates_start:command_candidates_start]
        bundled_marker = 'candidates.append(os.path.join(SCRIPT_DIR, "services_checker", "app.py"))'
        external_marker = "if _services_checker_external_sources_enabled():"
        _assert(bundled_marker in file_candidates_source, "bundled Services Checker candidate missing")
        _assert(external_marker in file_candidates_source, "external Services Checker fallback must be opt-in")
        _assert(
            file_candidates_source.index(bundled_marker) < file_candidates_source.index(external_marker),
            "bundled Services Checker must be preferred over external sources",
        )

    windows_launcher = Path.home() / "Downloads" / "Androidchecker.cmd"
    if windows_launcher.exists():
        launcher_source = windows_launcher.read_text(encoding="utf-8", errors="ignore")
        _assert("Shared drives\\IndieZ - Tester" in launcher_source, "Windows launcher must resolve the shared Drive app")
        _assert("SERVICE_APP_PATH" in launcher_source, "Windows launcher override missing")

    service_source = (ROOT / "services_checker" / "app.py").read_text(encoding="utf-8")
    fallback_source = (ROOT / "services_checker" / "axml_fallback.py").read_text(encoding="utf-8")
    _assert("def _load_axml_printer()" in service_source, "Services Checker AXML retry loader is missing")
    _assert("def _load_fallback_axml_printer()" in service_source, "Services Checker dependency-free fallback loader is missing")
    _assert('importlib.import_module("androguard.core.axml")' in service_source, "Services Checker AXML import is not explicit")
    _assert("_live_remote_preset_payloads" in service_source, "Services Checker must keep the latest downloaded payload in memory")
    _assert("refreshed_digests" in service_source, "Services Checker must expose downloaded preset hashes")
    _assert("loaded_preset_files" in service_source, "Services Checker must expose the preset copy used by the response")
    _assert("preset_revision" in service_source, "Services Checker must expose a client-visible preset revision")
    _assert("class AXMLPrinter" in fallback_source, "Services Checker dependency-free AXML fallback is missing")
    _assert((ROOT / "services_checker" / "axml_fallback.py").is_file(), "Services Checker fallback file is missing")
    _assert(".container h2.text-xl" in service_source, "Services Checker typography override is missing")
    _assert(".results-section h3, .results-section h4 { font-size: 0.78rem" in service_source, "Services Checker result headings are too large")
    _assert(
        ".results-section p, .results-section li, .results-section .comparison-item { font-size: 0.68rem"
        in service_source,
        "Services Checker comparison rows are too large",
    )
    _assert("value.join('\\\\n')" in service_source, "bundled Services Checker newline escape is invalid")
    key_path = ROOT / "services_checker" / "my-key.keystore"
    _assert(key_path.is_file() and key_path.stat().st_size > 0, "bundled test keystore is missing")


def test_services_checker_resource_contract_is_platform_neutral() -> None:
    source = (ROOT / "Log_checker.py").read_text(encoding="utf-8")
    helper_start = source.index("def _services_checker_resource_roots")
    helper_end = source.index("def _prepare_services_checker_runtime")
    helper_source = source[helper_start:helper_end].lower()
    for forbidden in ("sys.platform", "platform.system", "os.name", '"darwin"', '"windows"'):
        _assert(forbidden not in helper_source, f"resource resolver must not infer platform from {forbidden}")

    previous_root = os.environ.get("EVENTINSPECTOR_BUNDLED_SERVICES_ROOT")
    previous_bundletool = os.environ.get("EVENTINSPECTOR_BUNDLETOOL_PATH")
    try:
        for layout in (
            Path("mac-artifact") / "Contents" / "Resources" / "services_checker",
            Path("windows-artifact") / "_internal" / "services_checker",
        ):
            with tempfile.TemporaryDirectory(prefix="eventchecker-resource-contract-") as temp_dir:
                resource_root = Path(temp_dir) / layout
                resource_root.mkdir(parents=True)
                (resource_root / "bundletool-all-1.18.1.jar").write_bytes(b"bundletool")
                (resource_root / "my-key.keystore").write_bytes(b"keystore")
                app_path = resource_root / "app.py"
                app_path.write_text("# test app\n", encoding="utf-8")
                os.environ["EVENTINSPECTOR_BUNDLED_SERVICES_ROOT"] = str(resource_root)
                os.environ.pop("EVENTINSPECTOR_BUNDLETOOL_PATH", None)

                resolved_root = lc._services_checker_bundled_resource_root(str(app_path))
                _assert_equal(
                    Path(resolved_root).resolve(),
                    resource_root.resolve(),
                    f"resource resolver must use the declared {layout} layout",
                )
                lc._prepare_services_checker_runtime(str(app_path))
                _assert_equal(
                    Path(os.environ["EVENTINSPECTOR_BUNDLED_SERVICES_ROOT"]).resolve(),
                    resource_root.resolve(),
                    "runtime must pin the complete resource root",
                )
                _assert_equal(
                    Path(os.environ["EVENTINSPECTOR_BUNDLETOOL_PATH"]).resolve(),
                    (resource_root / "bundletool-all-1.18.1.jar").resolve(),
                    "runtime must pin bundletool inside the same resource root",
                )
    finally:
        if previous_root is None:
            os.environ.pop("EVENTINSPECTOR_BUNDLED_SERVICES_ROOT", None)
        else:
            os.environ["EVENTINSPECTOR_BUNDLED_SERVICES_ROOT"] = previous_root
        if previous_bundletool is None:
            os.environ.pop("EVENTINSPECTOR_BUNDLETOOL_PATH", None)
        else:
            os.environ["EVENTINSPECTOR_BUNDLETOOL_PATH"] = previous_bundletool


def test_native_download_contract() -> None:
    source = (ROOT / "desktop_app.py").read_text(encoding="utf-8")
    configure_marker = "webview.settings['ALLOW_DOWNLOADS'] = True"
    _assert(configure_marker in source, "native WebView downloads must be enabled")
    configure_index = source.index("_configure_webview_downloads()")
    window_index = source.index("webview.create_window(")
    _assert(configure_index < window_index, "download setting must be applied before creating the WebView")

    service_path = ROOT / "services_checker" / "app.py"
    service_source = service_path.read_text(encoding="utf-8", errors="ignore")
    _assert("def _resolve_bundletool_path()" in service_source, "AAB converter bundletool resolver is missing")
    _assert("def _resolve_keystore_path()" in service_source, "AAB converter keystore resolver is missing")
    _assert(
        "bundletool_path = _resolve_bundletool_path()" in service_source,
        "AAB conversion must resolve bundletool at request time",
    )
    _assert(
        "final_ks_path = _resolve_keystore_path()" in service_source,
        "AAB conversion must resolve the keystore at request time",
    )
    _assert(
        "BUNDLETOOL_PATH" not in service_source[service_source.index("def convert_aab_to_apk"):service_source.index("@app.route('/download/<filename>')")],
        "AAB conversion route must not use the import-time bundletool path",
    )
    _assert(
        "def _cleanup_stale_conversion_artifacts()" in service_source,
        "AAB conversion must clean abandoned generated artefacts",
    )
    _assert(
        "_cleanup_stale_conversion_artifacts()" in service_source[service_source.index("def convert_aab_to_apk"):service_source.index("@app.route('/download/<filename>')")],
        "AAB conversion must clean abandoned artefacts before bundletool runs",
    )
    _assert(
        "_remove_generated_file(source_path)" in service_source,
        "Native APK save must remove the generated source after copying",
    )
    _assert("@app.route('/download/<filename>')" in service_source, "Services Checker download route is missing")
    _assert("as_attachment=True" in service_source, "Services Checker download route must return an attachment")
    _assert(
        "@app.route('/save_download/<filename>', methods=['POST'])" in service_source,
        "Services Checker native save route is missing",
    )
    _assert(
        "def _copy_generated_file_to_downloads" in service_source,
        "AAB conversion must provide a platform-native Downloads save helper",
    )
    conversion_start = service_source.index("def convert_aab_to_apk()")
    download_route_start = service_source.index("@app.route('/download/<filename>')")
    conversion_source = service_source[conversion_start:download_route_start]
    _assert(
        "def _write_zip_member_to_downloads" in service_source,
        "AAB conversion must have a direct Downloads writer",
    )
    _assert(
        "download_apk_path, saved_filename = _write_zip_member_to_downloads(" in conversion_source,
        "AAB conversion must write universal.apk directly to Downloads",
    )
    _assert(
        "'saved_to_downloads': True" in conversion_source,
        "AAB conversion response must confirm the direct Downloads save",
    )
    _assert(
        "session.setdefault('downloadable_files', {})[final_apk_filename]" not in conversion_source,
        "Final APK must not be staged in the Flask session download cache",
    )
    _assert(
        "final_apk_savelocation" not in conversion_source,
        "Final APK must not be staged in UPLOAD_FOLDER",
    )
    _assert(
        "'apk_download_url'" not in conversion_source,
        "AAB conversion must not fall back to a temporary browser download URL",
    )
    _assert(
        "const saved_to_downloads" not in service_source,
        "AAB conversion UI must require the direct Downloads result",
    )
    _assert(
        "APK saved directly to Downloads:" in service_source,
        "AAB conversion UI must show the direct Downloads result",
    )


def test_gradle_comparison_hides_unlisted_preset_rows() -> None:
    service_source = (ROOT / "services_checker" / "app.py").read_text(encoding="utf-8", errors="ignore")
    _assert(
        'if (archiveType !== "Gradle") {' in service_source,
        "Gradle comparison must hide mapped libraries outside the selected preset",
    )
    _assert(
        "INFO (Found in ' + archiveType + ', not in expected list)" in service_source,
        "non-Gradle comparisons must retain informational unlisted-library rows",
    )
    _assert(
        "gradle_mapping = _load_gradle_lib_mapping()" in service_source
        and "found_versions = scan_gradle_for_versions(gradle_content, gradle_mapping)" in service_source,
        "Gradle scanner must continue scanning the complete loaded mapping",
    )


TESTS: List[Callable[[], None]] = [
    test_manifest_contract,
    test_manifest_payload_integrity,
    test_canonical_update_channel_contract,
    test_no_pre_v25_release_artifacts,
    test_package_code_mapping,
    test_installation_id_state_machine,
    test_installation_id_log_parsing,
    test_sdk_exact_contracts,
    test_cloudx_sdk_adapter_metadata,
    test_max_sdk_logs,
    test_sdk_check_preset_contract,
    test_rendered_sdk_preset_javascript_contract,
    test_default_ad_event_contract,
    test_installation_id_copy_contract,
    test_sdk_failed_groups_sort_first,
    test_release_build_marker,
    test_rewarded_bidding_filter_contract,
    test_price_rotation_exact_parser,
    test_load_ads_provider_contract,
    test_load_ads_max_contract,
    test_levelplay_impression_data_callback_contract,
    test_ascendx_cloudx_callback_contract,
    test_release_payload_sync,
    test_update_candidate_does_not_downgrade,
    test_services_checker_gradle_mapping_contract,
    test_services_checker_live_preset_refresh_after_restart,
    test_services_checker_git_value_reload_from_real_commit,
    test_sdk_preset_git_value_reload_from_real_commit,
    test_canonical_manifest_build_contract,
    test_update_flow_canonical_v25,
    test_windows_portable_update_staging,
    test_build_scripts_clean_outputs,
    test_windows_release_build_version_contract,
    test_windows_update_http_smoke,
    test_windows_update_recovery_script,
    test_services_checker_bridge_contract,
    test_services_checker_resource_contract_is_platform_neutral,
    test_native_download_contract,
    test_gradle_comparison_hides_unlisted_preset_rows,
]


def run_tests(selected: Iterable[Callable[[], None]]) -> List[HarnessResult]:
    results: List[HarnessResult] = []
    for test in selected:
        try:
            test()
            results.append(HarnessResult(test.__name__, True, "ok"))
        except Exception as exc:  # pragma: no cover - harness should surface the failure directly
            results.append(HarnessResult(test.__name__, False, str(exc)))
    return results


def main() -> int:
    parser = argparse.ArgumentParser(description="EventInspector local harness")
    parser.add_argument("--json", action="store_true", help="Emit JSON results")
    parser.add_argument("--only", nargs="*", default=[], help="Run only named tests")
    args = parser.parse_args()

    selected = TESTS
    if args.only:
        wanted = set(args.only)
        selected = [test for test in TESTS if test.__name__ in wanted]
        missing = wanted - {test.__name__ for test in selected}
        if missing:
            print(f"Unknown test name(s): {', '.join(sorted(missing))}", file=sys.stderr)
            return 2

    results = run_tests(selected)
    failures = [result for result in results if not result.passed]

    if args.json:
        print(json.dumps([result.__dict__ for result in results], indent=2, ensure_ascii=False))
    else:
        for result in results:
            status = "PASS" if result.passed else "FAIL"
            print(f"{status} {result.name}: {result.message}")

    if failures:
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
