import json
import re
import subprocess
import threading
import time
import requests
import html
from collections import deque
from flask import Flask, render_template_string, request, jsonify, Response
from flask_socketio import SocketIO
from flask_cors import CORS
from openpyxl import load_workbook
import os
import sys
import shutil
import logging
from pathlib import Path

# Khởi tạo ứng dụng Flask và SocketIO
app = Flask(__name__)
CORS(app)
try:
    socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')
except ValueError:
    # Fallback for environments where 'threading' isn't accepted
    socketio = SocketIO(app, cors_allowed_origins="*")

# --- CẤU HÌNH LOAD ADS (GOOGLE SHEET) ---
G_SHEET_URL = "https://script.google.com/macros/s/AKfycbyLMM9nLAjS9Zhwr4-J6ikjqBSpO7ZCNaNeHKTsfKltiIa0OniDBSrzjvqfvpg87Epl/exec"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

def _resolve_default_params_path():
    env_path = os.getenv("DEFAULT_PARAMS_XLSX_PATH")
    if env_path and os.path.exists(env_path):
        return env_path

    filename = "Default event + Default Params.xlsx"

    update_dir = os.getenv("EVENTINSPECTOR_UPDATE_DIR")
    if update_dir:
        upd = os.path.join(update_dir, filename)
        if os.path.exists(upd):
            return upd

    candidates = []

    # 1) Project folder (dev)
    candidates.append(os.path.join(SCRIPT_DIR, filename))

    # 2) Installed app folder (packaged)
    if getattr(sys, "frozen", False):
        exe_dir = os.path.dirname(sys.executable)
        candidates.append(os.path.join(exe_dir, filename))

        # macOS: Resources folder inside .app bundle
        if sys.platform == "darwin":
            resources_dir = os.path.abspath(os.path.join(exe_dir, "..", "Resources"))
            candidates.append(os.path.join(resources_dir, filename))

    # 3) PyInstaller temp bundle path
    meipass = getattr(sys, "_MEIPASS", None)
    if meipass:
        candidates.append(os.path.join(meipass, filename))

    # 4) Common user locations
    candidates.extend([
        os.path.join(os.path.expanduser("~/Downloads"), filename),
        os.path.join(os.path.expanduser("~/Documents"), filename),
    ])

    for p in candidates:
        if p and os.path.exists(p):
            return p

    # If packaged, try to copy bundled file into app folder for user to edit
    if getattr(sys, "frozen", False) and meipass:
        bundled = os.path.join(meipass, filename)
        if os.path.exists(bundled):
            try:
                target = os.path.join(os.path.dirname(sys.executable), filename)
                shutil.copyfile(bundled, target)
                return target
            except Exception:
                if sys.platform == "darwin":
                    try:
                        resources_dir = os.path.abspath(os.path.join(os.path.dirname(sys.executable), "..", "Resources"))
                        os.makedirs(resources_dir, exist_ok=True)
                        target = os.path.join(resources_dir, filename)
                        shutil.copyfile(bundled, target)
                        return target
                    except Exception:
                        pass

    return candidates[0]

DEFAULT_PARAMS_XLSX = _resolve_default_params_path()
DEFAULT_PARAM_FILL = "FFFCE5CD"
DEFAULT_REMOTE_MANIFEST_URL = "https://raw.githubusercontent.com/trucbm/Eventchecker/main/Updates/remote_manifest.json"


def _runtime_app_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return SCRIPT_DIR


def _resolve_profiles_dir():
    base_dir = _runtime_app_dir()
    profiles_dir = os.path.join(base_dir, "profiles")
    try:
        os.makedirs(profiles_dir, exist_ok=True)
        return profiles_dir
    except Exception:
        fallback = os.path.join(SCRIPT_DIR, "profiles")
        os.makedirs(fallback, exist_ok=True)
        return fallback


PROFILE_DIR = _resolve_profiles_dir()
active_profile_name = None
active_profile_path = None
active_profile_game_name = ""


def _user_data_dir():
    if os.name == "nt":
        base = os.getenv("LOCALAPPDATA") or os.path.expanduser("~")
        return os.path.join(base, "EventInspector")
    if sys.platform == "darwin":
        return os.path.join(os.path.expanduser("~/Library/Application Support"), "EventInspector")
    return os.path.join(os.path.expanduser("~"), ".eventinspector")


def _normalize_remote_update_config():
    try:
        user_dir = _user_data_dir()
        os.makedirs(user_dir, exist_ok=True)
        cfg_path = os.path.join(user_dir, "remote_update_config.json")
        cfg = {}
        if os.path.exists(cfg_path):
            try:
                with open(cfg_path, "r", encoding="utf-8") as f:
                    cfg = json.load(f)
            except Exception:
                cfg = {}
        cfg["enabled"] = True
        cfg["manifest_url"] = DEFAULT_REMOTE_MANIFEST_URL
        cfg["timeout_sec"] = 10
        cfg["min_interval_sec"] = 0
        with open(cfg_path, "w", encoding="utf-8") as f:
            json.dump(cfg, f, indent=2)
    except Exception as e:
        print(f"WARNING: Failed to normalize updater config: {e}")

def _resolve_adb():
    adb_env = os.getenv("ADB_PATH")
    if adb_env:
        return adb_env

    # Common SDK locations
    sdk_roots = [
        os.getenv("ANDROID_HOME"),
        os.getenv("ANDROID_SDK_ROOT"),
        os.path.expanduser("~/Library/Android/sdk"),
        os.path.expanduser("~/Android/Sdk"),
    ]
    for root in [r for r in sdk_roots if r]:
        candidate = os.path.join(root, "platform-tools", "adb")
        if os.path.exists(candidate):
            return candidate

    extra_paths = "/opt/homebrew/bin:/usr/local/bin:/usr/bin:/bin:/usr/sbin:/sbin"
    current_path = os.environ.get("PATH", "")
    if extra_paths not in current_path:
        os.environ["PATH"] = f"{extra_paths}:{current_path}" if current_path else extra_paths

    found = shutil.which("adb")
    return found or "adb"

ADB_EXECUTABLE = _resolve_adb()

if os.name == 'nt':
    creation_flags = 0x08000000  # CREATE_NO_WINDOW
else:
    creation_flags = 0

# Bảng tra cứu tên thiết bị
DEVICE_NAMES = {
    "R94Y40MZJ5T": "SS A16",
    "RFCX30DPW6D": "SS A35",
    "FMAEAYHIQ8FEYXPN": "OPPO A18",
    "ZY226DFRH2": "Moto Z4",
    "R5CR1282PAK": "SS S21",
    "F6QCCAGIRSQOVGFQ": "Redmi A3",
    "7ec7bca6": "Xiaomi 13 Lite"
}

# --- DỮ LIỆU TOÀN CỤC ---

# Giới hạn cache để UI không giữ quá nhiều log trong RAM.
MAX_LOAD_ADS_LOGS = 1000
MAX_VALIDATOR_LOGS = 1500
MAX_SPECIFIC_EVENT_LOGS = 1500
MAX_CALLBACK_AD_LOGS = 1500
MAX_ADREVENUE_LOGS = 1500

# 1. Dữ liệu cho Tab Load Ads
load_ads_events = deque(maxlen=MAX_LOAD_ADS_LOGS)
unique_load_ads = set()

# 2. Dữ liệu cho Tab Load Ads Ext
load_ads_ext_events = deque(maxlen=MAX_LOAD_ADS_LOGS)
unique_load_ads_ext = set()

# Trạng thái Recording (Google Sheet) - RIÊNG BIỆT CHO TỪNG TAB
recording_states = {
    "LoadAds": {"is_recording": False, "current_sheet": None},
    "LoadAdsExt": {"is_recording": False, "current_sheet": None}
}

# 3. Dữ liệu cho Tab Validator
validator_results = deque(maxlen=MAX_VALIDATOR_LOGS)
required_params = []  # Manual extra params from UI
default_params = []   # Default params from sheet (apply to all events)
event_specific_params = {}  # event_name -> list of params
validator_active = False

# 4. Dữ liệu cho Tab Specific Event
event_log_cache = deque(maxlen=MAX_SPECIFIC_EVENT_LOGS)
specific_event_results = []
specific_event_name_filters = []
specific_event_params_filters = []

# 5. Dữ liệu cho Tab Package Logcat
package_log_cache = deque(maxlen=15000)
target_package_name = ""
active_package_pids = {}
active_logcat_processes = {}

# 6. Dữ liệu cho Tab Callback & Ads Event
callback_ad_logs = deque(maxlen=MAX_CALLBACK_AD_LOGS)

# 7. Dữ liệu cho Tab AdRevenue
adrevenue_log_cache = []
adrevenue_logs = deque(maxlen=MAX_ADREVENUE_LOGS)
adrevenue_params_to_validate = []

# 8. Dữ liệu cho Tab SDK Check
sdk_check_search_list = []
sdk_check_results = {}
sdk_check_input_list = []

# Dữ liệu hệ thống chung
active_log_readers = {}
connected_devices_info = []
is_paused = False
lock = threading.Lock()
incomplete_impression_logs = {} # Buffer cho logs bị ngắt dòng
adb_error_counter = 0


# --- REGEX PATTERNS ---

# Pattern cho Load Ads (Unity)
UNITY_TRACKING_PATTERN = re.compile(r'\[\s*Tracking\s*\]\s*TrackingService->Track:\s*(\{"eventName":"ad_impression".*)')

# Pattern cho Load Ads Ext (AppMetrica)
METRICA_TRACKING_PATTERN = re.compile(r'Event sent: ad_impression with value\s*(\{.*\})')

# Patterns cũ của Log Checker
OLD_EVENT_LOG_PATTERN = re.compile(r'\[\s*Tracking\s*\]\s*TrackingService->Track:\s*(\{"eventName":.*)')
CALLBACK_LOG_PATTERN = re.compile(r"(_OnImpressionDataReadyEvent|LevelPlayInterstitialAdListener|LevelPlayBannerAdViewListener|LevelPlayRewardedAdListener)")
ADREVENUE_LOG_PATTERN = re.compile(r"AdRevenue Received:\s*AdRevenue\{(.*)\}")
SDK_CHECK_SEARCH_PATTERN = re.compile(r'"search_pattern"\s*:\s*["\'](.*?)["\']')
GADSME_SERVICE_KEYWORD = "GadsmeService->"
ADVERTY5_KEYWORD = "Adverty5"

# Mapping tên hiển thị cho Callback
CALLBACK_DISPLAY_NAMES = {
    "LevelPlayInterstitialAdListener": "Interstitial",
    "LevelPlayBannerAdViewListener": "Banner",
    "LevelPlayRewardedAdListener": "Rewarded",
    "_OnImpressionDataReadyEvent": "Impression Data"
}

def get_device_name(device_id):
    return DEVICE_NAMES.get(device_id, device_id)

# --- HELPER FUNCTIONS FOR FORMATTING ---
def format_json_html(data):
    """Format JSON object to HTML string with indentation and colors"""
    try:
        if isinstance(data, str):
            # Try to parse string as JSON if possible
            if data.strip().startswith('{') or data.strip().startswith('['):
                try:
                    data = json.loads(data)
                except: pass
        
        # If it's a dict or list, dump it pretty
        if isinstance(data, (dict, list)):
            # ensure_ascii=False để hiển thị tiếng Việt đúng
            json_str = json.dumps(data, indent=2, ensure_ascii=False)
            return f'<pre class="text-xs bg-gray-50 p-2 rounded border border-gray-200 overflow-x-auto font-mono text-gray-700">{json_str}</pre>'
        return str(data)
    except:
        return str(data)


def format_param_issue_html(title, items, color_class, chunk_size=4):
    if not items:
        return ""
    ordered = [html.escape(str(x)) for x in sorted(items)]
    lines = "<br>".join(f"&nbsp;&nbsp;{item}" for item in ordered)
    return (
        f'<div class="{color_class} text-xs mb-2 break-words leading-5">'
        f'<div>{html.escape(title)}:</div>'
        f'<div class="font-normal">{lines}</div>'
        f'</div>'
    )

def extract_json_object_from_text(text):
    """Extract first JSON object substring from text by brace matching."""
    try:
        start_idx = text.find('{')
        if start_idx == -1:
            return None
        open_braces = 0
        for i in range(start_idx, len(text)):
            if text[i] == '{':
                open_braces += 1
            elif text[i] == '}':
                open_braces -= 1
                if open_braces == 0:
                    return text[start_idx:i+1]
        return None
    except:
        return None

def load_default_params_config():
    """Load default params + event-specific params from XLSX."""
    global default_params, event_specific_params, active_profile_game_name
    path = active_profile_path or DEFAULT_PARAMS_XLSX
    if not path or not os.path.exists(path):
        print(f"INFO: Default params sheet not found: {path}")
        default_params = []
        event_specific_params = {}
        active_profile_game_name = ""
        return
    try:
        wb = load_workbook(path)
        ws = wb.active
        event_map = {}
        current_event = None
        header_row = 1

        if str(ws.cell(1, 2).value or "").strip().lower() == "game":
            active_profile_game_name = str(ws.cell(1, 3).value or "").strip()
            header_row = 2
        else:
            active_profile_game_name = ""

        for r in range(header_row + 1, ws.max_row + 1):
            event_val = ws.cell(r, 2).value
            param_cell = ws.cell(r, 3)
            param_val = param_cell.value

            if event_val:
                current_event = str(event_val).strip()
                if current_event:
                    event_map.setdefault(current_event, {"specific": [], "default": []})

            if current_event and param_val:
                param = str(param_val).strip()
                fill = param_cell.fill
                fg = fill.fgColor.rgb if fill and fill.fgColor else None
                is_default = (fg == DEFAULT_PARAM_FILL)
                if is_default:
                    event_map[current_event]["default"].append(param)
                else:
                    event_map[current_event]["specific"].append(param)

        # Build unique default params list
        seen = set()
        merged_default = []
        for data in event_map.values():
            for p in data["default"]:
                if p not in seen:
                    seen.add(p)
                    merged_default.append(p)

        default_params = merged_default
        event_specific_params = {k: v["specific"] for k, v in event_map.items()}
        print(f"INFO: Loaded default params: {len(default_params)}; events: {len(event_specific_params)}")
    except Exception as e:
        print(f"ERROR: Failed to load default params sheet: {e}")
        default_params = []
        event_specific_params = {}
        active_profile_game_name = ""


def _sanitize_profile_filename(filename):
    name = os.path.basename((filename or "").strip())
    name = name.replace("\\", "_").replace("/", "_")
    if not name.lower().endswith(".xlsx"):
        raise ValueError("Only .xlsx files are supported")
    return name


def _list_profile_names():
    try:
        files = [
            p.name for p in Path(PROFILE_DIR).glob("*.xlsx")
            if p.is_file()
        ]
        return sorted(files, key=lambda x: x.lower())
    except Exception:
        return []


def _ensure_default_profile_seed():
    source = DEFAULT_PARAMS_XLSX
    if not source or not os.path.exists(source):
        return
    target = os.path.join(PROFILE_DIR, os.path.basename(source))
    if os.path.abspath(source) == os.path.abspath(target):
        return
    if not os.path.exists(target):
        try:
            shutil.copyfile(source, target)
        except Exception as e:
            print(f"WARNING: Failed to seed default profile: {e}")


def _set_active_profile(profile_name=None):
    global active_profile_name, active_profile_path
    _ensure_default_profile_seed()
    names = _list_profile_names()
    if not names:
        active_profile_name = None
        active_profile_path = None
        load_default_params_config()
        return False

    selected = None
    if profile_name:
        clean = _sanitize_profile_filename(profile_name)
        if clean in names:
            selected = clean

    if not selected:
        preferred = os.path.basename(DEFAULT_PARAMS_XLSX) if DEFAULT_PARAMS_XLSX else None
        if preferred in names:
            selected = preferred
        else:
            selected = names[0]

    active_profile_name = selected
    active_profile_path = os.path.join(PROFILE_DIR, selected)
    load_default_params_config()
    return True


def _profile_payload():
    return {
        "profiles": _list_profile_names(),
        "current_profile": active_profile_name,
        "game_name": active_profile_game_name,
        "profile_dir": PROFILE_DIR,
        "default_event_names": sorted(event_specific_params.keys()),
    }

def _levenshtein_distance_limit(a, b, limit=2):
    """Compute Levenshtein distance with early exit if > limit."""
    if a == b:
        return 0
    if not a or not b:
        return max(len(a), len(b))
    if abs(len(a) - len(b)) > limit:
        return limit + 1

    # Initialize previous row
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        curr = [i]
        min_in_row = curr[0]
        for j, cb in enumerate(b, 1):
            cost = 0 if ca == cb else 1
            curr_val = min(
                prev[j] + 1,      # deletion
                curr[j - 1] + 1,  # insertion
                prev[j - 1] + cost  # substitution
            )
            curr.append(curr_val)
            if curr_val < min_in_row:
                min_in_row = curr_val
        if min_in_row > limit:
            return limit + 1
        prev = curr
    return prev[-1]

# --- GIAO DIỆN WEB (HTML/JS) ---
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="vi">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <link rel="icon" href="data:,"> <!-- Fix lỗi Favicon 404 -->
    <title>Event Inspector V2.0.0(26)</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/socket.io/4.7.4/socket.io.js"></script>
    <style>
        body { font-family: 'Inter', sans-serif; }
        .log-cell { max-width: 500px; word-wrap: break-word; font-family: monospace; font-size: 0.75rem; color: #6b7280; }
        .message-cell { white-space: nowrap; }
        #packageLogTable { table-layout: fixed; width: 100%; }
        #packageLogTable col.col-time { width: 110px; }
        #packageLogTable col.col-tag { width: 90px; }
        .details-cell { font-family: monospace; font-size: 0.8rem; line-height: 1.4; min-width: 260px; max-width: 640px; white-space: normal; overflow-wrap: anywhere; word-break: break-word; }
        .tag-cell { max-width: 90px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
        .tag-header { max-width: 90px; }
        .time-cell { white-space: nowrap; width: 1%; max-width: 110px; font-size: 0.75rem; }
        .time-header { white-space: nowrap; width: 1%; max-width: 110px; }
        .resizable { position: relative; }
        .resizer {
            position: absolute;
            right: -3px;
            top: 0;
            width: 12px;
            height: 100%;
            cursor: col-resize;
            user-select: none;
            z-index: 20;
        }
        .resizer:hover { background: rgba(59, 130, 246, 0.15); }
        #packageLogTableBody tr.selected { background-color: #bfdbfe !important; }
        .resizer.disabled { cursor: not-allowed; background: transparent; }
        .details-cell pre { margin: 0; white-space: pre-wrap; overflow-wrap: anywhere; word-break: break-word; }
        @keyframes pulse { 50% { opacity: .6; } }
        .animate-pulse-green { animation: pulse 2s cubic-bezier(0.4, 0, 0.6, 1) infinite; }
        .animate-record { animation: pulse 1.5s infinite; }
        .tab-btn { transition: all 0.2s ease-in-out; }
        .tab-btn.active { border-color: #4f46e5; color: #4f46e5; background-color: #eef2ff; }
        .log-row.selected { background-color: #dbeafe; }
        .log-cell,
        .details-cell,
        .details-cell *,
        #loadAdsTableBody td,
        #loadAdsExtTableBody td,
        #validatorTableBody td,
        #specificEventTableBody td,
        #adRevenueTableBody td,
        #callbackAdTableBody td,
        #sdkCheckTableBody td {
            user-select: text;
            -webkit-user-select: text;
        }
        
        /* Custom scrollbar for pre blocks */
        pre::-webkit-scrollbar { height: 6px; width: 6px; }
        pre::-webkit-scrollbar-thumb { background: #cbd5e1; border-radius: 3px; }
        pre::-webkit-scrollbar-track { background: #f1f5f9; }
    </style>
</head>
<body class="bg-gray-100 text-gray-800 h-screen overflow-hidden">

    <div class="container mx-auto p-3 sm:p-4 lg:p-6 h-full flex flex-col">
        <!-- HEADER -->
        <div class="bg-white rounded-xl shadow-md p-4 mb-4 flex-shrink-0">
            <div class="flex justify-between items-center flex-wrap gap-4">
                <div class="flex items-center gap-3">
                    <div>
                        <div class="flex items-center gap-2.5">
                            <h1 class="text-xl font-bold text-gray-700">Event Inspector</h1>
                            <span class="text-xs font-semibold bg-indigo-100 text-indigo-700 px-2 py-1 rounded-full">v2.0.0(26)</span>
                        </div>
                        <p class="text-sm text-gray-500">Integrates Load Ads & Event Validation.</p>
                    </div>
                    <button id="restartAppBtn" class="bg-blue-500 hover:bg-blue-600 text-white text-sm font-semibold py-2 px-3 rounded-lg transition-colors shadow-sm">Refresh</button>
                </div>
                <div class="flex items-center gap-4">
                    <div class="flex items-center gap-2">
                        <button id="pauseBtn" class="text-sm font-semibold py-2 px-3 rounded-lg transition-colors shadow-sm bg-yellow-500 hover:bg-yellow-600 text-white">Pause</button>
                        <button id="clearAllBtn" class="text-sm font-semibold py-2 px-3 rounded-lg transition-colors shadow-sm bg-red-500 hover:bg-red-600 text-white">Clear All</button>
                    </div>
                    <div class="text-right p-2.5 rounded-lg bg-gray-50 border min-w-[250px]">
                        <p class="text-sm font-semibold text-gray-700 mb-1 border-b pb-1">Connected Devices:</p>
                        <div id="deviceList" class="text-sm text-gray-600">
                             <p class="text-orange-500">WAITING...</p>
                        </div>
                    </div>
                     <div>
                        <label for="deviceFilter" class="block text-xs font-medium text-gray-700">Filter by Device:</label>
                        <select id="deviceFilter" class="mt-1 block w-full pl-3 pr-10 py-1.5 text-sm border-gray-300 focus:outline-none focus:ring-indigo-500 focus:border-indigo-500 rounded-md">
                            <option value="all">All Devices</option>
                        </select>
                    </div>
                </div>
            </div>
        </div>

        <!-- TABS & ACTIONS -->
        <div class="mb-3 flex-shrink-0">
            <div class="flex justify-between items-end border-b border-gray-200">
                <div class="flex flex-wrap">
                    <button id="tabBtnLoadAds" class="tab-btn active text-sm font-semibold py-2 px-4 -mb-px border-b-2 border-transparent" onclick="switchTab('LoadAds')">Load Ads</button>
                    <button id="tabBtnLoadAdsExt" class="tab-btn text-sm font-semibold py-2 px-4 -mb-px border-b-2 border-transparent" onclick="switchTab('LoadAdsExt')">Load Ads Ext (CP, KN)</button>
                    
                    <button id="tabBtnValidator" class="tab-btn text-sm font-semibold py-2 px-4 -mb-px border-b-2 border-transparent" onclick="switchTab('Validator')">Default Events/Params</button>
                    <button id="tabBtnSpecific" class="tab-btn text-sm font-semibold py-2 px-4 -mb-px border-b-2 border-transparent" onclick="switchTab('Specific')">Specific Validator</button>
                    <button id="tabBtnAdRevenue" class="tab-btn text-sm font-semibold py-2 px-4 -mb-px border-b-2 border-transparent" onclick="switchTab('AdRevenue')">AdRevenue</button>
                    <button id="tabBtnCallbackAd" class="tab-btn text-sm font-semibold py-2 px-4 -mb-px border-b-2 border-transparent" onclick="switchTab('CallbackAd')">CallBack & Ads</button>
                    <button id="tabBtnSdkCheck" class="tab-btn text-sm font-semibold py-2 px-4 -mb-px border-b-2 border-transparent" onclick="switchTab('SdkCheck')">SDK Check</button>
                    <button id="tabBtnPackage" class="tab-btn text-sm font-semibold py-2 px-4 -mb-px border-b-2 border-transparent" onclick="switchTab('Package')">Package Logcat</button>
                </div>
            </div>
        </div>

        <!-- CONTENT -->
        <div class="flex-grow overflow-auto">

            <!-- TAB 1: Load Ads -->
            <div id="tabContentLoadAds">
                 <div class="bg-white rounded-xl shadow-md p-4">
                    <div class="flex items-center gap-2 bg-gray-50 p-2.5 rounded-lg border mb-3">
                        <span class="text-sm font-semibold text-gray-700">Record Load Ads:</span>
                        <input type="text" id="sheetName_LoadAds" placeholder="Tên Sheet..." class="border p-2 rounded text-sm w-48 outline-none">
                        <button id="btnRecord_LoadAds" onclick="toggleRecord('LoadAds')" class="bg-indigo-600 hover:bg-indigo-700 text-white font-bold py-2 px-4 rounded shadow text-sm">Start Record</button>
                    </div>

                    <h2 class="text-lg font-semibold mb-3">Load Ads Logs</h2>
                    <div class="overflow-x-auto">
                        <table id="packageLogTable" class="min-w-full bg-white">
                            <colgroup>
                                <col class="col-time">
                                <col class="col-tag">
                                <col class="col-message">
                            </colgroup>
                            <thead class="bg-gray-50 sticky top-0 z-10">
                                <tr>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Device</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Ad_source</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Format</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Raw Log</th>
                                </tr>
                            </thead>
                            <tbody id="loadAdsTableBody" class="divide-y divide-gray-200"></tbody>
                        </table>
                    </div>
                </div>
            </div>

            <!-- TAB 2: Load Ads Ext (CP, KN)-->
            <div id="tabContentLoadAdsExt" class="hidden">
                 <div class="bg-white rounded-xl shadow-md p-4">
                    <div class="flex items-center gap-2 bg-gray-50 p-2.5 rounded-lg border mb-3">
                        <span class="text-sm font-semibold text-gray-700">Record Load Ads Ext:</span>
                        <input type="text" id="sheetName_LoadAdsExt" placeholder="Tên Sheet..." class="border p-2 rounded text-sm w-48 outline-none">
                        <button id="btnRecord_LoadAdsExt" onclick="toggleRecord('LoadAdsExt')" class="bg-indigo-600 hover:bg-indigo-700 text-white font-bold py-2 px-4 rounded shadow text-sm">Start Record</button>
                    </div>

                    <h2 class="text-lg font-semibold mb-3">Load Ads Ext Logs</h2>
                    <div class="overflow-x-auto">
                        <table class="min-w-full bg-white">
                            <thead class="bg-gray-50 sticky top-0 z-10">
                                <tr>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Device</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Ad_source</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Format</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Raw Log</th>
                                </tr>
                            </thead>
                            <tbody id="loadAdsExtTableBody" class="divide-y divide-gray-200"></tbody>
                        </table>
                    </div>
                </div>
            </div>

            <!-- TAB 3: Validator -->
            <div id="tabContentValidator" class="hidden">
                <div class="bg-white rounded-xl shadow-md p-4 mb-4">
                    <div class="grid grid-cols-1 lg:grid-cols-3 gap-5">
                        <div class="lg:col-span-1">
                            <div class="space-y-4">
                                <div>
                                    <div class="flex items-center gap-3 mb-2">
                                        <label for="profileSelect" class="text-xs font-medium text-gray-700">Game Profile:</label>
                                        <p id="profileGameText" class="text-xs font-medium text-indigo-700"></p>
                                    </div>
                                    <div class="space-y-3">
                                        <select id="profileSelect" class="w-full h-10 px-3 border rounded-md shadow-sm text-sm"></select>
                                        <input type="file" id="profileFileInput" accept=".xlsx" class="hidden">
                                        <div class="flex flex-wrap items-center gap-3">
                                            <button id="importProfileBtn" class="bg-slate-700 hover:bg-slate-800 text-white font-semibold text-sm px-4 rounded-lg h-10 min-w-[132px]">Import Profile</button>
                                            <button id="reloadProfileBtn" class="bg-slate-200 hover:bg-slate-300 text-gray-800 font-semibold text-sm px-4 rounded-lg h-10 min-w-[132px]">Reload Profile</button>
                                        </div>
                                    </div>
                                </div>
                                <div>
                                    <label for="validatorEventFilterInput" class="block text-xs font-medium text-gray-700 mb-1">Filter by Event Name:</label>
                                    <input type="text" id="validatorEventFilterInput" class="w-full p-2 border rounded-md shadow-sm" placeholder="Type to filter events...">
                                </div>
                                <div>
                                    <label for="paramInput" class="block text-xs font-medium text-gray-700 mb-1">Required Parameters (one per line):</label>
                                    <textarea id="paramInput" rows="6" class="w-full p-2 border rounded-md shadow-sm" placeholder="session_id\nfirst_open_time..."></textarea>
                                </div>
                                <div class="flex items-center gap-3">
                                    <button id="startValidationBtn" class="bg-indigo-600 hover:bg-indigo-700 text-white font-bold text-sm px-5 rounded-lg h-10">Start Checking</button>
                                    <button id="clearValidatorFilterBtn" class="bg-gray-200 hover:bg-gray-300 text-gray-800 font-semibold text-sm px-5 rounded-lg h-10">Clear Filter</button>
                                </div>
                            </div>
                        </div>
                        <div class="lg:col-span-2">
                            <div class="text-xs font-semibold text-gray-700 mb-2">Default Events Status</div>
                            <div id="defaultEventStatusList" class="grid grid-cols-1 sm:grid-cols-2 lg:grid-cols-3 gap-2 text-xs"></div>
                        </div>
                    </div>
                </div>
                <div class="bg-white rounded-xl shadow-md p-4">
                    <div class="overflow-x-auto">
                        <table class="min-w-full bg-white">
                           <thead class="bg-gray-50 sticky top-0 z-10">
                                <tr>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Device</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Status</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Event Name</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Details</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Raw Log</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Action</th>
                                </tr>
                            </thead>
                            <tbody id="validatorTableBody" class="divide-y divide-gray-200"></tbody>
                        </table>
                    </div>
                </div>
            </div>

            <!-- TAB 4: Specific -->
            <div id="tabContentSpecific" class="hidden">
                <div class="bg-white rounded-xl shadow-md p-4 mb-4">
                    <div class="grid grid-cols-1 md:grid-cols-2 gap-4 items-start">
                         <div>
                            <label for="specificEventInput" class="block text-xs font-medium text-gray-700 mb-1">Filter by Event Names:</label>
                            <textarea id="specificEventInput" rows="4" class="w-full p-2 border rounded-md shadow-sm" placeholder="Leave empty to show all events..."></textarea>
                         </div>
                         <div>
                            <label for="specificParamInput" class="block text-xs font-medium text-gray-700 mb-1">Validate Parameters:</label>
                            <textarea id="specificParamInput" rows="4" class="w-full p-2 border rounded-md shadow-sm" placeholder="Leave empty to display all params..."></textarea>
                         </div>
                    </div>
                </div>
                <div class="bg-white rounded-xl shadow-md p-4">
                    <div class="overflow-x-auto">
                        <table class="min-w-full bg-white">
                           <thead class="bg-gray-50 sticky top-0 z-10">
                                <tr>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Device</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Status</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Event Name</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Details</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Raw Log</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Action</th>
                                </tr>
                            </thead>
                            <tbody id="specificEventTableBody" class="divide-y divide-gray-200"></tbody>
                        </table>
                    </div>
                </div>
            </div>

            <!-- TAB 5: AdRevenue -->
            <div id="tabContentAdRevenue" class="hidden">
                <div class="bg-white rounded-xl shadow-md p-4 mb-4">
                     <div class="grid grid-cols-1 md:grid-cols-2 gap-4 items-start">
                        <div>
                            <label for="adRevenueParamInput" class="block text-xs font-medium text-gray-700 mb-1">Validate Parameters:</label>
                            <textarea id="adRevenueParamInput" rows="6" class="w-full p-2 border rounded-md shadow-sm" placeholder="adRevenue\ncurrency\npayload..."></textarea>
                        </div>
                        <div>
                            <label for="adRevenueFilterInput" class="block text-xs font-medium text-gray-700 mb-1">Filter logs by text:</label>
                            <input type="text" id="adRevenueFilterInput" class="w-full p-2 border rounded-md shadow-sm" placeholder="Search in raw log...">
                        </div>
                    </div>
                </div>
                <div class="bg-white rounded-xl shadow-md p-4">
                    <div class="overflow-x-auto">
                        <table class="min-w-full bg-white">
                           <thead class="bg-gray-50 sticky top-0 z-10">
                                <tr>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Device</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Status</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Event Name</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Details</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Raw Log</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Action</th>
                                </tr>
                            </thead>
                            <tbody id="adRevenueTableBody" class="divide-y divide-gray-200"></tbody>
                        </table>
                    </div>
                </div>
            </div>

            <!-- TAB 6: CallbackAd -->
            <div id="tabContentCallbackAd" class="hidden">
                 <div class="bg-white rounded-xl shadow-md p-4">
                    <div class="mb-3 grid grid-cols-1 md:grid-cols-5 gap-4 items-end">
                        <div>
                            <label class="block text-xs font-medium text-gray-700">Filter by Type:</label>
                            <div class="mt-2 flex flex-wrap gap-x-4 gap-y-2">
                                <div class="flex items-center">
                                    <input id="callbackTypeAll" name="callbackType" type="radio" value="all" checked class="h-4 w-4 text-indigo-600 border-gray-300 focus:ring-indigo-500">
                                    <label for="callbackTypeAll" class="ml-2 block text-sm text-gray-900">All</label>
                                </div>
                                <div class="flex items-center">
                                    <input id="callbackTypeCallback" name="callbackType" type="radio" value="callback" class="h-4 w-4 text-indigo-600 border-gray-300 focus:ring-indigo-500">
                                    <label for="callbackTypeCallback" class="ml-2 block text-sm text-gray-900">Callback</label>
                                </div>
                                <div class="flex items-center">
                                    <input id="callbackTypeGadsme" name="callbackType" type="radio" value="gadsme_callback" class="h-4 w-4 text-indigo-600 border-gray-300 focus:ring-indigo-500">
                                    <label for="callbackTypeGadsme" class="ml-2 block text-sm text-gray-900">Callback Gadsme</label>
                                </div>
                                <div class="flex items-center">
                                    <input id="callbackTypeAdverty5" name="callbackType" type="radio" value="adverty5_callback" class="h-4 w-4 text-indigo-600 border-gray-300 focus:ring-indigo-500">
                                    <label for="callbackTypeAdverty5" class="ml-2 block text-sm text-gray-900">Callback Adverty5</label>
                                </div>
                                <div class="flex items-center">
                                    <input id="callbackTypeAdEvent" name="callbackType" type="radio" value="ad_event" class="h-4 w-4 text-indigo-600 border-gray-300 focus:ring-indigo-500">
                                    <label for="callbackTypeAdEvent" class="ml-2 block text-sm text-gray-900">Ad Event</label>
                                </div>
                            </div>
                        </div>
                        <div class="md:col-span-2 md:pl-6">
                            <label for="callbackAdFilterInput" class="block text-sm font-medium text-gray-700">Filter (in raw log):</label>
                            <input type="text" id="callbackAdFilterInput" class="mt-1 block w-full p-2 border border-gray-300 rounded-md shadow-sm focus:ring-indigo-500 focus:border-indigo-500" placeholder="Search...">
                        </div>
                    </div>
                    <div class="overflow-x-auto">
                        <table class="min-w-full bg-white">
                            <thead class="bg-gray-50 sticky top-0 z-10">
                                <tr>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Device</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Type</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Event / Key</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Details</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Raw Log</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Action</th>
                                </tr>
                            </thead>
                            <tbody id="callbackAdTableBody" class="divide-y divide-gray-200"></tbody>
                        </table>
                    </div>
                </div>
            </div>

            <!-- TAB 7: SDK Check -->
            <div id="tabContentSdkCheck" class="hidden">
                <div class="bg-white rounded-xl shadow-md p-4 mb-4">
                    <h2 class="text-lg font-semibold mb-2">SDK Check Setup</h2>
                    <div class="grid grid-cols-1 md:grid-cols-3 gap-4 items-end">
                         <div class="md:col-span-2">
                            <label for="sdkCheckInput" class="block text-xs font-medium text-gray-700 mb-1">SDKs to Check (dán nội dung):</label>
                            <textarea id="sdkCheckInput" rows="8" class="w-full p-2 border rounded-md shadow-sm font-mono text-sm" placeholder="AppLovin\n&quot;Adapter 4.3.54&quot;, &quot;search_pattern&quot;: &quot;Adapter 4.3.54&quot;\n..."></textarea>
                         </div>
                         <div>
                            <button id="startSdkCheckBtn" class="bg-indigo-600 hover:bg-indigo-700 text-white text-sm font-semibold py-2 px-4 rounded-lg w-full h-10">Start Checking</button>
                         </div>
                    </div>
                </div>
                <div class="bg-white rounded-xl shadow-md p-4">
                    <h2 class="text-lg font-semibold mb-2">SDK Check Results</h2>
                    <div class="overflow-x-auto">
                        <table class="min-w-full bg-white">
                            <tbody id="sdkCheckTableBody" class="divide-y divide-gray-200"></tbody>
                        </table>
                    </div>
                </div>
            </div>
            
            <!-- TAB 8: Package -->
            <div id="tabContentPackage" class="hidden">
                 <div class="bg-white rounded-xl shadow-md p-4 mb-4">
                     <div class="grid grid-cols-1 lg:grid-cols-4 gap-4 items-start">
                        <div>
                            <label for="packageIdInput" class="block text-xs font-medium text-gray-700 mb-1">Package ID:</label>
                            <input type="text" id="packageIdInput" class="w-full p-2 border rounded-md shadow-sm" placeholder="com.example.app">
                            <div class="flex justify-center mt-3">
                                <button id="startPackageLogBtn" class="bg-indigo-600 hover:bg-indigo-700 text-white font-bold py-2 px-6 rounded-lg h-10">Start</button>
                            </div>
                            <div class="flex items-center mt-4 space-x-4">
                                <div class="flex items-center">
                                    <input id="showErrorsOnly" type="checkbox" class="h-4 w-4 text-indigo-600 border-gray-300 rounded focus:ring-indigo-500">
                                    <label for="showErrorsOnly" class="ml-2 block text-sm text-gray-900">Show errors only</label>
                                </div>
                                <div class="flex items-center">
                                    <input id="autoScroll" type="checkbox" checked class="h-4 w-4 text-indigo-600 border-gray-300 rounded focus:ring-indigo-500">
                                    <label for="autoScroll" class="ml-2 block text-sm text-gray-900">Auto-scroll</label>
                                </div>
                            </div>
                        </div>
                        <div>
                            <label class="block text-xs font-medium text-gray-700 mb-1">Quick Select:</label>
                            <div class="grid grid-cols-1 gap-1 text-sm text-gray-700">
                                <label class="inline-flex items-center gap-2">
                                    <input type="checkbox" class="package-id-checkbox h-4 w-4 text-indigo-600 border-gray-300 rounded focus:ring-indigo-500" value="com.indiez.nonogram">
                                    <span>NG - com.indiez.nonogram</span>
                                </label>
                                <label class="inline-flex items-center gap-2">
                                    <input type="checkbox" class="package-id-checkbox h-4 w-4 text-indigo-600 border-gray-300 rounded focus:ring-indigo-500" value="com.indiez.train.miner">
                                    <span>TM - com.indiez.train.miner</span>
                                </label>
                                <label class="inline-flex items-center gap-2">
                                    <input type="checkbox" class="package-id-checkbox h-4 w-4 text-indigo-600 border-gray-300 rounded focus:ring-indigo-500" value="com.indiez.idletycoon.horse.racing">
                                    <span>HR - com.indiez.idletycoon.horse.racing</span>
                                </label>
                                <label class="inline-flex items-center gap-2">
                                    <input type="checkbox" class="package-id-checkbox h-4 w-4 text-indigo-600 border-gray-300 rounded focus:ring-indigo-500" value="com.indiez.solitaire.food">
                                    <span>SC - com.indiez.solitaire.food</span>
                                </label>
                            </div>
                        </div>
                        <div class="lg:col-span-2">
                            <div class="grid grid-cols-2 gap-3 items-start">
                                <div>
                                    <label for="packageTagFilterInput" class="block text-xs font-medium text-gray-700 mb-1">Tag Filter:</label>
                                    <input type="text" id="packageTagFilterInput" class="w-full p-2 border rounded-md shadow-sm" placeholder="Tag...">
                                    <div class="mt-2 grid grid-cols-2 gap-4 text-sm text-gray-700">
                                        <label class="inline-flex items-center gap-2">
                                            <input type="radio" name="tagQuickFilter" value="" checked class="h-4 w-4 text-indigo-600 border-gray-300 focus:ring-indigo-500">
                                            <span>All</span>
                                        </label>
                                        <label class="inline-flex items-center gap-2">
                                            <input type="radio" name="tagQuickFilter" value="appsflyer" class="h-4 w-4 text-indigo-600 border-gray-300 focus:ring-indigo-500">
                                            <span>Appsflyer</span>
                                        </label>
                                        <label class="inline-flex items-center gap-2">
                                            <input type="radio" name="tagQuickFilter" value="integrationhelper" class="h-4 w-4 text-indigo-600 border-gray-300 focus:ring-indigo-500">
                                            <span>Integrationhelper</span>
                                        </label>
                                        <label class="inline-flex items-center gap-2">
                                            <input type="radio" name="tagQuickFilter" value="appmetrica" class="h-4 w-4 text-indigo-600 border-gray-300 focus:ring-indigo-500">
                                            <span>Appmetrica</span>
                                        </label>
                                    </div>
                                </div>
                                <div>
                                    <label for="packageFilterInput" class="block text-xs font-medium text-gray-700 mb-1">Message Filter 1:</label>
                                    <input type="text" id="packageFilterInput" class="w-full p-2 border rounded-md shadow-sm" placeholder="Search text 1...">
                                    <div class="mt-2">
                                        <label for="packageFilterInput2" class="block text-xs font-medium text-gray-700 mb-1">Message Filter 2:</label>
                                        <input type="text" id="packageFilterInput2" class="w-full p-2 border rounded-md shadow-sm" placeholder="Search text 2...">
                                    </div>
                                </div>
                            </div>
                        </div>
                        <div>
                             <button id="copyPackageBtn" class="hidden text-sm font-semibold py-2 px-3 rounded-lg transition-colors shadow-sm bg-blue-500 hover:bg-blue-600 text-white">Copy Selected</button>
                        </div>
                    </div>
                </div>
                <div class="bg-white rounded-xl shadow-md p-4">
                    <h2 class="text-lg font-semibold mb-2">Package Log Stream</h2>
                    <div id="packageLogContainer" class="overflow-auto overflow-x-auto" style="height: 50vh;">
                        <table class="min-w-full bg-white">
                           <thead class="bg-gray-50 sticky top-0 z-10">
                                <tr>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-2 border-b time-header resizable col-time">Time<div class="resizer" data-col="time"></div></th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 pr-1 pl-2 border-b tag-header resizable col-tag">Tag<div class="resizer" data-col="tag"></div></th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 pl-1 pr-3 border-b resizable col-message">Message<div class="resizer" data-col="message"></div></th>
                                </tr>
                            </thead>
                            <tbody id="packageLogTableBody" class="divide-y divide-gray-200"></tbody>
                        </table>
                    </div>
                </div>
            </div>

        </div>
    </div>
    
    <!-- MODALS -->
    <div id="jsonModal" class="fixed inset-0 z-50 flex items-center justify-center bg-black bg-opacity-50 hidden">
        <div class="bg-white rounded-xl shadow-lg p-6 w-1/2 h-2/3 flex flex-col">
            <div class="flex justify-between items-center border-b pb-3 mb-4">
                <h2 class="text-xl font-bold text-gray-800">Formatted JSON</h2>
                <button id="closeJsonModal" class="text-gray-500 hover:text-gray-800 text-2xl font-bold">&times;</button>
            </div>
            <div class="flex-grow overflow-auto bg-gray-800 text-white p-4 rounded-md">
                <pre><code id="jsonContent"></code></pre>
            </div>
        </div>
    </div>

    <div id="logDetailModal" class="fixed inset-0 z-50 flex items-center justify-center bg-black bg-opacity-50 hidden">
        <div class="bg-white rounded-xl shadow-lg p-6 w-3/4 h-2/3 flex flex-col">
            <div class="flex justify-between items-center border-b pb-3 mb-4">
                <h2 class="text-xl font-bold text-gray-800">Log Details</h2>
                <div class="flex items-center gap-3">
                    <button id="convertLogJsonBtn" class="text-sm bg-indigo-600 hover:bg-indigo-700 text-white font-semibold py-2 px-3 rounded">Convert to JSON</button>
                    <button id="closeLogDetailModal" class="text-gray-500 hover:text-gray-800 text-2xl font-bold">&times;</button>
                </div>
            </div>
            <div class="flex-grow overflow-auto bg-gray-50 p-4 rounded-md border">
                <pre id="logDetailContent" class="text-sm font-mono whitespace-pre-wrap text-gray-800"></pre>
            </div>
        </div>
    </div>

    <!-- SCRIPTS -->
    <script>
        const socket = io();
        let currentTab = 'LoadAds';
        let selectedDevice = 'all';

        // --- Common Elements ---
        const deviceListEl = document.getElementById('deviceList');
        const pauseBtn = document.getElementById('pauseBtn');
        const jsonModal = document.getElementById('jsonModal');
        const closeJsonModal = document.getElementById('closeJsonModal');
        const jsonContent = document.getElementById('jsonContent');
        const logDetailModal = document.getElementById('logDetailModal');
        const closeLogDetailModal = document.getElementById('closeLogDetailModal');
        const logDetailContent = document.getElementById('logDetailContent');
        const convertLogJsonBtn = document.getElementById('convertLogJsonBtn');
        const deviceFilter = document.getElementById('deviceFilter');
        const clearAllBtn = document.getElementById('clearAllBtn');
        const restartAppBtn = document.getElementById('restartAppBtn');

        // --- Tab Logic ---
        function switchTab(tabName) {
            currentTab = tabName;
            // Hide all contents with Safety Check
            ['tabContentLoadAds', 'tabContentLoadAdsExt', 'tabContentValidator', 'tabContentSpecific', 'tabContentAdRevenue', 'tabContentCallbackAd', 'tabContentSdkCheck', 'tabContentPackage'].forEach(id => {
                const el = document.getElementById(id);
                if (el) el.classList.add('hidden');
                else console.warn('Missing element ID:', id);
            });
            // Deactivate all buttons
            document.querySelectorAll('.tab-btn').forEach(btn => btn.classList.remove('active'));
            
            // Activate selected with Safety Check
            const contentEl = document.getElementById('tabContent' + tabName);
            if (contentEl) contentEl.classList.remove('hidden');
            
            const btnEl = document.getElementById('tabBtn' + tabName);
            if (btnEl) btnEl.classList.add('active');
            
            socket.emit('change_tab', { tab_name: tabName });
        }

        // --- Global Control Logic ---
        pauseBtn.addEventListener('click', () => {
            // Optimistic toggle for resizer usability
            const willPause = pauseBtn.textContent === 'Pause';
            setResizerEnabled(willPause);
            isPausedClient = willPause;
            socket.emit('toggle_pause');
        });
        clearAllBtn.addEventListener('click', () => {
            if (confirm('Are you sure you want to clear ALL logs?')) {
                socket.emit('clear_all_logs');
            }
        });

        restartAppBtn?.addEventListener('click', () => {
            if (confirm('Restart app now to check for updates?')) {
                fetch('/restart_app', { method: 'POST' });
            }
        });
        
        // --- Recording Logic (Updated for separate tabs) ---
        function toggleRecord(tabName) {
            const btn = document.getElementById('btnRecord_' + tabName);
            const input = document.getElementById('sheetName_' + tabName);
            
            if (!btn.classList.contains('bg-red-600')) {
                const name = input.value.trim();
                socket.emit('toggle_record', { sheet_name: name, tab_name: tabName });
            } else { 
                socket.emit('toggle_record', { tab_name: tabName }); 
            }
        }

        socket.on('record_status', (s) => {
            // Update UI based on which tab triggered the status
            const tabName = s.tab_name;
            const btn = document.getElementById('btnRecord_' + tabName);
            if (btn) {
                if (s.is_recording) {
                    btn.textContent = 'Stop Recording: ' + (s.current_sheet || 'Running');
                    btn.className = 'bg-red-600 text-white font-bold py-2 px-4 rounded animate-record shadow-lg text-sm';
                } else {
                    btn.textContent = 'Start Record';
                    btn.className = 'bg-indigo-600 hover:bg-indigo-700 text-white font-bold py-2 px-4 rounded shadow text-sm';
                }
            }
        });

        let isPausedClient = false;
        function setResizerEnabled(enabled) {
            document.querySelectorAll('.resizer').forEach(r => {
                r.classList.toggle('disabled', !enabled);
                r.style.pointerEvents = enabled ? 'auto' : 'none';
            });
        }

        socket.on('pause_status', (data) => {
            isPausedClient = !!data.is_paused;
            if (data.is_paused) {
                pauseBtn.textContent = 'Resume';
                pauseBtn.classList.replace('bg-yellow-500', 'bg-blue-500');
            } else {
                pauseBtn.textContent = 'Pause';
                pauseBtn.classList.replace('bg-blue-500', 'bg-yellow-500');
            }
            setResizerEnabled(isPausedClient);
        });

        // --- Rendering Helpers ---
        function escapeHTML(str) {
            if (!str) return '';
            const p = document.createElement("p");
            p.textContent = str;
            return p.innerHTML.replace(/  /g, ' &nbsp;');
        }
        
        // FIX: Better attribute escaping for JSON buttons
        function escapeAttribute(str) {
            if (!str) return '{}';
            // Replace single quotes with HTML entity to prevent breaking the attribute
            return str.replace(/'/g, '&#39;');
        }

        function renderSimpleTable(id, data) {
            const tbody = document.getElementById(id);
            if (!tbody) return;
            const filtered = (selectedDevice === 'all') ? data : data.filter(e => e.device_id === selectedDevice);
            
            if (filtered.length === 0) {
                 tbody.innerHTML = '<tr><td colspan="4" class="text-center py-4 text-gray-400">Waiting for recording...</td></tr>';
                 return;
            }

            tbody.innerHTML = filtered.map(e => `
                <tr class="hover:bg-gray-50 border-b text-sm">
                    <td class="py-2 px-4 text-purple-700 font-medium">${e.device_name}</td>
                    <td class="py-2 px-4 text-blue-600 font-semibold">${e.ad_source}</td>
                    <td class="py-2 px-4 text-green-600 font-semibold">${e.ad_format}</td>
                    <td class="py-2 px-4 log-cell">${e.raw_log}</td>
                </tr>
            `).join('');
        }

        let validator_results_cache = [];
        let defaultEventNames = {{ default_event_names | tojson }};
        let defaultEventStatusEls = {};
        let currentProfileName = {{ current_profile_name | tojson }};

        function updateProfileStatus(payload) {
            const gameEl = document.getElementById('profileGameText');
            if (gameEl) {
                gameEl.textContent = payload.game_name || 'Unknown';
            }
        }

        function renderProfileOptions(payload) {
            const select = document.getElementById('profileSelect');
            if (!select) return;
            const profiles = payload.profiles || [];
            select.innerHTML = profiles.length
                ? profiles.map(name => `<option value="${escapeAttribute(name)}"${name === payload.current_profile ? ' selected' : ''}>${escapeHTML(name)}</option>`).join('')
                : '<option value="">No profiles</option>';
            select.disabled = profiles.length === 0;
            currentProfileName = payload.current_profile || '';
            defaultEventNames = payload.default_event_names || [];
            renderDefaultEventStatusList();
            updateDefaultEventStatus(validator_results_cache);
            updateProfileStatus(payload);
        }

        async function refreshProfiles() {
            const res = await fetch('/api/profiles');
            const payload = await res.json();
            renderProfileOptions(payload);
        }

        function renderDefaultEventStatusList() {
            const container = document.getElementById('defaultEventStatusList');
            if (!container) return;
            if (!defaultEventNames || defaultEventNames.length === 0) {
                container.innerHTML = '<div class="text-gray-400">No default events</div>';
                return;
            }
            container.innerHTML = defaultEventNames.map(name => `
                <div class="default-event-item flex items-center gap-2 px-3 py-2 rounded-lg border border-gray-200 bg-white shadow-sm cursor-pointer hover:bg-gray-50"
                     data-event-name="${escapeAttribute(name)}" title="Click to filter by this event">
                    <span class="event-status-icon text-gray-400 font-bold w-4 text-center" data-event="${name}" title="checking">...</span>
                    <span class="truncate font-medium text-gray-700" title="${escapeHTML(name)}">${escapeHTML(name)}</span>
                </div>
            `).join('');
            defaultEventStatusEls = {};
            container.querySelectorAll('.event-status-icon').forEach(el => {
                defaultEventStatusEls[el.dataset.event] = el;
            });
        }

        function updateDefaultEventStatus(results) {
            if (!defaultEventNames || defaultEventNames.length === 0) return;
            const statusMap = {};
            defaultEventNames.forEach(n => { statusMap[n] = 'pending'; });
            results.forEach(r => {
                if (!statusMap.hasOwnProperty(r.event_name)) return;
                if (r.status === 'FAILED') statusMap[r.event_name] = 'failed';
                else if (r.status === 'PASSED' && statusMap[r.event_name] !== 'failed') statusMap[r.event_name] = 'passed';
            });
            Object.entries(statusMap).forEach(([name, status]) => {
                const el = defaultEventStatusEls[name];
                if (!el) return;
                if (status === 'passed') {
                    el.textContent = '✓';
                    el.className = 'event-status-icon text-green-600 font-bold';
                    el.title = 'PASSED';
                } else if (status === 'failed') {
                    el.textContent = '✕';
                    el.className = 'event-status-icon text-red-600 font-bold';
                    el.title = 'FAILED';
                } else {
                    el.textContent = '...';
                    el.className = 'event-status-icon text-gray-400 font-bold';
                    el.title = 'checking';
                }
            });
        }

        function renderValidatorTable(results) {
             const tbody = document.getElementById('validatorTableBody');
             if (!tbody) return;
             const filterText = (document.getElementById('validatorEventFilterInput')?.value || '').toLowerCase().trim();
             const filtered = results.filter(r => {
                 if (selectedDevice !== 'all' && r.device_id !== selectedDevice) return false;
                 if (filterText && !(r.event_name || '').toLowerCase().includes(filterText)) return false;
                 return true;
             });
             
             if (filtered.length === 0) {
                 tbody.innerHTML = '<tr><td colspan="6" class="text-center py-4">Waiting...</td></tr>';
             } else {
                 tbody.innerHTML = filtered.map(res => `
                    <tr class="hover:bg-gray-50 border-b text-sm">
                        <td class="py-2 px-3 text-purple-700 text-sm">${res.device_name}</td>
                        <td class="py-2 px-3 text-sm font-semibold ${res.status === 'PASSED' ? 'text-green-600' : 'text-red-600'}">${res.status}</td>
                        <td class="py-2 px-3"><span class="event-name-link cursor-pointer text-sm font-medium text-indigo-700 hover:underline" data-event-name="${escapeAttribute(res.event_name)}">${res.event_name}</span></td>
                        <td class="py-2 px-3 details-cell text-sm">${res.details}</td>
                        <td class="py-2 px-3 log-cell text-sm">${res.raw_log}</td>
                        <td class="py-2 px-3"><button class="view-json-btn text-xs bg-indigo-100 hover:bg-indigo-200 text-indigo-700 font-medium py-1 px-2 rounded" data-json='${escapeAttribute(res.json_data)}'>View JSON</button></td>
                    </tr>
                 `).join('');
             }
        }

        renderDefaultEventStatusList();
        refreshProfiles().catch(() => {});

        // Click default event to fill filter input
        document.getElementById('defaultEventStatusList')?.addEventListener('click', (e) => {
            const item = e.target.closest('.default-event-item');
            if (!item) return;
            const name = item.getAttribute('data-event-name') || '';
            const input = document.getElementById('validatorEventFilterInput');
            if (input) {
                input.value = name;
                input.focus();
                renderValidatorTable(validator_results_cache);
            }
        });


        // Click event name in results table to fill filter input
        document.getElementById('validatorTableBody')?.addEventListener('click', (e) => {
            const el = e.target.closest('.event-name-link');
            if (!el) return;
            const name = el.getAttribute('data-event-name') || '';
            const input = document.getElementById('validatorEventFilterInput');
            if (input) {
                input.value = name;
                input.focus();
                renderValidatorTable(validator_results_cache);
            }
        });

        document.getElementById('profileSelect')?.addEventListener('change', async (e) => {
            const profileName = e.target.value;
            if (!profileName || profileName === currentProfileName) return;
            const res = await fetch('/api/profiles/select', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({ profile_name: profileName })
            });
            const payload = await res.json();
            if (!payload.ok) {
                alert(payload.error || 'Failed to switch profile');
                await refreshProfiles();
                return;
            }
            validator_results_cache = [];
            renderValidatorTable(validator_results_cache);
            renderProfileOptions(payload);
        });

        document.getElementById('reloadProfileBtn')?.addEventListener('click', async () => {
            const res = await fetch('/api/profiles/reload', { method: 'POST' });
            const payload = await res.json();
            if (!payload.ok) {
                alert(payload.error || 'Failed to reload profile');
                return;
            }
            renderProfileOptions(payload);
        });

        document.getElementById('importProfileBtn')?.addEventListener('click', () => {
            document.getElementById('profileFileInput')?.click();
        });

        document.getElementById('profileFileInput')?.addEventListener('change', async (e) => {
            const file = e.target.files && e.target.files[0];
            if (!file) return;
            const formData = new FormData();
            formData.append('profile_file', file);
            const res = await fetch('/api/profiles/import', {
                method: 'POST',
                body: formData
            });
            const payload = await res.json();
            e.target.value = '';
            if (!payload.ok) {
                alert(payload.error || 'Failed to import profile');
                return;
            }
            validator_results_cache = [];
            renderValidatorTable(validator_results_cache);
            renderProfileOptions(payload);
        });

        // --- Socket Listeners (Renderers) ---
        socket.on('update_load_ads', (d) => renderSimpleTable('loadAdsTableBody', d));
        socket.on('update_load_ads_ext', (d) => renderSimpleTable('loadAdsExtTableBody', d));
        socket.on('update_validator_table', (d) => {
            validator_results_cache = d || [];
            renderValidatorTable(validator_results_cache);
            updateDefaultEventStatus(validator_results_cache);
        });

        socket.on('update_specific_event_table', (d) => {
             const tbody = document.getElementById('specificEventTableBody');
             if (!tbody) return;
             const filtered = (selectedDevice === 'all') ? d : d.filter(r => r.device_id === selectedDevice);
             if(filtered.length === 0) { tbody.innerHTML = '<tr><td colspan="6" class="text-center py-4">Waiting...</td></tr>'; }
             else {
                 tbody.innerHTML = filtered.map(res => `<tr class="hover:bg-gray-50 border-b text-sm"><td class="py-2 px-3 text-purple-700 text-sm">${res.device_name}</td><td class="py-2 px-3 text-sm font-semibold ${res.status === 'PASSED'?'text-green-600':'text-red-600'}">${res.status}</td><td class="py-2 px-3"><span class="event-name-link cursor-pointer text-sm font-medium text-indigo-700 hover:underline" data-event-name="${escapeAttribute(res.event_name)}">${res.event_name}</span></td><td class="py-2 px-3 details-cell text-sm">${res.details}</td><td class="py-2 px-3 log-cell text-sm">${res.raw_log}</td><td class="py-2 px-3"><button class="view-json-btn text-xs bg-indigo-100 hover:bg-indigo-200 text-indigo-700 font-medium py-1 px-2 rounded" data-json='${escapeAttribute(res.json_data)}'>View JSON</button></td></tr>`).join('');
             }
        });

         socket.on('update_adrevenue_table', (d) => {
             const tbody = document.getElementById('adRevenueTableBody');
             if (!tbody) return;
             const filterText = document.getElementById('adRevenueFilterInput').value.toLowerCase();
             const filtered = d.filter(r => (selectedDevice === 'all' || r.device_id === selectedDevice) && (!filterText || r.raw_log.toLowerCase().includes(filterText)));
             if(filtered.length === 0) { tbody.innerHTML = '<tr><td colspan="6" class="text-center py-4">Waiting...</td></tr>'; }
             else {
                 tbody.innerHTML = filtered.map(res => `<tr class="hover:bg-gray-50 border-b text-sm"><td class="py-2 px-3 text-purple-700 text-sm">${res.device_name}</td><td class="py-2 px-3 text-sm font-semibold ${res.status === 'PASSED'?'text-green-600':'text-red-600'}">${res.status}</td><td class="py-2 px-3"><span class="event-name-link cursor-pointer text-sm font-medium text-indigo-700 hover:underline" data-event-name="${escapeAttribute(res.event_name)}">${res.event_name}</span></td><td class="py-2 px-3 details-cell text-sm">${res.details}</td><td class="py-2 px-3 log-cell text-sm">${res.raw_log}</td><td class="py-2 px-3"><button class="view-json-btn text-xs bg-indigo-100 hover:bg-indigo-200 text-indigo-700 font-medium py-1 px-2 rounded" data-json='${escapeAttribute(res.json_data)}'>View JSON</button></td></tr>`).join('');
             }
        });

        // ============================================
        // === FIXED: Callback Table with Client Filter ===
        // ============================================
        
        let lastCallbackData = [];

        socket.on('update_callback_ad_table', (d) => {
            lastCallbackData = d;
            renderCallbackTable();
        });

        function renderCallbackTable() {
            const d = lastCallbackData;
            const tbody = document.getElementById('callbackAdTableBody');
            if (!tbody) return;

            // Get filter values
            const typeFilter = document.querySelector('input[name="callbackType"]:checked').value;
            const textFilter = document.getElementById('callbackAdFilterInput').value.toLowerCase();

            const filtered = d.filter(r => {
                // Device filter
                if (selectedDevice !== 'all' && r.device_id !== selectedDevice) return false;
                
                // Type filter
                if (typeFilter === 'callback' && r.type === 'Ad Event') return false;
                if (typeFilter === 'gadsme_callback' && r.type !== 'Callback Gadsme') return false;
                if (typeFilter === 'adverty5_callback' && r.type !== 'Callback Adverty5') return false;
                if (typeFilter === 'ad_event' && r.type !== 'Ad Event') return false;
                
                // Text filter
                if (textFilter && !r.raw_log.toLowerCase().includes(textFilter)) return false;

                return true;
            });

             if(filtered.length === 0) { tbody.innerHTML = '<tr><td colspan="6" class="text-center py-4">Waiting...</td></tr>'; }
             else {
                 tbody.innerHTML = filtered.map(res => {
                     const nameLower = (res.event_name || '').toLowerCase();
                     const isFailed = nameLower.includes('failed');
                     const isImpression = nameLower.includes('onimpression') || nameLower.includes('_onimpression');
                     const eventClass = isFailed ? 'text-red-600' : (isImpression ? 'text-blue-600' : '');
                     return `<tr class="hover:bg-gray-50 border-b text-sm"><td class="py-2 px-3 text-purple-700 text-sm">${res.device_name}</td><td class="py-2 px-3 text-sm font-semibold ${res.type==='Ad Event'?'text-orange-600':'text-cyan-600'}">${res.type}</td><td class="py-2 px-3 text-sm font-medium ${eventClass}">${res.event_name}</td><td class="py-2 px-3 details-cell text-sm">${res.details}</td><td class="py-2 px-3 log-cell text-sm">${res.raw_log}</td><td class="py-2 px-3"><button class="view-json-btn text-xs bg-indigo-100 hover:bg-indigo-200 text-indigo-700 font-medium py-1 px-2 rounded" data-json='${escapeAttribute(res.json_data)}'>View JSON</button></td></tr>`;
                 }).join('');
             }
        }

        // Add listeners
        document.querySelectorAll('input[name="callbackType"]').forEach(r => r.addEventListener('change', renderCallbackTable));
        document.getElementById('callbackAdFilterInput').addEventListener('input', renderCallbackTable);

        socket.on('update_sdk_check_table', (data) => {
            const tbody = document.getElementById('sdkCheckTableBody');
            if (!tbody) return;
            tbody.innerHTML = data.map(res => {
                 let rowClass = (selectedDevice !== 'all' && res.status !== 'HEADER') ? 'pl-8' : '';
                 let statusText = '';
                 if (res.status === 'PASSED') statusText = '<span class="font-semibold text-green-600"> - PASSED</span>';
                 else if (res.status === 'NOT_FOUND') statusText = '<span class="font-semibold text-red-600"> - Not Found</span>';
                 else if (res.status === 'HEADER') rowClass += ' font-semibold text-sm text-indigo-600 bg-gray-50';
                 else if (res.status === 'LABEL') rowClass += ' font-medium text-sm text-gray-800 pt-2';
                 
                 // Filter
                 if (selectedDevice !== 'all' && res.device_id !== selectedDevice && res.status !== 'LABEL' && res.status !== 'WAITING') return '';
                 
                 return `<tr><td class="py-1 px-4 ${rowClass}"><pre style="font-family: monospace; margin: 0; white-space: pre-wrap;">${escapeHTML(res.display_text)}${statusText}</pre></td></tr>`;
            }).join('');
        });

        let lastPackageLogs = [];
        let selectedPackageRowKeys = new Set();
        function renderPackageLogTable() {
            const tbody = document.getElementById('packageLogTableBody');
            if (!tbody) return;
            const filterText = document.getElementById('packageFilterInput').value.toLowerCase();
            const filterText2 = document.getElementById('packageFilterInput2').value.toLowerCase();
            const tagFilter = document.getElementById('packageTagFilterInput').value.toLowerCase();
            const quickTag = document.querySelector('input[name="tagQuickFilter"]:checked')?.value || '';
            const errorsOnly = document.getElementById('showErrorsOnly').checked;
            
            const filtered = lastPackageLogs.filter(l => {
                if (selectedDevice !== 'all' && l.device_id !== selectedDevice) return false;
                if (errorsOnly && !l.is_error) return false;
                const messageHaystack = `${l.message || ''}`.toLowerCase();
                const tagHaystack = `${l.tag || ''}`.toLowerCase();
                if (quickTag && !tagHaystack.includes(quickTag)) return false;
                if (tagFilter && !tagHaystack.includes(tagFilter)) return false;
                if (filterText && !messageHaystack.includes(filterText)) return false;
                if (filterText2 && !messageHaystack.includes(filterText2)) return false;
                return true;
            });
            
            tbody.innerHTML = filtered.map((l, idx) => {
                const msgText = (l.message || l.log || '');
                const isErrorLevel = (l.level === 'E' || l.level === 'F');
                const rowClass = isErrorLevel ? 'text-red-500' : '';
                const msgClass = isErrorLevel ? 'text-red-500' : '';
                const rowKey = `${l.time_display || l.time || ''}||${l.tag || ''}||${msgText}`;
                const selectedClass = selectedPackageRowKeys.has(rowKey) ? 'selected' : '';
                return `<tr class="package-log-row hover:bg-gray-50 ${rowClass} ${selectedClass}" data-row-key="${encodeURIComponent(rowKey)}" data-row-index="${idx}"><td class="py-2 px-2 font-mono text-xs time-cell col-time">${l.time_display || l.time || ''}</td><td class="py-2 pr-1 pl-2 font-mono text-xs tag-cell col-tag" title="${escapeHTML(l.tag || '')}">${l.tag || ''}</td><td class="py-2 pl-1 pr-3 log-cell message-cell col-message ${msgClass}">${msgText}</td></tr>`;
            }).join('');
            if(document.getElementById('autoScroll').checked) document.getElementById('packageLogContainer').scrollTop = document.getElementById('packageLogContainer').scrollHeight;
        }

        socket.on('package_log_cache', (logs) => {
            lastPackageLogs = logs || [];
            renderPackageLogTable();
        });

        // --- JSON Modal Handler ---
        document.body.addEventListener('click', (e) => {
            if (e.target && e.target.classList.contains('view-json-btn')) {
                const jsonData = e.target.getAttribute('data-json'); // Use getAttribute for safety
                try {
                    const parsed = JSON.parse(jsonData);
                    jsonContent.textContent = JSON.stringify(parsed, null, 2);
                    jsonModal.classList.remove('hidden');
                } catch(err) { 
                    console.error(err);
                    alert('Invalid JSON data. Check console for details.'); 
                }
            }
        });
        closeJsonModal.addEventListener('click', () => jsonModal.classList.add('hidden'));
        closeLogDetailModal.addEventListener('click', () => logDetailModal.classList.add('hidden'));

        // --- Log Detail Modal (Package Log) ---
        function getSelectedPackageRows() {
            const selected = Array.from(document.querySelectorAll('#packageLogTableBody tr.package-log-row.selected'));
            if (selected.length > 0) return selected;
            // Fallback: rehydrate from saved keys
            const allRows = Array.from(document.querySelectorAll('#packageLogTableBody tr.package-log-row'));
            const byKey = new Map();
            allRows.forEach(r => {
                const rawKey = r.getAttribute('data-row-key') || '';
                try { byKey.set(decodeURIComponent(rawKey), r); } catch {}
            });
            const fromKeys = [];
            selectedPackageRowKeys.forEach(k => {
                if (byKey.has(k)) fromKeys.push(byKey.get(k));
            });
            return fromKeys;
        }

        function getRowText(row) {
            const cells = row.querySelectorAll('td');
            const parts = Array.from(cells).map(td => (td.textContent || '').trim());
            return parts.join(' ');
        }

        function getRowMessageText(row) {
            const msgCell = row.querySelector('td.col-message');
            return (msgCell ? msgCell.textContent : '').trim();
        }

        let logDetailRawText = '';
        let logDetailIsJsonView = false;

        function openLogDetailModal(rows) {
            const lines = rows.map(getRowText);
            logDetailRawText = lines.join('\\n');
            logDetailIsJsonView = false;
            logDetailContent.textContent = logDetailRawText;
            logDetailModal.classList.remove('hidden');
        }

        function extractJsonFromText(text) {
            if (!text) return null;
            const starts = [];
            for (let i = 0; i < text.length; i++) {
                const ch = text[i];
                if (ch === '{' || ch === '[') starts.push(i);
            }
            for (const start of starts) {
                let openBraces = 0;
                let openBrackets = 0;
                for (let i = start; i < text.length; i++) {
                    const ch = text[i];
                    if (ch === '{') openBraces++;
                    if (ch === '}') openBraces--;
                    if (ch === '[') openBrackets++;
                    if (ch === ']') openBrackets--;
                    if (openBraces === 0 && openBrackets === 0 && i > start) {
                        const candidate = text.slice(start, i + 1);
                        try {
                            JSON.parse(candidate);
                            return candidate;
                        } catch (e) {}
                        break;
                    }
                }
            }
            return null;
        }

        function extractEventName(text) {
            if (!text) return '';
            const m = text.match(/with\\s+name\\s+([\\w\\.:-]+)/i);
            if (m && m[1]) return m[1];
            return '';
        }

        function extractHeaderBeforeJson(text) {
            if (!text) return '';
            const idx = text.indexOf('{');
            if (idx === -1) return '';
            let header = text.slice(0, idx).trim();
            // Remove leading time/tag/prefix if present
            const i1 = header.indexOf('] : ');
            if (i1 !== -1) header = header.slice(i1 + 4).trim();
            const i2 = header.indexOf('] ');
            if (i2 !== -1) header = header.slice(i2 + 2).trim();
            return header;
        }

        function convertSelectedLogsToJson() {
            if (logDetailIsJsonView) {
                logDetailContent.textContent = logDetailRawText;
                logDetailIsJsonView = false;
                return;
            }
            const rows = getSelectedPackageRows();
            if (rows.length === 0) return;
            const outputs = rows.map((row, idx) => {
                const line = getRowText(row);
                const msg = getRowMessageText(row);
                const jsonStr = extractJsonFromText(msg || line);
                const evt = extractEventName(line);
                const headerLine = extractHeaderBeforeJson(line);
                if (jsonStr) {
                    try {
                        const parsed = JSON.parse(jsonStr);
                        const pretty = JSON.stringify(parsed, null, 2);
                        const metaLines = [];
                        if (headerLine) metaLines.push(`context: ${headerLine}`);
                        else if (evt) metaLines.push(`event_name: ${evt}`);
                        metaLines.push('');
                        metaLines.push('raw_log:');
                        metaLines.push(line);
                        metaLines.push('');
                        metaLines.push('extracted_json:');
                        metaLines.push(pretty);
                        return `--- #${idx + 1} (JSON) ---\\n${metaLines.join('\\n')}`;
                    } catch (e) {
                        return `--- #${idx + 1} ---\\n${line}`;
                    }
                }
                return `--- #${idx + 1} ---\\n${line}`;
            });
            logDetailContent.textContent = outputs.join('\\n\\n');
            logDetailIsJsonView = true;
        }

        convertLogJsonBtn.addEventListener('click', convertSelectedLogsToJson);

        document.getElementById('packageLogTableBody').addEventListener('dblclick', (e) => {
            const row = e.target.closest('tr.package-log-row');
            if (!row) return;
            isSelectingRows = false;
            dragStartIndex = null;
            // Ensure clicked row is included
            const rawKey = row.getAttribute('data-row-key') || '';
            try { selectedPackageRowKeys.add(decodeURIComponent(rawKey)); } catch {}
            const selected = getSelectedPackageRows();
            const rowsToShow = selected.length > 0 ? selected : [row];
            openLogDetailModal(rowsToShow);
        });

        document.addEventListener('keydown', (e) => {
            if (e.key !== 'Enter') return;
            const selected = getSelectedPackageRows();
            if (selected.length === 0) return;
            openLogDetailModal(selected);
        });
        
        // --- Device Status ---
        socket.on('device_status', (status) => {
            const currentFilter = deviceFilter.value;
            deviceFilter.innerHTML = '<option value="all">All Devices</option>';
            if (status.connected_devices) {
                status.connected_devices.forEach(d => {
                    const opt = document.createElement('option');
                    opt.value = d.id; opt.textContent = d.name;
                    deviceFilter.appendChild(opt);
                });
            }
            // Restore selection if exists
            if([...deviceFilter.options].some(o => o.value === currentFilter)) deviceFilter.value = currentFilter;

            if (status.connected_devices && status.connected_devices.length > 0) {
                 deviceListEl.innerHTML = '<ul class="list-disc list-inside text-left">' + 
                    status.connected_devices.map(d => `<li class="text-green-600 font-semibold animate-pulse-green">${d.id} - ${d.name}</li>`).join('') + 
                    '</ul>';
            } else {
                 deviceListEl.innerHTML = `<p class="text-orange-500">${status.message || 'Waiting...'}</p>`;
            }
        });
        
        // --- FIXED: Trigger refresh on device change to update all tables including callback
        deviceFilter.addEventListener('change', (e) => { 
            selectedDevice = e.target.value; 
            socket.emit('refresh_request'); 
            renderCallbackTable(); // Trigger client-side re-render immediately
        });

        // --- Specific Tab Logic ---
        function setValidationButtonState(isActive) {
            const btn = document.getElementById('startValidationBtn');
            if (!btn) return;
            if (isActive) {
                btn.textContent = 'Stop';
                btn.className = 'bg-red-500 hover:bg-red-600 text-white font-bold text-sm px-5 rounded-lg h-10';
            } else {
                btn.textContent = 'Start Checking';
                btn.className = 'bg-indigo-600 hover:bg-indigo-700 text-white font-bold text-sm px-5 rounded-lg h-10';
            }
        }

        document.getElementById('startValidationBtn').addEventListener('click', () => {
            const val = document.getElementById('paramInput').value;
            const params = val.split('\\n').map(p=>p.trim()).filter(p=>p);
            const btn = document.getElementById('startValidationBtn');
            const isStarting = btn && btn.textContent === 'Start Checking';
            if (isStarting) socket.emit('start_validation', params);
            else socket.emit('stop_validation');
        });

        document.getElementById('clearValidatorFilterBtn')?.addEventListener('click', () => {
            const eventInput = document.getElementById('validatorEventFilterInput');
            const paramInput = document.getElementById('paramInput');
            if (eventInput) eventInput.value = '';
            if (paramInput) paramInput.value = '';
            renderValidatorTable(validator_results_cache);
        });
        
        document.getElementById('validatorEventFilterInput').addEventListener('input', () => {
            renderValidatorTable(validator_results_cache);
        });

        socket.on('validator_status', (data) => {
            setValidationButtonState(!!(data && data.active));
        });

        const specificEventInput = document.getElementById('specificEventInput');
        const specificParamInput = document.getElementById('specificParamInput');
        const parseParamList = (text) => {
            if (!text) return [];
            return text
                .replace(/,/g, ' ')
                .split(/\\s+/)
                .map(p => p.trim())
                .filter(p => p);
        };
        const updateSpecific = () => socket.emit('update_specific_filter', { eventNames: specificEventInput.value.split('\\n').filter(p=>p.trim()), params: parseParamList(specificParamInput.value) });
        specificEventInput.addEventListener('input', updateSpecific);
        specificParamInput.addEventListener('input', updateSpecific);
        
        document.getElementById('adRevenueParamInput').addEventListener('input', (e) => socket.emit('update_adrevenue_filter', {
            params: e.target.value
                .split('\\n')
                .map(p => p.trim().replace(/^['"]+|['"]+$/g, ''))
                .filter(p => p)
        }));
        document.getElementById('adRevenueFilterInput').addEventListener('input', (e) => socket.emit('refresh_request')); // Trigger re-render
        document.getElementById('packageFilterInput2').addEventListener('input', () => socket.emit('refresh_request'));
        document.getElementById('packageTagFilterInput').addEventListener('input', () => {
            const allOpt = document.querySelector('input[name="tagQuickFilter"][value=""]');
            if (allOpt) allOpt.checked = true;
            socket.emit('refresh_request');
        });
        document.querySelectorAll('input[name="tagQuickFilter"]').forEach(r => r.addEventListener('change', (e) => {
            const tagInput = document.getElementById('packageTagFilterInput');
            if (tagInput && e.target.value) tagInput.value = '';
            socket.emit('refresh_request');
        }));
        document.getElementById('packageFilterInput').addEventListener('keydown', (e) => {
            if (e.key === 'Enter') { e.preventDefault(); renderPackageLogTable(); }
        });
        document.getElementById('packageFilterInput2').addEventListener('keydown', (e) => {
            if (e.key === 'Enter') { e.preventDefault(); renderPackageLogTable(); }
        });
        document.getElementById('packageTagFilterInput').addEventListener('keydown', (e) => {
            if (e.key === 'Enter') { e.preventDefault(); renderPackageLogTable(); }
        });
        document.getElementById('showErrorsOnly').addEventListener('change', renderPackageLogTable);

        // --- Row Selection for Package Log (click + drag) ---
        let isSelectingRows = false;
        let dragStartIndex = null;

        function updateSelectionRange(startIdx, endIdx) {
            const rows = Array.from(document.querySelectorAll('#packageLogTableBody tr.package-log-row'));
            if (rows.length === 0) return;
            const [minIdx, maxIdx] = startIdx <= endIdx ? [startIdx, endIdx] : [endIdx, startIdx];
            selectedPackageRowKeys.clear();
            rows.forEach(r => {
                const idx = parseInt(r.getAttribute('data-row-index') || '-1', 10);
                if (idx >= minIdx && idx <= maxIdx) {
                    const rawKey = r.getAttribute('data-row-key') || '';
                    try { selectedPackageRowKeys.add(decodeURIComponent(rawKey)); } catch (err) {}
                    r.classList.add('selected');
                } else {
                    r.classList.remove('selected');
                }
            });
        }

        document.getElementById('packageLogTableBody').addEventListener('mousedown', (e) => {
            const row = e.target.closest('tr.package-log-row');
            if (!row) return;
            if (e.detail && e.detail >= 2) return;
            if (document.activeElement && (document.activeElement.tagName === 'INPUT' || document.activeElement.tagName === 'TEXTAREA')) {
                document.activeElement.blur();
            }
            isSelectingRows = true;
            const idx = parseInt(row.getAttribute('data-row-index') || '0', 10);
            dragStartIndex = idx;
            updateSelectionRange(idx, idx);
            e.preventDefault();
        });

        document.getElementById('packageLogTableBody').addEventListener('mouseover', (e) => {
            if (!isSelectingRows || dragStartIndex === null) return;
            const row = e.target.closest('tr.package-log-row');
            if (!row) return;
            const idx = parseInt(row.getAttribute('data-row-index') || '0', 10);
            updateSelectionRange(dragStartIndex, idx);
        });

        document.addEventListener('mouseup', () => {
            isSelectingRows = false;
            dragStartIndex = null;
        });

        // --- Ctrl+C to copy selected package log rows ---
        document.addEventListener('keydown', async (e) => {
            if (!(e.ctrlKey || e.metaKey) || e.key.toLowerCase() !== 'c') return;
            const active = document.activeElement;
            if (active && (active.tagName === 'INPUT' || active.tagName === 'TEXTAREA')) return;
            const selectedRows = Array.from(document.querySelectorAll('#packageLogTableBody tr.package-log-row.selected'));
            if (selectedRows.length === 0) return;
            const lines = selectedRows.map(r => {
                const cells = r.querySelectorAll('td');
                const parts = Array.from(cells).map(td => (td.textContent || '').trim());
                return parts.join('\\t');
            });
            const textToCopy = lines.join('\\n');
            try {
                if (navigator.clipboard && navigator.clipboard.writeText) {
                    await navigator.clipboard.writeText(textToCopy);
                } else {
                    const ta = document.createElement('textarea');
                    ta.value = textToCopy;
                    ta.style.position = 'fixed';
                    ta.style.opacity = '0';
                    document.body.appendChild(ta);
                    ta.select();
                    document.execCommand('copy');
                    document.body.removeChild(ta);
                }
            } catch (err) {
                console.error('Copy failed', err);
            }
        });
        
        document.getElementById('startSdkCheckBtn').addEventListener('click', () => {
             const text = document.getElementById('sdkCheckInput').value;
             if(text) socket.emit('start_sdk_check', {text: text});
        });
        
        function setPackageControlsEnabled(enabled) {
            const packageIdInput = document.getElementById('packageIdInput');
            const packageCheckboxes = document.querySelectorAll('.package-id-checkbox');
            if (packageIdInput) packageIdInput.disabled = !enabled;
            packageCheckboxes.forEach(cb => cb.disabled = !enabled);
        }

        document.getElementById('startPackageLogBtn').addEventListener('click', (e) => {
             const isStarting = e.target.textContent === 'Start';
             const pkg = document.getElementById('packageIdInput').value;
             socket.emit('start_package_log', {package_id: isStarting ? pkg : ''});
             e.target.textContent = isStarting ? 'Stop' : 'Start';
             e.target.classList.toggle('bg-red-500');
             setPackageControlsEnabled(!isStarting);
        });

        // --- Column Resizer (Package Log Table) ---
        (function setupColumnResizers() {
            let active = null;
            let startX = 0;
            let startWidth = 0;

            function onMouseMove(e) {
                if (!active) return;
                const dx = e.clientX - startX;
                const newWidth = Math.max(50, startWidth + dx);
                const colClass = active.getAttribute('data-col');
                const colEl = document.querySelector(`.col-${colClass}`);
                if (colEl) colEl.style.width = newWidth + 'px';
            }

            function onMouseUp() {
                if (!active) return;
                document.removeEventListener('mousemove', onMouseMove);
                document.removeEventListener('mouseup', onMouseUp);
                document.body.style.cursor = '';
                document.body.style.userSelect = '';
                active = null;
            }

            document.addEventListener('mousedown', (e) => {
                const resizer = e.target.closest('.resizer');
                if (!resizer) return;
                if (!(isPausedClient || pauseBtn.textContent === 'Resume')) return;
                e.preventDefault();
                active = resizer;
                const colClass = resizer.getAttribute('data-col');
                const headerCell = document.querySelector(`th.${colClass}`);
                startX = e.clientX;
                startWidth = headerCell ? headerCell.offsetWidth : 0;
                    document.body.style.cursor = 'col-resize';
                    document.body.style.userSelect = 'none';
                    document.addEventListener('mousemove', onMouseMove);
                    document.addEventListener('mouseup', onMouseUp);
            });
            // Apply initial state
            setResizerEnabled(isPausedClient);
        })();

        // --- Package ID Quick Select ---
        const packageIdInput = document.getElementById('packageIdInput');
        const packageCheckboxes = document.querySelectorAll('.package-id-checkbox');

        packageCheckboxes.forEach(cb => {
            cb.addEventListener('change', (e) => {
                if (e.target.checked) {
                    packageCheckboxes.forEach(other => { if (other !== e.target) other.checked = false; });
                    packageIdInput.value = e.target.value;
                } else {
                    if (packageIdInput.value === e.target.value) packageIdInput.value = '';
                }
            });
        });

        // Manual input clears quick select
        packageIdInput.addEventListener('input', () => {
            packageCheckboxes.forEach(cb => cb.checked = false);
        });

        // Clear All should also stop package log and re-enable inputs
        clearAllBtn.addEventListener('click', () => {
            const btn = document.getElementById('startPackageLogBtn');
            if (btn && btn.textContent === 'Stop') {
                socket.emit('start_package_log', {package_id: ''});
                btn.textContent = 'Start';
                btn.classList.remove('bg-red-500');
            }
            setPackageControlsEnabled(true);
        });
        
    </script>
</body>
</html>
"""

# --- SERVER ROUTE (QUAN TRỌNG) ---
@app.route('/')
def index():
    return render_template_string(
        HTML_TEMPLATE,
        default_event_names=sorted(event_specific_params.keys()),
        current_profile_name=active_profile_name
    )


@app.get('/api/profiles')
def get_profiles():
    return jsonify(_profile_payload())


@app.post('/api/profiles/select')
def select_profile():
    data = request.get_json(silent=True) or {}
    profile_name = data.get('profile_name', '')
    if not profile_name:
        return jsonify({'ok': False, 'error': 'profile_name_required'}), 400
    try:
        if not _set_active_profile(profile_name):
            return jsonify({'ok': False, 'error': 'profile_not_found'}), 404
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400
    return jsonify({'ok': True, **_profile_payload()})


@app.post('/api/profiles/reload')
def reload_profile():
    if active_profile_name:
        _set_active_profile(active_profile_name)
    else:
        _set_active_profile()
    return jsonify({'ok': True, **_profile_payload()})


@app.post('/api/profiles/import')
def import_profile():
    upload = request.files.get('profile_file')
    if not upload or not upload.filename:
        return jsonify({'ok': False, 'error': 'profile_file_required'}), 400
    try:
        filename = _sanitize_profile_filename(upload.filename)
        target = os.path.join(PROFILE_DIR, filename)
        os.makedirs(PROFILE_DIR, exist_ok=True)
        upload.save(target)
        _set_active_profile(filename)
        return jsonify({'ok': True, **_profile_payload()})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@app.post('/restart_app')
def restart_app():
    cmd = os.getenv('EVENTINSPECTOR_RESTART_CMD')
    args = os.getenv('EVENTINSPECTOR_RESTART_ARGS', '')
    if not cmd:
        return jsonify({'ok': False, 'error': 'restart_cmd_missing'})
    argv = [cmd]
    if args:
        argv.append(args)
    try:
        subprocess.Popen(argv, creationflags=creation_flags)
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})
    os._exit(0)

# --- BACKEND LOGIC ---

def send_to_sheet(device_name, ad_source, ad_format, raw_log, log_type):
    """Gửi data lên Google Sheet nếu đang bật Recording cho loại Log cụ thể"""
    state = recording_states.get(log_type)
    if state and state["is_recording"] and state["current_sheet"]:
        payload = {
            "sheet_name": state["current_sheet"], 
            "device_name": device_name, 
            "ad_source": ad_source, 
            "ad_format": ad_format, 
            "raw_log": raw_log
        }
        threading.Thread(target=lambda: requests.post(G_SHEET_URL, json=payload, timeout=30)).start()

def process_load_ads_unity_log(line, device_id):
    """Xử lý log cho Tab 1: Load Ads (Unity)"""
    # CHỈ XỬ LÝ NẾU ĐANG GHI (RECORDING)
    if not recording_states["LoadAds"]["is_recording"]:
        return

    match = UNITY_TRACKING_PATTERN.search(line)
    if match:
        try:
            data = json.loads(match.group(1))
            params = data.get("e", {})
            src = params.get("ad_source")
            fmt = params.get("ad_format")
            if params.get("mediation_ad_unit_name") == "MREC": fmt = "MREC"
            
            if src and fmt:
                d_name = get_device_name(device_id)
                with lock:
                    if (device_id, src, fmt, "unity") not in unique_load_ads:
                        unique_load_ads.add((device_id, src, fmt, "unity"))
                        load_ads_events.append({
                            "device_id": device_id,
                            "device_name": d_name, 
                            "ad_source": src, 
                            "ad_format": fmt, 
                            "raw_log": line.strip()
                        })
                        socketio.emit('update_load_ads', list(load_ads_events))
                        
                        # Gửi với type "LoadAds"
                        send_to_sheet(d_name, src, fmt, line.strip(), "LoadAds")
        except: pass

def process_load_ads_ext_log(line, device_id):
    """Xử lý log cho Tab 2: Load Ads Ext (Metrica)"""
    # CHỈ XỬ LÝ NẾU ĐANG GHI (RECORDING)
    if not recording_states["LoadAdsExt"]["is_recording"]:
        return

    match = METRICA_TRACKING_PATTERN.search(line)
    if match:
        try:
            params = json.loads(match.group(1))
            src = params.get("ad_source")
            fmt = params.get("ad_format")
            if params.get("mediation_ad_unit_name") == "MREC": fmt = "MREC"
            
            if src and fmt:
                d_name = get_device_name(device_id)
                with lock:
                    if (device_id, src, fmt, "metrica") not in unique_load_ads_ext:
                        unique_load_ads_ext.add((device_id, src, fmt, "metrica"))
                        load_ads_ext_events.append({
                            "device_id": device_id,
                            "device_name": d_name, 
                            "ad_source": src, 
                            "ad_format": fmt, 
                            "raw_log": line.strip()
                        })
                        socketio.emit('update_load_ads_ext', list(load_ads_ext_events))
                        
                        # Gửi với type "LoadAdsExt"
                        send_to_sheet(d_name, src, fmt, line.strip(), "LoadAdsExt")
        except: pass

def find_and_parse_event(log_entry):
    """Parse log sự kiện chung (TrackingService->Track only)"""
    match = OLD_EVENT_LOG_PATTERN.search(log_entry)
    if match:
        try:
            data = json.loads(match.group(1))
            event_name = data.get("eventName")
            params = data.get("e", {})
            if event_name:
                return event_name, params, match.group(1)
        except: pass
    return None, None, None

def process_callback_and_ad_event_log(log_entry, device_id, event_name=None, actual_params=None, json_string=None):
    global incomplete_impression_logs
    if is_paused: return

    # --- 0. Process GadsmeService callbacks ---
    if GADSME_SERVICE_KEYWORD in log_entry:
        try:
            after_keyword = log_entry.split(GADSME_SERVICE_KEYWORD, 1)[1]
            method_part = after_keyword.split(':', 1)[0].strip()
            json_str = extract_json_object_from_text(after_keyword)
            details = "N/A"
            json_data_for_log = "{}"
            if json_str:
                try:
                    data = json.loads(json_str)
                    details = format_json_html(data)
                    json_data_for_log = json_str
                except:
                    details = f'<div class="text-xs font-mono break-all text-red-600">JSON Parse Error</div><div class="text-xs font-mono break-all">{json_str}</div>'
            with lock:
                callback_ad_logs.append({
                    "device_id": device_id,
                    "device_name": get_device_name(device_id),
                    "type": "Callback Gadsme",
                    "event_name": method_part or "GadsmeService",
                    "details": details,
                    "raw_log": log_entry.strip(),
                    "json_data": json_data_for_log
                })
                socketio.emit('update_callback_ad_table', list(callback_ad_logs))
            return
        except:
            pass

    # --- 0b. Process Adverty5 callbacks ---
    if ADVERTY5_KEYWORD in log_entry and "Adverty5" in log_entry and "->" in log_entry:
        try:
            after_keyword = log_entry.split("Adverty5", 1)[1]
            if "->" in after_keyword:
                method_part = after_keyword.split("->", 1)[1].split(":", 1)[0].strip()
            else:
                method_part = "Adverty5"
            json_str = extract_json_object_from_text(after_keyword)
            details = "N/A"
            json_data_for_log = "{}"
            if json_str:
                try:
                    data = json.loads(json_str)
                    details = format_json_html(data)
                    json_data_for_log = json_str
                except:
                    details = f'<div class="text-xs font-mono break-all text-red-600">JSON Parse Error</div><div class="text-xs font-mono break-all">{json_str}</div>'
            with lock:
                callback_ad_logs.append({
                    "device_id": device_id,
                    "device_name": get_device_name(device_id),
                    "type": "Callback Adverty5",
                    "event_name": method_part or "Adverty5",
                    "details": details,
                    "raw_log": log_entry.strip(),
                    "json_data": json_data_for_log
                })
                socketio.emit('update_callback_ad_table', list(callback_ad_logs))
            return
        except:
            pass
    
    # --- 1. HANDLING BUFFERED IMPRESSION DATA (Split Logs) ---
    # Check if we are currently buffering for this device
    with lock:
        current_buffer = incomplete_impression_logs.get(device_id, "")

    # If we are buffering, OR if this is a new Impression Data event
    if current_buffer or "_OnImpressionDataReadyEvent" in log_entry:
        with lock:
            # Re-read buffer inside lock to be safe (though simple logic here is fine)
            current_buffer = incomplete_impression_logs.get(device_id, "")
            
            # Case A: Start of new log
            if "_OnImpressionDataReadyEvent" in log_entry:
                # Reset buffer with current line
                current_buffer = log_entry
            else:
                # Case B: Continuation line
                current_buffer += "\n" + log_entry # Add newline to separate lines if needed, or just string concat

            # Try to find JSON
            # 1. Find first '{'
            start_idx = current_buffer.find('{')
            
            # If no '{' yet, just keep buffering (unless it's been too long?)
            if start_idx != -1:
                # 2. Count braces to find end
                open_braces = 0
                end_idx = -1
                for i in range(start_idx, len(current_buffer)):
                    if current_buffer[i] == '{': open_braces += 1
                    elif current_buffer[i] == '}': open_braces -= 1
                    
                    if open_braces == 0:
                        end_idx = i
                        break
                
                if end_idx != -1:
                    # Found complete JSON
                    json_str = current_buffer[start_idx : end_idx+1]
                    details = ""
                    json_data_for_log = "{}"
                    
                    try:
                        data = json.loads(json_str)
                        if 'impressionData' in data:
                            details = format_json_html(data['impressionData'])
                        else:
                            details = format_json_html(data)
                        json_data_for_log = json_str
                    except:
                        details = f'<div class="text-xs font-mono break-all text-red-600">JSON Parse Error</div><div class="text-xs font-mono break-all">{json_str}</div>'
                        json_data_for_log = "{}"

                    # Clear buffer
                    incomplete_impression_logs[device_id] = ""
                    
                    # Emit
                    callback_ad_logs.append({
                        "device_id": device_id, 
                        "device_name": get_device_name(device_id), 
                        "type": "Callback", 
                        "event_name": "Impression Data", 
                        "details": details, 
                        "raw_log": current_buffer.strip()[:200]+"...", 
                        "json_data": json_data_for_log
                    })
                    socketio.emit('update_callback_ad_table', list(callback_ad_logs))
                    return # Done processing this line/buffer
                else:
                     # JSON start found but not ended -> Update buffer and wait for next line
                     incomplete_impression_logs[device_id] = current_buffer
                     return # Consumed line
            else:
                # No JSON start found yet (e.g. log line is "_OnImpressionDataReadyEvent:" and JSON is on next line)
                incomplete_impression_logs[device_id] = current_buffer
                return # Consumed line

    # --- 2. Process "ad_" events ---
    if event_name and event_name.startswith("ad_"):
        try:
            details = format_json_html(actual_params) if actual_params else "No params"
            with lock:
                callback_ad_logs.append({"device_id": device_id, "device_name": get_device_name(device_id), "type": "Ad Event", "event_name": event_name, "details": details, "raw_log": log_entry.strip(), "json_data": json_string})
                socketio.emit('update_callback_ad_table', list(callback_ad_logs))
        except: pass
    
    # --- 3. Process Other Callbacks ---
    callback_match = CALLBACK_LOG_PATTERN.search(log_entry)
    if callback_match:
        found_key = callback_match.group(1)
        # Skip if it is impression data (handled above)
        if "_OnImpressionDataReadyEvent" in found_key: return

        details = "N/A"
        display_name = CALLBACK_DISPLAY_NAMES.get(found_key, found_key)
        json_data_for_log = "{}"
        
        if "Listener" in found_key:
             # Try parsing adInfo if present
             try:
                 if 'adInfo:' in log_entry:
                     # Simple parsing for adInfo string
                     parts = log_entry.split('adInfo:')[1].strip().split(',')
                     ad_info = {}
                     for part in parts:
                         if ':' in part:
                             k, v = part.split(':', 1)
                             ad_info[k.strip()] = v.strip()
                     details = format_json_html(ad_info)
                     json_data_for_log = json.dumps(ad_info)
                 else:
                     details = "Listener Fired (No adInfo)"
             except:
                 details = "Listener Fired"

        with lock:
            callback_ad_logs.append({"device_id": device_id, "device_name": get_device_name(device_id), "type": "Callback", "event_name": display_name, "details": details, "raw_log": log_entry.strip(), "json_data": json_data_for_log})
            socketio.emit('update_callback_ad_table', list(callback_ad_logs))

def process_event_validator_log(event_name, actual_params, json_string, log_entry, device_id):
    if is_paused or not validator_active: 
        return
    with lock:
        required_all = []
        required_all.extend(default_params)
        required_all.extend(required_params)
        specific = event_specific_params.get(event_name, [])
        required_all.extend(specific)

        missing = set(required_all) - set(actual_params.keys())
        status = "PASSED" if not missing else "FAILED"
        
        # New Formatting Logic: Highlight missing -> Show full JSON
        details_html = ""
        if not specific and event_specific_params:
            closest = None
            best_dist = 3
            for known in event_specific_params.keys():
                d = _levenshtein_distance_limit(event_name, known, limit=2)
                if d < best_dist:
                    best_dist = d
                    closest = known
                    if best_dist == 1:
                        break
            if closest and best_dist <= 2:
                details_html += f'<div class="text-orange-600 font-bold mb-2">Possible typo: "{event_name}" ~ "{closest}"</div>'
        if missing:
            details_html += format_param_issue_html("Missing", missing, "text-red-600")
        
        details_html += format_json_html(actual_params)
        
        validator_results.append({"device_id": device_id, "event_name": event_name, "device_name": get_device_name(device_id), "status": status, "details": details_html, "raw_log": log_entry.strip(), "json_data": json_string})
        socketio.emit('update_validator_table', list(validator_results))

def _apply_specific_filter_and_emit():
    global specific_event_results
    with lock:
        res = []
        for item in event_log_cache:
             try:
                 data = json.loads(item['json_data'])
                 evt = data.get('eventName')
                 params = data.get('e', {})
                 
                 # Name Filter
                 if specific_event_name_filters and not any(evt.startswith(f) for f in specific_event_name_filters):
                     continue
                 
                 # Param Validation (Default Params + Default Events)
                 status = "INFO"
                 details = ""
                 
                 required_all = []
                 required_all.extend(default_params)
                 # If user input params, combine with defaults
                 if specific_event_params_filters:
                     required_all.extend(specific_event_params_filters)
                 # Add event-specific params if event is in default event list
                 if evt in event_specific_params:
                     required_all.extend(event_specific_params.get(evt, []))

                 if required_all:
                     missing = set(required_all) - set(params.keys())
                     strange = set(params.keys()) - set(required_all)
                     status = "PASSED" if not missing else "FAILED"
                     
                     if missing:
                         details += format_param_issue_html("Missing", missing, "text-red-600")
                     
                     if strange:
                         details += format_param_issue_html("Strange", strange, "text-orange-600")
                     
                     # Show full JSON
                     details += format_json_html(params)
                 else:
                     # Show all params json
                     details = format_json_html(params)

                 res.append({"device_id": item['device_id'], "device_name": get_device_name(item['device_id']), "status": status, "event_name": evt, "details": details, "raw_log": item['log'], "json_data": item['json_data']})
             except: pass
        specific_event_results = res
    socketio.emit('update_specific_event_table', specific_event_results)

def cache_specific_event_log(event_name, params, json_string, log_entry, device_id):
    if is_paused: return
    with lock: event_log_cache.append({'log': log_entry, 'device_id': device_id, 'json_data': json_string})
    _apply_specific_filter_and_emit()

def _apply_adrevenue_filter_and_emit():
    normalized = [p.strip().strip('"').strip("'") for p in adrevenue_params_to_validate if p and p.strip()]
    rendered = []
    with lock:
        for item in adrevenue_logs:
            parsed_data = item.get("parsed_data") or {}
            payload_data = parsed_data.get("payload") if isinstance(parsed_data.get("payload"), dict) else {}

            missing = []
            for param in normalized:
                if param in parsed_data:
                    continue
                if param in payload_data:
                    continue
                missing.append(param)

            status = "INFO"
            if normalized:
                status = "PASSED" if not missing else "FAILED"

            summary_parts = []
            if normalized:
                if missing:
                    summary_parts.append(format_param_issue_html("Missing params", missing, "text-red-600", chunk_size=3))
                else:
                    summary_parts.append("<div class='mb-2 text-xs font-semibold text-green-600'>All requested params found</div>")

            details_html = ''.join(summary_parts) + format_json_html(parsed_data if parsed_data else item.get("raw_details", ""))
            rendered.append({
                **item,
                "status": status,
                "details": details_html,
            })

    socketio.emit('update_adrevenue_table', rendered)

def _emit_sdk_check_results():
    res = []
    with lock:
        for dev in connected_devices_info:
            res.append({"status": "HEADER", "display_text": f"--- {dev['name']} ---", "device_name": dev['name'], "device_id": dev['id']})
            for item in sdk_check_input_list:
                if item["type"] == "label": res.append({"status": "LABEL", "display_text": item["display_name"], "device_id": dev['id']})
                else:
                    status = "PASSED" if (dev['id'], item["search_pattern"]) in sdk_check_results else "NOT_FOUND"
                    res.append({"status": status, "display_text": item["display_name"], "device_id": dev['id']})
    socketio.emit('update_sdk_check_table', res)

# --- THREADS & PROCESSES ---

def adb_log_reader(device_id):
    print(f"INFO: Starting log reader for {device_id}")
    try:
        subprocess.run([ADB_EXECUTABLE, '-s', device_id, 'logcat', '-c'], creationflags=creation_flags)
        proc = subprocess.Popen([ADB_EXECUTABLE, '-s', device_id, 'logcat'], stdout=subprocess.PIPE, text=True, encoding='utf-8', errors='ignore', creationflags=creation_flags)
        
        for line in iter(proc.stdout.readline, ''):
            if not line: break
            
            # 1. Process Load Ads (Unity) - ONLY IF RECORDING
            process_load_ads_unity_log(line, device_id)
            
            # 2. Process Load Ads Ext (Metrica) - ONLY IF RECORDING
            process_load_ads_ext_log(line, device_id)
            
            # 3. Process SDK Check
            if not is_paused and sdk_check_search_list:
                found = False
                for item in sdk_check_search_list:
                    if item["search_pattern"] in line:
                        with lock:
                            if (device_id, item["search_pattern"]) not in sdk_check_results:
                                sdk_check_results[(device_id, item["search_pattern"])] = True
                                found = True
                if found: _emit_sdk_check_results()

            # 4. Process AdRevenue
            if "AdRevenue Received:" in line:
                with lock:
                    match = ADREVENUE_LOG_PATTERN.search(line)
                    if match:
                        content = match.group(1)
                        details_html = content
                        ad_data = {}
                        
                        # --- Logic parse AdRevenue string to JSON object ---
                        try:
                            # 1. Extract payload JSON string if exists
                            payload_obj = {}
                            payload_match = re.search(r'payload=(\{.*?\})(?:,|$)', content)
                            
                            clean_content = content
                            if payload_match:
                                payload_str = payload_match.group(1)
                                try:
                                    payload_obj = json.loads(payload_str)
                                except: pass
                                # Remove payload part to parse the rest easily
                                clean_content = content.replace(f'payload={payload_str}', '')
                            
                            # 2. Parse key=value pairs
                            # Split by comma and space, but careful with empty strings
                            parts = [p.strip() for p in clean_content.split(',') if p.strip()]
                            
                            for part in parts:
                                if '=' in part:
                                    k, v = part.split('=', 1)
                                    ad_data[k.strip()] = v.strip()
                            
                            # 3. Re-attach payload object
                            if payload_obj:
                                ad_data['payload'] = payload_obj
                                
                            details_html = format_json_html(ad_data)
                        except: 
                            pass # Fallback to raw content if parsing fails
                        
                        adrevenue_logs.append({
                            "device_id": device_id,
                            "device_name": get_device_name(device_id),
                            "status": "INFO",
                            "event_name": "AdRevenue",
                            "details": details_html,
                            "raw_details": content,
                            "raw_log": line.strip(),
                            "json_data": json.dumps(ad_data, ensure_ascii=False) if ad_data else "{}",
                            "parsed_data": ad_data,
                        })
                _apply_adrevenue_filter_and_emit()

            # 5. Process Callback & Events
            process_callback_and_ad_event_log(line, device_id)
            
            # 6. Parse Generic Events for Validators
            event_name, params, json_string = find_and_parse_event(line)
            if event_name:
                process_event_validator_log(event_name, params, json_string, line, device_id)
                cache_specific_event_log(event_name, params, json_string, line, device_id)
                # Call callback processor for "ad_" events
                process_callback_and_ad_event_log(line, device_id, event_name, params, json_string)

    except Exception as e: print(f"Error {device_id}: {e}")

def device_manager():
    global connected_devices_info
    while True:
        try:
            output = subprocess.run([ADB_EXECUTABLE, 'devices'], capture_output=True, text=True, creationflags=creation_flags).stdout
            ids = {l.split('\t')[0] for l in output.strip().split('\n')[1:] if '\tdevice' in l}
            
            with lock:
                for did in ids - set(active_log_readers.keys()):
                    t = threading.Thread(target=adb_log_reader, args=(did,), daemon=True)
                    active_log_readers[did] = t
                    t.start()
                
                for did in set(active_log_readers.keys()) - ids:
                    del active_log_readers[did]
                
                connected_devices_info = [{'id': i, 'name': get_device_name(i)} for i in ids]
                if connected_devices_info:
                    socketio.emit('device_status', {"connected_devices": connected_devices_info})
                else:
                    socketio.emit('device_status', {
                        "connected_devices": [],
                        "message": f"Waiting... (ADB: {ADB_EXECUTABLE})"
                    })
        except Exception as e:
            socketio.emit('device_status', {
                "connected_devices": [],
                "message": f"ADB error: {e}"
            })
        time.sleep(3)

def package_log_consumer(device_id, logcat_process):
    try:
        for line in iter(logcat_process.stdout.readline, ''):
            if not line: break
            if not is_paused:
                is_error = bool(re.search(r'^\S+\s+\S+\s+\d+\s+\d+\s+[EF]\s', line))
                time_str = ""
                time_display = ""
                level = ""
                tag = ""
                message = line.strip()
                # Try to parse standard logcat format: MM-DD HH:MM:SS.mmm PID TID LEVEL TAG: message
                m = re.match(r'^(\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}\.\d+)\s+\d+\s+\d+\s+([A-Z])\s+([^:]+):\s*(.*)$', line.strip())
                if m:
                    time_str = m.group(1)
                    level = m.group(2)
                    tag = m.group(3).strip()
                    message = m.group(4).strip()
                    # Prefer shorter time display to save width: HH:MM:SS.mmm
                    if ' ' in time_str:
                        time_display = time_str.split(' ', 1)[1]
                with lock:
                    package_log_cache.append({
                        'device_id': device_id,
                        'device_name': get_device_name(device_id),
                        'log': line.strip(),
                        'time': time_str,
                        'time_display': time_display or time_str,
                        'level': level,
                        'tag': tag,
                        'message': message,
                        'timestamp': time.time(),
                        'is_error': is_error
                    })
    except: pass

def package_pid_monitor():
    global active_package_pids
    while True:
        time.sleep(3)

        with lock:
            pkg = target_package_name
            devices_snapshot = list(connected_devices_info)

        if not pkg:
            with lock:
                for p in active_logcat_processes.values():
                    try:
                        p.terminate()
                    except Exception:
                        pass
                active_logcat_processes.clear()
                active_package_pids.clear()
            continue

        current_ids = {d['id'] for d in devices_snapshot}

        # Clean up disconnected devices
        with lock:
            for did in list(active_logcat_processes.keys()):
                if did not in current_ids:
                    try:
                        active_logcat_processes[did].terminate()
                    except Exception:
                        pass
                    active_logcat_processes.pop(did, None)
                    active_package_pids.pop(did, None)

        for device in devices_snapshot:
            did = device['id']
            try:
                res = subprocess.run([ADB_EXECUTABLE, '-s', did, 'shell', 'pidof', '-s', pkg],
                                     capture_output=True, text=True, creationflags=creation_flags)
                pid = res.stdout.strip()

                if not pid:
                    with lock:
                        if did in active_logcat_processes:
                            try:
                                active_logcat_processes[did].terminate()
                            except Exception:
                                pass
                            active_logcat_processes.pop(did, None)
                        active_package_pids.pop(did, None)
                    continue

                with lock:
                    proc = active_logcat_processes.get(did)
                    prev_pid = active_package_pids.get(did)

                restart = (pid != prev_pid) or (proc is None) or (proc.poll() is not None)

                if restart:
                    with lock:
                        if did in active_logcat_processes:
                            try:
                                active_logcat_processes[did].terminate()
                            except Exception:
                                pass
                        cmd = [ADB_EXECUTABLE, '-s', did, 'logcat', f'--pid={pid}']
                        proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, text=True, encoding='utf-8', errors='ignore', creationflags=creation_flags)
                        active_logcat_processes[did] = proc
                        active_package_pids[did] = pid
                        threading.Thread(target=package_log_consumer, args=(did, proc), daemon=True).start()
            except Exception:
                with lock:
                    if did in active_logcat_processes:
                        try:
                            active_logcat_processes[did].terminate()
                        except Exception:
                            pass
                        active_logcat_processes.pop(did, None)
                    active_package_pids.pop(did, None)


def package_log_emitter():
    while True:
        time.sleep(1)
        with lock:
            if not is_paused and target_package_name:
                now = time.time()
                while package_log_cache and now - package_log_cache[0]['timestamp'] > 1000: package_log_cache.popleft()
                socketio.emit('package_log_cache', list(package_log_cache))

# --- SOCKET HANDLERS ---
@socketio.on('change_tab')
def handle_change_tab(data): 
    # Sync logs for current tab on switch
    if data.get('tab_name') == 'LoadAds': socketio.emit('update_load_ads', list(load_ads_events))
    if data.get('tab_name') == 'LoadAdsExt': socketio.emit('update_load_ads_ext', list(load_ads_ext_events))
    if data.get('tab_name') == 'SdkCheck': _emit_sdk_check_results()
    if data.get('tab_name') == 'CallbackAd': socketio.emit('update_callback_ad_table', list(callback_ad_logs))

@socketio.on('toggle_record')
def tr(data):
    tab_name = data.get('tab_name')
    if not tab_name or tab_name not in recording_states:
        return

    current_state = recording_states[tab_name]
    
    if not current_state["is_recording"]:
        name = data.get('sheet_name', 'Log_Default').strip()
        try:
            res = requests.post(G_SHEET_URL, json={"action": "create_or_get_sheet", "sheet_name": name}, timeout=30)
            current_state.update({"is_recording": True, "current_sheet": res.text})
        except: 
            current_state.update({"is_recording": True, "current_sheet": name}) # Fallback
    else:
        current_state["is_recording"] = False
    
    # Emit status back with tab_name so UI knows which button to update
    socketio.emit('record_status', {
        "tab_name": tab_name,
        "is_recording": current_state["is_recording"],
        "current_sheet": current_state["current_sheet"]
    })

@socketio.on('toggle_pause')
def tp(): 
    global is_paused
    is_paused = not is_paused
    socketio.emit('pause_status', {'is_paused': is_paused})

@socketio.on('clear_all_logs')
def cl():
    with lock:
        load_ads_events.clear(); unique_load_ads.clear()
        load_ads_ext_events.clear(); unique_load_ads_ext.clear()
        validator_results.clear()
        specific_event_results.clear(); event_log_cache.clear()
        adrevenue_logs.clear(); callback_ad_logs.clear()
        package_log_cache.clear()
        # Clean SDK check
        sdk_check_results.clear()
        incomplete_impression_logs.clear()
        
    socketio.emit('update_load_ads', [])
    socketio.emit('update_load_ads_ext', [])
    socketio.emit('update_validator_table', [])
    socketio.emit('update_specific_event_table', [])
    socketio.emit('update_adrevenue_table', [])
    socketio.emit('update_callback_ad_table', [])
    socketio.emit('package_log_cache', [])
    _emit_sdk_check_results()

@socketio.on('start_validation')
def val(p): 
    global required_params, validator_active
    required_params = p or []
    validator_active = True
    validator_results.clear()
    socketio.emit('update_validator_table', [])
    socketio.emit('validator_status', {'active': True})

@socketio.on('stop_validation')
def stop_val():
    global validator_active
    validator_active = False
    socketio.emit('validator_status', {'active': False})

@socketio.on('update_specific_filter')
def usf(d):
    global specific_event_name_filters, specific_event_params_filters
    with lock:
        specific_event_name_filters = d.get('eventNames', [])
        specific_event_params_filters = d.get('params', [])
    _apply_specific_filter_and_emit()

@socketio.on('update_adrevenue_filter')
def uaf(d):
    global adrevenue_params_to_validate; adrevenue_params_to_validate = d.get('params', []); _apply_adrevenue_filter_and_emit()

@socketio.on('start_sdk_check')
def sdk_check(data):
    global sdk_check_search_list, sdk_check_results, sdk_check_input_list
    with lock:
        sdk_check_search_list = []
        sdk_check_results = {}
        sdk_check_input_list = []
        for line in data.get('text', '').splitlines():
            line=line.strip()
            if not line: continue
            match = SDK_CHECK_SEARCH_PATTERN.search(line)
            if match:
                 pat = match.group(1)
                 disp = line[:match.start()].rstrip(', ').strip().strip('"') or pat
                 item = {"type": "search", "display_name": disp, "search_pattern": pat}
                 sdk_check_search_list.append(item); sdk_check_input_list.append(item)
            else: sdk_check_input_list.append({"type": "label", "display_name": line})
    _emit_sdk_check_results()

@socketio.on('start_package_log')
def spl(d):
    global target_package_name
    pid = d.get('package_id', '').strip()
    with lock: target_package_name = pid; package_log_cache.clear(); socketio.emit('package_log_cache', [])

@socketio.on('refresh_request')
def refresh():
    socketio.emit('update_load_ads', list(load_ads_events))
    socketio.emit('update_load_ads_ext', list(load_ads_ext_events))
    socketio.emit('update_callback_ad_table', list(callback_ad_logs)) # Ensure callback data is refreshed
    # ... trigger others ...

@socketio.on('connect')
def connect(): 
    socketio.emit('pause_status', {'is_paused': is_paused})
    socketio.emit('validator_status', {'active': validator_active})
    # Sync recording buttons on connect
    for tab, state in recording_states.items():
        socketio.emit('record_status', {
            "tab_name": tab,
            "is_recording": state["is_recording"],
            "current_sheet": state["current_sheet"]
        })

# --- MAIN ---
def run_server(host="0.0.0.0", port=5001):
    _normalize_remote_update_config()
    _set_active_profile()
    threading.Thread(target=device_manager, daemon=True).start()
    threading.Thread(target=package_log_emitter, daemon=True).start()
    threading.Thread(target=package_pid_monitor, daemon=True).start()

    def safe_print(msg):
        try:
            print(msg)
        except UnicodeEncodeError:
            try:
                print(msg.encode("utf-8", "ignore").decode("utf-8"))
            except Exception:
                logging.info(msg)

    safe_print("===================================================")
    safe_print("  Event Inspector Started")
    safe_print(f"  - Access: http://{host}:{port}")
    safe_print("===================================================")
    socketio.run(
        app,
        host=host,
        port=port,
        use_reloader=False,
        allow_unsafe_werkzeug=True
    )

if __name__ == '__main__':
    run_server(host="0.0.0.0", port=5001)
