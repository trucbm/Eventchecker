import json
import re
import subprocess
import threading
import time
import requests
import html
from collections import deque
from flask import Flask, render_template_string, request, jsonify, Response
from flask_socketio import SocketIO
from flask_cors import CORS
from openpyxl import load_workbook
import os
import sys
import shutil
import logging
import sqlite3
try:
    import webview
except Exception:
    webview = None
from pathlib import Path
from queue import Empty, Queue

# Khởi tạo ứng dụng Flask và SocketIO
app = Flask(__name__)
CORS(app)
try:
    socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')
except ValueError:
    # Fallback for environments where 'threading' isn't accepted
    socketio = SocketIO(app, cors_allowed_origins="*")

# --- CẤU HÌNH LOAD ADS (GOOGLE SHEET) ---
G_SHEET_URL = "https://script.google.com/macros/s/AKfycbyLMM9nLAjS9Zhwr4-J6ikjqBSpO7ZCNaNeHKTsfKltiIa0OniDBSrzjvqfvpg87Epl/exec"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

def _resolve_default_params_path():
    env_path = os.getenv("DEFAULT_PARAMS_XLSX_PATH")
    if env_path and os.path.exists(env_path):
        return env_path

    filename = "Default event + Default Params.xlsx"

    update_dir = os.getenv("EVENTINSPECTOR_UPDATE_DIR")
    if update_dir:
        upd = os.path.join(update_dir, filename)
        if os.path.exists(upd):
            return upd

    candidates = []

    # 1) Project folder (dev)
    candidates.append(os.path.join(SCRIPT_DIR, filename))

    # 2) Installed app folder (packaged)
    if getattr(sys, "frozen", False):
        exe_dir = os.path.dirname(sys.executable)
        candidates.append(os.path.join(exe_dir, filename))

        # macOS: Resources folder inside .app bundle
        if sys.platform == "darwin":
            resources_dir = os.path.abspath(os.path.join(exe_dir, "..", "Resources"))
            candidates.append(os.path.join(resources_dir, filename))

    # 3) PyInstaller temp bundle path
    meipass = getattr(sys, "_MEIPASS", None)
    if meipass:
        candidates.append(os.path.join(meipass, filename))

    # 4) Common user locations
    candidates.extend([
        os.path.join(os.path.expanduser("~/Downloads"), filename),
        os.path.join(os.path.expanduser("~/Documents"), filename),
    ])

    for p in candidates:
        if p and os.path.exists(p):
            return p

    # If packaged, try to copy bundled file into app folder for user to edit
    if getattr(sys, "frozen", False) and meipass:
        bundled = os.path.join(meipass, filename)
        if os.path.exists(bundled):
            try:
                target = os.path.join(os.path.dirname(sys.executable), filename)
                shutil.copyfile(bundled, target)
                return target
            except Exception:
                if sys.platform == "darwin":
                    try:
                        resources_dir = os.path.abspath(os.path.join(os.path.dirname(sys.executable), "..", "Resources"))
                        os.makedirs(resources_dir, exist_ok=True)
                        target = os.path.join(resources_dir, filename)
                        shutil.copyfile(bundled, target)
                        return target
                    except Exception:
                        pass

    return candidates[0]

DEFAULT_PARAMS_XLSX = _resolve_default_params_path()
DEFAULT_PARAM_FILL = "FFFCE5CD"
REMOTE_UPDATE_CONFIG_FILENAME = "remote_update_config_v230.json"
DEFAULT_REMOTE_MANIFEST_URL = "https://raw.githubusercontent.com/trucbm/Eventchecker/main/Updates_2_3/remote_manifest.json"


def _runtime_app_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return SCRIPT_DIR


def _resolve_profiles_dir():
    base_dir = _runtime_app_dir()
    profiles_dir = os.path.join(base_dir, "profiles")
    try:
        os.makedirs(profiles_dir, exist_ok=True)
        return profiles_dir
    except Exception:
        fallback = os.path.join(SCRIPT_DIR, "profiles")
        os.makedirs(fallback, exist_ok=True)
        return fallback


PROFILE_DIR = _resolve_profiles_dir()
active_profile_name = None
active_profile_path = None
active_profile_game_name = ""


def _user_data_dir():
    if os.name == "nt":
        base = os.getenv("LOCALAPPDATA") or os.path.expanduser("~")
        return os.path.join(base, "EventInspector")
    if sys.platform == "darwin":
        return os.path.join(os.path.expanduser("~/Library/Application Support"), "EventInspector")
    return os.path.join(os.path.expanduser("~"), ".eventinspector")


def _package_history_dir():
    candidates = [
        os.path.join(_runtime_app_dir(), "package_log_history"),
        os.path.join(_user_data_dir(), "package_log_history"),
    ]
    for history_dir in candidates:
        try:
            os.makedirs(history_dir, exist_ok=True)
            test_file = os.path.join(history_dir, ".write_test")
            with open(test_file, "w", encoding="utf-8") as f:
                f.write("ok")
            os.remove(test_file)
            return history_dir
        except Exception:
            continue
    fallback = os.path.join(_user_data_dir(), "package_log_history")
    os.makedirs(fallback, exist_ok=True)
    return fallback


PACKAGE_LOG_DB_PATH = os.path.join(_package_history_dir(), "package_logs.sqlite3")


def _normalize_remote_update_config():
    try:
        user_dir = _user_data_dir()
        os.makedirs(user_dir, exist_ok=True)
        cfg_path = os.path.join(user_dir, REMOTE_UPDATE_CONFIG_FILENAME)
        cfg = {}
        if os.path.exists(cfg_path):
            try:
                with open(cfg_path, "r", encoding="utf-8") as f:
                    cfg = json.load(f)
            except Exception:
                cfg = {}
        cfg["enabled"] = True
        cfg["manifest_url"] = DEFAULT_REMOTE_MANIFEST_URL
        cfg["timeout_sec"] = 10
        cfg["min_interval_sec"] = 0
        with open(cfg_path, "w", encoding="utf-8") as f:
            json.dump(cfg, f, indent=2)
    except Exception as e:
        print(f"WARNING: Failed to normalize updater config: {e}")

def _resolve_adb():
    adb_env = os.getenv("ADB_PATH")
    if adb_env:
        return adb_env

    # Common SDK locations
    sdk_roots = [
        os.getenv("ANDROID_HOME"),
        os.getenv("ANDROID_SDK_ROOT"),
        os.path.expanduser("~/Library/Android/sdk"),
        os.path.expanduser("~/Android/Sdk"),
    ]
    for root in [r for r in sdk_roots if r]:
        candidate = os.path.join(root, "platform-tools", "adb")
        if os.path.exists(candidate):
            return candidate

    extra_paths = "/opt/homebrew/bin:/usr/local/bin:/usr/bin:/bin:/usr/sbin:/sbin"
    current_path = os.environ.get("PATH", "")
    if extra_paths not in current_path:
        os.environ["PATH"] = f"{extra_paths}:{current_path}" if current_path else extra_paths

    found = shutil.which("adb")
    return found or "adb"

ADB_EXECUTABLE = _resolve_adb()

if os.name == 'nt':
    creation_flags = 0x08000000  # CREATE_NO_WINDOW
else:
    creation_flags = 0

# Bảng tra cứu tên thiết bị
DEVICE_NAMES = {
    "R94Y40MZJ5T": "SS A16",
    "RFCX30DPW6D": "SS A35",
    "FMAEAYHIQ8FEYXPN": "OPPO A18",
    "ZY226DFRH2": "Moto Z4",
    "R5CR1282PAK": "SS S21",
    "F6QCCAGIRSQOVGFQ": "Redmi A3",
    "7ec7bca6": "Xiaomi 13 Lite"
}

# --- DỮ LIỆU TOÀN CỤC ---

# Giới hạn cache để UI không giữ quá nhiều log trong RAM.
MAX_LOAD_ADS_LOGS = 1000
MAX_VALIDATOR_LOGS = 1500
MAX_SPECIFIC_EVENT_LOGS = 1500
MAX_CALLBACK_AD_LOGS = 1500
MAX_ADREVENUE_LOGS = 1500

# 1. Dữ liệu cho Tab Load Ads
load_ads_events = deque(maxlen=MAX_LOAD_ADS_LOGS)
unique_load_ads = set()

# 2. Dữ liệu cho Tab Load Ads Ext
load_ads_ext_events = deque(maxlen=MAX_LOAD_ADS_LOGS)
unique_load_ads_ext = set()

# Trạng thái Recording (Google Sheet) - RIÊNG BIỆT CHO TỪNG TAB
recording_states = {
    "LoadAds": {"is_recording": False, "current_sheet": None},
    "LoadAdsExt": {"is_recording": False, "current_sheet": None}
}

# 3. Dữ liệu cho Tab Validator
validator_results = deque(maxlen=MAX_VALIDATOR_LOGS)
required_params = []  # Manual extra params from UI
default_params = []   # Default params from sheet (apply to all events)
event_specific_params = {}  # event_name -> list of params
validator_active = False

# 4. Dữ liệu cho Tab Specific Event
event_log_cache = deque(maxlen=MAX_SPECIFIC_EVENT_LOGS)
specific_event_results = []
specific_event_name_filters = []
specific_event_params_filters = []

# 5. Dữ liệu cho Tab Package Logcat
package_log_cache = deque(maxlen=15000)
PACKAGE_LOG_UI_MAX_ROWS = 8000
target_package_name = ""
active_package_pids = {}
active_logcat_processes = {}
active_package_log_session_id = None
package_log_db_queue = Queue()

# 6. Dữ liệu cho Tab Callback & Ads Event
callback_ad_logs = deque(maxlen=MAX_CALLBACK_AD_LOGS)

# 7. Dữ liệu cho Tab AdRevenue
adrevenue_log_cache = []
adrevenue_logs = deque(maxlen=MAX_ADREVENUE_LOGS)
adrevenue_default_params = []
adrevenue_source_params = {}

# 8. Dữ liệu cho Tab SDK Check
sdk_check_search_list = []
sdk_check_results = {}
sdk_check_input_list = []
sdk_check_active = False
sdk_check_expected_map = {}
sdk_check_runtime_state = {}
sdk_check_current_network = {}
sdk_check_expected_order = []
active_platform = "android"

# Dữ liệu hệ thống chung
active_log_readers = {}
connected_devices_info = []
is_paused = False
lock = threading.Lock()
incomplete_impression_logs = {} # Buffer cho logs bị ngắt dòng
adb_error_counter = 0


def _get_package_db_connection():
    conn = sqlite3.connect(PACKAGE_LOG_DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn


def _init_package_log_db():
    conn = _get_package_db_connection()
    try:
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS package_log_sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                package_id TEXT NOT NULL,
                started_at REAL NOT NULL,
                ended_at REAL,
                status TEXT NOT NULL DEFAULT 'running'
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS package_log_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                session_id INTEGER NOT NULL,
                created_at REAL NOT NULL,
                time_display TEXT,
                device_id TEXT,
                device_name TEXT,
                level TEXT,
                tag TEXT,
                message TEXT,
                raw_log TEXT,
                is_error INTEGER NOT NULL DEFAULT 0,
                FOREIGN KEY(session_id) REFERENCES package_log_sessions(id)
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_package_log_entries_session ON package_log_entries(session_id, id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_package_log_sessions_started ON package_log_sessions(started_at DESC)")
        conn.commit()
    finally:
        conn.close()


def _start_package_log_session(package_id):
    conn = _get_package_db_connection()
    try:
        cur = conn.execute(
            "INSERT INTO package_log_sessions (package_id, started_at, status) VALUES (?, ?, 'running')",
            (package_id, time.time()),
        )
        conn.commit()
        return cur.lastrowid
    finally:
        conn.close()


def _finish_package_log_session(session_id):
    if not session_id:
        return
    conn = _get_package_db_connection()
    try:
        conn.execute(
            "UPDATE package_log_sessions SET ended_at = ?, status = 'stopped' WHERE id = ?",
            (time.time(), session_id),
        )
        conn.commit()
    finally:
        conn.close()


def _package_log_db_writer():
    conn = _get_package_db_connection()
    batch = []
    last_flush = time.time()

    def flush():
        nonlocal batch, last_flush
        if not batch:
            return
        conn.executemany(
            """
            INSERT INTO package_log_entries
            (session_id, created_at, time_display, device_id, device_name, level, tag, message, raw_log, is_error)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            batch,
        )
        conn.commit()
        batch = []
        last_flush = time.time()

    try:
        while True:
            try:
                item = package_log_db_queue.get(timeout=0.5)
                batch.append(item)
                if len(batch) >= 100:
                    flush()
            except Empty:
                if batch and time.time() - last_flush >= 0.5:
                    flush()
    finally:
        flush()
        conn.close()


# --- REGEX PATTERNS ---

# Pattern cho Load Ads (Unity)
UNITY_TRACKING_PATTERN = re.compile(r'\[\s*Tracking\s*\]\s*TrackingService->Track:\s*(\{"eventName":"ad_impression".*)')

# Pattern cho Load Ads Ext (AppMetrica)
METRICA_TRACKING_PATTERN = re.compile(r'Event sent: ad_impression with value\s*(\{.*\})')
LOAD_ADS_EXT_ADREVENUE_PATTERN = re.compile(r'AdRevenue Received:\s*AdRevenue\{(.*)\}', re.IGNORECASE)
METRICA_REGULAR_EVENT_PATTERN = re.compile(
    r'Event received on service:\s*EVENT_TYPE_REGULAR\s+with name\s+([A-Za-z0-9_.$-]+)\s+with value\s*(\{.*\})'
)

# Patterns cũ của Log Checker
OLD_EVENT_LOG_PATTERN = re.compile(r'\[\s*Tracking\s*\]\s*TrackingService->Track:\s*(\{"eventName":.*)')
CALLBACK_LOG_PATTERN = re.compile(r"(_OnImpressionDataReadyEvent|_OnLevelPlayImpressionDataReadyEvent|LevelPlayInterstitialAdListener|LevelPlayBannerAdViewListener|LevelPlayRewardedAdListener|Receive Ironsource Impression Data LevelPlayImpressionData)")
ADREVENUE_LOG_PATTERN = re.compile(r"AdRevenue Received:\s*AdRevenue\{(.*)\}")
APPSFLYER_ADREVENUE_PATTERN = re.compile(r"\b(ADREVENUE)-\d+:\s*preparing data:\s*(\{.*\})", re.IGNORECASE)
SDK_CHECK_SEARCH_PATTERN = re.compile(r'"search_pattern"\s*:\s*["\'](.*?)["\']')
GADSME_SERVICE_KEYWORD = "[InPlayAds,Gadsme]"
ADVERTY5_KEYWORD = "[InPlayAds,Adverty]"
SDK_HEADER_PATTERN = re.compile(r'-{5,}\s*(.*?)\s*-{5,}')
SDK_VERSION_LINE_PATTERN = re.compile(r'SDK\s+Version\s*[-–—]\s*(.+)$', re.IGNORECASE)
SDK_ADAPTER_VERSION_LINE_PATTERN = re.compile(r'Adapter\s+Version\s*[-–—]\s*(.+)$', re.IGNORECASE)
SDK_ADAPTER_MISSING_PATTERN = re.compile(r'Adapter\s*[-–—]\s*MISSING\b', re.IGNORECASE)
SDK_VERIFICATION_PATTERN = re.compile(r'>{3,}\s*(.*?)\s*-\s*(VERIFIED|NOT VERIFIED)\b', re.IGNORECASE)

# Mapping tên hiển thị cho Callback
CALLBACK_DISPLAY_NAMES = {
    "LevelPlayInterstitialAdListener": "Interstitial",
    "LevelPlayBannerAdViewListener": "Banner",
    "LevelPlayRewardedAdListener": "Rewarded",
    "_OnImpressionDataReadyEvent": "Impression Data",
    "_OnLevelPlayImpressionDataReadyEvent": "LevelPlay Impression Data",
    "Receive Ironsource Impression Data LevelPlayImpressionData": "LevelPlayImpressionData"
}

def get_device_name(device_id):
    return DEVICE_NAMES.get(device_id, device_id)

def _normalize_sdk_search_text(text):
    if text is None:
        return ""
    normalized = str(text).lower()
    normalized = normalized.replace("–", "-").replace("—", "-").replace("−", "-")
    normalized = normalized.replace('"', '').replace("'", '')
    normalized = re.sub(r'\s+', ' ', normalized)
    return normalized.strip()


def _normalize_sdk_network_name(name):
    if not name:
        return ""
    normalized = str(name).lower()
    normalized = normalized.replace("&", "and")
    normalized = re.sub(r'[^a-z0-9]+', '', normalized)
    return normalized


SDK_NETWORK_ALIASES = {
    "bigo": "bigoads",
    "fyber": "digitalturbinefyber",
    "digitalturbine": "digitalturbinefyber",
    "google": "googleadmobandadmanager",
    "admob": "googleadmobandadmanager",
    "googleadmanager": "googleadmobandadmanager",
    "line": "lineads",
    "mytarget": "mytargetvkads",
    "vkads": "mytargetvkads",
    "pubmatic": "pubmaticopenwrap",
    "openwrap": "pubmaticopenwrap",
    "verve": "vervepubnative",
    "pubnative": "vervepubnative",
    "meta": "metaaudiencenetwork",
    "facebook": "facebooksdk",
    "firebase": "firebasecrashlytics",
    "crashlytics": "firebasecrashlytics",
}


def _match_sdk_expected_key(actual_name):
    actual_norm = _normalize_sdk_network_name(actual_name)
    if not actual_norm:
        return ""
    if actual_norm in sdk_check_expected_map:
        return actual_norm
    alias = SDK_NETWORK_ALIASES.get(actual_norm)
    if alias and alias in sdk_check_expected_map:
        return alias
    matches = []
    for key in sdk_check_expected_map.keys():
        if actual_norm in key or key in actual_norm:
            matches.append(key)
    if len(matches) == 1:
        return matches[0]
    return ""


def _extract_sdk_comparable_version(value, expected_value=""):
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    match = re.search(r'(\d+(?:\.\d+)+)', text)
    if not match:
        return text
    version = match.group(1)
    expected_text = str(expected_value or "").strip()
    expected_match = re.search(r'(\d+(?:\.\d+)+)', expected_text)
    if expected_match:
        expected_parts = expected_match.group(1).split('.')
        version_parts = version.split('.')
        if len(version_parts) >= len(expected_parts):
            return '.'.join(version_parts[:len(expected_parts)])
    return version


def _sdk_result_status(actual_value, expected_value):
    actual_text = str(actual_value or "").strip()
    expected_text = str(expected_value or "").strip()
    if not actual_text or actual_text.upper() in {"NOT FOUND", "MISSING"}:
        return "NOT_FOUND"
    actual_compare = _extract_sdk_comparable_version(actual_text, expected_text)
    expected_compare = _extract_sdk_comparable_version(expected_text)
    if expected_compare:
        return "PASSED" if actual_compare == expected_compare else "FAILED"
    return "FOUND"


def _sdk_state_network_keys(device_id):
    keys = []
    seen = set()
    device_state = sdk_check_runtime_state.get(device_id, {})
    for expected_key in sdk_check_expected_order:
        if expected_key not in seen:
            keys.append(expected_key)
            seen.add(expected_key)
    for network_key in device_state.keys():
        if network_key not in seen:
            keys.append(network_key)
            seen.add(network_key)
    return keys


def _ensure_sdk_expected_blocks_for_device(device_id):
    device_state = sdk_check_runtime_state.setdefault(device_id, {})
    for expected_key in sdk_check_expected_order:
        expected = sdk_check_expected_map.get(expected_key, {})
        if expected_key not in device_state:
            device_state[expected_key] = {
                "display_name": expected.get("display_name", expected_key),
                "sdk_version": "",
                "adapter_version": "",
                "adapter_missing": False,
                "verification": "",
                "expected_key": expected_key,
                "updated_at": time.time(),
            }


def _extract_json_field_version(line, field_name):
    pattern = re.compile(rf'"{re.escape(field_name)}"\s*:\s*"([^"]+)"', re.IGNORECASE)
    match = pattern.search(line)
    return match.group(1).strip() if match else ""


def _process_sdk_external_line(line, device_id):
    normalized_line = _normalize_sdk_search_text(line)
    changed = False

    def update_block(network_name, sdk_version=None, adapter_version=None, adapter_missing=False):
        nonlocal changed
        block = _sdk_check_block_for_device(device_id, network_name)
        if sdk_version is not None and block.get("sdk_version") != sdk_version:
            block["sdk_version"] = sdk_version
            changed = True
        if adapter_version is not None and (block.get("adapter_version") != adapter_version or block.get("adapter_missing")):
            block["adapter_version"] = adapter_version
            block["adapter_missing"] = False
            changed = True
        if adapter_missing and (not block.get("adapter_missing") or block.get("adapter_version")):
            block["adapter_missing"] = True
            block["adapter_version"] = ""
            changed = True

    if "advertysdk" in normalized_line:
        match = re.search(r'AdvertySDK\s+([0-9]+(?:\.[0-9]+)+)', line, re.IGNORECASE)
        if match:
            update_block("Adverty", adapter_version=match.group(1))
    if "gadsmeservice->initialize" in normalized_line:
        version = _extract_json_field_version(line, "Version")
        if version:
            update_block("Gadsme", adapter_version=version)
    if "audiomob" in normalized_line:
        match = re.search(r'Audiomob\s+v?([0-9]+(?:\.[0-9]+)+)', line, re.IGNORECASE)
        if match:
            update_block("AudioMob", adapter_version=match.group(1))
    if "adquality" in normalized_line:
        match = re.search(r'AdQuality\s+([0-9]+(?:\.[0-9]+)+)', line, re.IGNORECASE)
        if match:
            version = match.group(1)
            update_block("AdQuality", adapter_version=version, sdk_version=version)
    if "facebook unity sdk" in normalized_line:
        match = re.search(r'Facebook\s+Unity\s+SDK\s+v?([0-9]+(?:\.[0-9]+)+)', line, re.IGNORECASE)
        if match:
            update_block("Facebook SDK", adapter_version=match.group(1))
    if "fbandroidsdk/" in normalized_line:
        match = re.search(r'FBAndroidSDK/([0-9]+(?:\.[0-9]+)+)', line, re.IGNORECASE)
        if match:
            update_block("Facebook SDK", sdk_version=match.group(1))
    if "appmetrica" in normalized_line and "release type" in normalized_line and "version" in normalized_line:
        match = re.search(r'AppMetrica.*?Version\s+([0-9]+(?:\.[0-9]+)+)', line, re.IGNORECASE)
        if match:
            update_block("AppMetrica", sdk_version=match.group(1))
    if "appsflyer" in normalized_line:
        adapter_match = re.search(r'"pluginVersion"\s*:\s*"([0-9]+(?:\.[0-9]+)+)"', line, re.IGNORECASE)
        if adapter_match:
            update_block("Appsflyer", adapter_version=adapter_match.group(1))
        appsflyer_sdk = ""
        sdk_json_match = re.search(r'"AppsFlyer\.getSdkVersion\(\)"\s*:\s*"version\s*:\s*([0-9]+(?:\.[0-9]+)+)', line, re.IGNORECASE)
        if sdk_json_match:
            appsflyer_sdk = sdk_json_match.group(1)
        if not appsflyer_sdk:
            sdk_field = _extract_json_field_version(line, "AppsFlyer.getSdkVersion()")
            if sdk_field:
                sdk_match = re.search(r'version\s*:\s*([0-9]+(?:\.[0-9]+)+)', sdk_field, re.IGNORECASE)
                if sdk_match:
                    appsflyer_sdk = sdk_match.group(1)
        if not appsflyer_sdk and "appsflyer:" in normalized_line:
            match = re.search(r'AppsFlyer:\s*\(v?([0-9]+(?:\.[0-9]+)+)', line, re.IGNORECASE)
            if match:
                appsflyer_sdk = match.group(1)
        if appsflyer_sdk:
            update_block("Appsflyer", sdk_version=appsflyer_sdk)
    if "firebase crashlytics" in normalized_line and "initializing" in normalized_line:
        match = re.search(r'Firebase\s+Crashlytics\s+([0-9]+(?:\.[0-9]+)+)', line, re.IGNORECASE)
        if match:
            update_block("Firebase Crashlytics", sdk_version=match.group(1))

    return changed

# --- HELPER FUNCTIONS FOR FORMATTING ---
def format_json_html(data):
    """Format JSON object to HTML string with indentation and colors"""
    try:
        if isinstance(data, str):
            # Try to parse string as JSON if possible
            if data.strip().startswith('{') or data.strip().startswith('['):
                try:
                    data = json.loads(data)
                except: pass
        
        # If it's a dict or list, dump it pretty
        if isinstance(data, (dict, list)):
            # ensure_ascii=False để hiển thị tiếng Việt đúng
            json_str = json.dumps(data, indent=2, ensure_ascii=False)
            return f'<pre class="text-xs bg-gray-50 p-2 rounded border border-gray-200 overflow-x-auto font-mono text-gray-700">{html.escape(json_str)}</pre>'
        return str(data)
    except:
        return str(data)


def format_param_issue_html(title, items, color_class, chunk_size=4):
    if not items:
        return ""
    ordered = [html.escape(str(x)) for x in sorted(items)]
    lines = "<br>".join(f"&nbsp;&nbsp;{item}" for item in ordered)
    return (
        f'<div class="{color_class} text-xs mb-2 break-words leading-5">'
        f'<div>{html.escape(title)}:</div>'
        f'<div class="font-normal">{lines}</div>'
        f'</div>'
    )

def extract_json_object_from_text(text):
    """Extract first JSON object substring from text by brace matching."""
    try:
        start_idx = text.find('{')
        if start_idx == -1:
            return None
        open_braces = 0
        for i in range(start_idx, len(text)):
            if text[i] == '{':
                open_braces += 1
            elif text[i] == '}':
                open_braces -= 1
                if open_braces == 0:
                    return text[start_idx:i+1]
        return None
    except:
        return None

def split_top_level_csv(text):
    parts = []
    current = []
    depth = 0
    in_single = False
    in_double = False
    escape = False
    for ch in text:
        current.append(ch)
        if escape:
            escape = False
            continue
        if ch == '\\':
            escape = True
            continue
        if ch == "'" and not in_double:
            in_single = not in_single
            continue
        if ch == '"' and not in_single:
            in_double = not in_double
            continue
        if in_single or in_double:
            continue
        if ch in '{[(':
            depth += 1
        elif ch in '}])':
            depth = max(0, depth - 1)
        elif ch == ',' and depth == 0:
            current.pop()
            part = ''.join(current).strip()
            if part:
                parts.append(part)
            current = []
    tail = ''.join(current).strip()
    if tail:
        parts.append(tail)
    return parts


def parse_levelplay_impression_text(payload):
    payload = (payload or '').strip()
    marker = 'LevelPlayImpressionData'
    if payload.startswith(marker):
        payload = payload[len(marker):].strip()
    if not (payload.startswith('{') and payload.endswith('}')):
        return None
    inner = payload[1:-1].strip()
    result = {}
    for item in split_top_level_csv(inner):
        if '=' not in item:
            continue
        key, raw = item.split('=', 1)
        key = key.strip()
        raw = raw.strip()
        if raw.startswith("'") and raw.endswith("'"):
            value = raw[1:-1]
        elif raw in ('null', 'None'):
            value = None
        elif raw == '':
            value = ''
        else:
            try:
                if any(ch in raw for ch in '.eE'):
                    value = float(raw)
                else:
                    value = int(raw)
            except:
                value = raw
        result[key] = value
    return result or None


def parse_appmetrica_adrevenue_text(payload):
    payload = (payload or "").strip()
    marker = "AdRevenue{"
    if payload.startswith(marker) and payload.endswith("}"):
        payload = payload[len(marker):-1].strip()

    result = {}
    for item in split_top_level_csv(payload):
        if "=" not in item:
            continue
        key, raw = item.split("=", 1)
        key = key.strip()
        raw = raw.strip()

        if key == "payload" and raw.startswith("{") and raw.endswith("}"):
            try:
                result[key] = json.loads(raw)
            except:
                result[key] = raw
            continue

        if raw.startswith("'") and raw.endswith("'"):
            value = raw[1:-1]
        elif raw in ("null", "None", "<null>"):
            value = None
        elif raw == "":
            value = ""
        elif raw.lower() == "true":
            value = True
        elif raw.lower() == "false":
            value = False
        else:
            try:
                if any(ch in raw for ch in ".eE"):
                    value = float(raw)
                else:
                    value = int(raw)
            except:
                value = raw
        result[key] = value
    return result or None

def _read_profile_sheet(ws, allow_default_fill=True):
    event_map = {}
    current_event = None
    header_row = 1
    game_name = ""

    if str(ws.cell(1, 2).value or "").strip().lower() == "game":
        game_name = str(ws.cell(1, 3).value or "").strip()
        header_row = 2

    for r in range(header_row + 1, ws.max_row + 1):
        event_val = ws.cell(r, 2).value
        param_cell = ws.cell(r, 3)
        param_val = param_cell.value

        if event_val:
            current_event = str(event_val).strip()
            if current_event:
                event_map.setdefault(current_event, {"specific": [], "default": []})

        if current_event and param_val:
            param = str(param_val).strip()
            fill = param_cell.fill
            fg = fill.fgColor.rgb if fill and fill.fgColor else None
            is_default = allow_default_fill and (fg == DEFAULT_PARAM_FILL)
            bucket = "default" if is_default else "specific"
            event_map[current_event][bucket].append(param)

    seen = set()
    merged_default = []
    for data in event_map.values():
        for p in data["default"]:
            if p not in seen:
                seen.add(p)
                merged_default.append(p)
    return game_name, merged_default, {k: v["specific"] for k, v in event_map.items()}


def _normalize_adrevenue_sheet_key(name):
    text = re.sub(r'[^a-z0-9]+', '', str(name or "").lower())
    aliases = {
        "appmetrica": {"appmetrica", "adrevenueappmetrica", "appmetricaadrevenue"},
        "appsflyer": {"appsflyer", "adrevenueappsflyer", "appsflyeradrevenue"},
        "all": {"adrevenue", "all", "default", "common", "shared"},
    }
    for key, values in aliases.items():
        if text in values:
            return key
    return text


def _read_adrevenue_sheet(ws):
    """Read AdRevenue params from a dedicated sheet with Appmetrica/Appsflyer columns."""
    params_by_source = {"appmetrica": [], "appsflyer": []}
    source_cols = {"appmetrica": None, "appsflyer": None}
    header_row = None

    # The adrevenue sheet can have headers on row 1 or row 2 (e.g. B=AdRevenue, C=Appmetrica, D=AppsFlyer).
    # Find the first row near the top that declares Appmetrica / Appsflyer columns.
    for row in range(1, min(ws.max_row, 5) + 1):
        candidate_cols = {"appmetrica": None, "appsflyer": None}
        for col in range(1, ws.max_column + 1):
            header = _normalize_adrevenue_sheet_key(ws.cell(row, col).value)
            if header in candidate_cols and candidate_cols[header] is None:
                candidate_cols[header] = col
        if any(candidate_cols.values()):
            header_row = row
            source_cols = candidate_cols
            break

    if header_row is None:
        return params_by_source

    for row in range(header_row + 1, ws.max_row + 1):
        for source, col in source_cols.items():
            if not col:
                continue
            raw = ws.cell(row, col).value
            if raw is None:
                continue
            value = str(raw).strip()
            if not value:
                continue
            params_by_source[source].append(value)

    return params_by_source


def load_default_params_config():
    """Load default params, event-specific params, and AdRevenue params from XLSX."""
    global default_params, event_specific_params, active_profile_game_name
    global adrevenue_default_params, adrevenue_source_params
    path = active_profile_path or DEFAULT_PARAMS_XLSX
    if not path or not os.path.exists(path):
        print(f"INFO: Default params sheet not found: {path}")
        default_params = []
        event_specific_params = {}
        active_profile_game_name = ""
        adrevenue_default_params = []
        adrevenue_source_params = {}
        return
    try:
        wb = load_workbook(path)
        active_profile_game_name, default_params, event_specific_params = _read_profile_sheet(wb.active, allow_default_fill=True)

        revenue_defaults = []
        revenue_specific = {}
        revenue_sheet = None
        for ws in wb.worksheets:
            title = _normalize_adrevenue_sheet_key(ws.title)
            if title == "adrevenue":
                revenue_sheet = ws
                break
        if revenue_sheet is None:
            for ws in wb.worksheets[1:]:
                title = _normalize_adrevenue_sheet_key(ws.title)
                if "revenue" in title or title in {"appmetrica", "appsflyer", "all"}:
                    revenue_sheet = ws
                    break

        if revenue_sheet:
            revenue_specific = _read_adrevenue_sheet(revenue_sheet)

        adrevenue_default_params = revenue_defaults
        adrevenue_source_params = revenue_specific
        print(
            f"INFO: Loaded default params: {len(default_params)}; events: {len(event_specific_params)}; "
            f"adrevenue defaults: {len(adrevenue_default_params)}; adrevenue sources: {len(adrevenue_source_params)}"
        )
    except Exception as e:
        print(f"ERROR: Failed to load default params sheet: {e}")
        default_params = []
        event_specific_params = {}
        active_profile_game_name = ""
        adrevenue_default_params = []
        adrevenue_source_params = {}


def _sanitize_profile_filename(filename):
    name = os.path.basename((filename or "").strip())
    name = name.replace("\\", "_").replace("/", "_")
    if not name.lower().endswith(".xlsx"):
        raise ValueError("Only .xlsx files are supported")
    return name


def _list_profile_names():
    try:
        files = [
            p.name for p in Path(PROFILE_DIR).glob("*.xlsx")
            if p.is_file()
        ]
        return sorted(files, key=lambda x: x.lower())
    except Exception:
        return []


def _ensure_default_profile_seed():
    source = DEFAULT_PARAMS_XLSX
    if not source or not os.path.exists(source):
        return
    target = os.path.join(PROFILE_DIR, os.path.basename(source))
    if os.path.abspath(source) == os.path.abspath(target):
        return
    if not os.path.exists(target):
        try:
            shutil.copyfile(source, target)
        except Exception as e:
            print(f"WARNING: Failed to seed default profile: {e}")


def _set_active_profile(profile_name=None):
    global active_profile_name, active_profile_path
    _ensure_default_profile_seed()
    names = _list_profile_names()
    if not names:
        active_profile_name = None
        active_profile_path = None
        load_default_params_config()
        return False

    selected = None
    if profile_name:
        clean = _sanitize_profile_filename(profile_name)
        if clean in names:
            selected = clean

    if not selected:
        preferred = os.path.basename(DEFAULT_PARAMS_XLSX) if DEFAULT_PARAMS_XLSX else None
        if preferred in names:
            selected = preferred
        else:
            selected = names[0]

    active_profile_name = selected
    active_profile_path = os.path.join(PROFILE_DIR, selected)
    load_default_params_config()
    return True


def _profile_payload():
    return {
        "profiles": _list_profile_names(),
        "current_profile": active_profile_name,
        "game_name": active_profile_game_name,
        "profile_dir": PROFILE_DIR,
        "default_event_names": sorted(event_specific_params.keys()),
    }

def _levenshtein_distance_limit(a, b, limit=2):
    """Compute Levenshtein distance with early exit if > limit."""
    if a == b:
        return 0
    if not a or not b:
        return max(len(a), len(b))
    if abs(len(a) - len(b)) > limit:
        return limit + 1

    # Initialize previous row
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        curr = [i]
        min_in_row = curr[0]
        for j, cb in enumerate(b, 1):
            cost = 0 if ca == cb else 1
            curr_val = min(
                prev[j] + 1,      # deletion
                curr[j - 1] + 1,  # insertion
                prev[j - 1] + cost  # substitution
            )
            curr.append(curr_val)
            if curr_val < min_in_row:
                min_in_row = curr_val
        if min_in_row > limit:
            return limit + 1
        prev = curr
    return prev[-1]

# --- GIAO DIỆN WEB (HTML/JS) ---
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="vi">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <link rel="icon" href="data:,"> <!-- Fix lỗi Favicon 404 -->
    <title>Event Inspector V2.0.0(52)</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/socket.io/4.7.4/socket.io.js"></script>
    <style>
        body { font-family: 'Inter', sans-serif; }
        .log-cell { max-width: 500px; word-wrap: break-word; font-family: monospace; font-size: 0.75rem; color: #6b7280; }
        .message-cell { white-space: nowrap; }
        #packageLogTable { table-layout: fixed; width: 100%; }
        #packageLogTable col.col-time { width: 110px; }
        #packageLogTable col.col-tag { width: 90px; }
        .details-cell { font-family: monospace; font-size: 0.8rem; line-height: 1.4; min-width: 260px; max-width: 640px; white-space: normal; overflow-wrap: anywhere; word-break: break-word; }
        .tag-cell { max-width: 90px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
        .tag-header { max-width: 90px; }
        .time-cell { white-space: nowrap; width: 1%; max-width: 110px; font-size: 0.75rem; }
        .time-header { white-space: nowrap; width: 1%; max-width: 110px; }
        .resizable { position: relative; }
        .resizer {
            position: absolute;
            right: -3px;
            top: 0;
            width: 12px;
            height: 100%;
            cursor: col-resize;
            user-select: none;
            z-index: 20;
        }
        .resizer:hover { background: rgba(59, 130, 246, 0.15); }
        #packageLogTableBody tr.selected, #packageHistoryTableBody tr.selected { background-color: #bfdbfe !important; }
        .resizer.disabled { cursor: not-allowed; background: transparent; }
        .details-cell pre { margin: 0; white-space: pre-wrap; overflow-wrap: anywhere; word-break: break-word; }
        .adrevenue-panel { border: 1px solid #e5e7eb; background: #f9fafb; border-radius: 0.5rem; padding: 0.5rem; }
        .adrevenue-details-panel { max-height: none; overflow: visible; }
        .adrevenue-details-panel pre { margin: 0 !important; background: transparent !important; border: 0 !important; padding: 0 !important; border-radius: 0 !important; overflow: visible !important; }
        .adrevenue-raw-panel { height: 16rem; overflow: auto; white-space: pre-wrap; overflow-wrap: anywhere; word-break: break-word; }
        #logDetailContent, #logDetailContent * { user-select: text; -webkit-user-select: text; }
        @keyframes pulse { 50% { opacity: .6; } }
        .animate-pulse-green { animation: pulse 2s cubic-bezier(0.4, 0, 0.6, 1) infinite; }
        .animate-record { animation: pulse 1.5s infinite; }
        .tab-btn { transition: all 0.2s ease-in-out; }
        .tab-btn.active { border-color: #4f46e5; color: #4f46e5; background-color: #eef2ff; }
        .log-row.selected { background-color: #dbeafe; }
        .log-cell,
        .details-cell,
        .details-cell *,
        #loadAdsTableBody td,
        #loadAdsExtTableBody td,
        #validatorTableBody td,
        #specificEventTableBody td,
        #adRevenueTableBody td,
        #callbackAdTableBody td,
        #sdkCheckTableBody td {
            user-select: text;
            -webkit-user-select: text;
        }
        
        /* Custom scrollbar for pre blocks */
        pre::-webkit-scrollbar { height: 6px; width: 6px; }
        pre::-webkit-scrollbar-thumb { background: #cbd5e1; border-radius: 3px; }
        pre::-webkit-scrollbar-track { background: #f1f5f9; }
    </style>
</head>
<body class="bg-gray-100 text-gray-800 h-screen overflow-hidden">

    <div class="container mx-auto p-3 sm:p-4 lg:p-6 h-full flex flex-col">
        <!-- HEADER -->
        <div class="bg-white rounded-xl shadow-md p-4 mb-4 flex-shrink-0">
            <div class="flex justify-between items-center flex-wrap gap-4">
                <div class="flex items-center gap-3">
                    <div>
                        <div class="flex items-center gap-2.5">
                            <h1 class="text-xl font-bold text-gray-700">Event Inspector</h1>
                            <span class="text-xs font-semibold bg-indigo-100 text-indigo-700 px-2 py-1 rounded-full">v2.3.0(3)</span>
                        </div>
                        <p class="text-sm text-gray-500">Integrates Load Ads & Event Validation.</p>
                    </div>
                    <div class="flex items-center gap-2">
                    <button id="restartAppBtn" class="bg-blue-500 hover:bg-blue-600 text-white text-sm font-semibold py-2 px-3 rounded-lg transition-colors shadow-sm">Check Update</button>
                    <button id="manualRestartBtn" class="bg-slate-500 hover:bg-slate-600 text-white text-sm font-semibold py-2 px-3 rounded-lg transition-colors shadow-sm">Restart</button>
                    <button id="platformBtn" class="bg-white hover:bg-gray-50 text-slate-700 border border-slate-300 text-sm font-semibold py-2 px-3 rounded-lg transition-colors shadow-sm">Platform: Android</button>
                </div>
                </div>
                <div class="flex items-center gap-4">
                    <div class="flex items-center gap-2">
                        <button id="pauseBtn" class="text-sm font-semibold py-2 px-3 rounded-lg transition-colors shadow-sm bg-yellow-500 hover:bg-yellow-600 text-white">Pause</button>
                        <button id="clearAllBtn" class="text-sm font-semibold py-2 px-3 rounded-lg transition-colors shadow-sm bg-red-500 hover:bg-red-600 text-white">Clear All</button>
                    </div>
                    <div class="text-right p-2.5 rounded-lg bg-gray-50 border min-w-[250px]">
                        <p class="text-sm font-semibold text-gray-700 mb-1 border-b pb-1">Connected Devices:</p>
                        <div id="deviceList" class="text-sm text-gray-600">
                             <p class="text-orange-500">WAITING...</p>
                        </div>
                    </div>
                     <div>
                        <label for="deviceFilter" class="block text-xs font-medium text-gray-700">Filter by Device:</label>
                        <select id="deviceFilter" class="mt-1 block w-full pl-3 pr-10 py-1.5 text-sm border-gray-300 focus:outline-none focus:ring-indigo-500 focus:border-indigo-500 rounded-md">
                            <option value="all">All Devices</option>
                        </select>
                    </div>
                </div>
            </div>
        </div>

        <div id="platformModal" class="fixed inset-0 z-50 hidden items-center justify-center bg-black/40 px-4">
            <div class="bg-white rounded-xl shadow-xl border w-full max-w-md p-5">
                <h2 class="text-lg font-bold text-slate-800 mb-2">Select Platform</h2>
                <div class="grid grid-cols-1 sm:grid-cols-2 gap-3 mt-4">
                    <button data-platform-choice="android" class="platform-choice-btn rounded-lg border border-slate-200 bg-slate-50 hover:bg-indigo-50 hover:border-indigo-300 text-slate-800 font-semibold py-4 px-4 transition-colors">Android</button>
                    <button data-platform-choice="ios" class="platform-choice-btn rounded-lg border border-slate-200 bg-slate-50 hover:bg-indigo-50 hover:border-indigo-300 text-slate-800 font-semibold py-4 px-4 transition-colors">iOS</button>
                </div>
            </div>
        </div>

        <!-- TABS & ACTIONS -->
        <div class="mb-3 flex-shrink-0">
            <div class="flex justify-between items-end border-b border-gray-200">
                <div class="flex flex-wrap">
                    <button id="tabBtnLoadAdsExt" class="tab-btn active text-sm font-semibold py-2 px-4 -mb-px border-b-2 border-transparent" onclick="switchTab('LoadAdsExt')">Load Ads Ironsource</button>
                    
                    <button id="tabBtnValidator" class="tab-btn text-sm font-semibold py-2 px-4 -mb-px border-b-2 border-transparent" onclick="switchTab('Validator')">Default Events/Params</button>
                    <button id="tabBtnSpecific" class="tab-btn text-sm font-semibold py-2 px-4 -mb-px border-b-2 border-transparent" onclick="switchTab('Specific')">Specific Validator</button>
                    <button id="tabBtnAdRevenue" class="tab-btn text-sm font-semibold py-2 px-4 -mb-px border-b-2 border-transparent" onclick="switchTab('AdRevenue')">AdRevenue</button>
                    <button id="tabBtnCallbackAd" class="tab-btn text-sm font-semibold py-2 px-4 -mb-px border-b-2 border-transparent" onclick="switchTab('CallbackAd')">CallBack & Ads</button>
                    <button id="tabBtnSdkCheck" class="tab-btn text-sm font-semibold py-2 px-4 -mb-px border-b-2 border-transparent" onclick="switchTab('SdkCheck')">SDK Check</button>
                    <button id="tabBtnPackage" class="tab-btn text-sm font-semibold py-2 px-4 -mb-px border-b-2 border-transparent" onclick="switchTab('Package')">Package Logcat</button>
                </div>
            </div>
        </div>

        <!-- CONTENT -->
        <div class="flex-grow overflow-auto">

            <!-- TAB 1: Load Ads -->
            <div id="tabContentLoadAds" class="hidden">
                 <div class="bg-white rounded-xl shadow-md p-4">
                    <div class="flex items-center gap-2 bg-gray-50 p-2.5 rounded-lg border mb-3">
                        <span class="text-sm font-semibold text-gray-700">Record Load Ads:</span>
                        <input type="text" id="sheetName_LoadAds" placeholder="Tên Sheet..." class="border p-2 rounded text-sm w-48 outline-none">
                        <button id="btnRecord_LoadAds" onclick="toggleRecord('LoadAds')" class="bg-indigo-600 hover:bg-indigo-700 text-white font-semibold py-2 px-3 rounded shadow text-xs">Start Record</button>
                    </div>

                    <div class="overflow-x-auto">
                        <table id="packageLogTable" class="min-w-full bg-white">
                            <colgroup>
                                <col class="col-time">
                                <col class="col-tag">
                                <col class="col-message">
                            </colgroup>
                            <thead class="bg-gray-50 sticky top-0 z-10">
                                <tr>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Device</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Ad_network</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Format</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Raw Log</th>
                                </tr>
                            </thead>
                            <tbody id="loadAdsTableBody" class="divide-y divide-gray-200"></tbody>
                        </table>
                    </div>
                </div>
            </div>

            <!-- TAB 2: Load Ads Ironsource -->
            <div id="tabContentLoadAdsExt">
                 <div class="bg-white rounded-xl shadow-md p-4">
                    <div class="flex items-center gap-2 bg-gray-50 p-2.5 rounded-lg border mb-3">
                        <span class="text-sm font-semibold text-gray-700">Record Load Ads Ironsource:</span>
                        <input type="text" id="sheetName_LoadAdsExt" placeholder="Tên Sheet..." class="border p-2 rounded text-sm w-48 outline-none">
                        <button id="btnRecord_LoadAdsExt" onclick="toggleRecord('LoadAdsExt')" class="bg-indigo-600 hover:bg-indigo-700 text-white font-semibold py-2 px-3 rounded shadow text-xs">Start Record</button>
                    </div>

                    <div class="overflow-x-auto">
                        <table class="min-w-full w-full bg-white table-fixed">
                            <colgroup>
                                <col style="width:88px">
                                <col style="width:96px">
                                <col style="width:112px">
                                <col>
                            </colgroup>
                            <thead class="bg-gray-50 sticky top-0 z-10">
                                <tr>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-2 border-b">Device</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-2 border-b">Ad_network</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-2 border-b">Format</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Raw Log</th>
                                </tr>
                            </thead>
                            <tbody id="loadAdsExtTableBody" class="divide-y divide-gray-200"></tbody>
                        </table>
                    </div>
                </div>
            </div>

            <!-- TAB 3: Validator -->
            <div id="tabContentValidator" class="hidden">
                <div class="bg-white rounded-xl shadow-md p-4 mb-4">
                    <div class="grid grid-cols-1 lg:grid-cols-3 gap-4">
                        <div class="lg:col-span-1">
                            <div class="space-y-3">
                                <div>
                                    <div class="flex items-center gap-2 mb-1.5">
                                        <label for="profileSelect" class="text-[11px] font-medium text-gray-700">Game Profile:</label>
                                        <p id="profileGameText" class="text-[11px] font-medium text-indigo-700"></p>
                                    </div>
                                    <div class="space-y-2">
                                        <select id="profileSelect" class="w-full h-9 px-3 border rounded-md shadow-sm text-xs"></select>
                                        <input type="file" id="profileFileInput" accept=".xlsx" class="hidden">
                                        <div class="flex flex-wrap items-center gap-3">
                                            <button id="importProfileBtn" class="bg-slate-700 hover:bg-slate-800 text-white font-medium text-xs px-3 rounded-lg h-9 min-w-[116px]">Import Profile</button>
                                            <button id="reloadProfileBtn" class="bg-slate-200 hover:bg-slate-300 text-gray-800 font-medium text-xs px-3 rounded-lg h-9 min-w-[116px]">Reload Profile</button>
                                        </div>
                                    </div>
                                </div>
                                <div class="grid grid-cols-1 sm:grid-cols-2 gap-3">
                                    <div>
                                        <label for="validatorEventFilterInput" class="block text-[11px] font-medium text-gray-700 mb-1">Filter by Event Name:</label>
                                        <input type="text" id="validatorEventFilterInput" class="w-full h-9 px-3 border rounded-md shadow-sm text-xs" placeholder="Type to filter events...">
                                    </div>
                                    <div>
                                        <label for="validatorRawFilterInput" class="block text-[11px] font-medium text-gray-700 mb-1">Filter by Raw Log:</label>
                                        <input type="text" id="validatorRawFilterInput" class="w-full h-9 px-3 border rounded-md shadow-sm text-xs" placeholder="Search raw log...">
                                    </div>
                                </div>
                                <div class="flex items-center gap-3 pt-1">
                                    <button id="startValidationBtn" class="bg-indigo-600 hover:bg-indigo-700 text-white font-semibold text-xs px-4 rounded-lg h-9">Start Checking</button>
                                    <button id="clearValidatorFilterBtn" class="bg-gray-200 hover:bg-gray-300 text-gray-800 font-medium text-xs px-4 rounded-lg h-9">Clear Filter</button>
                                    <div class="flex items-center gap-3 text-[11px] text-gray-700 ml-1">
                                        <label class="inline-flex items-center gap-1.5"><input type="radio" name="validatorSourceFilter" value="all" checked> <span>All</span></label>
                                        <label class="inline-flex items-center gap-1.5"><input type="radio" name="validatorSourceFilter" value="appmetrica"> <span>Appmetrica</span></label>
                                        <label class="inline-flex items-center gap-1.5"><input type="radio" name="validatorSourceFilter" value="firebase"> <span>Firebase</span></label>
                                    </div>
                                </div>
                            </div>
                        </div>
                        <div class="lg:col-span-2">
                            <div class="text-xs font-semibold text-gray-700 mb-2">Default Events Status</div>
                            <div id="defaultEventStatusList" class="flex flex-wrap gap-1.5 text-xs"></div>
                        </div>
                    </div>
                </div>
                <div class="bg-white rounded-xl shadow-md p-4">
                    <div class="overflow-x-auto">
                        <table class="min-w-full bg-white">
                           <thead class="bg-gray-50 sticky top-0 z-10">
                                <tr>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Device</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Status</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Event Name</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Details</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Raw Log</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Action</th>
                                </tr>
                            </thead>
                            <tbody id="validatorTableBody" class="divide-y divide-gray-200"></tbody>
                        </table>
                    </div>
                </div>
            </div>

            <!-- TAB 4: Specific -->
            <div id="tabContentSpecific" class="hidden">
                <div class="bg-white rounded-xl shadow-md p-4 mb-4">
                    <div class="grid grid-cols-1 lg:grid-cols-[minmax(0,1fr)_150px_minmax(0,1fr)] gap-4 items-start">
                         <div>
                            <label for="specificEventInput" class="block text-xs font-medium text-gray-700 mb-1">Filter by Event Names:</label>
                            <textarea id="specificEventInput" rows="4" class="w-full p-2 border rounded-md shadow-sm" placeholder="Leave empty to show all events..."></textarea>
                         </div>
                         <div>
                            <label class="block text-xs font-medium text-gray-700 mb-1">Source Filter:</label>
                            <div class="flex flex-col gap-2 text-xs text-gray-700 pt-1">
                                <label class="inline-flex items-center gap-1.5">
                                    <input type="radio" name="specificSourceFilter" value="all" checked>
                                    <span>All</span>
                                </label>
                                <label class="inline-flex items-center gap-1.5">
                                    <input type="radio" name="specificSourceFilter" value="firebase">
                                    <span>Firebase</span>
                                </label>
                                <label class="inline-flex items-center gap-1.5">
                                    <input type="radio" name="specificSourceFilter" value="appmetrica">
                                    <span>Appmetrica</span>
                                </label>
                            </div>
                         </div>
                         <div>
                            <label for="specificParamInput" class="block text-xs font-medium text-gray-700 mb-1">Filter by Text:</label>
                            <textarea id="specificParamInput" rows="4" class="w-full p-2 border rounded-md shadow-sm" placeholder="Leave empty to show all raw logs..."></textarea>
                         </div>
                    </div>
                </div>
                <div class="bg-white rounded-xl shadow-md p-4">
                    <div class="overflow-x-auto">
                        <table class="min-w-full bg-white">
                           <thead class="bg-gray-50 sticky top-0 z-10">
                                <tr>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Device</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Status</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Event Name</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Details</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Raw Log</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Action</th>
                                </tr>
                            </thead>
                            <tbody id="specificEventTableBody" class="divide-y divide-gray-200"></tbody>
                        </table>
                    </div>
                </div>
            </div>

            <!-- TAB 5: AdRevenue -->
            <div id="tabContentAdRevenue" class="hidden">
                <div class="bg-white rounded-xl shadow-md p-4 mb-4">
                     <div class="grid grid-cols-1 md:grid-cols-[minmax(260px,380px)_minmax(0,1fr)] gap-4 items-start">
                        <div class="space-y-3">
                            <div class="flex items-center gap-3">
                                <label for="adRevenueProfileSelect" class="text-xs font-medium text-gray-700 whitespace-nowrap">Game Profile:</label>
                                <p id="adRevenueProfileGameText" class="text-xs font-medium text-indigo-700"></p>
                            </div>
                            <select id="adRevenueProfileSelect" class="w-full h-10 px-3 border rounded-md shadow-sm text-sm"></select>
                        </div>
                        <div class="space-y-3">
                            <div>
                                <label class="block text-xs font-medium text-gray-700 mb-1">Source Filter:</label>
                                <div class="flex flex-wrap items-center gap-x-4 gap-y-2 text-sm">
                                    <label class="inline-flex items-center whitespace-nowrap">
                                        <input name="adRevenueSourceFilter" type="radio" value="all" checked class="h-4 w-4 text-indigo-600 border-gray-300 focus:ring-indigo-500">
                                        <span class="ml-2 text-sm text-gray-900">All</span>
                                    </label>
                                    <label class="inline-flex items-center whitespace-nowrap">
                                        <input name="adRevenueSourceFilter" type="radio" value="appmetrica" class="h-4 w-4 text-indigo-600 border-gray-300 focus:ring-indigo-500">
                                        <span class="ml-2 text-sm text-gray-900">Appmetrica</span>
                                    </label>
                                    <label class="inline-flex items-center whitespace-nowrap">
                                        <input name="adRevenueSourceFilter" type="radio" value="appsflyer" class="h-4 w-4 text-indigo-600 border-gray-300 focus:ring-indigo-500">
                                        <span class="ml-2 text-sm text-gray-900">Appsflyer</span>
                                    </label>
                                </div>
                            </div>
                            <div>
                                <label for="adRevenueFilterInput" class="block text-xs font-medium text-gray-700 mb-1">Filter logs by text:</label>
                                <input type="text" id="adRevenueFilterInput" class="w-full p-2 border rounded-md shadow-sm" placeholder="Search in raw log...">
                            </div>
                        </div>
                    </div>
                </div>
                <div class="bg-white rounded-xl shadow-md p-4">
                    <div class="overflow-x-auto">
                        <table class="min-w-full bg-white">
                           <thead class="bg-gray-50 sticky top-0 z-10">
                                <tr>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Device</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Status</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Event Name</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Details</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Raw Log</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Action</th>
                                </tr>
                            </thead>
                            <tbody id="adRevenueTableBody" class="divide-y divide-gray-200"></tbody>
                        </table>
                    </div>
                </div>
            </div>

            <!-- TAB 6: CallbackAd -->
            <div id="tabContentCallbackAd" class="hidden">
                 <div class="bg-white rounded-xl shadow-md p-4">
                    <div class="mb-3 grid grid-cols-1 lg:grid-cols-[1fr_minmax(320px,420px)] gap-4 items-end">
                        <div>
                            <label class="block text-xs font-medium text-gray-700">Filter by Type:</label>
                            <div class="mt-2 flex flex-wrap lg:flex-nowrap items-center gap-x-4 gap-y-2 text-sm">
                                <label class="inline-flex items-center whitespace-nowrap">
                                    <input id="callbackTypeAll" name="callbackType" type="radio" value="all" checked class="h-4 w-4 text-indigo-600 border-gray-300 focus:ring-indigo-500">
                                    <span class="ml-2 text-sm text-gray-900">All</span>
                                </label>
                                <label class="inline-flex items-center whitespace-nowrap">
                                    <input id="callbackTypeCallback" name="callbackType" type="radio" value="callback" class="h-4 w-4 text-indigo-600 border-gray-300 focus:ring-indigo-500">
                                    <span class="ml-2 text-sm text-gray-900">Callback Levelplay</span>
                                </label>
                                <label class="inline-flex items-center whitespace-nowrap">
                                    <input id="callbackTypeGadsme" name="callbackType" type="radio" value="gadsme_callback" class="h-4 w-4 text-indigo-600 border-gray-300 focus:ring-indigo-500">
                                    <span class="ml-2 text-sm text-gray-900">Callback Gadsme</span>
                                </label>
                                <label class="inline-flex items-center whitespace-nowrap">
                                    <input id="callbackTypeAdverty5" name="callbackType" type="radio" value="adverty5_callback" class="h-4 w-4 text-indigo-600 border-gray-300 focus:ring-indigo-500">
                                    <span class="ml-2 text-sm text-gray-900">Callback Adverty5</span>
                                </label>
                                <label class="inline-flex items-center whitespace-nowrap">
                                    <input id="callbackTypeAdEvent" name="callbackType" type="radio" value="ad_event" class="h-4 w-4 text-indigo-600 border-gray-300 focus:ring-indigo-500">
                                    <span class="ml-2 text-sm text-gray-900">Ad Event</span>
                                </label>
                            </div>
                        </div>
                        <div class="lg:justify-self-end w-full lg:max-w-[420px]">
                            <label for="callbackAdFilterInput" class="block text-xs font-medium text-gray-700">Filter (in raw log):</label>
                            <input type="text" id="callbackAdFilterInput" class="mt-1 block w-full p-2 border border-gray-300 rounded-md shadow-sm focus:ring-indigo-500 focus:border-indigo-500" placeholder="Search...">
                        </div>
                    </div>
                    <div class="overflow-x-auto">
                        <table class="min-w-full bg-white">
                            <thead class="bg-gray-50 sticky top-0 z-10">
                                <tr>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Device</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Type</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Event / Key</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Details</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Raw Log</th>
                                    <th class="text-left text-sm font-semibold text-gray-600 py-2 px-3 border-b">Action</th>
                                </tr>
                            </thead>
                            <tbody id="callbackAdTableBody" class="divide-y divide-gray-200"></tbody>
                        </table>
                    </div>
                </div>
            </div>

            <!-- TAB 7: SDK Check -->
            <div id="tabContentSdkCheck" class="hidden">
                <div class="bg-white rounded-xl shadow-md p-4 mb-4">
                    <h2 class="text-lg font-semibold mb-2">SDK Check Setup</h2>
                    <div class="grid grid-cols-1 md:grid-cols-3 gap-4 items-end">
                         <div class="md:col-span-2">
                            <label for="sdkCheckInput" class="block text-xs font-medium text-gray-700 mb-1">SDKs to Check (dán nội dung):</label>
                            <textarea id="sdkCheckInput" rows="8" class="w-full p-2 border rounded-md shadow-sm font-mono text-sm" placeholder="AppLovin\n&quot;Adapter 4.3.54&quot;, &quot;search_pattern&quot;: &quot;Adapter 4.3.54&quot;\n..."></textarea>
                         </div>
                         <div>
                            <button id="startSdkCheckBtn" class="bg-indigo-600 hover:bg-indigo-700 text-white text-sm font-semibold py-2 px-4 rounded-lg w-full h-10">Start Checking</button>
                         </div>
                    </div>
                </div>
                <div class="bg-white rounded-xl shadow-md p-4">
                    <div class="flex items-center justify-between gap-3 mb-2">
                        <h2 class="text-lg font-semibold">SDK Check Results</h2>
                        <span id="sdkCheckPlatformBadge" class="text-xs font-semibold bg-slate-100 text-slate-700 px-3 py-1 rounded-full">Android Table</span>
                    </div>
                    <div id="sdkCheckAndroidPanel" class="sdk-check-panel overflow-x-auto" data-sdk-platform="android">
                        <div class="text-xs font-semibold text-slate-500 mb-2">Android SDK Check</div>
                        <table class="min-w-full bg-white">
                            <tbody id="sdkCheckTableBody" class="divide-y divide-gray-200"></tbody>
                        </table>
                    </div>
                    <div id="sdkCheckIosPanel" class="sdk-check-panel overflow-x-auto hidden" data-sdk-platform="ios">
                        <div class="text-xs font-semibold text-slate-500 mb-2">iOS SDK Check</div>
                        <table class="min-w-full bg-white">
                            <tbody id="sdkCheckIosTableBody" class="divide-y divide-gray-200">
                                <tr><td class="py-2 px-4 text-sm text-gray-500 italic">iOS table ready. Device connection will be added in the next step.</td></tr>
                            </tbody>
                        </table>
                    </div>
                </div>
            </div>
            
            <!-- TAB 8: Package -->
            <div id="tabContentPackage" class="hidden">
                 <div class="bg-white rounded-xl shadow-md p-5 mb-4">
                     <div class="grid grid-cols-1 lg:grid-cols-[0.85fr_1.7fr] gap-6 items-start">
                        <div class="max-w-md">
                            <label for="packageIdInput" class="block text-[11px] font-medium text-gray-700 mb-1">Package ID:</label>
                            <input type="text" id="packageIdInput" class="w-full p-2 text-[11px] border rounded-md shadow-sm" placeholder="com.example.app">
                            <div class="flex justify-center items-center gap-2 mt-3">
                                <button id="startPackageLogBtn" class="bg-indigo-600 hover:bg-indigo-700 text-white font-semibold text-xs py-2 px-4 rounded-lg h-9">Start</button>
                                <button id="openPackageHistoryBtn" class="text-sm font-semibold py-2 px-3 rounded-lg transition-colors shadow-sm bg-slate-600 hover:bg-slate-700 text-white h-9">Recorded Log</button>
                            </div>
                        </div>
                        <div class="max-w-[980px] min-h-[120px] flex flex-col justify-start">
                            <label class="block text-[11px] font-medium text-gray-700 mb-1">Quick Select:</label>
                            <div class="flex items-start gap-16 text-[11px] text-gray-700 pt-1">
                                <div class="flex flex-col gap-1 min-w-[320px]">
                                    <label class="inline-flex items-center gap-2">
                                        <input type="checkbox" class="package-id-checkbox h-4 w-4 text-indigo-600 border-gray-300 rounded focus:ring-indigo-500" value="com.indiez.nonogram">
                                        <span>NG - com.indiez.nonogram</span>
                                    </label>
                                    <label class="inline-flex items-center gap-2">
                                        <input type="checkbox" class="package-id-checkbox h-4 w-4 text-indigo-600 border-gray-300 rounded focus:ring-indigo-500" value="com.indiez.train.miner">
                                        <span>TM - com.indiez.train.miner</span>
                                    </label>
                                    <label class="inline-flex items-center gap-2">
                                        <input type="checkbox" class="package-id-checkbox h-4 w-4 text-indigo-600 border-gray-300 rounded focus:ring-indigo-500" value="com.indiez.idletycoon.horse.racing">
                                        <span>HR - com.indiez.idletycoon.horse.racing</span>
                                    </label>
                                    <label class="inline-flex items-center gap-2">
                                        <input type="checkbox" class="package-id-checkbox h-4 w-4 text-indigo-600 border-gray-300 rounded focus:ring-indigo-500" value="com.indiez.solitaire.food">
                                        <span>SC - com.indiez.solitaire.food</span>
                                    </label>
                                </div>
                                <div class="flex flex-col gap-1 min-w-[320px]">
                                    <label class="inline-flex items-center gap-2">
                                        <input type="checkbox" class="package-id-checkbox h-4 w-4 text-indigo-600 border-gray-300 rounded focus:ring-indigo-500" value="com.nostel.dot.line.puzzle">
                                        <span>KN - com.nostel.dot.line.puzzle</span>
                                    </label>
                                    <label class="inline-flex items-center gap-2">
                                        <input type="checkbox" class="package-id-checkbox h-4 w-4 text-indigo-600 border-gray-300 rounded focus:ring-indigo-500" value="com.nostel.parking.car">
                                        <span>CP - com.nostel.parking.car</span>
                                    </label>
                                    <label class="inline-flex items-center gap-2">
                                        <input type="checkbox" class="package-id-checkbox h-4 w-4 text-indigo-600 border-gray-300 rounded focus:ring-indigo-500" value="com.nostel.jigblock">
                                        <span>JS - com.nostel.jigblock</span>
                                    </label>
                                    <label class="inline-flex items-center gap-2">
                                        <input type="checkbox" class="package-id-checkbox h-4 w-4 text-indigo-600 border-gray-300 rounded focus:ring-indigo-500" value="tap.monster.block.away">
                                        <span>TP - tap.monster.block.away</span>
                                    </label>
                                </div>
                            </div>
                        </div>
                    </div>
                </div>
            </div>

        </div>
    </div>
    
    <!-- MODALS -->
    <div id="jsonModal" class="fixed inset-0 z-50 flex items-center justify-center bg-black bg-opacity-50 hidden">
        <div class="bg-white rounded-xl shadow-lg p-6 w-1/2 h-2/3 flex flex-col">
            <div class="flex justify-between items-center border-b pb-3 mb-4">
                <h2 class="text-xl font-bold text-gray-800">Formatted JSON</h2>
                <button id="closeJsonModal" class="text-gray-500 hover:text-gray-800 text-2xl font-bold">&times;</button>
            </div>
            <div class="flex-grow overflow-auto bg-gray-800 text-white p-4 rounded-md">
                <pre><code id="jsonContent"></code></pre>
            </div>
        </div>
    </div>

    <div id="packageHistoryModal" class="fixed inset-0 z-50 flex items-center justify-center bg-black bg-opacity-50 hidden">
        <div class="bg-white rounded-xl shadow-lg p-6 w-[88vw] h-[82vh] flex flex-col">
            <div class="flex justify-between items-start border-b pb-3 mb-4">
                <div class="flex items-center gap-3 flex-wrap">
                    <h2 class="text-xl font-bold text-gray-800">Recorded Package Logs</h2>
                    <button id="loadPackageHistoryBtn" class="bg-slate-600 hover:bg-slate-700 text-white font-semibold text-xs py-2 px-4 rounded-lg h-9">Load</button>
                    <button id="loadMorePackageHistoryBtn" class="bg-slate-100 hover:bg-slate-200 text-slate-700 font-semibold text-xs py-2 px-4 rounded-lg h-9 hidden">Load More</button>
                    <button id="refreshPackageSessionsBtn" class="bg-gray-200 hover:bg-gray-300 text-gray-800 font-semibold text-xs py-2 px-4 rounded-lg h-9">Refresh Sessions</button>
                    <button id="exportPackageHistoryAllBtn" class="bg-emerald-500 hover:bg-emerald-600 text-white font-semibold text-xs py-2 px-4 rounded-lg h-9">Export Full</button>
                    <button id="exportPackageHistoryFilteredBtn" class="bg-emerald-600 hover:bg-emerald-700 text-white font-semibold text-xs py-2 px-4 rounded-lg h-9">Export Filtered</button>
                </div>
                <button id="closePackageHistoryModal" class="text-gray-500 hover:text-gray-800 text-2xl font-bold">&times;</button>
            </div>
            <div class="flex flex-col gap-4">
                <div class="grid grid-cols-1 xl:grid-cols-[340px_160px_160px_1fr_1fr_1fr] gap-4 items-end">
                    <div>
                        <label for="packageHistorySessionSelect" class="block text-[11px] font-medium text-gray-700 mb-1">Saved Sessions:</label>
                        <select id="packageHistorySessionSelect" class="w-full p-2 text-[11px] border rounded-md shadow-sm"></select>
                    </div>
                    <div>
                        <label class="block text-[11px] font-medium text-gray-700 mb-1">&nbsp;</label>
                        <button id="clearPackageHistoryBtn" class="w-full bg-red-50 hover:bg-red-100 text-red-700 border border-red-200 font-semibold text-xs py-2 px-4 rounded-lg h-9">Clear Data</button>
                    </div>
                    <div>
                        <label class="block text-[11px] font-medium text-gray-700 mb-1">&nbsp;</label>
                        <button id="clearPackageHistoryFiltersBtn" class="w-full bg-gray-100 hover:bg-gray-200 text-gray-700 border border-gray-200 font-semibold text-xs py-2 px-4 rounded-lg h-9">Clear Filter</button>
                    </div>
                    <div>
                        <label for="packageHistoryFilterInput" class="block text-[11px] font-medium text-gray-700 mb-1">Filter 1:</label>
                        <input type="text" id="packageHistoryFilterInput" class="w-full p-2 text-[11px] border rounded-md shadow-sm" placeholder="Search saved log text 1...">
                    </div>
                    <div>
                        <label for="packageHistoryFilterInput2" class="block text-[11px] font-medium text-gray-700 mb-1">Filter 2:</label>
                        <input type="text" id="packageHistoryFilterInput2" class="w-full p-2 text-[11px] border rounded-md shadow-sm" placeholder="Search saved log text 2...">
                    </div>
                    <div>
                        <label for="packageHistoryFilterInput3" class="block text-[11px] font-medium text-gray-700 mb-1">Filter 3:</label>
                        <input type="text" id="packageHistoryFilterInput3" class="w-full p-2 text-[11px] border rounded-md shadow-sm" placeholder="Search saved log text 3...">
                    </div>
                </div>
                <p id="packageHistoryMeta" class="text-[11px] text-gray-500">No session selected.</p>
            </div>
            <div id="packageHistoryTableWrap" class="overflow-auto border rounded-md mt-3 flex-1">
                <table class="min-w-full bg-white">
                    <thead class="bg-gray-50 sticky top-0 z-10">
                        <tr>
                            <th class="text-left text-xs font-semibold text-gray-600 py-2 px-2 border-b">Time</th>
                            <th class="text-left text-xs font-semibold text-gray-600 py-2 px-2 border-b">Device</th>
                            <th class="text-left text-xs font-semibold text-gray-600 py-2 px-2 border-b w-[88px] max-w-[88px]">Tag</th>
                            <th class="text-left text-xs font-semibold text-gray-600 py-2 px-2 border-b">Message</th>
                        </tr>
                    </thead>
                    <tbody id="packageHistoryTableBody" class="divide-y divide-gray-200"></tbody>
                </table>
            </div>
        </div>
    </div>

    <div id="packageStreamModal" class="fixed inset-0 z-50 flex items-center justify-center bg-black bg-opacity-50 hidden">
        <div class="bg-white rounded-xl shadow-lg p-6 w-[94vw] h-[88vh] flex flex-col">
            <div class="border-b pb-3 mb-4">
                <div class="grid grid-cols-1 xl:grid-cols-[420px_1fr_420px] gap-6 items-start">
                    <div class="min-w-0">
                        <h2 class="text-xl font-bold text-gray-800 mb-4">Package Log Stream</h2>
                        <label for="packageTagFilterInput" class="block text-[11px] font-medium text-gray-700 mb-1">Tag Filter:</label>
                        <input type="text" id="packageTagFilterInput" class="w-full p-2 border rounded-md shadow-sm" placeholder="Tag...">
                        <div class="flex items-center gap-3 mt-4">
                            <button id="pausePackageLogBtn" class="text-sm font-semibold py-2 px-3 rounded-lg transition-colors shadow-sm bg-amber-500 hover:bg-amber-600 text-white">Pause</button>
                            <button id="stopPackageLogBtn" class="text-sm font-semibold py-2 px-3 rounded-lg transition-colors shadow-sm bg-red-500 hover:bg-red-600 text-white">Stop</button>
                        </div>
                    </div>
                    <div class="min-w-0">
                        <div class="flex flex-wrap items-center gap-4 mb-4">
                            <label class="inline-flex items-center gap-2">
                                <input id="showErrorsOnly" type="checkbox" class="h-4 w-4 text-indigo-600 border-gray-300 rounded focus:ring-indigo-500">
                                <span class="block text-xs text-gray-900">Show errors only</span>
                            </label>
                            <label class="inline-flex items-center gap-2">
                                <input id="showWarningsOnly" type="checkbox" class="h-4 w-4 text-indigo-600 border-gray-300 rounded focus:ring-indigo-500">
                                <span class="block text-xs text-gray-900">Show only Warning</span>
                            </label>
                            <label class="inline-flex items-center gap-2">
                                <input id="autoScroll" type="checkbox" checked class="h-4 w-4 text-indigo-600 border-gray-300 rounded focus:ring-indigo-500">
                                <span class="block text-xs text-gray-900">Auto-scroll</span>
                            </label>
                        </div>
                        <div class="pt-4 border-t grid grid-cols-1 gap-y-2 text-xs text-gray-700">
                            <label class="inline-flex items-center gap-2">
                                <input type="radio" name="tagQuickFilter" value="" checked class="h-4 w-4 text-indigo-600 border-gray-300 focus:ring-indigo-500">
                                <span>All</span>
                            </label>
                            <label class="inline-flex items-center gap-2">
                                <input type="radio" name="tagQuickFilter" value="integrationhelper" class="h-4 w-4 text-indigo-600 border-gray-300 focus:ring-indigo-500">
                                <span>Integrationhelper</span>
                            </label>
                            <label class="inline-flex items-center gap-2">
                                <input type="radio" name="tagQuickFilter" value="appsflyer" class="h-4 w-4 text-indigo-600 border-gray-300 focus:ring-indigo-500">
                                <span>Appsflyer</span>
                            </label>
                            <label class="inline-flex items-center gap-2">
                                <input type="radio" name="tagQuickFilter" value="appmetrica" class="h-4 w-4 text-indigo-600 border-gray-300 focus:ring-indigo-500">
                                <span>Appmetrica</span>
                            </label>
                        </div>
                    </div>
                    <div class="min-w-0">
                        <label for="packageFilterInput" class="block text-[11px] font-medium text-gray-700 mb-1">Message Filter 1:</label>
                        <input type="text" id="packageFilterInput" class="w-full p-2 border rounded-md shadow-sm" placeholder="Search text 1...">
                        <div class="mt-3">
                            <label for="packageFilterInput2" class="block text-[11px] font-medium text-gray-700 mb-1">Message Filter 2:</label>
                            <input type="text" id="packageFilterInput2" class="w-full p-2 border rounded-md shadow-sm" placeholder="Search text 2...">
                        </div>
                    </div>
                </div>
            </div>
            <div id="packageLogContainer" class="overflow-auto overflow-x-auto flex-1 border rounded-md" style="min-height: 0;">
                <table class="min-w-full bg-white">
                   <thead class="bg-gray-50 sticky top-0 z-10">
                        <tr>
                            <th class="text-left text-xs font-semibold text-gray-600 py-1.5 px-2 border-b time-header resizable col-time">Time<div class="resizer" data-col="time"></div></th>
                            <th class="text-left text-xs font-semibold text-gray-600 py-1.5 pr-1 pl-2 border-b tag-header resizable col-tag">Tag<div class="resizer" data-col="tag"></div></th>
                            <th class="text-left text-xs font-semibold text-gray-600 py-1.5 pl-1 pr-3 border-b resizable col-message">Message<div class="resizer" data-col="message"></div></th>
                        </tr>
                    </thead>
                    <tbody id="packageLogTableBody" class="divide-y divide-gray-200"></tbody>
                </table>
            </div>
        </div>
    </div>

    <div id="logDetailModal" class="fixed inset-0 z-50 flex items-center justify-center bg-black bg-opacity-50 hidden">
        <div class="bg-white rounded-xl shadow-lg p-6 w-3/4 h-2/3 flex flex-col">
            <div class="flex justify-between items-center border-b pb-3 mb-4">
                <h2 class="text-xl font-bold text-gray-800">Log Details</h2>
                <div class="flex items-center gap-3">
                    <button id="convertLogJsonBtn" class="text-sm bg-indigo-600 hover:bg-indigo-700 text-white font-semibold py-2 px-3 rounded">Convert to JSON</button>
                    <button id="closeLogDetailModal" class="text-gray-500 hover:text-gray-800 text-2xl font-bold">&times;</button>
                </div>
            </div>
            <div class="flex-grow overflow-auto bg-gray-50 p-4 rounded-md border">
                <pre id="logDetailContent" class="text-sm font-mono whitespace-pre-wrap text-gray-800 select-text"></pre>
            </div>
        </div>
    </div>

    <!-- SCRIPTS -->
    <script>
        const socket = io();
        let currentTab = 'LoadAdsExt';
        let selectedDevice = 'all';

        // --- Common Elements ---
        const deviceListEl = document.getElementById('deviceList');
        const pauseBtn = document.getElementById('pauseBtn');
        const jsonModal = document.getElementById('jsonModal');
        const closeJsonModal = document.getElementById('closeJsonModal');
        const jsonContent = document.getElementById('jsonContent');
        const logDetailModal = document.getElementById('logDetailModal');
        const closeLogDetailModal = document.getElementById('closeLogDetailModal');
        const logDetailContent = document.getElementById('logDetailContent');
        const convertLogJsonBtn = document.getElementById('convertLogJsonBtn');
        const deviceFilter = document.getElementById('deviceFilter');
        const clearAllBtn = document.getElementById('clearAllBtn');
        const restartAppBtn = document.getElementById('restartAppBtn');
        const platformBtn = document.getElementById('platformBtn');
        const platformModal = document.getElementById('platformModal');
        let activePlatform = localStorage.getItem('eventInspectorPlatform') || '';

        function platformLabel(platform) {
            return platform === 'ios' ? 'iOS' : 'Android';
        }

        function showPlatformModal() {
            platformModal?.classList.remove('hidden');
            platformModal?.classList.add('flex');
        }

        function hidePlatformModal() {
            platformModal?.classList.add('hidden');
            platformModal?.classList.remove('flex');
        }

        function syncPlatformUi() {
            const platform = activePlatform === 'ios' ? 'ios' : 'android';
            document.querySelectorAll('.sdk-check-panel').forEach(panel => {
                panel.classList.toggle('hidden', panel.getAttribute('data-sdk-platform') !== platform);
            });
            const sdkBadge = document.getElementById('sdkCheckPlatformBadge');
            if (sdkBadge) sdkBadge.textContent = `${platformLabel(platform)} Table`;
        }

        function resetRuntimeUiForPlatformSwitch() {
            if (typeof sdkCheckRunning !== 'undefined') {
                sdkCheckRunning = false;
                const sdkBtn = document.getElementById('startSdkCheckBtn');
                if (sdkBtn) sdkBtn.textContent = 'Start Checking';
            }
            if (typeof setPackagePauseState === 'function') setPackagePauseState(false);
            if (typeof closePackageStreamModal === 'function') closePackageStreamModal();
            if (typeof setPackageRunningState === 'function') setPackageRunningState(false);
            if (typeof resetPackageLogUiState === 'function') resetPackageLogUiState();
            document.getElementById('sdkCheckTableBody')?.replaceChildren();
            const iosBody = document.getElementById('sdkCheckIosTableBody');
            if (iosBody) iosBody.innerHTML = '<tr><td class="py-2 px-4 text-sm text-gray-500 italic">iOS table ready. Device connection will be added in the next step.</td></tr>';
            syncPlatformUi();
        }

        function setActivePlatform(platform, persist = true) {
            activePlatform = platform === 'ios' ? 'ios' : 'android';
            if (persist) localStorage.setItem('eventInspectorPlatform', activePlatform);
            if (platformBtn) platformBtn.textContent = `Platform: ${platformLabel(activePlatform)}`;
            syncPlatformUi();
            socket.emit('set_platform', { platform: activePlatform, reset: !!persist });
        }

        document.querySelectorAll('[data-platform-choice]').forEach(btn => {
            btn.addEventListener('click', () => {
                setActivePlatform(btn.getAttribute('data-platform-choice') || 'android');
                hidePlatformModal();
            });
        });

        platformBtn?.addEventListener('click', showPlatformModal);
        if (activePlatform) setActivePlatform(activePlatform, false);
        else showPlatformModal();

        // --- Tab Logic ---
        function switchTab(tabName) {
            currentTab = tabName;
            // Hide all contents with Safety Check
            ['tabContentLoadAds', 'tabContentLoadAdsExt', 'tabContentValidator', 'tabContentSpecific', 'tabContentAdRevenue', 'tabContentCallbackAd', 'tabContentSdkCheck', 'tabContentPackage'].forEach(id => {
                const el = document.getElementById(id);
                if (el) el.classList.add('hidden');
                else console.warn('Missing element ID:', id);
            });
            // Deactivate all buttons
            document.querySelectorAll('.tab-btn').forEach(btn => btn.classList.remove('active'));
            
            // Activate selected with Safety Check
            const contentEl = document.getElementById('tabContent' + tabName);
            if (contentEl) contentEl.classList.remove('hidden');
            
            const btnEl = document.getElementById('tabBtn' + tabName);
            if (btnEl) btnEl.classList.add('active');
            
            socket.emit('change_tab', { tab_name: tabName });
        }

        // --- Global Control Logic ---
        pauseBtn.addEventListener('click', () => {
            // Optimistic toggle for resizer usability
            const willPause = pauseBtn.textContent === 'Pause';
            setResizerEnabled(willPause);
            isPausedClient = willPause;
            socket.emit('toggle_pause');
        });
        clearAllBtn.addEventListener('click', () => {
            if (confirm('Are you sure you want to clear ALL logs?')) {
                socket.emit('clear_all_logs');
            }
        });

        restartAppBtn?.addEventListener('click', () => {
            if (restartAppBtn.disabled) return;
            const originalText = restartAppBtn.textContent;
            restartAppBtn.disabled = true;
            restartAppBtn.textContent = 'Checking...';
            fetch('/check_update', { method: 'POST' })
                .then(r => r.json())
                .then(data => {
                    if (data.status === 'updated') {
                        restartAppBtn.textContent = 'Restarting...';
                        return fetch('/restart_app', { method: 'POST' });
                    }
                    if (data.status === 'up_to_date') {
                        alert('Already up to date.');
                    } else if (data.error) {
                        alert('Update check failed: ' + data.error);
                    } else {
                        alert('Update check failed.');
                    }
                })
                .catch(err => {
                    alert('Update check failed: ' + err);
                })
                .finally(() => {
                    restartAppBtn.disabled = false;
                    restartAppBtn.textContent = originalText;
                });
        });

        const manualRestartBtn = document.getElementById('manualRestartBtn');
        manualRestartBtn?.addEventListener('click', () => {
            if (manualRestartBtn.disabled) return;
            const originalText = manualRestartBtn.textContent;
            manualRestartBtn.disabled = true;
            manualRestartBtn.textContent = 'Restarting...';
            fetch('/restart_app', { method: 'POST' })
                .catch(err => {
                    alert('Restart failed: ' + err);
                })
                .finally(() => {
                    manualRestartBtn.disabled = false;
                    manualRestartBtn.textContent = originalText;
                });
        });
        
        // --- Recording Logic (Updated for separate tabs) ---
        function toggleRecord(tabName) {
            const btn = document.getElementById('btnRecord_' + tabName);
            const input = document.getElementById('sheetName_' + tabName);
            
            if (!btn.classList.contains('bg-red-600')) {
                const name = input.value.trim();
                socket.emit('toggle_record', { sheet_name: name, tab_name: tabName });
            } else { 
                socket.emit('toggle_record', { tab_name: tabName }); 
            }
        }

        socket.on('record_status', (s) => {
            // Update UI based on which tab triggered the status
            const tabName = s.tab_name;
            const btn = document.getElementById('btnRecord_' + tabName);
            if (btn) {
                if (s.is_recording) {
                    btn.textContent = 'Stop Recording: ' + (s.current_sheet || 'Running');
                    btn.className = 'bg-red-600 text-white font-bold py-2 px-4 rounded animate-record shadow-lg text-sm';
                } else {
                    btn.textContent = 'Start Record';
                    btn.className = 'bg-indigo-600 hover:bg-indigo-700 text-white font-bold py-2 px-4 rounded shadow text-sm';
                }
            }
        });

        let isPausedClient = false;
        function setResizerEnabled(enabled) {
            document.querySelectorAll('.resizer').forEach(r => {
                r.classList.toggle('disabled', !enabled);
                r.style.pointerEvents = enabled ? 'auto' : 'none';
            });
        }

        socket.on('pause_status', (data) => {
            isPausedClient = !!data.is_paused;
            if (data.is_paused) {
                pauseBtn.textContent = 'Resume';
                pauseBtn.classList.replace('bg-yellow-500', 'bg-blue-500');
            } else {
                pauseBtn.textContent = 'Pause';
                pauseBtn.classList.replace('bg-blue-500', 'bg-yellow-500');
            }
            setResizerEnabled(isPausedClient);
        });

        // --- Rendering Helpers ---
        function escapeHTML(str) {
            if (!str) return '';
            const p = document.createElement("p");
            p.textContent = str;
            return p.innerHTML.replace(/  /g, ' &nbsp;');
        }
        
        // FIX: Better attribute escaping for JSON buttons
        function escapeAttribute(str) {
            if (!str) return '{}';
            // Replace single quotes with HTML entity to prevent breaking the attribute
            return str.replace(/'/g, '&#39;');
        }

        function renderSimpleTable(id, data) {
            const tbody = document.getElementById(id);
            if (!tbody) return;
            const filtered = (selectedDevice === 'all') ? data : data.filter(e => e.device_id === selectedDevice);
            
            if (filtered.length === 0) {
                 tbody.innerHTML = '<tr><td colspan="4" class="text-center py-4 text-gray-400">Waiting for recording...</td></tr>';
                 return;
            }

            const useAdNetwork = id === 'loadAdsExtTableBody';
            tbody.innerHTML = filtered.map(e => `
                <tr class="hover:bg-gray-50 border-b text-sm">
                    <td class="py-2 px-2 text-purple-700 text-sm font-medium whitespace-nowrap">${e.device_name}</td>
                    <td class="py-2 px-2 text-blue-600 text-sm font-medium whitespace-nowrap">${useAdNetwork ? (e.ad_network || e.ad_source || '') : (e.ad_source || e.ad_network || '')}</td>
                    <td class="py-2 px-2 text-green-600 text-sm font-medium whitespace-nowrap">${e.ad_format}</td>
                    <td class="py-2 px-3 log-cell text-xs font-normal text-gray-600">${escapeHTML(e.raw_log || '')}</td>
                </tr>
            `).join('');
        }

        let validator_results_cache = [];
        let defaultEventNames = {{ default_event_names | tojson }};
        let defaultEventStatusEls = {};
        let currentProfileName = {{ current_profile_name | tojson }};

        function updateProfileStatus(payload) {
            const gameEl = document.getElementById('profileGameText');
            if (gameEl) {
                gameEl.textContent = payload.game_name || 'Unknown';
            }
        }

        function renderProfileOptions(payload) {
            const profiles = payload.profiles || [];
            [document.getElementById('profileSelect'), document.getElementById('adRevenueProfileSelect')].forEach(select => {
                if (!select) return;
                select.innerHTML = profiles.length
                    ? profiles.map(name => `<option value="${escapeAttribute(name)}"${name === payload.current_profile ? ' selected' : ''}>${escapeHTML(name)}</option>`).join('')
                    : '<option value="">No profiles</option>';
                select.disabled = profiles.length === 0;
            });
            currentProfileName = payload.current_profile || '';
            defaultEventNames = payload.default_event_names || [];
            renderDefaultEventStatusList();
            updateDefaultEventStatus(validator_results_cache);
            updateProfileStatus(payload);
            const adRevenueGameEl = document.getElementById('adRevenueProfileGameText');
            if (adRevenueGameEl) {
                adRevenueGameEl.textContent = payload.game_name || 'Unknown';
            }
        }

        async function refreshProfiles() {
            const res = await fetch('/api/profiles');
            const payload = await res.json();
            renderProfileOptions(payload);
        }

        function renderDefaultEventStatusList() {
            const container = document.getElementById('defaultEventStatusList');
            if (!container) return;
            if (!defaultEventNames || defaultEventNames.length === 0) {
                container.innerHTML = '<div class="text-gray-400">No default events</div>';
                return;
            }
            container.innerHTML = defaultEventNames.map(name => `
                <div class="default-event-item inline-flex items-center gap-1.5 px-2 py-1 rounded-md border border-gray-200 bg-white cursor-pointer hover:bg-gray-50"
                     data-event-name="${escapeAttribute(name)}" title="Click to filter by this event">
                    <span class="event-status-icon text-gray-400 font-bold w-4 text-center" data-event="${name}" title="checking">...</span>
                    <span class="truncate max-w-[180px] text-xs font-medium text-gray-700" title="${escapeHTML(name)}">${escapeHTML(name)}</span>
                </div>
            `).join('');
            defaultEventStatusEls = {};
            container.querySelectorAll('.event-status-icon').forEach(el => {
                defaultEventStatusEls[el.dataset.event] = el;
            });
        }

        function updateDefaultEventStatus(results) {
            if (!defaultEventNames || defaultEventNames.length === 0) return;
            const statusMap = {};
            defaultEventNames.forEach(n => { statusMap[n] = 'pending'; });
            results.forEach(r => {
                if (!statusMap.hasOwnProperty(r.event_name)) return;
                if (r.status === 'FAILED') statusMap[r.event_name] = 'failed';
                else if (r.status === 'PASSED') statusMap[r.event_name] = 'passed';
            });
            Object.entries(statusMap).forEach(([name, status]) => {
                const el = defaultEventStatusEls[name];
                if (!el) return;
                if (status === 'passed') {
                    el.textContent = '✓';
                    el.className = 'event-status-icon text-green-600 font-bold';
                    el.title = 'PASSED';
                } else if (status === 'failed') {
                    el.textContent = '✕';
                    el.className = 'event-status-icon text-red-600 font-bold';
                    el.title = 'FAILED';
                } else {
                    el.textContent = '...';
                    el.className = 'event-status-icon text-gray-400 font-bold';
                    el.title = 'checking';
                }
            });
        }

        function renderValidatorTable(results) {
             const tbody = document.getElementById('validatorTableBody');
             if (!tbody) return;
             const filterText = (document.getElementById('validatorEventFilterInput')?.value || '').toLowerCase().trim();
             const rawFilterText = (document.getElementById('validatorRawFilterInput')?.value || '').toLowerCase().trim();
             const sourceFilter = document.querySelector('input[name="validatorSourceFilter"]:checked')?.value || 'all';
             const filtered = results.filter(r => {
                 if (selectedDevice !== 'all' && r.device_id !== selectedDevice) return false;
                 if (filterText && !(r.event_name || '').toLowerCase().includes(filterText)) return false;
                 if (rawFilterText && !(r.raw_log || '').toLowerCase().includes(rawFilterText)) return false;
                 if (sourceFilter !== 'all' && (r.source || 'firebase') !== sourceFilter) return false;
                 return true;
             });
             
             if (filtered.length === 0) {
                 tbody.innerHTML = '<tr><td colspan="6" class="text-center py-4">Waiting...</td></tr>';
             } else {
                 tbody.innerHTML = filtered.map(res => `
                    <tr class="hover:bg-gray-50 border-b text-sm">
                        <td class="py-2 px-3 text-purple-700 text-sm">${res.device_name}</td>
                        <td class="py-2 px-3 text-sm font-semibold ${res.status === 'PASSED' ? 'text-green-600' : 'text-red-600'}">${res.status}</td>
                        <td class="py-2 px-3"><span class="event-name-link cursor-pointer text-sm font-medium text-indigo-700 hover:underline" data-event-name="${escapeAttribute(res.event_name)}">${res.event_name}</span></td>
                        <td class="py-2 px-3 details-cell text-sm">${res.details}</td>
                        <td class="py-2 px-3 log-cell text-xs font-normal text-gray-600">${escapeHTML(res.raw_log || '')}</td>
                        <td class="py-2 px-3"><button class="view-json-btn text-xs bg-indigo-100 hover:bg-indigo-200 text-indigo-700 font-medium py-1 px-2 rounded" data-json='${escapeAttribute(res.json_data)}'>View JSON</button></td>
                    </tr>
                 `).join('');
             }
        }

        renderDefaultEventStatusList();
        refreshProfiles().catch(() => {});

        // Click default event to fill filter input
        document.getElementById('defaultEventStatusList')?.addEventListener('click', (e) => {
            const item = e.target.closest('.default-event-item');
            if (!item) return;
            const name = item.getAttribute('data-event-name') || '';
            const input = document.getElementById('validatorEventFilterInput');
            if (input) {
                input.value = name;
                input.focus();
                renderValidatorTable(validator_results_cache);
            }
        });


        // Click event name in results table to fill filter input
        document.getElementById('validatorTableBody')?.addEventListener('click', (e) => {
            const el = e.target.closest('.event-name-link');
            if (!el) return;
            const name = el.getAttribute('data-event-name') || '';
            const input = document.getElementById('validatorEventFilterInput');
            if (input) {
                input.value = name;
                input.focus();
                renderValidatorTable(validator_results_cache);
            }
        });

        document.getElementById('specificEventTableBody')?.addEventListener('click', (e) => {
            const el = e.target.closest('.event-name-link');
            if (!el) return;
            const name = el.getAttribute('data-event-name') || '';
            const input = document.getElementById('specificEventInput');
            if (input) {
                input.value = name;
                input.focus();
                socket.emit('update_specific_filter', {
                    eventNames: [name],
                    params: []
                });
                renderSpecificEventTable();
            }
        });

        async function switchSharedProfile(profileName) {
            if (!profileName || profileName === currentProfileName) return;
            const res = await fetch('/api/profiles/select', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({ profile_name: profileName })
            });
            const payload = await res.json();
            if (!payload.ok) {
                alert(payload.error || 'Failed to switch profile');
                await refreshProfiles();
                return;
            }
            validator_results_cache = [];
            renderValidatorTable(validator_results_cache);
            renderProfileOptions(payload);
        }

        document.getElementById('profileSelect')?.addEventListener('change', async (e) => {
            await switchSharedProfile(e.target.value);
        });

        document.getElementById('adRevenueProfileSelect')?.addEventListener('change', async (e) => {
            await switchSharedProfile(e.target.value);
        });

        document.getElementById('reloadProfileBtn')?.addEventListener('click', async () => {
            const res = await fetch('/api/profiles/reload', { method: 'POST' });
            const payload = await res.json();
            if (!payload.ok) {
                alert(payload.error || 'Failed to reload profile');
                return;
            }
            renderProfileOptions(payload);
        });

        document.getElementById('importProfileBtn')?.addEventListener('click', () => {
            document.getElementById('profileFileInput')?.click();
        });

        document.getElementById('profileFileInput')?.addEventListener('change', async (e) => {
            const file = e.target.files && e.target.files[0];
            if (!file) return;
            const formData = new FormData();
            formData.append('profile_file', file);
            const res = await fetch('/api/profiles/import', {
                method: 'POST',
                body: formData
            });
            const payload = await res.json();
            e.target.value = '';
            if (!payload.ok) {
                alert(payload.error || 'Failed to import profile');
                return;
            }
            validator_results_cache = [];
            renderValidatorTable(validator_results_cache);
            renderProfileOptions(payload);
        });

        // --- Socket Listeners (Renderers) ---
        socket.on('update_load_ads', (d) => renderSimpleTable('loadAdsTableBody', d));
        socket.on('update_load_ads_ext', (d) => renderSimpleTable('loadAdsExtTableBody', d));
        socket.on('update_validator_table', (d) => {
            validator_results_cache = d || [];
            renderValidatorTable(validator_results_cache);
            updateDefaultEventStatus(validator_results_cache);
        });

        let lastSpecificEventData = [];

        socket.on('update_specific_event_table', (d) => {
             lastSpecificEventData = d || [];
             renderSpecificEventTable();
        });

        function renderSpecificEventTable() {
             const tbody = document.getElementById('specificEventTableBody');
             if (!tbody) return;
             const sourceFilter = document.querySelector('input[name="specificSourceFilter"]:checked')?.value || 'all';
             const textFilter = (document.getElementById('specificParamInput')?.value || '').trim().toLowerCase();
             const filtered = lastSpecificEventData.filter(r => {
                 if (selectedDevice !== 'all' && r.device_id !== selectedDevice) return false;
                 if (sourceFilter !== 'all' && (r.source || 'firebase') !== sourceFilter) return false;
                 if (textFilter && !(r.raw_log || '').toLowerCase().includes(textFilter)) return false;
                 return true;
             });
             if(filtered.length === 0) { tbody.innerHTML = '<tr><td colspan="6" class="text-center py-4">Waiting...</td></tr>'; }
             else {
                 tbody.innerHTML = filtered.map(res => `<tr class="hover:bg-gray-50 border-b text-sm"><td class="py-2 px-3 text-purple-700 text-sm">${res.device_name}</td><td class="py-2 px-3 text-sm font-semibold ${res.status === 'PASSED'?'text-green-600':'text-red-600'}">${res.status}</td><td class="py-2 px-3"><span class="event-name-link cursor-pointer text-sm font-medium text-indigo-700 hover:underline" data-event-name="${escapeAttribute(res.event_name)}">${res.event_name}</span></td><td class="py-2 px-3 details-cell text-sm">${res.details}</td><td class="py-2 px-3 log-cell text-xs font-normal text-gray-600">${escapeHTML(res.raw_log || '')}</td><td class="py-2 px-3"><button class="view-json-btn text-xs bg-indigo-100 hover:bg-indigo-200 text-indigo-700 font-medium py-1 px-2 rounded" data-json='${escapeAttribute(res.json_data)}'>View JSON</button></td></tr>`).join('');
             }
        }
        let lastAdRevenueData = [];

         socket.on('update_adrevenue_table', (d) => {
             lastAdRevenueData = d || [];
             renderAdRevenueTable();
        });

        function renderAdRevenueTable() {
            const tbody = document.getElementById('adRevenueTableBody');
            if (!tbody) return;
            const filterText = document.getElementById('adRevenueFilterInput').value.toLowerCase();
            const sourceFilter = document.querySelector('input[name="adRevenueSourceFilter"]:checked')?.value || 'all';
            const filtered = lastAdRevenueData.filter(r => {
                if (selectedDevice !== 'all' && r.device_id !== selectedDevice) return false;
                if (sourceFilter !== 'all' && (r.source || '').toLowerCase() !== sourceFilter) return false;
                if (filterText && !(r.raw_log || '').toLowerCase().includes(filterText)) return false;
                return true;
            });
            if(filtered.length === 0) { tbody.innerHTML = '<tr><td colspan="6" class="text-center py-4">Waiting...</td></tr>'; }
            else {
                tbody.innerHTML = filtered.map(res => `<tr class="hover:bg-gray-50 border-b text-sm"><td class="py-2 px-3 text-purple-700 text-sm align-top whitespace-nowrap">${res.device_name}</td><td class="py-2 px-3 text-sm font-semibold ${res.status === 'PASSED'?'text-green-600':(res.status === 'FAILED'?'text-red-600':'text-orange-500')} align-top whitespace-nowrap">${res.status}</td><td class="py-2 px-3 align-top"><span class="event-name-link cursor-pointer text-sm font-medium text-indigo-700 hover:underline" data-event-name="${escapeAttribute(res.event_name)}">${res.event_name}</span></td><td class="py-2 px-3 details-cell text-sm align-top"><div class="adrevenue-panel adrevenue-details-panel">${res.details}</div></td><td class="py-2 px-3 log-cell text-xs font-normal text-gray-600 align-top"><div class="adrevenue-panel adrevenue-raw-panel">${escapeHTML(res.raw_log || '')}</div></td><td class="py-2 px-3 align-top whitespace-nowrap"><button class="view-json-btn text-xs bg-indigo-100 hover:bg-indigo-200 text-indigo-700 font-medium py-1 px-2 rounded" data-json='${escapeAttribute(res.json_data)}'>View JSON</button></td></tr>`).join('');
            }
        }

        // ============================================
        // === FIXED: Callback Table with Client Filter ===
        // ============================================
        
        let lastCallbackData = [];

        socket.on('update_callback_ad_table', (d) => {
            lastCallbackData = d;
            renderCallbackTable();
        });

        function renderCallbackTable() {
            const d = lastCallbackData;
            const tbody = document.getElementById('callbackAdTableBody');
            if (!tbody) return;

            // Get filter values
            const typeFilter = document.querySelector('input[name="callbackType"]:checked').value;
            const textFilter = document.getElementById('callbackAdFilterInput').value.toLowerCase();

            const filtered = d.filter(r => {
                // Device filter
                if (selectedDevice !== 'all' && r.device_id !== selectedDevice) return false;
                
                // Type filter
                if (typeFilter === 'callback' && (r.type === 'Ad Event' || r.type === 'Callback Gadsme' || r.type === 'Callback Adverty5')) return false;
                if (typeFilter === 'gadsme_callback' && r.type !== 'Callback Gadsme') return false;
                if (typeFilter === 'adverty5_callback' && r.type !== 'Callback Adverty5') return false;
                if (typeFilter === 'ad_event' && r.type !== 'Ad Event') return false;
                
                // Text filter
                if (textFilter && !r.raw_log.toLowerCase().includes(textFilter)) return false;

                return true;
            });

             if(filtered.length === 0) { tbody.innerHTML = '<tr><td colspan="6" class="text-center py-4">Waiting...</td></tr>'; }
             else {
                 tbody.innerHTML = filtered.map(res => {
                     const nameLower = (res.event_name || '').toLowerCase();
                     const isFailed = nameLower.includes('failed');
                     const isImpression = nameLower.includes('onimpression') || nameLower.includes('_onimpression');
                     const eventClass = isFailed ? 'text-red-600' : (isImpression ? 'text-blue-600' : '');
                     return `<tr class="hover:bg-gray-50 border-b text-sm"><td class="py-2 px-3 text-purple-700 text-sm">${res.device_name}</td><td class="py-2 px-3 text-sm font-semibold ${res.type==='Ad Event'?'text-orange-600':'text-cyan-600'}">${res.type}</td><td class="py-2 px-3 text-sm font-medium ${eventClass}">${res.event_name}</td><td class="py-2 px-3 details-cell text-sm">${res.details}</td><td class="py-2 px-3 log-cell text-xs font-normal text-gray-600">${escapeHTML(res.raw_log || '')}</td><td class="py-2 px-3"><button class="view-json-btn text-xs bg-indigo-100 hover:bg-indigo-200 text-indigo-700 font-medium py-1 px-2 rounded" data-json='${escapeAttribute(res.json_data)}'>View JSON</button></td></tr>`;
                 }).join('');
             }
        }

        // Add listeners
        document.querySelectorAll('input[name="callbackType"]').forEach(r => r.addEventListener('change', renderCallbackTable));
        document.getElementById('callbackAdFilterInput').addEventListener('input', renderCallbackTable);

        socket.on('update_sdk_check_table', (data) => {
            const tbody = document.getElementById(activePlatform === 'ios' ? 'sdkCheckIosTableBody' : 'sdkCheckTableBody');
            if (!tbody) return;
            tbody.innerHTML = data.map(res => {
                 let rowClass = (selectedDevice !== 'all' && res.status !== 'HEADER') ? 'pl-8' : '';
                 let statusText = '';
                 if (res.status === 'PASSED') statusText = '<span class="font-semibold text-green-600"> - PASSED</span>';
                 else if (res.status === 'FAILED') statusText = '<span class="font-semibold text-red-600"> - FAILED</span>';
                 else if (res.status === 'NOT_FOUND') statusText = '<span class="font-semibold text-red-600"> - Not Found</span>';
                 else if (res.status === 'FOUND') statusText = '<span class="font-semibold text-amber-600"> - FOUND</span>';
                 else if (res.status === 'HEADER') rowClass += ' font-semibold text-sm text-indigo-600 bg-gray-50';
                 else if (res.status === 'LABEL') rowClass += ' font-bold text-base text-indigo-700 pt-3';
                 else if (res.status === 'SECTION') rowClass += ' font-bold text-sm text-slate-500 pt-4 border-t border-gray-200';
                 else if (res.status === 'WAITING') rowClass += ' text-sm text-gray-500 italic';
                 
                 // Filter
                 if (selectedDevice !== 'all' && res.device_id !== selectedDevice && res.status !== 'LABEL' && res.status !== 'SECTION' && res.status !== 'WAITING') return '';
                 
                 return `<tr><td class="py-1 px-4 ${rowClass}"><pre style="font-family: monospace; margin: 0; white-space: pre-wrap;">${escapeHTML(res.display_text)}${statusText}</pre></td></tr>`;
            }).join('');
        });

        socket.on('runtime_reset', () => {
            resetRuntimeUiForPlatformSwitch();
        });

        let lastPackageLogs = [];
        let selectedPackageRowKeys = new Set();
        let lastPackageFilterSignature = '';
        let lastPackageRenderedCount = 0;
        let lastPackageFirstRowKey = '';
        let packageHistorySessions = [];
        let packageUiPaused = false;
        let pausedPackageSnapshot = [];

        function getPackageRowKey(l) {
            const msgText = (l.message || l.log || '');
            return `${l.time_display || l.time || ''}||${l.tag || ''}||${msgText}`;
        }

        function getPackageFilterState() {
            return {
                selectedDevice,
                filterText: document.getElementById('packageFilterInput').value.toLowerCase(),
                filterText2: document.getElementById('packageFilterInput2').value.toLowerCase(),
                tagFilter: document.getElementById('packageTagFilterInput').value.toLowerCase(),
                quickTag: document.querySelector('input[name="tagQuickFilter"]:checked')?.value || '',
                errorsOnly: document.getElementById('showErrorsOnly').checked,
                warningsOnly: document.getElementById('showWarningsOnly').checked,
            };
        }

        function filterPackageLogs(logs, state) {
            return logs.filter(l => {
                if (state.selectedDevice !== 'all' && l.device_id !== state.selectedDevice) return false;
                const isWarningLevel = l.level === 'W';
                if (state.errorsOnly && state.warningsOnly) {
                    if (!(l.is_error || isWarningLevel)) return false;
                } else if (state.errorsOnly) {
                    if (!l.is_error) return false;
                } else if (state.warningsOnly) {
                    if (!isWarningLevel) return false;
                }
                const messageHaystack = `${l.message || ''}`.toLowerCase();
                const tagHaystack = `${l.tag || ''}`.toLowerCase();
                if (state.quickTag && !tagHaystack.includes(state.quickTag)) return false;
                if (state.tagFilter && !tagHaystack.includes(state.tagFilter)) return false;
                if (state.filterText && !messageHaystack.includes(state.filterText)) return false;
                if (state.filterText2 && !messageHaystack.includes(state.filterText2)) return false;
                return true;
            });
        }

        function packageRowHtml(l, idx) {
            const msgText = (l.message || l.log || '');
            const isErrorLevel = (l.level === 'E' || l.level === 'F');
            const isWarningLevel = l.level === 'W';
            const rowClass = isErrorLevel ? 'text-red-500' : (isWarningLevel ? 'text-amber-500' : '');
            const msgClass = isErrorLevel ? 'text-red-500' : (isWarningLevel ? 'text-amber-500' : '');
            const rowKey = getPackageRowKey(l);
            const selectedClass = selectedPackageRowKeys.has(rowKey) ? 'selected' : '';
            return `<tr class="package-log-row hover:bg-gray-50 ${rowClass} ${selectedClass}" data-row-key="${encodeURIComponent(rowKey)}" data-row-index="${idx}"><td class="py-1.5 px-2 font-mono text-[11px] leading-4 time-cell col-time">${escapeHTML(l.time_display || l.time || '')}</td><td class="py-1.5 pr-1 pl-2 font-mono text-[11px] leading-4 tag-cell col-tag" title="${escapeHTML(l.tag || '')}">${escapeHTML(l.tag || '')}</td><td class="py-1.5 pl-1 pr-3 font-mono text-[11px] leading-4 log-cell message-cell col-message ${msgClass}">${escapeHTML(msgText)}</td></tr>`;
        }

        function getPackageSourceLogs(state) {
            return packageUiPaused ? pausedPackageSnapshot : lastPackageLogs;
        }

        function renderPackageLogTable(forceFull = false) {
            const tbody = document.getElementById('packageLogTableBody');
            if (!tbody) return;
            const state = getPackageFilterState();
            const sourceLogs = getPackageSourceLogs(state);
            const signature = JSON.stringify(state);
            const canAppendOnly =
                !packageUiPaused &&
                !forceFull &&
                signature === lastPackageFilterSignature &&
                sourceLogs.length >= lastPackageRenderedCount &&
                (lastPackageRenderedCount === 0 ||
                    (sourceLogs[0] && getPackageRowKey(sourceLogs[0]) === lastPackageFirstRowKey));

            if (canAppendOnly && sourceLogs.length > lastPackageRenderedCount) {
                const appendedLogs = filterPackageLogs(sourceLogs.slice(lastPackageRenderedCount), state);
                if (appendedLogs.length > 0) {
                    const startIdx = tbody.querySelectorAll('tr.package-log-row').length;
                    tbody.insertAdjacentHTML('beforeend', appendedLogs.map((l, idx) => packageRowHtml(l, startIdx + idx)).join(''));
                }
            } else {
                const filtered = filterPackageLogs(sourceLogs, state);
                tbody.innerHTML = filtered.map((l, idx) => packageRowHtml(l, idx)).join('');
            }

            lastPackageFilterSignature = signature;
            lastPackageRenderedCount = sourceLogs.length;
            lastPackageFirstRowKey = sourceLogs[0] ? getPackageRowKey(sourceLogs[0]) : '';
            if(document.getElementById('autoScroll').checked) document.getElementById('packageLogContainer').scrollTop = document.getElementById('packageLogContainer').scrollHeight;
        }

        socket.on('package_log_cache', (logs) => {
            lastPackageLogs = logs || [];
            if (!packageUiPaused) renderPackageLogTable();
        });

        function renderPackageHistorySessions(payload) {
            const select = document.getElementById('packageHistorySessionSelect');
            const meta = document.getElementById('packageHistoryMeta');
            if (!select || !meta) return;
            const prev = select.value;
            packageHistorySessions = payload.sessions || [];
            select.innerHTML = '';
            if (packageHistorySessions.length === 0) {
                packageHistoryLoadedRows = [];
                packageHistoryOffset = 0;
                packageHistoryHasMore = false;
                updatePackageHistoryLoadMoreButton();
                const opt = document.createElement('option');
                opt.value = '';
                opt.textContent = 'No saved sessions';
                select.appendChild(opt);
                meta.textContent = 'No saved package-log sessions yet.';
                document.getElementById('packageHistoryTableBody').innerHTML = '';
                return;
            }
            packageHistorySessions.forEach((s, idx) => {
                const opt = document.createElement('option');
                opt.value = String(s.id);
                opt.textContent = `#${s.id} - ${s.package_id} - ${s.started_label}`;
                if ((!prev && idx === 0) || prev === String(s.id)) opt.selected = true;
                select.appendChild(opt);
            });
            const current = packageHistorySessions.find(s => String(s.id) === select.value) || packageHistorySessions[0];
            if (current) {
                meta.textContent = `Selected: ${current.package_id} | Started: ${current.started_label} | Status: ${current.status} | Rows: ${current.row_count}`;
            }
        }

        function getPackageHistoryRowKey(row) {
            return [row.time_display || '', row.device_name || row.device_id || '', row.tag || '', row.raw_log || row.message || ''].join('||');
        }

        function renderPackageHistoryRows(rows, append = false) {
            const tbody = document.getElementById('packageHistoryTableBody');
            if (!tbody) return;
            if ((!rows || rows.length === 0) && !append) {
                tbody.innerHTML = '<tr><td colspan="4" class="py-3 px-2 text-xs text-gray-500">No saved logs found for this filter.</td></tr>';
                return;
            }
            if (!rows || rows.length === 0) return;
            if (!append) tbody.innerHTML = '';
            const startIdx = append ? tbody.querySelectorAll('tr.package-history-row').length : 0;
            const html = rows.map((row, idx) => {
                const rowKey = getPackageHistoryRowKey(row);
                const selectedClass = selectedPackageHistoryRowKeys.has(rowKey) ? 'selected' : '';
                return `
                <tr class="package-history-row hover:bg-gray-50 ${selectedClass}" data-row-key="${encodeURIComponent(rowKey)}" data-row-index="${startIdx + idx}">
                    <td class="py-2 px-2 font-mono text-[11px] text-gray-700 align-top whitespace-nowrap">${escapeHTML(row.time_display || '')}</td>
                    <td class="py-2 px-2 text-[11px] text-gray-700 align-top whitespace-nowrap">${escapeHTML(row.device_name || row.device_id || '')}</td>
                    <td class="py-2 px-2 font-mono text-[11px] text-gray-700 align-top whitespace-nowrap max-w-[88px] w-[88px] overflow-hidden text-ellipsis" title="${escapeHTML(row.tag || '')}">${escapeHTML(row.tag || '')}</td>
                    <td class="py-2 px-2 font-mono text-[11px] text-gray-700 align-top whitespace-pre-wrap break-all">${escapeHTML(row.raw_log || row.message || '')}</td>
                </tr>
            `}).join('');
            if (append) tbody.insertAdjacentHTML('beforeend', html);
            else tbody.innerHTML = html;
        }

        function loadPackageHistorySessions() {
            fetch('/api/package-log/sessions')
                .then(r => r.json())
                .then(data => {
                    if (!data.ok) throw new Error(data.error || 'failed_to_load_sessions');
                    renderPackageHistorySessions(data);
                })
                .catch(err => {
                    document.getElementById('packageHistoryMeta').textContent = `Failed to load sessions: ${err}`;
                });
        }

        function loadSelectedPackageHistory(append = false) {
            if (packageHistoryLoading) return;
            const select = document.getElementById('packageHistorySessionSelect');
            const keyword1 = document.getElementById('packageHistoryFilterInput')?.value || '';
            const keyword2 = document.getElementById('packageHistoryFilterInput2')?.value || '';
            const keyword3 = document.getElementById('packageHistoryFilterInput3')?.value || '';
            if (!select || !select.value) {
                packageHistoryLoadedRows = [];
                packageHistoryOffset = 0;
                packageHistoryHasMore = false;
                updatePackageHistoryLoadMoreButton();
                renderPackageHistoryRows([]);
                return;
            }
            const offset = append ? packageHistoryOffset : 0;
            const query = new URLSearchParams({
                session_id: select.value,
                q1: keyword1,
                q2: keyword2,
                q3: keyword3,
                offset: String(offset),
                limit: String(PACKAGE_HISTORY_PAGE_SIZE),
            });
            packageHistoryLoading = true;
            updatePackageHistoryLoadMoreButton();
            fetch(`/api/package-log/logs?${query.toString()}`)
                .then(r => r.json())
                .then(data => {
                    if (!data.ok) throw new Error(data.error || 'failed_to_load_logs');
                    const current = packageHistorySessions.find(s => String(s.id) === String(data.session_id));
                    const meta = document.getElementById('packageHistoryMeta');
                    if (!append) {
                        packageHistoryLoadedRows = [];
                        packageHistoryOffset = 0;
                        clearPackageHistorySelection();
                    }
                    packageHistoryLoadedRows = append ? packageHistoryLoadedRows.concat(data.rows || []) : (data.rows || []);
                    packageHistoryOffset = packageHistoryLoadedRows.length;
                    packageHistoryHasMore = !!data.has_more;
                    updatePackageHistoryLoadMoreButton();
                    if (meta && current) {
                        const filters = [keyword1, keyword2, keyword3].filter(Boolean).join(' | ');
                        meta.textContent = `Selected: ${current.package_id} | Started: ${current.started_label} | Status: ${current.status} | Showing: ${packageHistoryLoadedRows.length} / ${data.total_rows}${filters ? ` | Filter: ${filters}` : ''}`;
                    }
                    renderPackageHistoryRows(data.rows || [], append);
                })
                .catch(err => {
                    document.getElementById('packageHistoryMeta').textContent = `Failed to load saved logs: ${err}`;
                    packageHistoryHasMore = false;
                    updatePackageHistoryLoadMoreButton();
                })
                .finally(() => {
                    packageHistoryLoading = false;
                    updatePackageHistoryLoadMoreButton();
                });
        }

        function exportSelectedPackageHistory(filteredOnly = true) {
            const select = document.getElementById('packageHistorySessionSelect');
            const keyword1 = document.getElementById('packageHistoryFilterInput')?.value || '';
            const keyword2 = document.getElementById('packageHistoryFilterInput2')?.value || '';
            const keyword3 = document.getElementById('packageHistoryFilterInput3')?.value || '';
            if (!select || !select.value) {
                alert('Please select a recorded session first.');
                return;
            }
            fetch('/api/package-log/export', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({
                    session_id: Number(select.value),
                    q1: filteredOnly ? keyword1 : '',
                    q2: filteredOnly ? keyword2 : '',
                    q3: filteredOnly ? keyword3 : '',
                    filtered_only: filteredOnly,
                }),
            })
                .then(r => r.json())
                .then(data => {
                    if (data.cancelled) return;
                    if (!data.ok) throw new Error(data.error || 'failed_to_export_logs');
                    alert(`Exported ${data.row_count} rows to:\n${data.path}`);
                })
                .catch(err => {
                    alert(`Failed to export recorded log: ${err}`);
                });
        }

        function clearAllRecordedPackageHistory() {
            const confirmed = confirm('Clear all recorded package logs from the database? This cannot be undone.');
            if (!confirmed) return;
            fetch('/api/package-log/clear', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({ clear_all: true }),
            })
                .then(r => r.json())
                .then(data => {
                    if (!data.ok) throw new Error(data.error || 'failed_to_clear_package_logs');
                    packageHistorySessions = [];
                    const select = document.getElementById('packageHistorySessionSelect');
                    if (select) {
                        select.innerHTML = '<option value="">No saved sessions</option>';
                    }
                    packageHistoryLoadedRows = [];
                    packageHistoryOffset = 0;
                    packageHistoryHasMore = false;
                    updatePackageHistoryLoadMoreButton();
                    document.getElementById('packageHistoryFilterInput').value = '';
                    document.getElementById('packageHistoryFilterInput2').value = '';
                    document.getElementById('packageHistoryFilterInput3').value = '';
                    document.getElementById('packageHistoryMeta').textContent = 'No saved package-log sessions yet.';
                    renderPackageHistoryRows([]);
                    alert(`Cleared ${data.deleted_sessions} session(s) and ${data.deleted_rows} log row(s).`);
                })
                .catch(err => {
                    alert(`Failed to clear recorded logs: ${err}`);
                });
        }

        function clearRecordedPackageHistoryFilters(reload = true) {
            const filter1 = document.getElementById('packageHistoryFilterInput');
            const filter2 = document.getElementById('packageHistoryFilterInput2');
            const filter3 = document.getElementById('packageHistoryFilterInput3');
            if (filter1) filter1.value = '';
            if (filter2) filter2.value = '';
            if (filter3) filter3.value = '';
            if (reload) loadSelectedPackageHistory(false);
        }

        // --- JSON Modal Handler ---
        document.body.addEventListener('click', (e) => {
            if (e.target && e.target.classList.contains('view-json-btn')) {
                const jsonData = e.target.getAttribute('data-json'); // Use getAttribute for safety
                try {
                    const parsed = JSON.parse(jsonData);
                    jsonContent.textContent = JSON.stringify(parsed, null, 2);
                    jsonModal.classList.remove('hidden');
                } catch(err) { 
                    console.error(err);
                    alert('Invalid JSON data. Check console for details.'); 
                }
            }
        });
        closeJsonModal.addEventListener('click', () => jsonModal.classList.add('hidden'));
        closeLogDetailModal.addEventListener('click', () => logDetailModal.classList.add('hidden'));

        // --- Log Detail Modal (Package Log) ---
        let selectedPackageHistoryRowKeys = new Set();
        let isSelectingHistoryRows = false;
        let historyDragStartIndex = null;
        const PACKAGE_HISTORY_PAGE_SIZE = 2000;
        let packageHistoryLoadedRows = [];
        let packageHistoryOffset = 0;
        let packageHistoryHasMore = false;
        let packageHistoryLoading = false;

        function updatePackageHistoryLoadMoreButton() {
            const btn = document.getElementById('loadMorePackageHistoryBtn');
            if (!btn) return;
            btn.classList.toggle('hidden', !packageHistoryHasMore);
            btn.disabled = !packageHistoryHasMore || packageHistoryLoading;
            btn.textContent = packageHistoryLoading ? 'Loading...' : 'Load More';
        }

        function getSelectedPackageRows() {
            const selected = Array.from(document.querySelectorAll('#packageLogTableBody tr.package-log-row.selected'));
            if (selected.length > 0) return selected;
            // Fallback: rehydrate from saved keys
            const allRows = Array.from(document.querySelectorAll('#packageLogTableBody tr.package-log-row'));
            const byKey = new Map();
            allRows.forEach(r => {
                const rawKey = r.getAttribute('data-row-key') || '';
                try { byKey.set(decodeURIComponent(rawKey), r); } catch {}
            });
            const fromKeys = [];
            selectedPackageRowKeys.forEach(k => {
                if (byKey.has(k)) fromKeys.push(byKey.get(k));
            });
            return fromKeys;
        }

        function getSelectedPackageHistoryRows() {
            const selected = Array.from(document.querySelectorAll('#packageHistoryTableBody tr.package-history-row.selected'));
            if (selected.length > 0) return selected;
            const allRows = Array.from(document.querySelectorAll('#packageHistoryTableBody tr.package-history-row'));
            const byKey = new Map();
            allRows.forEach(r => {
                const rawKey = r.getAttribute('data-row-key') || '';
                try { byKey.set(decodeURIComponent(rawKey), r); } catch {}
            });
            const fromKeys = [];
            selectedPackageHistoryRowKeys.forEach(k => {
                if (byKey.has(k)) fromKeys.push(byKey.get(k));
            });
            return fromKeys;
        }

        function clearPackageHistorySelection() {
            selectedPackageHistoryRowKeys.clear();
            document.querySelectorAll('#packageHistoryTableBody tr.package-history-row.selected').forEach(r => r.classList.remove('selected'));
            isSelectingHistoryRows = false;
            historyDragStartIndex = null;
        }

        function getActiveSelectedLogRows() {
            if (packageHistoryModal && !packageHistoryModal.classList.contains('hidden')) {
                const historyRows = getSelectedPackageHistoryRows();
                if (historyRows.length > 0) return historyRows;
            }
            return getSelectedPackageRows();
        }

        function getRowText(row) {
            const cells = row.querySelectorAll('td');
            const parts = Array.from(cells).map(td => (td.textContent || '').trim());
            return parts.join(' ');
        }

        function getRowMessageText(row) {
            const msgCell = row.querySelector('td.col-message');
            return (msgCell ? msgCell.textContent : '').trim();
        }

        let logDetailRawText = '';
        let logDetailIsJsonView = false;

        function openLogDetailModal(rows) {
            const lines = rows.map(getRowText);
            logDetailRawText = lines.join('\\n');
            logDetailIsJsonView = false;
            logDetailContent.textContent = logDetailRawText;
            logDetailModal.classList.remove('hidden');
        }

        function extractJsonFromText(text) {
            if (!text) return null;
            const starts = [];
            for (let i = 0; i < text.length; i++) {
                const ch = text[i];
                if (ch === '{' || ch === '[') starts.push(i);
            }
            for (const start of starts) {
                let openBraces = 0;
                let openBrackets = 0;
                for (let i = start; i < text.length; i++) {
                    const ch = text[i];
                    if (ch === '{') openBraces++;
                    if (ch === '}') openBraces--;
                    if (ch === '[') openBrackets++;
                    if (ch === ']') openBrackets--;
                    if (openBraces === 0 && openBrackets === 0 && i > start) {
                        const candidate = text.slice(start, i + 1);
                        try {
                            JSON.parse(candidate);
                            return candidate;
                        } catch (e) {}
                        break;
                    }
                }
            }
            return null;
        }

        function extractEventName(text) {
            if (!text) return '';
            const m = text.match(/with\\s+name\\s+([\\w\\.:-]+)/i);
            if (m && m[1]) return m[1];
            return '';
        }

        function extractHeaderBeforeJson(text) {
            if (!text) return '';
            const idx = text.indexOf('{');
            if (idx === -1) return '';
            let header = text.slice(0, idx).trim();
            // Remove leading time/tag/prefix if present
            const i1 = header.indexOf('] : ');
            if (i1 !== -1) header = header.slice(i1 + 4).trim();
            const i2 = header.indexOf('] ');
            if (i2 !== -1) header = header.slice(i2 + 2).trim();
            return header;
        }

        function convertSelectedLogsToJson() {
            if (logDetailIsJsonView) {
                logDetailContent.textContent = logDetailRawText;
                logDetailIsJsonView = false;
                return;
            }
            const rows = getActiveSelectedLogRows();
            if (rows.length === 0) return;
            const outputs = rows.map((row, idx) => {
                const line = getRowText(row);
                const msg = getRowMessageText(row);
                const jsonStr = extractJsonFromText(msg || line);
                const evt = extractEventName(line);
                const headerLine = extractHeaderBeforeJson(line);
                if (jsonStr) {
                    try {
                        const parsed = JSON.parse(jsonStr);
                        const pretty = JSON.stringify(parsed, null, 2);
                        const metaLines = [];
                        if (headerLine) metaLines.push(`context: ${headerLine}`);
                        else if (evt) metaLines.push(`event_name: ${evt}`);
                        metaLines.push('');
                        metaLines.push('raw_log:');
                        metaLines.push(line);
                        metaLines.push('');
                        metaLines.push('extracted_json:');
                        metaLines.push(pretty);
                        return `--- #${idx + 1} (JSON) ---\\n${metaLines.join('\\n')}`;
                    } catch (e) {
                        return `--- #${idx + 1} ---\\n${line}`;
                    }
                }
                return `--- #${idx + 1} ---\\n${line}`;
            });
            logDetailContent.textContent = outputs.join('\\n\\n');
            logDetailIsJsonView = true;
        }

        convertLogJsonBtn.addEventListener('click', convertSelectedLogsToJson);

        document.getElementById('packageLogTableBody').addEventListener('dblclick', (e) => {
            const row = e.target.closest('tr.package-log-row');
            if (!row) return;
            isSelectingRows = false;
            dragStartIndex = null;
            // Ensure clicked row is included
            const rawKey = row.getAttribute('data-row-key') || '';
            try { selectedPackageRowKeys.add(decodeURIComponent(rawKey)); } catch {}
            const selected = getActiveSelectedLogRows();
            const rowsToShow = selected.length > 0 ? selected : [row];
            openLogDetailModal(rowsToShow);
        });

        document.addEventListener('keydown', (e) => {
            if (e.key !== 'Enter') return;
            const selected = getActiveSelectedLogRows();
            if (selected.length === 0) return;
            openLogDetailModal(selected);
        });
        
        // --- Device Status ---
        socket.on('device_status', (status) => {
            const currentFilter = deviceFilter.value;
            deviceFilter.innerHTML = '<option value="all">All Devices</option>';
            if (status.connected_devices) {
                status.connected_devices.forEach(d => {
                    const opt = document.createElement('option');
                    opt.value = d.id; opt.textContent = d.name;
                    deviceFilter.appendChild(opt);
                });
            }
            // Restore selection if exists
            if([...deviceFilter.options].some(o => o.value === currentFilter)) deviceFilter.value = currentFilter;

            if (status.connected_devices && status.connected_devices.length > 0) {
                 deviceListEl.innerHTML = '<ul class="list-disc list-inside text-left">' + 
                    status.connected_devices.map(d => `<li class="text-green-600 font-semibold animate-pulse-green">${d.id} - ${d.name}</li>`).join('') + 
                    '</ul>';
            } else {
                 deviceListEl.innerHTML = `<p class="text-orange-500">${status.message || 'Waiting...'}</p>`;
            }
        });

        socket.on('platform_status', (status) => {
            const platform = status?.platform === 'ios' ? 'ios' : 'android';
            activePlatform = platform;
            if (platformBtn) platformBtn.textContent = `Platform: ${platformLabel(platform)}`;
            syncPlatformUi();
        });
        
        // --- FIXED: Trigger refresh on device change to update all tables including callback
        deviceFilter.addEventListener('change', (e) => { 
            selectedDevice = e.target.value; 
            socket.emit('refresh_request'); 
            renderCallbackTable(); // Trigger client-side re-render immediately
            renderAdRevenueTable();
        });

        // --- Specific Tab Logic ---
        function setValidationButtonState(isActive) {
            const btn = document.getElementById('startValidationBtn');
            if (!btn) return;
            if (isActive) {
                btn.textContent = 'Stop';
                btn.className = 'bg-red-500 hover:bg-red-600 text-white font-semibold text-xs px-4 rounded-lg h-9';
            } else {
                btn.textContent = 'Start Checking';
                btn.className = 'bg-indigo-600 hover:bg-indigo-700 text-white font-semibold text-xs px-4 rounded-lg h-9';
            }
        }

        document.getElementById('startValidationBtn').addEventListener('click', () => {
            const btn = document.getElementById('startValidationBtn');
            const isStarting = btn && btn.textContent === 'Start Checking';
            if (isStarting) socket.emit('start_validation', []);
            else socket.emit('stop_validation');
        });

        document.getElementById('clearValidatorFilterBtn')?.addEventListener('click', () => {
            const eventInput = document.getElementById('validatorEventFilterInput');
            const rawInput = document.getElementById('validatorRawFilterInput');
            if (eventInput) eventInput.value = '';
            if (rawInput) rawInput.value = '';
            const allRadio = document.querySelector('input[name="validatorSourceFilter"][value="all"]');
            if (allRadio) allRadio.checked = true;
            renderValidatorTable(validator_results_cache);
        });
        
        document.getElementById('validatorEventFilterInput').addEventListener('input', () => {
            renderValidatorTable(validator_results_cache);
        });
        document.getElementById('validatorRawFilterInput').addEventListener('input', () => {
            renderValidatorTable(validator_results_cache);
        });

        document.querySelectorAll('input[name="validatorSourceFilter"]').forEach(r => {
            r.addEventListener('change', () => renderValidatorTable(validator_results_cache));
        });

        socket.on('validator_status', (data) => {
            setValidationButtonState(!!(data && data.active));
        });

        const specificEventInput = document.getElementById('specificEventInput');
        const specificParamInput = document.getElementById('specificParamInput');
        const updateSpecific = () => {
            socket.emit('update_specific_filter', {
                eventNames: specificEventInput.value.split('\\n').filter(p => p.trim()),
                params: []
            });
            renderSpecificEventTable();
        };
        specificEventInput.addEventListener('input', updateSpecific);
        specificParamInput.addEventListener('input', renderSpecificEventTable);
        document.querySelectorAll('input[name="specificSourceFilter"]').forEach(r => {
            r.addEventListener('change', () => renderSpecificEventTable());
        });
        
        document.getElementById('adRevenueFilterInput').addEventListener('input', renderAdRevenueTable);
        document.querySelectorAll('input[name="adRevenueSourceFilter"]').forEach(r => r.addEventListener('change', renderAdRevenueTable));
        let packageFilterRenderTimer = null;
        const schedulePackageRender = () => {
            clearTimeout(packageFilterRenderTimer);
            packageFilterRenderTimer = setTimeout(() => renderPackageLogTable(true), 120);
        };
        document.getElementById('packageFilterInput').addEventListener('input', schedulePackageRender);
        document.getElementById('packageFilterInput2').addEventListener('input', schedulePackageRender);
        document.getElementById('packageTagFilterInput').addEventListener('input', () => {
            const allOpt = document.querySelector('input[name="tagQuickFilter"][value=""]');
            if (allOpt) allOpt.checked = true;
            schedulePackageRender();
        });
        document.querySelectorAll('input[name="tagQuickFilter"]').forEach(r => r.addEventListener('change', (e) => {
            const tagInput = document.getElementById('packageTagFilterInput');
            if (tagInput && e.target.value) tagInput.value = '';
            renderPackageLogTable(true);
        }));
        document.getElementById('packageFilterInput').addEventListener('keydown', (e) => {
            if (e.key === 'Enter') { e.preventDefault(); renderPackageLogTable(true); }
        });
        document.getElementById('packageFilterInput2').addEventListener('keydown', (e) => {
            if (e.key === 'Enter') { e.preventDefault(); renderPackageLogTable(true); }
        });
        document.getElementById('packageTagFilterInput').addEventListener('keydown', (e) => {
            if (e.key === 'Enter') { e.preventDefault(); renderPackageLogTable(true); }
        });
        document.getElementById('showErrorsOnly').addEventListener('change', () => renderPackageLogTable(true));
        document.getElementById('showWarningsOnly').addEventListener('change', () => renderPackageLogTable(true));
        document.getElementById('loadPackageHistoryBtn').addEventListener('click', () => loadSelectedPackageHistory(false));
        document.getElementById('loadMorePackageHistoryBtn').addEventListener('click', () => loadSelectedPackageHistory(true));
        document.getElementById('exportPackageHistoryFilteredBtn').addEventListener('click', () => exportSelectedPackageHistory(true));
        document.getElementById('exportPackageHistoryAllBtn').addEventListener('click', () => exportSelectedPackageHistory(false));
        document.getElementById('refreshPackageSessionsBtn').addEventListener('click', loadPackageHistorySessions);
        document.getElementById('clearPackageHistoryBtn').addEventListener('click', clearAllRecordedPackageHistory);
        document.getElementById('clearPackageHistoryFiltersBtn').addEventListener('click', () => clearRecordedPackageHistoryFilters(true));
        document.getElementById('packageHistorySessionSelect').addEventListener('change', () => {
            clearRecordedPackageHistoryFilters(false);
            loadSelectedPackageHistory(false);
        });
        document.getElementById('packageHistoryFilterInput').addEventListener('keydown', (e) => {
            if (e.key === 'Enter') {
                e.preventDefault();
                loadSelectedPackageHistory(false);
            }
        });
        document.getElementById('packageHistoryFilterInput2').addEventListener('keydown', (e) => {
            if (e.key === 'Enter') {
                e.preventDefault();
                loadSelectedPackageHistory(false);
            }
        });
        document.getElementById('packageHistoryFilterInput3').addEventListener('keydown', (e) => {
            if (e.key === 'Enter') {
                e.preventDefault();
                loadSelectedPackageHistory(false);
            }
        });
        const packageHistoryModal = document.getElementById('packageHistoryModal');
        const openPackageHistoryBtn = document.getElementById('openPackageHistoryBtn');
        const closePackageHistoryModal = document.getElementById('closePackageHistoryModal');
        openPackageHistoryBtn?.addEventListener('click', () => {
            packageHistoryModal?.classList.remove('hidden');
            clearRecordedPackageHistoryFilters(false);
            loadPackageHistorySessions();
        });
        closePackageHistoryModal?.addEventListener('click', () => {
            packageHistoryModal?.classList.add('hidden');
        });
        packageHistoryModal?.addEventListener('click', (e) => {
            if (e.target === packageHistoryModal) {
                clearPackageHistorySelection();
                packageHistoryModal.classList.add('hidden');
                return;
            }
            if (!e.target.closest('#packageHistoryTableBody tr.package-history-row') && !e.target.closest('#logDetailModal')) {
                clearPackageHistorySelection();
            }
        });

        const packageHistoryTableWrap = document.getElementById('packageHistoryTableWrap');
        packageHistoryTableWrap?.addEventListener('scroll', () => {
            if (!packageHistoryHasMore || packageHistoryLoading) return;
            const remaining = packageHistoryTableWrap.scrollHeight - packageHistoryTableWrap.scrollTop - packageHistoryTableWrap.clientHeight;
            if (remaining < 120) {
                loadSelectedPackageHistory(true);
            }
        });

        ['packageHistorySessionSelect','loadPackageHistoryBtn','loadMorePackageHistoryBtn','refreshPackageSessionsBtn','exportPackageHistoryAllBtn','exportPackageHistoryFilteredBtn','clearPackageHistoryBtn','clearPackageHistoryFiltersBtn','packageHistoryFilterInput','packageHistoryFilterInput2','packageHistoryFilterInput3','closePackageHistoryModal'].forEach(id => {
            document.getElementById(id)?.addEventListener('focus', clearPackageHistorySelection);
            document.getElementById(id)?.addEventListener('click', () => { clearPackageHistorySelection(); });
        });

        // --- Row Selection for Package Log (click + drag) ---
        let isSelectingRows = false;
        let dragStartIndex = null;

        function updateSelectionRange(startIdx, endIdx) {
            const rows = Array.from(document.querySelectorAll('#packageLogTableBody tr.package-log-row'));
            if (rows.length === 0) return;
            const [minIdx, maxIdx] = startIdx <= endIdx ? [startIdx, endIdx] : [endIdx, startIdx];
            selectedPackageRowKeys.clear();
            rows.forEach(r => {
                const idx = parseInt(r.getAttribute('data-row-index') || '-1', 10);
                if (idx >= minIdx && idx <= maxIdx) {
                    const rawKey = r.getAttribute('data-row-key') || '';
                    try { selectedPackageRowKeys.add(decodeURIComponent(rawKey)); } catch (err) {}
                    r.classList.add('selected');
                } else {
                    r.classList.remove('selected');
                }
            });
        }

        document.getElementById('packageLogTableBody').addEventListener('mousedown', (e) => {
            const row = e.target.closest('tr.package-log-row');
            if (!row) return;
            if (e.detail && e.detail >= 2) return;
            if (document.activeElement && (document.activeElement.tagName === 'INPUT' || document.activeElement.tagName === 'TEXTAREA')) {
                document.activeElement.blur();
            }
            isSelectingRows = true;
            const idx = parseInt(row.getAttribute('data-row-index') || '0', 10);
            dragStartIndex = idx;
            updateSelectionRange(idx, idx);
            e.preventDefault();
        });

        document.getElementById('packageLogTableBody').addEventListener('mouseover', (e) => {
            if (!isSelectingRows || dragStartIndex === null) return;
            const row = e.target.closest('tr.package-log-row');
            if (!row) return;
            const idx = parseInt(row.getAttribute('data-row-index') || '0', 10);
            updateSelectionRange(dragStartIndex, idx);
        });

        function updateHistorySelectionRange(startIdx, endIdx) {
            const rows = Array.from(document.querySelectorAll('#packageHistoryTableBody tr.package-history-row'));
            if (rows.length === 0) return;
            const [minIdx, maxIdx] = startIdx <= endIdx ? [startIdx, endIdx] : [endIdx, startIdx];
            selectedPackageHistoryRowKeys.clear();
            rows.forEach(r => {
                const idx = parseInt(r.getAttribute('data-row-index') || '-1', 10);
                if (idx >= minIdx && idx <= maxIdx) {
                    const rawKey = r.getAttribute('data-row-key') || '';
                    try { selectedPackageHistoryRowKeys.add(decodeURIComponent(rawKey)); } catch (err) {}
                    r.classList.add('selected');
                } else {
                    r.classList.remove('selected');
                }
            });
        }

        document.getElementById('packageHistoryTableBody').addEventListener('mousedown', (e) => {
            const row = e.target.closest('tr.package-history-row');
            if (!row) return;
            if (e.detail && e.detail >= 2) return;
            if (document.activeElement && (document.activeElement.tagName === 'INPUT' || document.activeElement.tagName === 'TEXTAREA')) {
                document.activeElement.blur();
            }
            isSelectingHistoryRows = true;
            const idx = parseInt(row.getAttribute('data-row-index') || '0', 10);
            historyDragStartIndex = idx;
            updateHistorySelectionRange(idx, idx);
            e.preventDefault();
        });

        document.getElementById('packageHistoryTableBody').addEventListener('mouseover', (e) => {
            if (!isSelectingHistoryRows || historyDragStartIndex === null) return;
            const row = e.target.closest('tr.package-history-row');
            if (!row) return;
            const idx = parseInt(row.getAttribute('data-row-index') || '0', 10);
            updateHistorySelectionRange(historyDragStartIndex, idx);
        });

        document.getElementById('packageHistoryTableBody').addEventListener('dblclick', (e) => {
            const row = e.target.closest('tr.package-history-row');
            if (!row) return;
            isSelectingHistoryRows = false;
            historyDragStartIndex = null;
            const rawKey = row.getAttribute('data-row-key') || '';
            try { selectedPackageHistoryRowKeys.add(decodeURIComponent(rawKey)); } catch {}
            const selected = getSelectedPackageHistoryRows();
            openLogDetailModal(selected.length > 0 ? selected : [row]);
        });

        document.addEventListener('mouseup', () => {
            isSelectingRows = false;
            dragStartIndex = null;
            isSelectingHistoryRows = false;
            historyDragStartIndex = null;
        });

        // --- Ctrl+C to copy selected package log rows ---
        document.addEventListener('keydown', async (e) => {
            if (!(e.ctrlKey || e.metaKey) || e.key.toLowerCase() !== 'c') return;
            const active = document.activeElement;
            if (active && (active.tagName === 'INPUT' || active.tagName === 'TEXTAREA')) return;
            const selectedRows = getActiveSelectedLogRows();
            if (selectedRows.length === 0) return;
            const lines = selectedRows.map(r => {
                const cells = r.querySelectorAll('td');
                const parts = Array.from(cells).map(td => (td.textContent || '').trim());
                return parts.join('\\t');
            });
            const textToCopy = lines.join('\\n');
            try {
                if (navigator.clipboard && navigator.clipboard.writeText) {
                    await navigator.clipboard.writeText(textToCopy);
                } else {
                    const ta = document.createElement('textarea');
                    ta.value = textToCopy;
                    ta.style.position = 'fixed';
                    ta.style.opacity = '0';
                    document.body.appendChild(ta);
                    ta.select();
                    document.execCommand('copy');
                    document.body.removeChild(ta);
                }
            } catch (err) {
                console.error('Copy failed', err);
            }
        });
        
        let sdkCheckRunning = false;
        document.getElementById('startSdkCheckBtn').addEventListener('click', () => {
             const btn = document.getElementById('startSdkCheckBtn');
             if (sdkCheckRunning) {
                 socket.emit('stop_sdk_check');
                 sdkCheckRunning = false;
                 btn.textContent = 'Start Checking';
                 return;
             }
             const text = document.getElementById('sdkCheckInput').value;
             if(text) {
                 socket.emit('start_sdk_check', {text: text});
                 sdkCheckRunning = true;
                 btn.textContent = 'Stop Checking';
             }
        });
        
        function setPackageControlsEnabled(enabled) {
            const packageIdInput = document.getElementById('packageIdInput');
            const packageCheckboxes = document.querySelectorAll('.package-id-checkbox');
            if (packageIdInput) packageIdInput.disabled = !enabled;
            packageCheckboxes.forEach(cb => cb.disabled = !enabled);
        }

        const packageStreamModal = document.getElementById('packageStreamModal');
        const startPackageLogBtn = document.getElementById('startPackageLogBtn');
        const pausePackageLogBtn = document.getElementById('pausePackageLogBtn');
        const stopPackageLogBtn = document.getElementById('stopPackageLogBtn');

        function setPackageRunningState(isRunning) {
            if (!startPackageLogBtn) return;
            startPackageLogBtn.textContent = isRunning ? 'Running...' : 'Start';
            startPackageLogBtn.disabled = isRunning;
            startPackageLogBtn.classList.toggle('opacity-60', isRunning);
            startPackageLogBtn.classList.toggle('cursor-not-allowed', isRunning);
            setPackageControlsEnabled(!isRunning);
        }

        function openPackageStreamModal() {
            packageStreamModal?.classList.remove('hidden');
        }

        function closePackageStreamModal() {
            packageStreamModal?.classList.add('hidden');
        }

        function resetPackageLogUiState() {
            const packageIdInput = document.getElementById('packageIdInput');
            const tagInput = document.getElementById('packageTagFilterInput');
            const filter1 = document.getElementById('packageFilterInput');
            const filter2 = document.getElementById('packageFilterInput2');
            const showErrorsOnly = document.getElementById('showErrorsOnly');
            const showWarningsOnly = document.getElementById('showWarningsOnly');
            const autoScroll = document.getElementById('autoScroll');
            const allQuickTag = document.querySelector('input[name="tagQuickFilter"][value=""]');
            const packageContainer = document.getElementById('packageLogContainer');

            if (packageIdInput) packageIdInput.value = '';
            document.querySelectorAll('.package-id-checkbox').forEach(cb => { cb.checked = false; });
            if (tagInput) tagInput.value = '';
            if (filter1) filter1.value = '';
            if (filter2) filter2.value = '';
            if (showErrorsOnly) showErrorsOnly.checked = false;
            if (showWarningsOnly) showWarningsOnly.checked = false;
            if (autoScroll) autoScroll.checked = true;
            if (allQuickTag) allQuickTag.checked = true;
            packageSelectedRowIndex = null;
            packageDragAnchor = null;
            packageLastClickedIndex = null;
            packageUiPaused = false;
            pausedPackageSnapshot = [];
            if (packageContainer) packageContainer.scrollTop = 0;
            renderPackageLogTable(true);
        }

        function setPackagePauseState(isPaused) {
            packageUiPaused = isPaused;
            if (pausePackageLogBtn) {
                pausePackageLogBtn.textContent = isPaused ? 'Resume' : 'Pause';
                pausePackageLogBtn.classList.toggle('bg-amber-500', !isPaused);
                pausePackageLogBtn.classList.toggle('hover:bg-amber-600', !isPaused);
                pausePackageLogBtn.classList.toggle('bg-emerald-500', isPaused);
                pausePackageLogBtn.classList.toggle('hover:bg-emerald-600', isPaused);
            }
            if (isPaused) {
                pausedPackageSnapshot = filterPackageLogs(lastPackageLogs, getPackageFilterState());
            } else {
                pausedPackageSnapshot = [];
            }
            renderPackageLogTable(true);
        }

        startPackageLogBtn?.addEventListener('click', () => {
             const pkg = document.getElementById('packageIdInput').value;
             setPackagePauseState(false);
             openPackageStreamModal();
             socket.emit('start_package_log', {package_id: pkg});
             setPackageRunningState(true);
        });

        pausePackageLogBtn?.addEventListener('click', () => {
            setPackagePauseState(!packageUiPaused);
        });

        stopPackageLogBtn?.addEventListener('click', () => {
            socket.emit('start_package_log', {package_id: ''});
            setPackagePauseState(false);
            closePackageStreamModal();
            setPackageRunningState(false);
            resetPackageLogUiState();
        });

        // --- Column Resizer (Package Log Table) ---
        (function setupColumnResizers() {
            let active = null;
            let startX = 0;
            let startWidth = 0;

            function onMouseMove(e) {
                if (!active) return;
                const dx = e.clientX - startX;
                const newWidth = Math.max(50, startWidth + dx);
                const colClass = active.getAttribute('data-col');
                const colEl = document.querySelector(`.col-${colClass}`);
                if (colEl) colEl.style.width = newWidth + 'px';
            }

            function onMouseUp() {
                if (!active) return;
                document.removeEventListener('mousemove', onMouseMove);
                document.removeEventListener('mouseup', onMouseUp);
                document.body.style.cursor = '';
                document.body.style.userSelect = '';
                active = null;
            }

            document.addEventListener('mousedown', (e) => {
                const resizer = e.target.closest('.resizer');
                if (!resizer) return;
                if (!(isPausedClient || pauseBtn.textContent === 'Resume')) return;
                e.preventDefault();
                active = resizer;
                const colClass = resizer.getAttribute('data-col');
                const headerCell = document.querySelector(`th.${colClass}`);
                startX = e.clientX;
                startWidth = headerCell ? headerCell.offsetWidth : 0;
                    document.body.style.cursor = 'col-resize';
                    document.body.style.userSelect = 'none';
                    document.addEventListener('mousemove', onMouseMove);
                    document.addEventListener('mouseup', onMouseUp);
            });
            // Apply initial state
            setResizerEnabled(isPausedClient);
        })();

        // --- Package ID Quick Select ---
        const packageIdInput = document.getElementById('packageIdInput');
        const packageCheckboxes = document.querySelectorAll('.package-id-checkbox');

        packageCheckboxes.forEach(cb => {
            cb.addEventListener('change', (e) => {
                if (e.target.checked) {
                    packageCheckboxes.forEach(other => { if (other !== e.target) other.checked = false; });
                    packageIdInput.value = e.target.value;
                } else {
                    if (packageIdInput.value === e.target.value) packageIdInput.value = '';
                }
            });
        });

        // Manual input clears quick select
        packageIdInput.addEventListener('input', () => {
            packageCheckboxes.forEach(cb => cb.checked = false);
        });

        // Clear All should also stop package log and re-enable inputs
        clearAllBtn.addEventListener('click', () => {
            const btn = document.getElementById('startPackageLogBtn');
            if (btn && btn.textContent !== 'Start') {
                socket.emit('start_package_log', {package_id: ''});
                setPackagePauseState(false);
                closePackageStreamModal();
                setPackageRunningState(false);
                resetPackageLogUiState();
            }
            setPackageControlsEnabled(true);
        });

        loadPackageHistorySessions();
        
    </script>
</body>
</html>
"""

# --- SERVER ROUTE (QUAN TRỌNG) ---
@app.route('/')
def index():
    return render_template_string(
        HTML_TEMPLATE,
        default_event_names=sorted(event_specific_params.keys()),
        current_profile_name=active_profile_name
    )


@app.get('/api/profiles')
def get_profiles():
    return jsonify(_profile_payload())


@app.post('/api/profiles/select')
def select_profile():
    data = request.get_json(silent=True) or {}
    profile_name = data.get('profile_name', '')
    if not profile_name:
        return jsonify({'ok': False, 'error': 'profile_name_required'}), 400
    try:
        if not _set_active_profile(profile_name):
            return jsonify({'ok': False, 'error': 'profile_not_found'}), 404
        _apply_adrevenue_filter_and_emit()
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400
    return jsonify({'ok': True, **_profile_payload()})


@app.post('/api/profiles/reload')
def reload_profile():
    if active_profile_name:
        _set_active_profile(active_profile_name)
    else:
        _set_active_profile()
    _apply_adrevenue_filter_and_emit()
    return jsonify({'ok': True, **_profile_payload()})


@app.post('/api/profiles/import')
def import_profile():
    upload = request.files.get('profile_file')
    if not upload or not upload.filename:
        return jsonify({'ok': False, 'error': 'profile_file_required'}), 400
    try:
        filename = _sanitize_profile_filename(upload.filename)
        target = os.path.join(PROFILE_DIR, filename)
        os.makedirs(PROFILE_DIR, exist_ok=True)
        upload.save(target)
        _set_active_profile(filename)
        _apply_adrevenue_filter_and_emit()
        return jsonify({'ok': True, **_profile_payload()})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@app.post('/check_update')
def check_update():
    try:
        import remote_update
    except Exception as e:
        return jsonify({'ok': False, 'status': 'error', 'error': f'updater_unavailable: {e}'})
    try:
        result = remote_update.check_for_updates()
        return jsonify(result)
    except Exception as e:
        return jsonify({'ok': False, 'status': 'error', 'error': str(e)})


@app.post('/restart_app')
def restart_app():
    cmd = os.getenv('EVENTINSPECTOR_RESTART_CMD')
    args = os.getenv('EVENTINSPECTOR_RESTART_ARGS', '')
    if not cmd:
        return jsonify({'ok': False, 'error': 'restart_cmd_missing'})
    argv = [cmd]
    if args:
        argv.append(args)
    try:
        subprocess.Popen(argv, creationflags=creation_flags)
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})
    os._exit(0)


@app.get('/api/package-log/sessions')
def package_log_sessions_api():
    conn = _get_package_db_connection()
    try:
        rows = conn.execute(
            """
            SELECT s.id, s.package_id, s.started_at, s.ended_at, s.status, COUNT(e.id) AS row_count
            FROM package_log_sessions s
            LEFT JOIN package_log_entries e ON e.session_id = s.id
            GROUP BY s.id
            ORDER BY s.id DESC
            LIMIT 50
            """
        ).fetchall()
        sessions = []
        for row in rows:
            started_label = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row["started_at"])) if row["started_at"] else ""
            sessions.append({
                "id": row["id"],
                "package_id": row["package_id"],
                "started_at": row["started_at"],
                "ended_at": row["ended_at"],
                "started_label": started_label,
                "status": row["status"],
                "row_count": row["row_count"] or 0,
            })
        return jsonify({"ok": True, "sessions": sessions})
    finally:
        conn.close()


@app.get('/api/package-log/logs')
def package_log_rows_api():
    session_id = request.args.get('session_id', type=int)
    keyword1 = (request.args.get('q1') or request.args.get('q') or '').strip().lower()
    keyword2 = (request.args.get('q2') or '').strip().lower()
    keyword3 = (request.args.get('q3') or '').strip().lower()
    offset = max(request.args.get('offset', default=0, type=int) or 0, 0)
    limit = request.args.get('limit', default=2000, type=int) or 2000
    limit = max(1, min(limit, 5000))
    if not session_id:
        return jsonify({"ok": False, "error": "session_id_required"}), 400
    conn = _get_package_db_connection()
    try:
        base_where = """
            FROM package_log_entries
            WHERE session_id = ?
        """
        params = [session_id]

        def search_clause():
            return """
              AND (
                lower(coalesce(raw_log, '')) LIKE ?
                OR lower(coalesce(message, '')) LIKE ?
                OR lower(coalesce(tag, '')) LIKE ?
                OR lower(coalesce(device_name, '')) LIKE ?
              )
            """

        if keyword1:
            base_where += search_clause()
            like1 = f"%{keyword1}%"
            params.extend([like1, like1, like1, like1])
        if keyword2:
            base_where += search_clause()
            like2 = f"%{keyword2}%"
            params.extend([like2, like2, like2, like2])
        if keyword3:
            base_where += search_clause()
            like3 = f"%{keyword3}%"
            params.extend([like3, like3, like3, like3])

        total_rows = conn.execute(f"SELECT COUNT(*) AS c {base_where}", params).fetchone()["c"] or 0
        sql = f"""
            SELECT time_display, device_id, device_name, tag, message, raw_log
            {base_where}
            ORDER BY id DESC
            LIMIT ? OFFSET ?
        """
        query_params = params + [limit, offset]
        rows = conn.execute(sql, query_params).fetchall()
        rows = list(reversed(rows))
        loaded_count = offset + len(rows)
        return jsonify({
            "ok": True,
            "session_id": session_id,
            "rows": [dict(r) for r in rows],
            "offset": offset,
            "limit": limit,
            "total_rows": total_rows,
            "has_more": loaded_count < total_rows,
        })
    finally:
        conn.close()


@app.post('/api/package-log/clear')
def package_log_clear_api():
    payload = request.get_json(silent=True) or {}
    if not payload.get("clear_all"):
        return jsonify({"ok": False, "error": "confirmation_required"}), 400
    conn = _get_package_db_connection()
    try:
        session_count = conn.execute("SELECT COUNT(*) AS c FROM package_log_sessions").fetchone()["c"] or 0
        row_count = conn.execute("SELECT COUNT(*) AS c FROM package_log_entries").fetchone()["c"] or 0
        conn.execute("DELETE FROM package_log_entries")
        conn.execute("DELETE FROM package_log_sessions")
        conn.commit()
        try:
            conn.execute("VACUUM")
        except Exception:
            pass
        return jsonify({
            "ok": True,
            "deleted_sessions": session_count,
            "deleted_rows": row_count,
        })
    finally:
        conn.close()


@app.post('/api/package-log/export')
def package_log_export_api():
    payload = request.get_json(silent=True) or {}
    try:
        session_id = int(payload.get('session_id'))
    except Exception:
        return jsonify({"ok": False, "error": "session_id_required"}), 400

    keyword1 = str(payload.get('q1') or '').strip().lower()
    keyword2 = str(payload.get('q2') or '').strip().lower()
    keyword3 = str(payload.get('q3') or '').strip().lower()
    filtered_only = bool(payload.get('filtered_only', True))
    conn = _get_package_db_connection()
    try:
        session = conn.execute(
            "SELECT id, package_id, started_at FROM package_log_sessions WHERE id = ?",
            (session_id,),
        ).fetchone()
        if not session:
            return jsonify({"ok": False, "error": "session_not_found"}), 404

        sql = """
            SELECT raw_log, message
            FROM package_log_entries
            WHERE session_id = ?
        """
        params = [session_id]

        def search_clause():
            return """
              AND (
                lower(coalesce(raw_log, '')) LIKE ?
                OR lower(coalesce(message, '')) LIKE ?
                OR lower(coalesce(tag, '')) LIKE ?
                OR lower(coalesce(device_name, '')) LIKE ?
              )
            """

        if keyword1:
            like1 = f"%{keyword1}%"
            sql += search_clause()
            params.extend([like1, like1, like1, like1])
        if keyword2:
            like2 = f"%{keyword2}%"
            sql += search_clause()
            params.extend([like2, like2, like2, like2])
        if keyword3:
            like3 = f"%{keyword3}%"
            sql += search_clause()
            params.extend([like3, like3, like3, like3])

        sql += " ORDER BY id ASC"
        rows = conn.execute(sql, params).fetchall()

        export_dir = os.path.join(_package_history_dir(), "exports")
        os.makedirs(export_dir, exist_ok=True)
        started_at = session["started_at"] or time.time()
        started_label = time.strftime("%Y%m%d_%H%M%S", time.localtime(started_at))
        package_stub = re.sub(r'[^A-Za-z0-9._-]+', '_', session["package_id"] or f"session_{session_id}")
        suffix = "filtered" if filtered_only and (keyword1 or keyword2 or keyword3) else "full"
        filename = f"package_log_{session_id}_{package_stub}_{started_label}_{suffix}.log"
        path = os.path.join(export_dir, filename)

        chosen_path = None
        try:
            if webview and getattr(webview, "windows", None):
                selected = webview.windows[0].create_file_dialog(
                    webview.SAVE_DIALOG,
                    directory=export_dir,
                    save_filename=filename,
                    file_types=("Log files (*.log)", "All files (*.*)"),
                )
                if selected:
                    chosen_path = selected[0] if isinstance(selected, (list, tuple)) else selected
        except Exception:
            logging.exception("Package log save dialog failed")

        if chosen_path:
            path = chosen_path
        elif webview and getattr(webview, "windows", None):
            return jsonify({"ok": False, "cancelled": True, "session_id": session_id})

        with open(path, "w", encoding="utf-8") as f:
            for row in rows:
                text = row["raw_log"] or row["message"] or ""
                f.write(text.rstrip("\n"))
                f.write("\n")

        return jsonify({"ok": True, "path": path, "row_count": len(rows), "session_id": session_id})
    finally:
        conn.close()

# --- BACKEND LOGIC ---

def send_to_sheet(device_name, ad_source, ad_format, raw_log, log_type):
    """Gửi data lên Google Sheet nếu đang bật Recording cho loại Log cụ thể"""
    state = recording_states.get(log_type)
    if state and state["is_recording"] and state["current_sheet"]:
        payload = {
            "sheet_name": state["current_sheet"], 
            "device_name": device_name, 
            "ad_source": ad_source, 
            "ad_format": ad_format, 
            "raw_log": raw_log
        }
        threading.Thread(target=lambda: requests.post(G_SHEET_URL, json=payload, timeout=30)).start()

def process_load_ads_unity_log(line, device_id):
    """Xử lý log cho Tab 1: Load Ads (Unity)"""
    # CHỈ XỬ LÝ NẾU ĐANG GHI (RECORDING)
    if not recording_states["LoadAds"]["is_recording"]:
        return

    match = UNITY_TRACKING_PATTERN.search(line)
    if match:
        try:
            data = json.loads(match.group(1))
            params = data.get("e", {})
            src = params.get("ad_source")
            fmt = params.get("ad_format")
            if params.get("mediation_ad_unit_name") == "MREC": fmt = "MREC"
            
            if src and fmt:
                d_name = get_device_name(device_id)
                with lock:
                    if (device_id, src, fmt, "unity") not in unique_load_ads:
                        unique_load_ads.add((device_id, src, fmt, "unity"))
                        load_ads_events.append({
                            "device_id": device_id,
                            "device_name": d_name, 
                            "ad_source": src, 
                            "ad_format": fmt, 
                            "raw_log": line.strip()
                        })
                        socketio.emit('update_load_ads', list(load_ads_events))
                        
                        # Gửi với type "LoadAds"
                        send_to_sheet(d_name, src, fmt, line.strip(), "LoadAds")
        except: pass

def process_load_ads_ext_log(line, device_id):
    """Xử lý log cho Tab 2: Load Ads Ext (AppMetrica AdRevenue)"""
    # CHỈ XỬ LÝ NẾU ĐANG GHI (RECORDING)
    if not recording_states["LoadAdsExt"]["is_recording"]:
        return

    match = LOAD_ADS_EXT_ADREVENUE_PATTERN.search(line)
    if match:
        try:
            parsed = parse_appmetrica_adrevenue_text(f"AdRevenue{{{match.group(1)}}}") or {}
            payload = parsed.get("payload") if isinstance(parsed.get("payload"), dict) else {}

            ad_network = parsed.get("adNetwork") or payload.get("ad_network")
            fmt = parsed.get("adType") or payload.get("ad_format")
            if fmt and str(fmt).strip().lower() == "mrec":
                fmt = "MREC"

            if ad_network and fmt:
                d_name = get_device_name(device_id)
                with lock:
                    if (device_id, ad_network, fmt, "metrica") not in unique_load_ads_ext:
                        unique_load_ads_ext.add((device_id, ad_network, fmt, "metrica"))
                        load_ads_ext_events.append({
                            "device_id": device_id,
                            "device_name": d_name, 
                            "ad_network": ad_network,
                            "ad_format": fmt, 
                            "raw_log": line.strip()
                        })
                        socketio.emit('update_load_ads_ext', list(load_ads_ext_events))
                        
                        # Gửi với type "LoadAdsExt"
                        send_to_sheet(d_name, ad_network, fmt, line.strip(), "LoadAdsExt")
        except: pass

def find_and_parse_event(log_entry):
    """Parse log sự kiện chung từ TrackingService->Track và AppMetrica regular event."""
    if 'TrackingService->Track:' in log_entry:
        try:
            after_keyword = log_entry.split('TrackingService->Track:', 1)[1]
            json_str = extract_json_object_from_text(after_keyword)
            if not json_str:
                match = OLD_EVENT_LOG_PATTERN.search(log_entry)
                json_str = match.group(1) if match else None
            if json_str:
                data = json.loads(json_str)
                event_name = (
                    data.get('eventName')
                    or data.get('EventName')
                    or data.get('event_name')
                )
                params = data.get('e')
                if params is None:
                    params = data.get('params', {})
                if isinstance(params, str):
                    try:
                        params = json.loads(params)
                    except Exception:
                        params = {}
                if event_name and isinstance(params, dict):
                    wrapped = {
                        'eventName': event_name,
                        'e': params,
                        'source': 'firebase',
                    }
                    return event_name, params, json.dumps(wrapped, ensure_ascii=False)
        except Exception:
            pass

    match = METRICA_REGULAR_EVENT_PATTERN.search(log_entry)
    if match:
        try:
            event_name = match.group(1)
            params = json.loads(match.group(2))
            wrapped = {
                'eventName': event_name,
                'e': params,
                'source': 'appmetrica',
            }
            return event_name, params, json.dumps(wrapped, ensure_ascii=False)
        except Exception:
            pass
    return None, None, None

def process_callback_and_ad_event_log(log_entry, device_id, event_name=None, actual_params=None, json_string=None):
    global incomplete_impression_logs
    if is_paused: return

    # --- 0. Process Gadsme callbacks ---
    if GADSME_SERVICE_KEYWORD in log_entry:
        try:
            after_keyword = log_entry.split(GADSME_SERVICE_KEYWORD, 1)[1]
            if "->" in after_keyword:
                method_part = after_keyword.split("->", 1)[1].split(":", 1)[0].strip()
            else:
                method_part = after_keyword.split(":", 1)[0].strip() or "Gadsme"
            json_str = extract_json_object_from_text(after_keyword)
            details = "N/A"
            json_data_for_log = "{}"
            if json_str:
                try:
                    data = json.loads(json_str)
                    details = format_json_html(data)
                    json_data_for_log = json_str
                except:
                    details = f'<div class="text-xs font-mono break-all text-red-600">JSON Parse Error</div><div class="text-xs font-mono break-all">{json_str}</div>'
            with lock:
                callback_ad_logs.append({
                    "device_id": device_id,
                    "device_name": get_device_name(device_id),
                    "type": "Callback Gadsme",
                    "event_name": method_part or "Gadsme",
                    "details": details,
                    "raw_log": log_entry.strip(),
                    "json_data": json_data_for_log
                })
                socketio.emit('update_callback_ad_table', list(callback_ad_logs))
            return
        except:
            pass

    # --- 0b. Process Adverty callbacks ---
    if ADVERTY5_KEYWORD in log_entry:
        try:
            after_keyword = log_entry.split(ADVERTY5_KEYWORD, 1)[1]
            if "->" in after_keyword:
                method_part = after_keyword.split("->", 1)[1].split(":", 1)[0].strip()
            else:
                method_part = after_keyword.split(":", 1)[0].strip() or "Adverty"
            json_str = extract_json_object_from_text(after_keyword)
            details = "N/A"
            json_data_for_log = "{}"
            if json_str:
                try:
                    data = json.loads(json_str)
                    details = format_json_html(data)
                    json_data_for_log = json_str
                except:
                    details = f'<div class="text-xs font-mono break-all text-red-600">JSON Parse Error</div><div class="text-xs font-mono break-all">{json_str}</div>'
            with lock:
                callback_ad_logs.append({
                    "device_id": device_id,
                    "device_name": get_device_name(device_id),
                    "type": "Callback Adverty5",
                    "event_name": method_part or "Adverty",
                    "details": details,
                    "raw_log": log_entry.strip(),
                    "json_data": json_data_for_log
                })
                socketio.emit('update_callback_ad_table', list(callback_ad_logs))
            return
        except:
            pass
    
    # --- 1. HANDLING BUFFERED IMPRESSION DATA (Split Logs) ---
    # Check if we are currently buffering for this device
    with lock:
        current_buffer = incomplete_impression_logs.get(device_id, "")

    current_split_key = ""
    if "_OnLevelPlayImpressionDataReadyEvent" in current_buffer:
        current_split_key = "_OnLevelPlayImpressionDataReadyEvent"
    elif "_OnImpressionDataReadyEvent" in current_buffer:
        current_split_key = "_OnImpressionDataReadyEvent"

    split_impression_key = ""
    if "_OnLevelPlayImpressionDataReadyEvent" in log_entry:
        split_impression_key = "_OnLevelPlayImpressionDataReadyEvent"
    elif "_OnImpressionDataReadyEvent" in log_entry:
        split_impression_key = "_OnImpressionDataReadyEvent"

    is_split_impression_start = bool(split_impression_key)

    # If we are buffering, OR if this is a new split impression event
    if current_buffer or is_split_impression_start:
        with lock:
            # Re-read buffer inside lock to be safe (though simple logic here is fine)
            current_buffer = incomplete_impression_logs.get(device_id, "")
            current_split_key = ""
            if "_OnLevelPlayImpressionDataReadyEvent" in current_buffer:
                current_split_key = "_OnLevelPlayImpressionDataReadyEvent"
            elif "_OnImpressionDataReadyEvent" in current_buffer:
                current_split_key = "_OnImpressionDataReadyEvent"
            
            # Case A: Start of new log
            if current_buffer and current_split_key and split_impression_key == current_split_key:
                # Continuation line may repeat the same callback prefix; append instead of
                # resetting the buffer so two split lines become one merged record.
                current_buffer += "\n" + log_entry
            elif is_split_impression_start:
                # Reset buffer with current line
                current_buffer = log_entry
                current_split_key = split_impression_key
            elif current_split_key and current_split_key in log_entry:
                # Case B: Continuation line
                current_buffer += "\n" + log_entry # Add newline to separate lines if needed, or just string concat
            else:
                # Buffer existed but this line is not the matching continuation.
                # Keep the old buffer for its own pair and let this line continue
                # through the normal parsing path below.
                incomplete_impression_logs[device_id] = current_buffer
                current_buffer = ""

            parse_buffer = current_buffer
            if current_split_key:
                callback_marker = f"LevelPlayAdService->{current_split_key}:"
                buffer_parts = []
                for buffer_line in current_buffer.splitlines():
                    cleaned_line = buffer_line.strip()
                    if callback_marker in cleaned_line:
                        cleaned_line = cleaned_line.split(callback_marker, 1)[1].strip()
                    buffer_parts.append(cleaned_line)
                parse_buffer = ''.join(buffer_parts)

            if not parse_buffer:
                pass
            else:
                # Try to find JSON
            # 1. Find first '{'
                start_idx = parse_buffer.find('{')
            
            # If no '{' yet, just keep buffering (unless it's been too long?)
            if parse_buffer and start_idx != -1:
                # 2. Count braces to find end
                open_braces = 0
                end_idx = -1
                for i in range(start_idx, len(parse_buffer)):
                    if parse_buffer[i] == '{': open_braces += 1
                    elif parse_buffer[i] == '}': open_braces -= 1
                    
                    if open_braces == 0:
                        end_idx = i
                        break
                
                if end_idx != -1:
                    # Found complete JSON
                    json_str = parse_buffer[start_idx : end_idx+1]
                    details = ""
                    json_data_for_log = "{}"
                    display_name = (
                        "LevelPlay Impression Data"
                        if "_OnLevelPlayImpressionDataReadyEvent" in current_buffer
                        else "Impression Data"
                    )
                    
                    try:
                        data = json.loads(json_str)
                        if 'impressionData' in data:
                            details_target = data['impressionData']
                        else:
                            details_target = data
                        details = format_json_html(details_target)
                        json_data_for_log = json.dumps(details_target, ensure_ascii=False)
                    except:
                        details = f'<div class="text-xs font-mono break-all text-red-600">JSON Parse Error</div><div class="text-xs font-mono break-all">{json_str}</div>'
                        json_data_for_log = "{}"

                    # Clear buffer
                    incomplete_impression_logs[device_id] = ""
                    
                    # Emit
                    callback_ad_logs.append({
                        "device_id": device_id, 
                        "device_name": get_device_name(device_id), 
                        "type": "Callback", 
                        "event_name": display_name, 
                        "details": details, 
                        "raw_log": current_buffer.strip(), 
                        "json_data": json_data_for_log
                    })
                    socketio.emit('update_callback_ad_table', list(callback_ad_logs))
                    return # Done processing this line/buffer
                else:
                     # JSON start found but not ended -> Update buffer and wait for next line
                     incomplete_impression_logs[device_id] = current_buffer
                     return # Consumed line
            elif current_buffer:
                # No JSON start found yet (e.g. log line is "_OnImpressionDataReadyEvent:" and JSON is on next line)
                incomplete_impression_logs[device_id] = current_buffer
                return # Consumed line

    # --- 2. Process "ad_" events ---
    if event_name and event_name.startswith("ad_"):
        try:
            details = format_json_html(actual_params) if actual_params else "No params"
            with lock:
                callback_ad_logs.append({"device_id": device_id, "device_name": get_device_name(device_id), "type": "Ad Event", "event_name": event_name, "details": details, "raw_log": log_entry.strip(), "json_data": json_string})
                socketio.emit('update_callback_ad_table', list(callback_ad_logs))
        except: pass
    
    # --- 3. Process Other Callbacks ---
    callback_match = CALLBACK_LOG_PATTERN.search(log_entry)
    if callback_match:
        found_key = callback_match.group(1)
        # Skip if it is impression data (handled above)
        if "_OnImpressionDataReadyEvent" in found_key or "_OnLevelPlayImpressionDataReadyEvent" in found_key: return

        details = "N/A"
        display_name = CALLBACK_DISPLAY_NAMES.get(found_key, found_key)
        json_data_for_log = "{}"
        
        if found_key == "Receive Ironsource Impression Data LevelPlayImpressionData":
             try:
                 payload = log_entry.split(found_key, 1)[1].strip()
                 parsed_data = parse_levelplay_impression_text(payload)
                 if parsed_data:
                     details = format_json_html(parsed_data)
                     json_data_for_log = json.dumps(parsed_data, ensure_ascii=False)
                 else:
                     details = f'<div class="text-xs font-mono break-all">{html.escape(payload)}</div>' if payload else "Ironsource Impression Data"
             except:
                 details = "Ironsource Impression Data"
        elif "Listener" in found_key:
             # Try parsing adInfo if present
             try:
                 if 'adInfo:' in log_entry:
                     # Simple parsing for adInfo string
                     parts = log_entry.split('adInfo:')[1].strip().split(',')
                     ad_info = {}
                     for part in parts:
                         if ':' in part:
                             k, v = part.split(':', 1)
                             ad_info[k.strip()] = v.strip()
                     details = format_json_html(ad_info)
                     json_data_for_log = json.dumps(ad_info)
                 else:
                     details = "Listener Fired (No adInfo)"
             except:
                 details = "Listener Fired"

        with lock:
            callback_ad_logs.append({"device_id": device_id, "device_name": get_device_name(device_id), "type": "Callback", "event_name": display_name, "details": details, "raw_log": log_entry.strip(), "json_data": json_data_for_log})
            socketio.emit('update_callback_ad_table', list(callback_ad_logs))

def process_event_validator_log(event_name, actual_params, json_string, log_entry, device_id):
    if is_paused or not validator_active: 
        return
    source = "firebase"
    try:
        payload = json.loads(json_string)
        source = payload.get("source", "firebase")
    except:
        pass
    with lock:
        required_all = []
        required_all.extend(default_params)
        required_all.extend(required_params)
        specific = event_specific_params.get(event_name, [])
        required_all.extend(specific)

        required_set = set(required_all)
        actual_set = set(actual_params.keys())
        missing = sorted(required_set - actual_set)
        strange = sorted(actual_set - required_set) if required_set else []
        status = "PASSED" if not missing and not strange else "FAILED"
        
        # New Formatting Logic: Highlight missing -> Show full JSON
        details_html = ""
        if not specific and event_specific_params:
            closest = None
            best_dist = 3
            for known in event_specific_params.keys():
                d = _levenshtein_distance_limit(event_name, known, limit=2)
                if d < best_dist:
                    best_dist = d
                    closest = known
                    if best_dist == 1:
                        break
            if closest and best_dist <= 2:
                details_html += f'<div class="text-orange-600 font-bold mb-2">Possible typo: "{event_name}" ~ "{closest}"</div>'
        if missing:
            details_html += format_param_issue_html("Missing", missing, "text-red-600")
        if strange:
            details_html += format_param_issue_html("Strange", strange, "text-orange-600")
        
        details_html += format_json_html(actual_params)
        
        validator_results.append({"device_id": device_id, "event_name": event_name, "device_name": get_device_name(device_id), "status": status, "details": details_html, "raw_log": log_entry.strip(), "json_data": json_string, "source": source})
        socketio.emit('update_validator_table', list(validator_results))

def _apply_specific_filter_and_emit():
    global specific_event_results
    with lock:
        res = []
        for item in event_log_cache:
             try:
                 data = json.loads(item['json_data'])
                 evt = data.get('eventName')
                 params = data.get('e', {})
                 
                 source = data.get('source', 'firebase')

                 # Name Filter
                 if specific_event_name_filters and not any(evt.startswith(f) for f in specific_event_name_filters):
                     continue
                 
                 # Param Validation (Default Params + Default Events)
                 status = "INFO"
                 details = ""
                 
                 required_all = []
                 required_all.extend(default_params)
                 # If user input params, combine with defaults
                 if specific_event_params_filters:
                     required_all.extend(specific_event_params_filters)
                 # Add event-specific params if event is in default event list
                 if evt in event_specific_params:
                     required_all.extend(event_specific_params.get(evt, []))

                 if required_all:
                     required_set = set(required_all)
                     actual_set = set(params.keys())
                     missing = sorted(required_set - actual_set)
                     strange = sorted(actual_set - required_set)
                     status = "PASSED" if not missing and not strange else "FAILED"
                     
                     if missing:
                         details += format_param_issue_html("Missing", missing, "text-red-600")
                     
                     if strange:
                         details += format_param_issue_html("Strange", strange, "text-orange-600")
                     
                     # Show full JSON
                     details += format_json_html(params)
                 else:
                     # Show all params json
                     details = format_json_html(params)

                 res.append({"device_id": item['device_id'], "device_name": get_device_name(item['device_id']), "status": status, "event_name": evt, "details": details, "raw_log": item['log'], "json_data": item['json_data'], "source": source})
             except: pass
        specific_event_results = res
    socketio.emit('update_specific_event_table', specific_event_results)

def cache_specific_event_log(event_name, params, json_string, log_entry, device_id):
    if is_paused: return
    with lock: event_log_cache.append({'log': log_entry, 'device_id': device_id, 'json_data': json_string})
    _apply_specific_filter_and_emit()

def _apply_adrevenue_filter_and_emit():
    rendered = []
    with lock:
        for item in adrevenue_logs:
            parsed_data = item.get("parsed_data") or {}
            source = (item.get("source") or "appmetrica").lower()

            if source == "appsflyer":
                ad_network = parsed_data.get("ad_network") if isinstance(parsed_data.get("ad_network"), dict) else {}
                payload_data = ad_network.get("payload") if isinstance(ad_network.get("payload"), dict) else {}
                custom_params = payload_data.get("custom_parameters") if isinstance(payload_data.get("custom_parameters"), dict) else {}
                validate_maps = [ad_network, payload_data, custom_params]
                details_target = ad_network if ad_network else parsed_data
            else:
                payload_data = parsed_data.get("payload") if isinstance(parsed_data.get("payload"), dict) else {}
                validate_maps = [parsed_data, payload_data]
                details_target = parsed_data if parsed_data else item.get("raw_details", "")

            required_all = list(adrevenue_default_params)
            normalized_source = _normalize_adrevenue_sheet_key(source)
            for alias in (
                normalized_source,
                _normalize_adrevenue_sheet_key(item.get("event_name", "")),
                "all",
            ):
                required_all.extend(adrevenue_source_params.get(alias, []))
            seen_required = []
            seen_set = set()
            for param in required_all:
                if param not in seen_set:
                    seen_set.add(param)
                    seen_required.append(param)

            missing = []
            for param in seen_required:
                if any(isinstance(m, dict) and param in m for m in validate_maps):
                    continue
                missing.append(param)

            actual_keys = set()
            for candidate in validate_maps:
                if isinstance(candidate, dict):
                    actual_keys.update(candidate.keys())
            strange = sorted(actual_keys - set(seen_required)) if seen_required else []

            status = "INFO"
            if seen_required:
                status = "PASSED" if not missing else "FAILED"

            summary_parts = []
            if seen_required:
                if missing:
                    summary_parts.append(format_param_issue_html("Missing", missing, "text-red-600", chunk_size=1))
                else:
                    summary_parts.append("<div class='mb-2 text-xs font-medium text-green-600'>All required params found</div>")
                if strange:
                    summary_parts.append(format_param_issue_html("Strange", strange, "text-orange-600", chunk_size=1))

            details_html = ''.join(summary_parts) + format_json_html(details_target)
            rendered.append({
                **item,
                "status": status,
                "details": details_html,
            })

    socketio.emit('update_adrevenue_table', rendered)

def _emit_sdk_check_results():
    res = []
    with lock:
        if active_platform == "ios":
            res.append({"status": "HEADER", "display_text": "--- iOS ---", "device_name": "iOS", "device_id": "ios"})
            res.append({"status": "WAITING", "display_text": "iOS SDK table ready. Device connection will be added in the next step.", "device_id": "ios"})
            socketio.emit('update_sdk_check_table', res)
            return

        def append_sdk_network_rows(device_id, network_key, block):
            if _normalize_sdk_network_name(block.get("display_name")) == "googleplayservices":
                return
            expected_key = block.get("expected_key") or _match_sdk_expected_key(block.get("display_name"))
            expected = sdk_check_expected_map.get(expected_key, {})
            res.append({"status": "LABEL", "display_text": block.get("display_name", network_key), "device_id": device_id})

            actual_sdk = block.get("sdk_version", "")
            expected_sdk = expected.get("sdk", "")
            actual_adapter = block.get("adapter_version", "")
            if not actual_adapter and block.get("adapter_missing"):
                actual_adapter = "MISSING"
            expected_adapter = expected.get("adapter", "")

            if bool(expected_sdk) ^ bool(expected_adapter):
                if expected_sdk:
                    actual_single = actual_sdk
                    expected_single = expected_sdk
                else:
                    actual_single = actual_adapter
                    expected_single = expected_adapter
                single_status = _sdk_result_status(actual_single, expected_single)
                actual_single_display = actual_single
                if actual_single_display and actual_single_display.upper() not in {"NOT FOUND", "MISSING"}:
                    actual_single_display = _extract_sdk_comparable_version(actual_single_display, expected_single)
                res.append({
                    "status": single_status,
                    "display_text": f"Version  Actual: {actual_single_display or 'NOT FOUND'}  Expected: {expected_single}",
                    "device_id": device_id
                })
            else:
                sdk_status = _sdk_result_status(actual_sdk, expected_sdk)
                actual_sdk_display = actual_sdk
                if actual_sdk_display and actual_sdk_display.upper() not in {"NOT FOUND", "MISSING"}:
                    actual_sdk_display = _extract_sdk_comparable_version(actual_sdk_display, expected_sdk)
                res.append({
                    "status": sdk_status,
                    "display_text": f"SDK  Actual: {actual_sdk_display or 'NOT FOUND'}  Expected: {expected_sdk}",
                    "device_id": device_id
                })

                adapter_status = _sdk_result_status(actual_adapter, expected_adapter)
                actual_adapter_display = actual_adapter
                if actual_adapter_display and actual_adapter_display.upper() not in {"NOT FOUND", "MISSING"}:
                    actual_adapter_display = _extract_sdk_comparable_version(actual_adapter_display, expected_adapter)
                res.append({
                    "status": adapter_status,
                    "display_text": f"Adapter  Actual: {actual_adapter_display or 'NOT FOUND'}  Expected: {expected_adapter}",
                    "device_id": device_id
                })

        for dev in connected_devices_info:
            res.append({"status": "HEADER", "display_text": f"--- {dev['name']} ---", "device_name": dev['name'], "device_id": dev['id']})
            device_state = sdk_check_runtime_state.get(dev['id'], {})
            if not device_state and not sdk_check_expected_order:
                res.append({"status": "WAITING", "display_text": "Waiting for IntegrationHelper logs...", "device_id": dev['id']})
                continue

            _ensure_sdk_expected_blocks_for_device(dev['id'])
            device_state = sdk_check_runtime_state.get(dev['id'], {})
            in_list_keys = [key for key in _sdk_state_network_keys(dev['id']) if key in sdk_check_expected_map]
            not_in_list_keys = [key for key in _sdk_state_network_keys(dev['id']) if key not in sdk_check_expected_map]

            for network_key in in_list_keys:
                append_sdk_network_rows(dev['id'], network_key, device_state.get(network_key, {}))

            if not_in_list_keys:
                res.append({"status": "SECTION", "display_text": "----------- Not in List -------", "device_id": dev['id']})
                for network_key in not_in_list_keys:
                    append_sdk_network_rows(dev['id'], network_key, device_state.get(network_key, {}))
    socketio.emit('update_sdk_check_table', res)


def _sdk_check_block_for_device(device_id, network_name):
    device_state = sdk_check_runtime_state.setdefault(device_id, {})
    raw_network_key = _normalize_sdk_network_name(network_name)
    expected_key = _match_sdk_expected_key(network_name) or raw_network_key
    storage_key = expected_key or raw_network_key
    block = device_state.get(storage_key)
    if not block:
        expected = sdk_check_expected_map.get(expected_key, {})
        block = {
            "display_name": network_name.strip() or expected.get("display_name", network_name.strip()),
            "sdk_version": "",
            "adapter_version": "",
            "adapter_missing": False,
            "verification": "",
            "expected_key": expected_key,
            "updated_at": time.time(),
        }
        if expected.get("display_name"):
            block["display_name"] = expected.get("display_name")
        device_state[storage_key] = block
    else:
        expected_key = block.get("expected_key") or _match_sdk_expected_key(network_name) or raw_network_key
        expected = sdk_check_expected_map.get(expected_key, {})
        block["display_name"] = expected.get("display_name") or network_name.strip() or block["display_name"] or expected.get("display_name", network_name.strip())
        block["expected_key"] = expected_key
        block["updated_at"] = time.time()
    return block

# --- THREADS & PROCESSES ---

def adb_log_reader(device_id):
    print(f"INFO: Starting log reader for {device_id}")
    try:
        subprocess.run([ADB_EXECUTABLE, '-s', device_id, 'logcat', '-c'], creationflags=creation_flags)
        proc = subprocess.Popen([ADB_EXECUTABLE, '-s', device_id, 'logcat'], stdout=subprocess.PIPE, text=True, encoding='utf-8', errors='ignore', creationflags=creation_flags)
        
        for line in iter(proc.stdout.readline, ''):
            if not line: break
            
            # 1. Process Load Ads (Unity) - ONLY IF RECORDING
            process_load_ads_unity_log(line, device_id)
            
            # 2. Process Load Ads Ext (Metrica) - ONLY IF RECORDING
            process_load_ads_ext_log(line, device_id)
            
            # 3. Process SDK Check
            if not is_paused and sdk_check_active:
                changed = False
                with lock:
                    if "IntegrationHelper" in line:
                        current_network = sdk_check_current_network.get(device_id, "")
                        header_match = SDK_HEADER_PATTERN.search(line)
                        if header_match:
                            current_network = header_match.group(1).strip()
                            _sdk_check_block_for_device(device_id, current_network)
                            sdk_check_current_network[device_id] = current_network
                            changed = True
                        else:
                            sdk_match = SDK_VERSION_LINE_PATTERN.search(line)
                            adapter_match = SDK_ADAPTER_VERSION_LINE_PATTERN.search(line)
                            adapter_missing_match = SDK_ADAPTER_MISSING_PATTERN.search(line)
                            verification_match = SDK_VERIFICATION_PATTERN.search(line)

                            if verification_match:
                                current_network = verification_match.group(1).strip()
                                block = _sdk_check_block_for_device(device_id, current_network)
                                status = verification_match.group(2).strip().upper()
                                if block.get("verification") != status:
                                    block["verification"] = status
                                    changed = True
                                sdk_check_current_network[device_id] = current_network

                            target_network = current_network
                            if target_network:
                                block = _sdk_check_block_for_device(device_id, target_network)
                                if sdk_match:
                                    sdk_version = sdk_match.group(1).strip()
                                    if block.get("sdk_version") != sdk_version:
                                        block["sdk_version"] = sdk_version
                                        changed = True
                                if adapter_match:
                                    adapter_version = adapter_match.group(1).strip()
                                    if block.get("adapter_version") != adapter_version or block.get("adapter_missing"):
                                        block["adapter_version"] = adapter_version
                                        block["adapter_missing"] = False
                                        changed = True
                                elif adapter_missing_match:
                                    if not block.get("adapter_missing") or block.get("adapter_version"):
                                        block["adapter_missing"] = True
                                        block["adapter_version"] = ""
                                        changed = True

                    if _process_sdk_external_line(line, device_id):
                        changed = True

                if changed:
                    _emit_sdk_check_results()

            # 4. Process AdRevenue
            handled_adrevenue = False
            if "AdRevenue Received:" in line:
                with lock:
                    match = ADREVENUE_LOG_PATTERN.search(line)
                    if match:
                        content = match.group(1)
                        ad_data = {}

                        try:
                            payload_obj = {}
                            payload_match = re.search(r'payload=(\{.*?\})(?:,|$)', content)

                            clean_content = content
                            if payload_match:
                                payload_str = payload_match.group(1)
                                try:
                                    payload_obj = json.loads(payload_str)
                                except:
                                    payload_obj = {}
                                clean_content = content.replace(f'payload={payload_str}', '')

                            parts = [p.strip() for p in clean_content.split(',') if p.strip()]
                            for part in parts:
                                if '=' in part:
                                    k, v = part.split('=', 1)
                                    ad_data[k.strip()] = v.strip()

                            if payload_obj:
                                ad_data['payload'] = payload_obj
                        except:
                            ad_data = {}

                        adrevenue_logs.append({
                            "device_id": device_id,
                            "device_name": get_device_name(device_id),
                            "status": "INFO",
                            "event_name": "AdRevenue - Appmetrica",
                            "source": "appmetrica",
                            "details": format_json_html(ad_data) if ad_data else content,
                            "raw_details": content,
                            "raw_log": line.strip(),
                            "json_data": json.dumps(ad_data, ensure_ascii=False) if ad_data else "{}",
                            "parsed_data": ad_data,
                        })
                        handled_adrevenue = True
                _apply_adrevenue_filter_and_emit()

            if (not handled_adrevenue) and "AppsFlyer" in line and "ADREVENUE-" in line and "preparing data:" in line:
                with lock:
                    match = APPSFLYER_ADREVENUE_PATTERN.search(line)
                    if match:
                        event_prefix = match.group(1).upper()
                        json_str = match.group(2)
                        appsflyer_data = {}
                        try:
                            appsflyer_data = json.loads(json_str)
                        except:
                            appsflyer_data = {}

                        ad_network_data = appsflyer_data.get("ad_network") if isinstance(appsflyer_data.get("ad_network"), dict) else {}
                        adrevenue_logs.append({
                            "device_id": device_id,
                            "device_name": get_device_name(device_id),
                            "status": "INFO",
                            "event_name": "AdRevenue - Appsflyer",
                            "source": "appsflyer",
                            "details": format_json_html(ad_network_data) if ad_network_data else format_json_html(appsflyer_data),
                            "raw_details": json_str,
                            "raw_log": line.strip(),
                            "json_data": json.dumps(appsflyer_data, ensure_ascii=False) if appsflyer_data else "{}",
                            "parsed_data": appsflyer_data,
                            "raw_event_prefix": event_prefix,
                        })
                        handled_adrevenue = True
                if handled_adrevenue:
                    _apply_adrevenue_filter_and_emit()

            # 5. Process Callback & Events
            process_callback_and_ad_event_log(line, device_id)
            
            # 6. Parse Generic Events for Validators
            event_name, params, json_string = find_and_parse_event(line)
            if event_name:
                process_event_validator_log(event_name, params, json_string, line, device_id)
                cache_specific_event_log(event_name, params, json_string, line, device_id)
                # Call callback processor for "ad_" events
                process_callback_and_ad_event_log(line, device_id, event_name, params, json_string)

    except Exception as e: print(f"Error {device_id}: {e}")

def device_manager():
    global connected_devices_info
    while True:
        try:
            output = subprocess.run([ADB_EXECUTABLE, 'devices'], capture_output=True, text=True, creationflags=creation_flags).stdout
            ids = {l.split('\t')[0] for l in output.strip().split('\n')[1:] if '\tdevice' in l}
            
            with lock:
                for did in ids - set(active_log_readers.keys()):
                    t = threading.Thread(target=adb_log_reader, args=(did,), daemon=True)
                    active_log_readers[did] = t
                    t.start()
                
                for did in set(active_log_readers.keys()) - ids:
                    del active_log_readers[did]
                
                connected_devices_info = [{'id': i, 'name': get_device_name(i)} for i in ids]
                if connected_devices_info:
                    socketio.emit('device_status', {"connected_devices": connected_devices_info})
                else:
                    socketio.emit('device_status', {
                        "connected_devices": [],
                        "message": f"Waiting... (ADB: {ADB_EXECUTABLE})"
                    })
        except Exception as e:
            socketio.emit('device_status', {
                "connected_devices": [],
                "message": f"ADB error: {e}"
            })
        time.sleep(3)

def package_log_consumer(device_id, logcat_process):
    try:
        for line in iter(logcat_process.stdout.readline, ''):
            if not line: break
            if not is_paused:
                is_error = bool(re.search(r'^\S+\s+\S+\s+\d+\s+\d+\s+[EF]\s', line))
                time_str = ""
                time_display = ""
                level = ""
                tag = ""
                message = line.strip()
                # Try to parse standard logcat format: MM-DD HH:MM:SS.mmm PID TID LEVEL TAG: message
                m = re.match(r'^(\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}\.\d+)\s+\d+\s+\d+\s+([A-Z])\s+([^:]+):\s*(.*)$', line.strip())
                if m:
                    time_str = m.group(1)
                    level = m.group(2)
                    tag = m.group(3).strip()
                    message = m.group(4).strip()
                    # Prefer shorter time display to save width: HH:MM:SS.mmm
                    if ' ' in time_str:
                        time_display = time_str.split(' ', 1)[1]
                with lock:
                    package_log_cache.append({
                        'device_id': device_id,
                        'device_name': get_device_name(device_id),
                        'log': line.strip(),
                        'time': time_str,
                        'time_display': time_display or time_str,
                        'level': level,
                        'tag': tag,
                        'message': message,
                        'timestamp': time.time(),
                        'is_error': is_error
                    })
                    session_id = active_package_log_session_id
                if session_id:
                    try:
                        package_log_db_queue.put_nowait((
                            session_id,
                            time.time(),
                            time_display or time_str,
                            device_id,
                            get_device_name(device_id),
                            level,
                            tag,
                            message,
                            line.strip(),
                            1 if is_error else 0,
                        ))
                    except Exception:
                        pass
    except: pass

def package_pid_monitor():
    global active_package_pids
    while True:
        time.sleep(3)

        with lock:
            pkg = target_package_name
            devices_snapshot = list(connected_devices_info)

        if not pkg:
            with lock:
                for p in active_logcat_processes.values():
                    try:
                        p.terminate()
                    except Exception:
                        pass
                active_logcat_processes.clear()
                active_package_pids.clear()
            continue

        current_ids = {d['id'] for d in devices_snapshot}

        # Clean up disconnected devices
        with lock:
            for did in list(active_logcat_processes.keys()):
                if did not in current_ids:
                    try:
                        active_logcat_processes[did].terminate()
                    except Exception:
                        pass
                    active_logcat_processes.pop(did, None)
                    active_package_pids.pop(did, None)

        for device in devices_snapshot:
            did = device['id']
            try:
                res = subprocess.run([ADB_EXECUTABLE, '-s', did, 'shell', 'pidof', '-s', pkg],
                                     capture_output=True, text=True, creationflags=creation_flags)
                pid = res.stdout.strip()

                if not pid:
                    with lock:
                        if did in active_logcat_processes:
                            try:
                                active_logcat_processes[did].terminate()
                            except Exception:
                                pass
                            active_logcat_processes.pop(did, None)
                        active_package_pids.pop(did, None)
                    continue

                with lock:
                    proc = active_logcat_processes.get(did)
                    prev_pid = active_package_pids.get(did)

                restart = (pid != prev_pid) or (proc is None) or (proc.poll() is not None)

                if restart:
                    with lock:
                        if did in active_logcat_processes:
                            try:
                                active_logcat_processes[did].terminate()
                            except Exception:
                                pass
                        cmd = [ADB_EXECUTABLE, '-s', did, 'logcat', f'--pid={pid}']
                        proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, text=True, encoding='utf-8', errors='ignore', creationflags=creation_flags)
                        active_logcat_processes[did] = proc
                        active_package_pids[did] = pid
                        threading.Thread(target=package_log_consumer, args=(did, proc), daemon=True).start()
            except Exception:
                with lock:
                    if did in active_logcat_processes:
                        try:
                            active_logcat_processes[did].terminate()
                        except Exception:
                            pass
                        active_logcat_processes.pop(did, None)
                    active_package_pids.pop(did, None)


def package_log_emitter():
    while True:
        time.sleep(1)
        with lock:
            if not is_paused and target_package_name:
                now = time.time()
                while package_log_cache and now - package_log_cache[0]['timestamp'] > 1000: package_log_cache.popleft()
                ui_rows = list(package_log_cache)[-PACKAGE_LOG_UI_MAX_ROWS:]
                socketio.emit('package_log_cache', ui_rows)

# --- SOCKET HANDLERS ---
@socketio.on('change_tab')
def handle_change_tab(data): 
    # Sync logs for current tab on switch
    if data.get('tab_name') == 'LoadAds': socketio.emit('update_load_ads', list(load_ads_events))
    if data.get('tab_name') == 'LoadAdsExt': socketio.emit('update_load_ads_ext', list(load_ads_ext_events))
    if data.get('tab_name') == 'SdkCheck': _emit_sdk_check_results()
    if data.get('tab_name') == 'CallbackAd': socketio.emit('update_callback_ad_table', list(callback_ad_logs))

def _reset_runtime_for_platform_switch():
    global is_paused, validator_active, sdk_check_active, sdk_check_current_network
    global target_package_name, active_package_log_session_id
    global specific_event_name_filters, specific_event_params_filters
    with lock:
        is_paused = False
        validator_active = False
        sdk_check_active = False
        sdk_check_search_list.clear()
        sdk_check_results.clear()
        sdk_check_input_list.clear()
        sdk_check_expected_map.clear()
        sdk_check_runtime_state.clear()
        sdk_check_current_network = {}
        sdk_check_expected_order.clear()

        load_ads_events.clear(); unique_load_ads.clear()
        load_ads_ext_events.clear(); unique_load_ads_ext.clear()
        validator_results.clear()
        specific_event_name_filters = []
        specific_event_params_filters = []
        specific_event_results.clear(); event_log_cache.clear()
        adrevenue_logs.clear(); adrevenue_log_cache.clear()
        callback_ad_logs.clear(); incomplete_impression_logs.clear()
        package_log_cache.clear(); active_package_pids.clear()
        target_package_name = ""
        if active_package_log_session_id:
            _finish_package_log_session(active_package_log_session_id)
            active_package_log_session_id = None

        for state in recording_states.values():
            state["is_recording"] = False

    socketio.emit('pause_status', {'is_paused': False})
    socketio.emit('validator_status', {'active': False})
    for tab, state in recording_states.items():
        socketio.emit('record_status', {
            "tab_name": tab,
            "is_recording": False,
            "current_sheet": state.get("current_sheet", "")
        })
    socketio.emit('update_load_ads', [])
    socketio.emit('update_load_ads_ext', [])
    socketio.emit('update_validator_table', [])
    socketio.emit('update_specific_event_table', [])
    socketio.emit('update_adrevenue_table', [])
    socketio.emit('update_callback_ad_table', [])
    socketio.emit('package_log_cache', [])
    socketio.emit('runtime_reset', {})
    _emit_sdk_check_results()

@socketio.on('set_platform')
def set_platform(data):
    global active_platform
    platform = (data or {}).get('platform', 'android')
    active_platform = 'ios' if platform == 'ios' else 'android'
    if (data or {}).get('reset'):
        _reset_runtime_for_platform_switch()
    socketio.emit('platform_status', {'platform': active_platform})

@socketio.on('toggle_record')
def tr(data):
    tab_name = data.get('tab_name')
    if not tab_name or tab_name not in recording_states:
        return

    current_state = recording_states[tab_name]
    
    if not current_state["is_recording"]:
        name = data.get('sheet_name', 'Log_Default').strip()
        try:
            res = requests.post(G_SHEET_URL, json={"action": "create_or_get_sheet", "sheet_name": name}, timeout=30)
            current_state.update({"is_recording": True, "current_sheet": res.text})
        except: 
            current_state.update({"is_recording": True, "current_sheet": name}) # Fallback
    else:
        current_state["is_recording"] = False
    
    # Emit status back with tab_name so UI knows which button to update
    socketio.emit('record_status', {
        "tab_name": tab_name,
        "is_recording": current_state["is_recording"],
        "current_sheet": current_state["current_sheet"]
    })

@socketio.on('toggle_pause')
def tp(): 
    global is_paused
    is_paused = not is_paused
    socketio.emit('pause_status', {'is_paused': is_paused})

@socketio.on('clear_all_logs')
def cl():
    with lock:
        load_ads_events.clear(); unique_load_ads.clear()
        load_ads_ext_events.clear(); unique_load_ads_ext.clear()
        validator_results.clear()
        specific_event_results.clear(); event_log_cache.clear()
        adrevenue_logs.clear(); callback_ad_logs.clear()
        package_log_cache.clear()
        # Clean SDK check
        sdk_check_results.clear()
        incomplete_impression_logs.clear()
        
    socketio.emit('update_load_ads', [])
    socketio.emit('update_load_ads_ext', [])
    socketio.emit('update_validator_table', [])
    socketio.emit('update_specific_event_table', [])
    socketio.emit('update_adrevenue_table', [])
    socketio.emit('update_callback_ad_table', [])
    socketio.emit('package_log_cache', [])
    _emit_sdk_check_results()

@socketio.on('start_validation')
def val(p): 
    global required_params, validator_active
    required_params = p or []
    validator_active = True
    validator_results.clear()
    socketio.emit('update_validator_table', [])
    socketio.emit('validator_status', {'active': True})

@socketio.on('stop_validation')
def stop_val():
    global validator_active
    validator_active = False
    socketio.emit('validator_status', {'active': False})

@socketio.on('update_specific_filter')
def usf(d):
    global specific_event_name_filters, specific_event_params_filters
    with lock:
        specific_event_name_filters = d.get('eventNames', [])
        specific_event_params_filters = d.get('params', [])
    _apply_specific_filter_and_emit()

@socketio.on('update_adrevenue_filter')
def uaf(_d=None):
    _apply_adrevenue_filter_and_emit()

@socketio.on('start_sdk_check')
def sdk_check(data):
    global sdk_check_search_list, sdk_check_results, sdk_check_input_list, sdk_check_active, sdk_check_expected_map, sdk_check_runtime_state, sdk_check_current_network, sdk_check_expected_order
    with lock:
        sdk_check_search_list = []
        sdk_check_results = {}
        sdk_check_input_list = []
        sdk_check_expected_map = {}
        sdk_check_runtime_state = {}
        sdk_check_current_network = {}
        sdk_check_expected_order = []
        lines = [line.rstrip('\r') for line in data.get('text', '').splitlines() if line.strip()]
        if lines:
            header = [col.strip() for col in lines[0].split('\t')]
            if len(header) >= 3 and header[0].lower().startswith('ads network'):
                for line in lines[1:]:
                    cols = [col.strip() for col in line.split('\t')]
                    if not cols or not cols[0]:
                        continue
                    network = cols[0]
                    adapter = cols[1] if len(cols) > 1 else ""
                    sdk = cols[2] if len(cols) > 2 else ""
                    log_search = cols[3] if len(cols) > 3 else ""
                    expected_key = _normalize_sdk_network_name(network)
                    sdk_check_expected_map[expected_key] = {
                        "display_name": network,
                        "adapter": adapter,
                        "sdk": sdk,
                        "log_search": log_search,
                    }
                    sdk_check_expected_order.append(expected_key)
            else:
                for line in lines:
                    match = SDK_CHECK_SEARCH_PATTERN.search(line)
                    if match:
                        pat = match.group(1)
                        disp = line[:match.start()].rstrip(', ').strip().strip('"') or pat
                        item = {
                            "type": "search",
                            "display_name": disp,
                            "search_pattern": pat,
                            "search_pattern_normalized": _normalize_sdk_search_text(pat),
                        }
                        sdk_check_search_list.append(item)
                        sdk_check_input_list.append(item)
                    else:
                        sdk_check_input_list.append({"type": "label", "display_name": line})
        sdk_check_active = True
    _emit_sdk_check_results()


@socketio.on('stop_sdk_check')
def stop_sdk_check():
    global sdk_check_active, sdk_check_current_network
    with lock:
        sdk_check_active = False
        sdk_check_current_network = {}
    _emit_sdk_check_results()

@socketio.on('start_package_log')
def spl(d):
    global target_package_name, active_package_log_session_id
    pid = d.get('package_id', '').strip()
    with lock:
        if active_package_log_session_id:
            _finish_package_log_session(active_package_log_session_id)
            active_package_log_session_id = None
        target_package_name = pid
        if pid:
            active_package_log_session_id = _start_package_log_session(pid)
        package_log_cache.clear()
        socketio.emit('package_log_cache', [])
    socketio.emit('package_log_cache', [])

@socketio.on('refresh_request')
def refresh():
    socketio.emit('update_load_ads', list(load_ads_events))
    socketio.emit('update_load_ads_ext', list(load_ads_ext_events))
    socketio.emit('update_callback_ad_table', list(callback_ad_logs)) # Ensure callback data is refreshed
    # ... trigger others ...

@socketio.on('connect')
def connect(): 
    socketio.emit('pause_status', {'is_paused': is_paused})
    socketio.emit('validator_status', {'active': validator_active})
    socketio.emit('platform_status', {'platform': active_platform})
    # Sync recording buttons on connect
    for tab, state in recording_states.items():
        socketio.emit('record_status', {
            "tab_name": tab,
            "is_recording": state["is_recording"],
            "current_sheet": state["current_sheet"]
        })

# --- MAIN ---
def run_server(host="0.0.0.0", port=5001):
    _normalize_remote_update_config()
    _set_active_profile()
    _init_package_log_db()
    threading.Thread(target=device_manager, daemon=True).start()
    threading.Thread(target=package_log_emitter, daemon=True).start()
    threading.Thread(target=package_pid_monitor, daemon=True).start()
    threading.Thread(target=_package_log_db_writer, daemon=True).start()

    def safe_print(msg):
        try:
            print(msg)
        except UnicodeEncodeError:
            try:
                print(msg.encode("utf-8", "ignore").decode("utf-8"))
            except Exception:
                logging.info(msg)

    safe_print("===================================================")
    safe_print("  Event Inspector Started")
    safe_print(f"  - Access: http://{host}:{port}")
    safe_print("===================================================")
    socketio.run(
        app,
        host=host,
        port=port,
        use_reloader=False,
        allow_unsafe_werkzeug=True
    )

if __name__ == '__main__':
    run_server(host="0.0.0.0", port=5001)
